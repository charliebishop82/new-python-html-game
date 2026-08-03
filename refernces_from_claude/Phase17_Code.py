# Phase 17: contextual UI help and comprehensive in-code documentation

# FILE: admin.py
"""Local admin application for operations, diagnostics, balance, and support."""
# admin.py  (Phase 8)
# Separate Flask app for admin tools.
# Run with: flask --app admin:create_admin_app run --port 5001
# Localhost only — never expose publicly.

import math
import json
import logging
import os
import re
import sqlite3
import uuid
from datetime import datetime

from flask import (Flask, render_template, request, redirect,
                   url_for, g, jsonify, Response)
from werkzeug.security import generate_password_hash

import config_defaults as cfg
from database import (execute, execute_one, execute_write,
                      exclusive_transaction, init_db, close_db,
                      get_all_settings)

logger = logging.getLogger(__name__)
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")


# ─────────────────────────────────────────────────────────────────────────────
# APP FACTORY
# ─────────────────────────────────────────────────────────────────────────────

def create_admin_app() -> Flask:
    """Handle the create admin app workflow."""
    init_db()
    app = Flask(__name__, template_folder="templates")
    app.secret_key = cfg.SECRET_KEY + "-admin"
    app.teardown_appcontext(close_db)

    @app.before_request
    def check_admin_auth():
        """Handle the check admin auth workflow."""
        if not ADMIN_PASSWORD:
            return None
        auth = request.authorization
        if not auth or auth.password != ADMIN_PASSWORD:
            return Response(
                "Admin access required.", 401,
                {"WWW-Authenticate": 'Basic realm="Admin"'}
            )
        return None

    _register_routes(app)
    return app


def _register_routes(app: Flask):
    """Provide the internal register routes operation used by this module."""
    app.add_url_rule("/", "admin_root", lambda: redirect(url_for("admin_index")))
    app.add_url_rule("/admin",                        "admin_index",        admin_index)
    app.add_url_rule("/admin/import",                 "admin_import",       admin_import,        methods=["GET","POST"])
    app.add_url_rule("/admin/players",                "admin_players",      admin_players)
    app.add_url_rule("/admin/players/<int:pid>",      "admin_player_detail",admin_player_detail)
    app.add_url_rule("/admin/players/<int:pid>/ban",  "admin_ban",          admin_ban,           methods=["POST"])
    app.add_url_rule("/admin/players/<int:pid>/retire", "admin_retire_player", admin_retire_player, methods=["POST"])
    app.add_url_rule("/admin/players/<int:pid>/edit", "admin_edit",         admin_edit,          methods=["POST"])
    app.add_url_rule("/admin/config",                 "admin_config",       admin_config,        methods=["GET","POST"])
    app.add_url_rule("/admin/reset/midnight",         "admin_midnight",     admin_midnight,      methods=["POST"])
    app.add_url_rule("/admin/reset/full",             "admin_full_reset",   admin_full_reset,    methods=["POST"])
    app.add_url_rule("/admin/logs",                   "admin_logs",         admin_logs)
    app.add_url_rule("/admin/players/<int:pid>/activity", "admin_player_activity", admin_player_activity)
    app.add_url_rule("/admin/health",                 "admin_health",       admin_health)
    app.add_url_rule("/admin/items",                  "admin_items",        admin_items)
    app.add_url_rule("/admin/items/<item_type>/<int:item_id>/edit", "admin_item_edit", admin_item_edit, methods=["POST"])
    app.add_url_rule("/admin/analytics",              "admin_analytics",    admin_analytics)
    app.add_url_rule("/admin/npcs",                   "admin_npcs",         admin_npcs,          methods=["GET","POST"])
    app.add_url_rule("/admin/npcs/<int:pid>/edit",    "admin_npc_edit",     admin_npc_edit,      methods=["POST"])
    app.add_url_rule("/admin/npcs/<int:pid>/run",     "admin_npc_run",      admin_npc_run,       methods=["POST"])
    app.add_url_rule("/admin/npcs/<int:pid>/spend-ap", "admin_npc_spend_ap", admin_npc_spend_ap, methods=["POST"])
    app.add_url_rule("/admin/npcs/<int:pid>/retire",  "admin_npc_retire",   admin_npc_retire,    methods=["POST"])
    app.add_url_rule("/admin/npcs/<int:pid>/inventory/grant", "admin_npc_grant", admin_npc_grant, methods=["POST"])
    app.add_url_rule("/admin/npcs/<int:pid>/inventory/<int:inv_id>/remove", "admin_npc_remove", admin_npc_remove, methods=["POST"])


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def admin_index():
    """Render or process the index administrative workflow."""
    import os
    player_count  = execute_one("SELECT COUNT(*) as cnt FROM players WHERE is_banned = 0")["cnt"]
    active_combat = execute_one("SELECT COUNT(*) as cnt FROM combat_sessions WHERE status='ACTIVE'")["cnt"]
    pending_import = os.path.exists(cfg.PENDING_IMPORT_PATH)
    queue_failed   = execute_one("SELECT COUNT(*) as cnt FROM action_queue WHERE status='FAILED'")["cnt"]
    boss_count     = execute_one("SELECT COUNT(*) as cnt FROM bosses WHERE is_active=1")["cnt"]
    special_pool   = execute_one(
        "SELECT COUNT(*) as cnt FROM special_item_registry WHERE status='IN_POOL'"
    )["cnt"]

    recent_errors = []
    if os.path.exists(cfg.IMPORT_ERROR_LOG):
        with open(cfg.IMPORT_ERROR_LOG) as f:
            recent_errors = f.readlines()[-10:]

    return render_template("admin/dashboard.html",
        player_count=player_count,
        active_combat=active_combat,
        pending_import=pending_import,
        queue_failed=queue_failed,
        boss_count=boss_count,
        special_pool=special_pool,
        recent_errors=recent_errors,
        now=datetime.utcnow().isoformat()
    )


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT
# ─────────────────────────────────────────────────────────────────────────────

def admin_import():
    """Render or process the import administrative workflow."""
    import os, shutil
    feedback = error = None

    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename.endswith(".xlsx"):
            error = "Please upload a valid .xlsx file."
        else:
            os.makedirs(os.path.dirname(cfg.PENDING_IMPORT_PATH), exist_ok=True)
            file.save(cfg.PENDING_IMPORT_PATH)
            feedback = f"File staged at {cfg.PENDING_IMPORT_PATH}. Will be applied at next midnight reset."

    pending = os.path.exists(cfg.PENDING_IMPORT_PATH)
    return render_template("admin/import.html",
                           pending=pending, feedback=feedback, error=error)


# ─────────────────────────────────────────────────────────────────────────────
# PLAYERS
# ─────────────────────────────────────────────────────────────────────────────

def admin_players():
    """Render or process the players administrative workflow."""
    players = execute(
        """SELECT p.*, ps.pvp_kills, ps.times_reduced_to_1hp,
                  CASE WHEN np.player_id IS NULL THEN 0 ELSE 1 END AS is_npc
           FROM players p
           LEFT JOIN player_stats ps ON ps.player_id = p.id
           LEFT JOIN npc_profiles np ON np.player_id = p.id
           ORDER BY p.level DESC, p.xp DESC"""
    )
    return render_template("admin/players.html", players=players)


def _audit(action: str, target_type: str, target_id=None, reason=None, details=None):
    """Provide the internal audit operation used by this module."""
    execute_write(
        """INSERT INTO admin_audit_log(action,target_type,target_id,reason,details_json)
           VALUES(?,?,?,?,?)""",
        (action, target_type, target_id, reason, json.dumps(details or {}, default=str)[:8000])
    )


def admin_player_activity(pid: int):
    """Render or process the player activity administrative workflow."""
    player = execute_one("SELECT * FROM players WHERE id=?", (pid,))
    if not player:
        return redirect(url_for("admin_players"))
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    category = request.args.get("category", "").strip().upper()
    errors_only = request.args.get("errors_only") == "1"
    page = max(1, request.args.get("page", type=int, default=1))
    where, params = ["player_id=?"], [pid]
    if start: where.append("occurred_at>=?"); params.append(start)
    if end: where.append("occurred_at<?"); params.append(end + " 23:59:59")
    if category: where.append("category=?"); params.append(category)
    if errors_only: where.append("status='FAILED'")
    clause = " AND ".join(where)
    total = execute_one(f"SELECT COUNT(*) cnt FROM player_activity_log WHERE {clause}", tuple(params))["cnt"]
    rows = execute(
        f"SELECT * FROM player_activity_log WHERE {clause} ORDER BY id DESC LIMIT 100 OFFSET ?",
        (*params, (page - 1) * 100)
    )
    categories = execute("SELECT DISTINCT category FROM player_activity_log WHERE player_id=? ORDER BY category", (pid,))
    return render_template("admin/player_activity.html", player=player, rows=rows,
                           categories=categories, page=page, total=total,
                           start=start, end=end, category=category, errors_only=errors_only)


def admin_health():
    """Render or process the health administrative workflow."""
    stats = {
        "failed_actions": execute_one("SELECT COUNT(*) cnt FROM action_queue WHERE status='FAILED'")["cnt"],
        "processing_actions": execute_one("SELECT COUNT(*) cnt FROM action_queue WHERE status='PROCESSING'")["cnt"],
        "active_combats": execute_one("SELECT COUNT(*) cnt FROM combat_sessions WHERE status='ACTIVE'")["cnt"],
        "stuck_flags": execute_one("""SELECT COUNT(*) cnt FROM players p WHERE p.in_combat=1 AND NOT EXISTS
                                      (SELECT 1 FROM combat_sessions c WHERE c.status='ACTIVE' AND
                                       (c.attacker_player_id=p.id OR c.defender_player_id=p.id))""")["cnt"],
        "orphan_equipment": execute_one("""SELECT COUNT(*) cnt FROM players p WHERE
            (p.equipped_weapon_id IS NOT NULL AND NOT EXISTS(SELECT 1 FROM inventory_items i WHERE i.id=p.equipped_weapon_id AND i.player_id=p.id)) OR
            (p.equipped_armor_id IS NOT NULL AND NOT EXISTS(SELECT 1 FROM inventory_items i WHERE i.id=p.equipped_armor_id AND i.player_id=p.id)) OR
            (p.equipped_special_id IS NOT NULL AND NOT EXISTS(SELECT 1 FROM inventory_items i WHERE i.id=p.equipped_special_id AND i.player_id=p.id))""")["cnt"],
        "special_mismatches": execute_one("""SELECT COUNT(*) cnt FROM special_item_registry r
            WHERE (r.status='IN_INVENTORY' AND (r.current_owner_player_id IS NULL OR r.inventory_item_id IS NULL))
               OR (r.status='IN_POOL' AND (r.current_owner_player_id IS NOT NULL OR r.inventory_item_id IS NOT NULL))""")["cnt"],
    }
    scheduler_runs = execute("SELECT * FROM scheduler_run_log ORDER BY id DESC LIMIT 30")
    failures = execute("""SELECT q.*,p.character_name FROM action_queue q JOIN players p ON p.id=q.player_id
                          WHERE q.status='FAILED' ORDER BY q.id DESC LIMIT 30""")
    audits = execute("SELECT * FROM admin_audit_log ORDER BY id DESC LIMIT 30")
    return render_template("admin/health.html", stats=stats, scheduler_runs=scheduler_runs,
                           failures=failures, audits=audits)


ITEM_TABLES = {"weapon": "weapons", "armor": "armor", "special": "special_items"}
ITEM_EDIT_FIELDS = {
    "weapon": ("name","is_active","level","weapon_type","damage_die","damage_type","str_bonus","end_bonus","agi_bonus","lck_bonus","per_bonus","credit_cost","drop_chance","starting_durability"),
    "armor": ("name","is_active","level","ac_bonus","res_blade","res_blunt","res_ballistic","res_energy","res_arcane","res_explosive","res_venom","str_bonus","end_bonus","agi_bonus","lck_bonus","per_bonus","credit_cost","drop_chance","starting_durability"),
    "special": ("name","is_active","associated_to","association_type","str_bonus","end_bonus","agi_bonus","lck_bonus","per_bonus","initiative_bonus","extra_attack","crit_chance_bonus","crit_dmg_multiplier","ac_bonus","credit_cost","drop_chance","starting_durability","steal_bonus","xp_multiplier","credit_multiplier","bonus_ap","hp_regen_bonus","durability_reduction","shop_discount","sell_bonus","encounter_bonus"),
}


def admin_items():
    """Render or process the items administrative workflow."""
    selected = request.args.get("type", "weapon").lower()
    if selected not in ITEM_TABLES: selected = "weapon"
    items = execute(f"SELECT * FROM {ITEM_TABLES[selected]} ORDER BY is_active DESC,name")
    registry = execute("""SELECT r.*,s.name,p.character_name FROM special_item_registry r
                          JOIN special_items s ON s.id=r.special_item_id
                          LEFT JOIN players p ON p.id=r.current_owner_player_id ORDER BY s.name""")
    return render_template("admin/items.html", items=items, selected=selected,
                           fields=ITEM_EDIT_FIELDS[selected], registry=registry)


def admin_item_edit(item_type: str, item_id: int):
    """Render or process the item edit administrative workflow."""
    item_type = item_type.lower()
    if item_type not in ITEM_TABLES:
        return redirect(url_for("admin_items", error="Unknown item type."))
    table, allowed = ITEM_TABLES[item_type], ITEM_EDIT_FIELDS[item_type]
    current = execute_one(f"SELECT * FROM {table} WHERE id=?", (item_id,))
    if not current:
        return redirect(url_for("admin_items", type=item_type, error="Item not found."))
    changes = {}
    for field in allowed:
        raw = request.form.get(field)
        if raw is None: continue
        old = current.get(field)
        try:
            value = int(raw) if isinstance(old, int) else float(raw) if isinstance(old, float) else raw.strip()
        except ValueError:
            return redirect(url_for("admin_items", type=item_type, error=f"Invalid value for {field}."))
        if value != old: changes[field] = value
    reason = request.form.get("reason", "").strip()
    if not reason:
        return redirect(url_for("admin_items", type=item_type, error="A balancing reason is required."))
    if changes:
        try:
            with exclusive_transaction():
                execute_write(f"UPDATE {table} SET " + ",".join(f"{f}=?" for f in changes) + " WHERE id=?",
                              (*changes.values(), item_id))
                _audit("EDIT_ITEM", item_type.upper(), item_id, reason,
                       {f: {"from": current.get(f), "to": v} for f, v in changes.items()})
        except sqlite3.IntegrityError as exc:
            return redirect(url_for("admin_items", type=item_type, error=str(exc)))
    return redirect(url_for("admin_items", type=item_type, feedback=f"Updated {current['name']}."))


def admin_analytics():
    """Render or process the analytics administrative workflow."""
    action_counts = execute("""SELECT action,status,COUNT(*) cnt FROM player_activity_log
                               GROUP BY action,status ORDER BY cnt DESC LIMIT 30""")
    economy = execute_one("""SELECT COUNT(*) players,COALESCE(SUM(credits),0) credits,
                              COALESCE(AVG(credits),0) avg_credits,COALESCE(AVG(level),0) avg_level
                              FROM players WHERE is_banned=0""")
    combats = execute("SELECT combat_type,result,COUNT(*) cnt FROM combat_sessions GROUP BY combat_type,result ORDER BY cnt DESC")
    item_events = execute("SELECT event_type,COUNT(*) cnt FROM item_history GROUP BY event_type ORDER BY cnt DESC LIMIT 25")
    npc_decisions = execute("SELECT decision,COUNT(*) cnt FROM npc_action_log GROUP BY decision ORDER BY cnt DESC")
    random_events = execute("""SELECT flavor_text,COUNT(*) cnt FROM daily_feed WHERE event_category='RANDOM_EVENT'
                               GROUP BY flavor_text ORDER BY cnt DESC LIMIT 20""")
    return render_template("admin/analytics.html", action_counts=action_counts, economy=economy,
                           combats=combats, item_events=item_events,
                           npc_decisions=npc_decisions, random_events=random_events)
def admin_player_detail(pid: int):
    """Render or process the player detail administrative workflow."""
    player    = execute_one("SELECT * FROM players WHERE id = ?", (pid,))
    if not player:
        return redirect(url_for("admin_players"))
    stats     = execute_one("SELECT * FROM player_stats WHERE player_id = ?", (pid,))
    inventory = execute(
        "SELECT ii.*, "
        "CASE ii.item_type WHEN 'WEAPON' THEN w.name WHEN 'ARMOR' THEN a.name ELSE si.name END as item_name "
        "FROM inventory_items ii "
        "LEFT JOIN weapons w ON w.id = ii.item_id AND ii.item_type='WEAPON' "
        "LEFT JOIN armor   a ON a.id = ii.item_id AND ii.item_type='ARMOR' "
        "LEFT JOIN special_items si ON si.id = ii.item_id AND ii.item_type='SPECIAL' "
        "WHERE ii.player_id = ?", (pid,)
    )
    history   = execute(
        "SELECT * FROM item_history WHERE player_id = ? ORDER BY occurred_at DESC LIMIT 50", (pid,)
    )
    boss_kills = execute(
        """SELECT b.name, bi.kill_count, bi.discovered_at
           FROM boss_instances bi JOIN bosses b ON b.id = bi.boss_id
           WHERE bi.player_id = ? ORDER BY bi.kill_count DESC""", (pid,)
    )
    return render_template("admin/player_detail.html",
                           player=player, stats=stats,
                           inventory=inventory, history=history,
                           boss_kills=boss_kills,
                           feedback=request.args.get("feedback"),
                           error=request.args.get("error"))


# ─────────────────────────────────────────────────────────────────────────────
# BAN
# ─────────────────────────────────────────────────────────────────────────────

def admin_ban(pid: int):
    """Render or process the ban administrative workflow."""
    action = request.form.get("action", "ban")

    if action == "unban":
        with exclusive_transaction():
            execute_write("UPDATE players SET is_banned = 0 WHERE id = ?", (pid,))
            _audit("UNBAN_PLAYER", "PLAYER", pid)
        return redirect(url_for("admin_player_detail", pid=pid, feedback="Player unbanned."))

    # Ban: wipe credits, remove gear, return specials to pool, clear in_combat
    player = execute_one("SELECT * FROM players WHERE id = ?", (pid,))
    if not player:
        return redirect(url_for("admin_players"))

    with exclusive_transaction():
        execute_write("UPDATE players SET is_banned = 1, credits = 0, in_combat = 0, "
                      "equipped_weapon_id = NULL, equipped_armor_id = NULL, "
                      "equipped_special_id = NULL WHERE id = ?", (pid,))

        # Return any special items to pool
        specials = execute(
            "SELECT * FROM inventory_items WHERE player_id = ? AND item_type = 'SPECIAL'", (pid,)
        )
        for s in specials:
            execute_write(
                """UPDATE special_item_registry
                   SET status='IN_POOL', current_owner_player_id=NULL,
                       inventory_item_id=NULL, last_released_method='BANNED',
                       updated_at=?
                   WHERE special_item_id=?""",
                (datetime.utcnow().isoformat(), s["item_id"])
            )

        # Delete all inventory
        execute_write("DELETE FROM inventory_items WHERE player_id = ?", (pid,))

        # Cancel any active combat sessions
        execute_write(
            """UPDATE combat_sessions SET status='CANCELLED', result='CANCELLED'
               WHERE (attacker_player_id=? OR defender_player_id=?) AND status='ACTIVE'""",
            (pid, pid)
        )
        _audit("BAN_PLAYER", "PLAYER", pid)

    logger.info("Admin: banned player id=%d", pid)
    return redirect(url_for("admin_player_detail", pid=pid, feedback="Player banned."))


def admin_retire_player(pid: int):
    """Permanently retire a character without deleting its history."""
    player = execute_one("SELECT * FROM players WHERE id=?", (pid,))
    if not player:
        return redirect(url_for("admin_players"))
    if player.get("retired_at"):
        return redirect(url_for("admin_player_detail", pid=pid, error="Character is already retired."))
    now = datetime.utcnow().isoformat()
    with exclusive_transaction():
        sessions = execute(
            """SELECT id,attacker_player_id,defender_player_id FROM combat_sessions
               WHERE status='ACTIVE' AND (attacker_player_id=? OR defender_player_id=?)""", (pid, pid)
        )
        for combat in sessions:
            other = (combat["defender_player_id"] if combat["attacker_player_id"] == pid
                     else combat["attacker_player_id"])
            execute_write("UPDATE combat_sessions SET status='CANCELLED',result='PLAYER_RETIRED',resolved_at=? WHERE id=?",
                          (now, combat["id"]))
            if other:
                execute_write("UPDATE players SET in_combat=0 WHERE id=?", (other,))
        specials = execute("SELECT id,item_id FROM inventory_items WHERE player_id=? AND item_type='SPECIAL'", (pid,))
        for item in specials:
            execute_write(
                """UPDATE special_item_registry SET status='IN_POOL',current_owner_player_id=NULL,
                   inventory_item_id=NULL,last_released_method='PLAYER_RETIRED',updated_at=?
                   WHERE special_item_id=?""", (now, item["item_id"])
            )
        execute_write("DELETE FROM inventory_items WHERE player_id=? AND item_type='SPECIAL'", (pid,))
        execute_write(
            """UPDATE players SET retired_at=?,is_banned=1,in_combat=0,equipped_special_id=NULL
               WHERE id=?""", (now, pid)
        )
        execute_write("UPDATE npc_profiles SET enabled=0,retired=1 WHERE player_id=?", (pid,))
        _audit("RETIRE_PLAYER", "PLAYER", pid)
    logger.info("Admin: retired player id=%d", pid)
    return redirect(url_for("admin_player_detail", pid=pid, feedback="Character retired; unique specials returned to the pool."))


# ─────────────────────────────────────────────────────────────────────────────
# EDIT PLAYER
# ─────────────────────────────────────────────────────────────────────────────

def admin_edit(pid: int):
    """Manual field edits — credits, HP, AP, stat overrides."""
    fields = {}
    for field in ("credits", "current_hp", "current_ap",
                  "str_stat", "end_stat", "agi_stat", "lck_stat", "per_stat", "level", "xp"):
        val = request.form.get(field)
        if val is not None and val.strip() != "":
            try:
                fields[field] = max(0, int(val))
            except ValueError:
                pass

    if fields:
        sets   = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [pid]
        with exclusive_transaction():
            execute_write(f"UPDATE players SET {sets} WHERE id = ?", values)
            _audit("EDIT_PLAYER", "PLAYER", pid, details=fields)
        logger.info("Admin: edited player id=%d fields=%s", pid, list(fields.keys()))

    return redirect(url_for("admin_player_detail", pid=pid, feedback="Player updated."))


# NPC MANAGEMENT
def admin_npcs():
    """Render or process the npcs administrative workflow."""
    if request.method == "POST":
        try:
            pid = _create_npc(request.form)
            return redirect(url_for("admin_npcs", feedback=f"NPC created (player {pid})."))
        except (ValueError, RuntimeError) as exc:
            return redirect(url_for("admin_npcs", error=str(exc)))

    npcs = execute(
        """SELECT p.*,np.*,c.name class_name,
                  (SELECT COUNT(*) FROM inventory_items ii WHERE ii.player_id=p.id AND ii.item_type='SPECIAL') special_count
           FROM npc_profiles np JOIN players p ON p.id=np.player_id
           LEFT JOIN classes c ON c.id=p.class_id ORDER BY np.retired,p.level DESC,p.character_name"""
    )
    classes = execute("SELECT id,name FROM classes WHERE is_active=1 ORDER BY name")
    items = execute(
        """SELECT 'WEAPON' item_type,id,name,level,credit_cost FROM weapons WHERE is_active=1
           UNION ALL SELECT 'ARMOR',id,name,level,credit_cost FROM armor WHERE is_active=1
           UNION ALL SELECT 'SPECIAL',id,name,1 AS level,credit_cost FROM special_items WHERE is_active=1
           ORDER BY item_type,level,name"""
    )
    logs = execute(
        """SELECT nal.*,p.character_name FROM npc_action_log nal JOIN players p ON p.id=nal.player_id
           ORDER BY nal.id DESC LIMIT 50"""
    )
    inventory = execute(
        """SELECT ii.*,CASE ii.item_type WHEN 'WEAPON' THEN w.name WHEN 'ARMOR' THEN a.name ELSE si.name END item_name
           FROM inventory_items ii LEFT JOIN weapons w ON w.id=ii.item_id AND ii.item_type='WEAPON'
           LEFT JOIN armor a ON a.id=ii.item_id AND ii.item_type='ARMOR'
           LEFT JOIN special_items si ON si.id=ii.item_id AND ii.item_type='SPECIAL'
           WHERE ii.player_id IN (SELECT player_id FROM npc_profiles) ORDER BY ii.player_id,ii.item_type"""
    )
    return render_template("admin/npcs.html", npcs=npcs, classes=classes, items=items,
                           inventory=inventory, logs=logs)


def admin_npc_edit(pid: int):
    """Render or process the npc edit administrative workflow."""
    fields = {}
    for name in ("player_hunter", "boss_killer", "hoarder", "thief", "aggression",
                 "self_preservation", "repair_tendency"):
        fields[name] = max(0, min(100, request.form.get(name, type=int, default=0)))
    fields["enabled"] = 1 if request.form.get("enabled") else 0
    with exclusive_transaction():
        execute_write(
            """UPDATE npc_profiles SET player_hunter=?,boss_killer=?,hoarder=?,thief=?,aggression=?,
               self_preservation=?,repair_tendency=?,enabled=? WHERE player_id=?""",
            (*fields.values(), pid)
        )
    return redirect(url_for("admin_npcs", feedback="NPC behavior updated."))


def admin_npc_run(pid: int):
    """Render or process the npc run administrative workflow."""
    from npc import run_npc_turn
    try:
        result = run_npc_turn(pid)
        with exclusive_transaction():
            _audit("RUN_NPC_TURN", "PLAYER", pid, details=result)
        return redirect(url_for("admin_npcs", feedback=f"NPC turn: {result['decision']} - {result['result']}"))
    except Exception as exc:
        logger.exception("Manual NPC turn failed for %d", pid)
        return redirect(url_for("admin_npcs", error=str(exc)))


def admin_npc_spend_ap(pid: int):
    """Render or process the npc spend ap administrative workflow."""
    from npc import spend_npc_ap_now
    try:
        result = spend_npc_ap_now(pid)
        with exclusive_transaction():
            _audit("SPEND_NPC_AP", "PLAYER", pid, details=result)
        return redirect(url_for(
            "admin_npcs",
            feedback=(f"NPC ran {result['decisions']} decision(s) and spent "
                      f"{result['ap_spent']} AP; {result['ap_remaining']} AP remains.")
        ))
    except Exception as exc:
        logger.exception("Immediate NPC AP spending failed for %d", pid)
        return redirect(url_for("admin_npcs", error=str(exc)))


def admin_npc_retire(pid: int):
    """Render or process the npc retire administrative workflow."""
    from npc import retire_npc
    retire_npc(pid)
    with exclusive_transaction():
        _audit("RETIRE_NPC", "PLAYER", pid)
    return redirect(url_for("admin_npcs", feedback="NPC retired; unique specials returned to the pool."))


def admin_npc_grant(pid: int):
    """Render or process the npc grant administrative workflow."""
    try:
        item_type, raw_item_id = request.form.get("item_key", "").upper().split(":", 1)
        item_id = int(raw_item_id)
    except (TypeError, ValueError):
        return redirect(url_for("admin_npcs", error="Invalid inventory item."))
    table = {"WEAPON": "weapons", "ARMOR": "armor", "SPECIAL": "special_items"}.get(item_type)
    if not table or not item_id:
        return redirect(url_for("admin_npcs", error="Invalid inventory item."))
    item = execute_one(f"SELECT * FROM {table} WHERE id=? AND is_active=1", (item_id,))
    if not item:
        return redirect(url_for("admin_npcs", error="Content item not found."))
    if item_type == "SPECIAL":
        registry = execute_one("SELECT * FROM special_item_registry WHERE special_item_id=?", (item_id,))
        if not registry or registry["status"] != "IN_POOL":
            return redirect(url_for("admin_npcs", error="That special item is not currently available."))
    with exclusive_transaction():
        inv_id = execute_write(
            """INSERT INTO inventory_items(player_id,item_type,item_id,current_durability,acquired_method)
               VALUES(?,?,?,?, 'ADMIN_GRANT')""", (pid, item_type, item_id, item.get("starting_durability", 100))
        )
        execute_write(
            """INSERT INTO item_history(player_id,item_type,item_id,item_name,event_type)
               VALUES(?,?,?,?, 'ADMIN_GRANTED')""", (pid, item_type, item_id, item["name"])
        )
        if item_type == "SPECIAL":
            execute_write(
                """UPDATE special_item_registry SET status='IN_INVENTORY',current_owner_player_id=?,
                   inventory_item_id=?,last_acquired_method='ADMIN_GRANT',updated_at=? WHERE special_item_id=?""",
                (pid, inv_id, datetime.utcnow().isoformat(), item_id)
            )
        _audit("GRANT_ITEM", "PLAYER", pid, details={"item_type": item_type, "item_id": item_id})
    return redirect(url_for("admin_npcs", feedback=f"Granted {item['name']}."))


def admin_npc_remove(pid: int, inv_id: int):
    """Render or process the npc remove administrative workflow."""
    item = execute_one("SELECT * FROM inventory_items WHERE id=? AND player_id=?", (inv_id, pid))
    if not item:
        return redirect(url_for("admin_npcs", error="Inventory item not found."))
    with exclusive_transaction():
        execute_write(
            """UPDATE players SET equipped_weapon_id=CASE WHEN equipped_weapon_id=? THEN NULL ELSE equipped_weapon_id END,
               equipped_armor_id=CASE WHEN equipped_armor_id=? THEN NULL ELSE equipped_armor_id END,
               equipped_special_id=CASE WHEN equipped_special_id=? THEN NULL ELSE equipped_special_id END WHERE id=?""",
            (inv_id, inv_id, inv_id, pid)
        )
        execute_write("DELETE FROM inventory_items WHERE id=?", (inv_id,))
        if item["item_type"] == "SPECIAL":
            execute_write(
                """UPDATE special_item_registry SET status='IN_POOL',current_owner_player_id=NULL,
                   inventory_item_id=NULL,last_released_method='ADMIN_REMOVED',updated_at=? WHERE special_item_id=?""",
                (datetime.utcnow().isoformat(), item["item_id"])
            )
        _audit("REMOVE_ITEM", "PLAYER", pid, details={"inventory_id": inv_id})
    return redirect(url_for("admin_npcs", feedback="Inventory item removed."))


def _create_npc(form) -> int:
    """Provide the internal create npc operation used by this module."""
    name = form.get("character_name", "").strip()
    if not name:
        raise ValueError("NPC character name is required.")
    class_id = form.get("class_id", type=int)
    cls = execute_one("SELECT * FROM classes WHERE id=? AND is_active=1", (class_id,))
    if not cls:
        raise ValueError("Select an active class.")
    level = max(2, min(3, form.get("level", type=int, default=3)))
    scores = _npc_scores_from_form(form)
    # Balanced creation allocation plus one earned point per level after first.
    stat_points = get_all_settings().get("STARTING_STAT_POINTS", cfg.STARTING_STAT_POINTS)
    alloc = [0, 0, 0, 0, 0]
    for i in range(stat_points): alloc[i % 5] += 1
    for i in range(level - 1): alloc[i % 5] += 1
    stats = [1 + cls[key] + alloc[i] for i, key in enumerate(
        ("str_bonus", "end_bonus", "agi_bonus", "lck_bonus", "per_bonus"))]
    max_hp = 10 + stats[1] + 5 * level
    current_ap = get_all_settings().get("BASE_DAILY_AP", cfg.BASE_DAILY_AP) + math.floor(stats[1] / 2)
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_") or "npc"
    username = f"npc_{slug}_{uuid.uuid4().hex[:6]}"
    email = f"{username}@npc.local"
    with exclusive_transaction():
        pid = execute_write(
            """INSERT INTO players(username,password_hash,email,character_name,sex,class_id,
               str_stat,end_stat,agi_stat,lck_stat,per_stat,level,xp,current_hp,current_ap,credits)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (username, generate_password_hash(uuid.uuid4().hex), email, name, form.get("sex", "Other"),
             class_id, *stats, level, 0, max_hp, current_ap, max(100, cfg.STARTING_CREDITS))
        )
        execute_write("INSERT INTO player_stats(player_id) VALUES(?)", (pid,))
        execute_write(
            """INSERT INTO npc_profiles(player_id,player_hunter,boss_killer,hoarder,thief,aggression,
               self_preservation,repair_tendency) VALUES(?,?,?,?,?,?,?,?)""",
            (pid, *scores)
        )
    from routes.auth import _award_starter_gear
    from npc import _equip_best_core_items
    _award_starter_gear(pid)
    _equip_best_core_items(pid)
    return pid


def _npc_scores_from_form(form):
    # Presets populate the visible form in the browser. Always save the visible
    # values so an administrator can select a preset and then customize it.
    """Provide the internal npc scores from form operation used by this module."""
    return tuple(max(0, min(100, form.get(key, type=int, default=default)))
                 for key, default in (
                     ("player_hunter", 100), ("boss_killer", 0),
                     ("hoarder", 0), ("thief", 0), ("aggression", 85),
                     ("self_preservation", 35), ("repair_tendency", 55)))


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

def admin_config():
    """Render or process the config administrative workflow."""
    feedback = error = None

    if request.method == "POST":
        constant = request.form.get("constant_name", "").strip()
        value    = request.form.get("value", "").strip()
        if constant and value:
            with exclusive_transaction():
                execute_write(
                    """INSERT OR REPLACE INTO settings (constant_name, value, imported_at)
                       VALUES (?, ?, ?)""",
                    (constant, value, datetime.utcnow().isoformat())
                )
                _audit("EDIT_CONFIG", "SETTING", reason=constant, details={"value": value})
            feedback = f"Setting '{constant}' updated to '{value}'."
        else:
            error = "Both constant name and value are required."

    settings_rows = execute("SELECT * FROM settings ORDER BY constant_name")
    defaults      = {k: v for k, v in vars(cfg).items()
                     if k.isupper() and not k.startswith("_")}
    return render_template("admin/config.html",
                           settings_rows=settings_rows,
                           defaults=defaults,
                           feedback=feedback,
                           error=error)


# ─────────────────────────────────────────────────────────────────────────────
# MANUAL MIDNIGHT RESET
# ─────────────────────────────────────────────────────────────────────────────

def admin_midnight():
    """Render or process the midnight administrative workflow."""
    from scheduler import midnight_reset
    try:
        midnight_reset()
        return redirect(url_for("admin_index") + "?feedback=Midnight+reset+complete.")
    except Exception as e:
        logger.exception("Admin triggered midnight reset failed")
        return redirect(url_for("admin_index") + f"?error={str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# FULL GAME RESET
# ─────────────────────────────────────────────────────────────────────────────

def admin_full_reset():
    """Wipe all player data, reinitialise schema, re-import Excel if staged."""
    confirm = request.form.get("confirm", "")
    if confirm != "RESET":
        return redirect(
            url_for("admin_index") + "?error=Full+reset+requires+typing+RESET+to+confirm."
        )

    import sqlite3, os

    logger.warning("Admin: initiating FULL GAME RESET")

    # Drop all operational tables (content tables survive)
    operational = [
        "combat_buffs", "combat_logs", "combat_sessions",
        "boss_instances", "minion_instances", "boss_intel",
        "inventory_items", "item_history", "special_item_registry",
        "shop_listings", "daily_feed", "action_queue",
        "player_stats", "level_up_history", "status_effects", "players",
    ]
    conn = sqlite3.connect(cfg.DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")
    for tbl in operational:
        conn.execute(f"DROP TABLE IF EXISTS {tbl}")
    conn.execute("PRAGMA foreign_keys = ON")
    conn.commit()
    conn.close()

    # Reinitialise schema (recreates dropped tables)
    init_db()

    # Rebuild special_item_registry from existing special_items content
    with exclusive_transaction():
        specials = execute("SELECT id FROM special_items WHERE is_active = 1")
        for s in specials:
            execute_write(
                "INSERT OR IGNORE INTO special_item_registry (special_item_id, status) VALUES (?, 'IN_POOL')",
                (s["id"],)
            )

    # Re-import if staged file exists
    if os.path.exists(cfg.PENDING_IMPORT_PATH):
        from importer import run_import
        result = run_import(cfg.PENDING_IMPORT_PATH, full_reset=True)
        logger.info("Post-reset import: %s", result)

    logger.warning("Admin: FULL GAME RESET complete")
    return redirect(url_for("admin_index") + "?feedback=Full+game+reset+complete.")


# ─────────────────────────────────────────────────────────────────────────────
# LOGS
# ─────────────────────────────────────────────────────────────────────────────

def admin_logs():
    """Render or process the logs administrative workflow."""
    import os

    def read_tail(path: str, n: int = 100) -> list[str]:
        """Handle the read tail workflow."""
        if not os.path.exists(path):
            return []
        with open(path, "r") as f:
            return f.readlines()[-n:]

    import_errors = read_tail(cfg.IMPORT_ERROR_LOG)
    orphan_log    = read_tail(cfg.ORPHAN_LOG)
    failed_queue  = execute(
        "SELECT * FROM action_queue WHERE status='FAILED' ORDER BY created_at DESC LIMIT 50"
    )

    return render_template("admin/logs.html",
                           import_errors=import_errors,
                           orphan_log=orphan_log,
                           failed_queue=failed_queue)


################################################################################

# FILE: app.py
"""Main Flask application factory, request guards, and background-job setup."""
# app.py
# Main Flask application factory.

import logging
from datetime import datetime, timezone

from flask import Flask, session, redirect, url_for, g, request
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

import config_defaults as cfg
from database import get_db, close_db, init_db, get_player, get_all_settings
from queue_handler import startup_cleanup

logger = logging.getLogger(__name__)

_AUTH_EXEMPT = {
    "auth.login", "auth.login_post",
    "auth.register", "auth.register_post",
    "static",
}
_LEVELUP_EXEMPT = {"auth.levelup", "auth.levelup_post", "auth.logout", "static"}


def create_app() -> Flask:
    """Handle the create app workflow."""
    app = Flask(__name__)
    app.secret_key = cfg.SECRET_KEY

    init_db()
    startup_cleanup()

    app.teardown_appcontext(close_db)
    _register_blueprints(app)
    app.context_processor(_context_processor)
    app.before_request(_check_auth)
    app.before_request(_load_player)
    app.before_request(_check_levelup)
    app.before_request(_set_blackout_flag)
    _start_scheduler(app)

    return app


def _register_blueprints(app: Flask):
    """Provide the internal register blueprints operation used by this module."""
    from routes.auth        import bp as auth_bp
    from routes.dashboard   import bp as dashboard_bp
    from routes.actions     import bp as actions_bp
    from routes.combat      import bp as combat_bp
    from routes.shop        import bp as shop_bp
    from routes.blacksmith  import bp as blacksmith_bp
    from routes.character   import bp as character_bp
    from routes.scoreboards import bp as scoreboards_bp
    from routes.feeds       import bp as feeds_bp

    for bp in [auth_bp, dashboard_bp, actions_bp, combat_bp, shop_bp,
               blacksmith_bp, character_bp, scoreboards_bp, feeds_bp]:
        app.register_blueprint(bp)


def _context_processor() -> dict:
    """Provide the internal context processor operation used by this module."""
    player = g.get("player")
    if not player:
        return {}
    return {"player": player, "settings": get_all_settings()}


def _check_auth():
    """Provide the internal check auth operation used by this module."""
    if request.endpoint in _AUTH_EXEMPT:
        return None
    if not session.get("player_id"):
        return redirect(url_for("auth.login"))
    return None


def _load_player():
    """Load the authenticated player before route handlers access ``g.player``."""
    player_id = session.get("player_id")
    if not player_id:
        return None
    player = get_player(player_id)
    if player is None:
        session.clear()
        return redirect(url_for("auth.login"))
    g.player = player
    return None


def _check_levelup():
    """Provide the internal check levelup operation used by this module."""
    if request.endpoint in _LEVELUP_EXEMPT:
        return None
    player = g.get("player")
    if player and player.get("pending_levelup") and not player.get("in_combat"):
        return redirect(url_for("auth.levelup"))
    return None


def _set_blackout_flag():
    """Provide the internal set blackout flag operation used by this module."""
    settings = get_all_settings()
    blackout_mins = settings.get("MIDNIGHT_BLACKOUT_MINUTES", cfg.MIDNIGHT_BLACKOUT_MINUTES)
    now = datetime.now(timezone.utc)
    minutes_to_midnight = (24 * 60) - (now.hour * 60 + now.minute)
    g.blackout = (minutes_to_midnight <= blackout_mins)


def _start_scheduler(app: Flask):
    """Provide the internal start scheduler operation used by this module."""
    from scheduler import midnight_reset, ap_trickle
    from npc import run_due_npc_turns

    scheduler = BackgroundScheduler(timezone="UTC")
    scheduler.add_job(
        func=lambda: _run_with_context(app, midnight_reset),
        trigger=CronTrigger(hour=0, minute=0, timezone="UTC"),
        id="midnight_reset", name="Midnight Reset",
        replace_existing=True, misfire_grace_time=300,
    )
    scheduler.add_job(
        func=lambda: _run_with_context(app, ap_trickle),
        trigger=CronTrigger(hour="3,9,15,21", minute=0, timezone="UTC"),
        id="ap_trickle", name="AP Trickle",
        replace_existing=True, misfire_grace_time=300,
    )
    scheduler.add_job(
        func=lambda: _run_with_context(app, run_due_npc_turns),
        trigger=CronTrigger(minute="*/5", timezone="UTC"),
        id="npc_turns", name="NPC Turns",
        replace_existing=True, misfire_grace_time=240,
    )
    scheduler.start()
    logger.info("APScheduler started")


def _run_with_context(app: Flask, fn):
    """Provide the internal run with context operation used by this module."""
    with app.app_context():
        from database import execute_write, exclusive_transaction
        started = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        with exclusive_transaction():
            run_id = execute_write(
                "INSERT INTO scheduler_run_log(job_name,status,started_at) VALUES(?, 'RUNNING', ?)",
                (fn.__name__, started)
            )
        try:
            result = fn()
            with exclusive_transaction():
                execute_write(
                    "UPDATE scheduler_run_log SET status='SUCCESS',result_summary=?,finished_at=? WHERE id=?",
                    (str(result)[:2000], datetime.now(timezone.utc).replace(tzinfo=None).isoformat(), run_id)
                )
        except Exception as exc:
            with exclusive_transaction():
                execute_write(
                    "UPDATE scheduler_run_log SET status='FAILED',result_summary=?,finished_at=? WHERE id=?",
                    (str(exc)[:2000], datetime.now(timezone.utc).replace(tzinfo=None).isoformat(), run_id)
                )
            logger.exception("Scheduled job '%s' raised an exception", fn.__name__)


################################################################################

# FILE: config_defaults.py
"""Typed fallback values for gameplay settings that may be overridden in the database."""
# config_defaults.py
# Hardcoded fallback constants. database.get_setting() tries the settings DB
# table first, falls back here if the row is missing, and logs a warning.
# Deployment constants (paths, secret key) live here only — never in the DB.

import os

# ── Deployment constants (never in DB) ───────────────────────────────────────
PENDING_IMPORT_PATH      = "data/pending_import.xlsx"
REJECTED_IMPORT_PATH     = "data/logs/rejected/"
IMPORT_ERROR_LOG         = "data/logs/import_errors.log"
ORPHAN_LOG               = "data/logs/orphan_actions.log"
DB_PATH                  = "data/game.db"
TERMINAL_HISTORY_ENTRIES = 20
SECRET_KEY               = os.environ.get("GAME_SECRET_KEY", "dev-secret-change-in-production")

# ── Game constants (fallbacks if row missing from settings table) ─────────────
BASE_DAILY_AP                          = 20
AP_CARRYOVER_CAP                       = 40
AP_COST_BOSS                           = 3
AP_COST_PVP                            = 3
AP_COST_TAVERN                         = 2
AP_COST_BLACKSMITH                     = 2
AP_COST_SHOP                           = 1
AP_COST_ESCAPE                         = 1
TRICKLE_AP_AMOUNT                      = 3
TRICKLE_AP_INTERVAL_HOURS              = 6
COMBAT_EXTENSION_TIMEOUT               = 20
MIDNIGHT_BLACKOUT_MINUTES              = 10
STARTING_CREDITS                       = 50
STARTING_STAT_POINTS                   = 10
BASE_HP                                = 10
HP_PER_LEVEL                           = 5
END_HP_REGEN_DIVISOR                   = 2
TAVERN_HEAL_COST                       = 15
TAVERN_HEAL_PERCENT                    = 0.50
BRACE_HEAL_PERCENT                     = 0.25
BRACE_AC_BONUS_PERCENT                 = 0.25
BRACE_DODGE_BONUS                      = 5
MIDNIGHT_HEAL_PERCENT                  = 0.50
REPAIR_BASE_PERCENT                    = 0.50
REPAIR_LCK_MULTIPLIER                  = 2
REPAIR_LCK_CAP                         = 0.75
REPAIR_COST_PERCENT                    = 0.25
COMBAT_ROUNDS_DEFAULT                  = 4
COMBAT_ROUNDS_EXTENSION                = 4
COMBAT_WIN_HP_WEIGHT                   = 0.40
COMBAT_WIN_DMG_WEIGHT                  = 0.60
CREDIT_STEAL_PERCENT                   = 0.10
CREDIT_STEAL_LUCK_MULTIPLIER           = 2
ZERO_CREDIT_XP_BONUS                   = 25
STEAL_ACTION_CREDIT_PERCENT            = 0.20
STEAL_BOSS_CREDIT_MULTIPLIER           = 20
STEAL_SPECIAL_BASE_CHANCE              = 0.03
ESCAPE_CREDIT_DROP_CHANCE              = 0.10
INVENTORY_LIMIT                        = 10
OVERENCUMBERED_AP_MULTIPLIER           = 2
OVERENCUMBERED_AC_PENALTY              = 3
OVERENCUMBERED_ATTACK_PENALTY          = 3
SWAP_GEAR_ACCURACY_PENALTY             = 0.30
SWAP_GEAR_AC_PENALTY                   = 0.30
SHOP_WEAPONS_COUNT                     = 10
SHOP_ARMOR_COUNT                       = 10
SHOP_DISCOUNT_MAX                      = 0.50
RANDOM_EVENT_BASE_CHANCE               = 0.20
RANDOM_EVENT_MAX_CHANCE                = 0.60
RANDOM_EVENT_GOOD_BASE                 = 0.50
RANDOM_EVENT_GOOD_MAX                  = 0.90
RANDOM_EVENT_BAD_MIN                   = 0.10
RANDOM_EVENT_LCK_BONUS                 = 0.05
AP_PASSIVE_HP_REGEN                    = 1
CRIT_BASE_THRESHOLD                    = 20
CRIT_LCK_DIVISOR                       = 5
CRIT_MIN_THRESHOLD                     = 15
RESISTANCE_STACK_MIN_DAMAGE_PERCENT    = 0.25
XP_LOSS_DIVISOR                        = 3
SELL_PRICE_PERCENT                     = 0.50
COMBAT_PREF_BALANCED_SPLIT             = 0.50
COMBAT_PREF_OPPORTUNIST_SPLIT          = 0.50
WEALTH_TIER_POOR_MAX                   = 0.33
WEALTH_TIER_MIDDLE_MAX                 = 0.66
INACTIVE_DAYS_THRESHOLD                = 7
MINION_ENCOUNTER_CHANCE                = 0.50
CURSE_AP_REDUCTION                     = 0.20
SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND = 0.02
BOSS_LEVEL_WARNING_THRESHOLD           = 3
LOG_DAILY_ARCHIVE                      = False
LOG_ARCHIVE_PATH                       = "data/logs/daily/"

XP_CURVE = {
    2: 100,   3: 250,   4: 500,   5: 900,
    6: 1400,  7: 2000,  8: 2700,  9: 3500,
    10: 4400, 11: 5500, 12: 7000, 13: 9000,
    14: 12000, 15: 16000
}

# ── Type coercion helpers used by database.get_setting() ─────────────────────
SETTING_TYPES = {
    "BASE_DAILY_AP": int, "AP_CARRYOVER_CAP": int, "AP_COST_BOSS": int,
    "AP_COST_PVP": int, "AP_COST_TAVERN": int, "AP_COST_BLACKSMITH": int,
    "AP_COST_SHOP": int, "AP_COST_ESCAPE": int, "TRICKLE_AP_AMOUNT": int,
    "TRICKLE_AP_INTERVAL_HOURS": int, "COMBAT_EXTENSION_TIMEOUT": int,
    "MIDNIGHT_BLACKOUT_MINUTES": int, "STARTING_CREDITS": int,
    "STARTING_STAT_POINTS": int, "BASE_HP": int, "HP_PER_LEVEL": int,
    "END_HP_REGEN_DIVISOR": int, "TAVERN_HEAL_COST": int,
    "BRACE_DODGE_BONUS": int, "REPAIR_LCK_MULTIPLIER": int,
    "COMBAT_ROUNDS_DEFAULT": int, "COMBAT_ROUNDS_EXTENSION": int,
    "CREDIT_STEAL_LUCK_MULTIPLIER": int, "ZERO_CREDIT_XP_BONUS": int,
    "STEAL_BOSS_CREDIT_MULTIPLIER": int, "INVENTORY_LIMIT": int,
    "OVERENCUMBERED_AP_MULTIPLIER": int, "OVERENCUMBERED_AC_PENALTY": int,
    "OVERENCUMBERED_ATTACK_PENALTY": int, "SHOP_WEAPONS_COUNT": int,
    "SHOP_ARMOR_COUNT": int, "AP_PASSIVE_HP_REGEN": int,
    "CRIT_BASE_THRESHOLD": int, "CRIT_LCK_DIVISOR": int,
    "CRIT_MIN_THRESHOLD": int, "XP_LOSS_DIVISOR": int,
    "INACTIVE_DAYS_THRESHOLD": int, "BOSS_LEVEL_WARNING_THRESHOLD": int,
    "TERMINAL_HISTORY_ENTRIES": int,
    "TAVERN_HEAL_PERCENT": float, "BRACE_HEAL_PERCENT": float,
    "BRACE_AC_BONUS_PERCENT": float, "MIDNIGHT_HEAL_PERCENT": float,
    "REPAIR_BASE_PERCENT": float, "REPAIR_LCK_CAP": float,
    "REPAIR_COST_PERCENT": float, "COMBAT_WIN_HP_WEIGHT": float,
    "COMBAT_WIN_DMG_WEIGHT": float, "CREDIT_STEAL_PERCENT": float,
    "STEAL_ACTION_CREDIT_PERCENT": float, "STEAL_SPECIAL_BASE_CHANCE": float,
    "ESCAPE_CREDIT_DROP_CHANCE": float, "SWAP_GEAR_ACCURACY_PENALTY": float,
    "SWAP_GEAR_AC_PENALTY": float, "SHOP_DISCOUNT_MAX": float,
    "RANDOM_EVENT_BASE_CHANCE": float, "RANDOM_EVENT_MAX_CHANCE": float,
    "RANDOM_EVENT_GOOD_BASE": float, "RANDOM_EVENT_GOOD_MAX": float,
    "RANDOM_EVENT_BAD_MIN": float, "RANDOM_EVENT_LCK_BONUS": float,
    "RESISTANCE_STACK_MIN_DAMAGE_PERCENT": float, "SELL_PRICE_PERCENT": float,
    "COMBAT_PREF_BALANCED_SPLIT": float, "COMBAT_PREF_OPPORTUNIST_SPLIT": float,
    "WEALTH_TIER_POOR_MAX": float, "WEALTH_TIER_MIDDLE_MAX": float,
    "MINION_ENCOUNTER_CHANCE": float, "CURSE_AP_REDUCTION": float,
    "SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND": float,
    "LOG_DAILY_ARCHIVE": bool,
}


################################################################################

# FILE: database.py
"""SQLite connection, transaction, settings, and player-state helpers."""
# database.py
# Single point of contact for all DB operations.
# Provides: connection management, schema init, query helpers,
# exclusive transaction context manager, player/setting loaders.

import sqlite3
import logging
import math
import uuid
from contextlib import contextmanager
from flask import g
import config_defaults as cfg

logger = logging.getLogger(__name__)


def get_db() -> sqlite3.Connection:
    """Return the thread-local DB connection, creating it if needed."""
    if 'db' not in g:
        g.db = sqlite3.connect(
            cfg.DB_PATH,
            detect_types=sqlite3.PARSE_DECLTYPES,
            timeout=10,
        )
        g.db.row_factory = dict_factory
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA journal_mode = WAL")
    return g.db


def close_db(e=None):
    """Close thread-local DB connection. Registered as teardown_appcontext."""
    db = g.pop('db', None)
    if db is not None:
        db.close()


def dict_factory(cursor: sqlite3.Cursor, row: tuple) -> dict:
    """sqlite3 row_factory: rows as dicts keyed by column name."""
    return {col[0]: val for col, val in zip(cursor.description, row)}


def init_db():
    """Create all tables and indexes if they don't exist.
    Safe to call on an existing DB. Called at startup and after full reset."""
    import os
    os.makedirs("data/logs/rejected", exist_ok=True)
    os.makedirs("data/logs/daily",    exist_ok=True)

    schema_path = os.path.join(os.path.dirname(__file__), "schema.sql")
    with sqlite3.connect(cfg.DB_PATH) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        with open(schema_path, "r") as f:
            conn.executescript(f.read())
        # CREATE TABLE IF NOT EXISTS does not add new columns to an existing
        # installation. Keep small additive migrations safe and repeatable.
        npc_columns = {row[1] for row in conn.execute("PRAGMA table_info(npc_profiles)")}
        if "thief" not in npc_columns:
            conn.execute("ALTER TABLE npc_profiles ADD COLUMN thief INTEGER NOT NULL DEFAULT 0")
        player_columns = {row[1] for row in conn.execute("PRAGMA table_info(players)")}
        if "retired_at" not in player_columns:
            conn.execute("ALTER TABLE players ADD COLUMN retired_at TEXT")
    logger.info("Database initialised at %s", cfg.DB_PATH)


@contextmanager
def exclusive_transaction():
    """Context manager: BEGIN EXCLUSIVE ... COMMIT/ROLLBACK.
    Use for all write operations to prevent race conditions.

    Usage:
        with exclusive_transaction():
            execute_write("UPDATE players SET credits = ? WHERE id = ?", (amt, pid))
    """
    db = get_db()
    if db.in_transaction:
        savepoint = f"nested_{uuid.uuid4().hex}"
        db.execute(f"SAVEPOINT {savepoint}")
        try:
            yield
            db.execute(f"RELEASE SAVEPOINT {savepoint}")
        except Exception:
            db.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
            db.execute(f"RELEASE SAVEPOINT {savepoint}")
            raise
        return

    db.execute("BEGIN EXCLUSIVE")
    try:
        yield
        db.execute("COMMIT")
    except Exception:
        db.execute("ROLLBACK")
        raise


def execute(sql: str, params: tuple = ()) -> list[dict]:
    """Execute a SELECT, return all rows as list of dicts."""
    return get_db().execute(sql, params).fetchall()


def execute_one(sql: str, params: tuple = ()) -> dict | None:
    """Execute a SELECT, return first row as dict or None."""
    return get_db().execute(sql, params).fetchone()


def execute_write(sql: str, params: tuple = ()) -> int:
    """Execute INSERT/UPDATE/DELETE. Returns lastrowid or rowcount.
    Must be called inside exclusive_transaction()."""
    cursor = get_db().execute(sql, params)
    return cursor.lastrowid if cursor.lastrowid else cursor.rowcount


def get_player(player_id: int) -> dict | None:
    """Load player row with all derived computed fields attached.
    Called by context processor on every request."""
    player = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    if player is None:
        return None

    active_effects = execute(
        "SELECT effect_type, value FROM status_effects WHERE player_id = ?", (player_id,)
    )
    is_cursed = any(e["effect_type"] == "CURSED" for e in active_effects)
    modifiers = {"str": 0, "end": 0, "agi": 0, "lck": 0, "per": 0, "initiative": 0}
    names = {
        "STR": "str", "END": "end", "AGI": "agi", "LCK": "lck",
        "PER": "per", "INITIATIVE": "initiative",
    }
    for effect in active_effects:
        parts = effect["effect_type"].split("_")
        if len(parts) == 3 and parts[0] == "STAT" and parts[2] in names:
            modifiers[names[parts[2]]] += int(effect["value"])
    for column, key in (("str_stat", "str"), ("end_stat", "end"),
                        ("agi_stat", "agi"), ("lck_stat", "lck"),
                        ("per_stat", "per")):
        player[column] = max(1, player[column] + modifiers[key])
    player["initiative_modifier"] = modifiers["initiative"]
    settings = get_all_settings()

    base_daily_ap  = settings.get("BASE_DAILY_AP",            cfg.BASE_DAILY_AP)
    ap_cap         = settings.get("AP_CARRYOVER_CAP",         cfg.AP_CARRYOVER_CAP)
    inv_limit_base = settings.get("INVENTORY_LIMIT",          cfg.INVENTORY_LIMIT)
    inactive_days  = settings.get("INACTIVE_DAYS_THRESHOLD",  cfg.INACTIVE_DAYS_THRESHOLD)
    ap_regen       = settings.get("AP_PASSIVE_HP_REGEN",      cfg.AP_PASSIVE_HP_REGEN)
    end_divisor    = settings.get("END_HP_REGEN_DIVISOR",     cfg.END_HP_REGEN_DIVISOR)
    curse_red      = settings.get("CURSE_AP_REDUCTION",       cfg.CURSE_AP_REDUCTION)

    end   = player["end_stat"]
    level = player["level"]

    max_hp     = 10 + end + (5 * level)
    raw_max_ap = base_daily_ap + math.floor(end / 2)
    max_ap     = int(raw_max_ap * (1 - curse_red)) if is_cursed else raw_max_ap
    max_ap     = min(max_ap, ap_cap)
    inv_limit  = inv_limit_base + math.floor(player["str_stat"] / 2)
    passive_regen = ap_regen + math.floor(end / end_divisor)

    inv_count = execute_one(
        "SELECT COUNT(*) as cnt FROM inventory_items WHERE player_id = ?", (player_id,)
    )["cnt"]

    from datetime import datetime
    is_inactive = False
    if player["last_login_at"]:
        try:
            last = datetime.fromisoformat(player["last_login_at"])
            is_inactive = (datetime.utcnow() - last).days >= inactive_days
        except ValueError:
            pass

    hp_pct = (player["current_hp"] / max_hp * 100) if max_hp > 0 else 0
    if   hp_pct >= 76: hp_tier = "Healthy"
    elif hp_pct >= 51: hp_tier = "Wounded"
    elif hp_pct >= 26: hp_tier = "Hurt"
    else:              hp_tier = "Critical"

    player.update({
        "max_hp":            max_hp,
        "max_ap":            max_ap,
        "inventory_limit":   inv_limit,
        "inventory_count":   inv_count,
        "is_overencumbered": inv_count > inv_limit,
        "is_cursed":         is_cursed,
        "is_inactive":       is_inactive,
        "passive_regen":     passive_regen,
        "hp_tier":           hp_tier,
        "hp_pct":            round(hp_pct, 1),
    })
    return player


def get_player_equipped(player: dict) -> dict:
    """Load full weapon, armor, and special item rows for a player's equipped gear.
    Returns {'weapon': dict|None, 'armor': dict|None, 'special': dict|None}"""
    result = {"weapon": None, "armor": None, "special": None}
    for slot, col, table in [
        ("weapon",  "equipped_weapon_id",  "weapons"),
        ("armor",   "equipped_armor_id",   "armor"),
        ("special", "equipped_special_id", "special_items"),
    ]:
        inv_id = player.get(col)
        if inv_id:
            inv_row = execute_one("SELECT * FROM inventory_items WHERE id = ?", (inv_id,))
            if inv_row:
                content = execute_one(f"SELECT * FROM {table} WHERE id = ?", (inv_row["item_id"],))
                if content:
                    result[slot] = {**content,
                                    "inv_id": inv_id,
                                    "current_durability": inv_row["current_durability"]}
    return result


def get_setting(constant_name: str, default=None):
    """Look up one constant from settings table; falls back to config_defaults."""
    row = execute_one("SELECT value FROM settings WHERE constant_name = ?", (constant_name,))
    if row is None:
        fallback = getattr(cfg, constant_name, default)
        logger.warning("Setting '%s' missing from DB — using fallback: %s", constant_name, fallback)
        return fallback
    raw = row["value"]
    target_type = cfg.SETTING_TYPES.get(constant_name)
    if target_type is bool:  return raw.upper() in ("TRUE", "1", "YES")
    if target_type is int:   return int(raw)
    if target_type is float: return float(raw)
    return raw


def get_all_settings() -> dict:
    """Return all settings as a typed dict. Cached on g per request."""
    if "settings_cache" in g:
        return g.settings_cache
    rows = execute("SELECT constant_name, value FROM settings")
    result = {}
    for row in rows:
        name, raw = row["constant_name"], row["value"]
        t = cfg.SETTING_TYPES.get(name)
        try:
            if t is bool:  result[name] = raw.upper() in ("TRUE", "1", "YES")
            elif t is int:   result[name] = int(raw)
            elif t is float: result[name] = float(raw)
            else:            result[name] = raw
        except (ValueError, TypeError):
            result[name] = raw
    g.settings_cache = result
    return result


################################################################################

# FILE: importer.py
"""Validate and atomically import the game-content Excel workbook."""
# importer.py  (Phase 7 — full implementation)
# Reads the staged Excel file, validates it, diffs against current DB content,
# and applies changes atomically at midnight reset.
# Called by scheduler.py (midnight) and admin.py (manual trigger / full reset).

import os
import math
import logging
import shutil
from datetime import datetime

from openpyxl import load_workbook

from database import execute, execute_one, execute_write, exclusive_transaction
import config_defaults as cfg

logger = logging.getLogger(__name__)

REQUIRED_SHEETS = {
    "Master", "Bosses", "Minions", "Weapons",
    "Armor", "SpecialItems", "Classes", "RandomEvents", "Settings"
}

DAMAGE_TYPES = ("Blade", "Blunt", "Ballistic", "Energy", "Arcane", "Explosive", "Venom")

VALID_BUFF_TYPES = {
    "AC_BONUS", "DMG_REDUCTION", "ATTACK_BONUS",
    "CRIT_BONUS", "RESISTANCE_TYPE", "HP_RESTORE"
}

VALID_EFFECT_TYPES = {
    "CREDITS", "ITEM_AT_LEVEL", "BONUS_AP", "DURABILITY_RESTORE_RANDOM",
    "SPECIAL_ITEM_FROM_POOL", "HP_LOSS", "DURABILITY_LOSS_RANDOM",
    "XP_LOSS", "AP_REDUCTION_PERCENT", "PROTAGONIST_ENCOUNTER",
    "STAT_BOOST_STR", "STAT_BOOST_END", "STAT_BOOST_AGI",
    "STAT_BOOST_LCK", "STAT_BOOST_PER", "STAT_BOOST_INITIATIVE",
    "STAT_PENALTY_STR", "STAT_PENALTY_END", "STAT_PENALTY_AGI",
    "STAT_PENALTY_LCK", "STAT_PENALTY_PER", "STAT_PENALTY_INITIATIVE",
}

# Intel-sensitive columns on the bosses table — changing these clears boss_intel
INTEL_SENSITIVE_COLS = {
    "res_blade", "res_blunt", "res_ballistic", "res_energy",
    "res_arcane", "res_explosive", "res_venom",
    "weak_blade", "weak_blunt", "weak_ballistic", "weak_energy",
    "weak_arcane", "weak_explosive", "weak_venom",
    "special_attack_damage_type", "special_buff_damage_type",
}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def run_import(filepath: str = None, full_reset: bool = False) -> dict:
    """Main entry point. Returns {'success': bool, 'errors': list, 'summary': dict}."""
    if filepath is None:
        filepath = cfg.PENDING_IMPORT_PATH

    if not os.path.exists(filepath):
        return {"success": False, "errors": ["No staged import file found."], "summary": {}}

    logger.info("Starting import from %s (full_reset=%s)", filepath, full_reset)

    try:
        raw_data = parse_workbook(filepath)
    except Exception as e:
        _reject_import(filepath, f"Failed to parse workbook: {e}")
        return {"success": False, "errors": [str(e)], "summary": {}}

    errors = validate(raw_data)
    if errors:
        _reject_import(filepath, "\n".join(errors))
        return {"success": False, "errors": errors, "summary": {}}

    try:
        changes = diff_content(raw_data, full_reset)
        with exclusive_transaction():
            summary = apply_changes(changes, full_reset)
            clear_stale_intel(changes)
            auto_populate_associated_to()
        # Move the processed file out of the way
        _archive_import(filepath)
        logger.info("Import complete: %s", summary)
        return {"success": True, "errors": [], "summary": summary}
    except Exception as e:
        logger.exception("Import failed during apply_changes")
        _reject_import(filepath, str(e))
        return {"success": False, "errors": [str(e)], "summary": {}}


# ─────────────────────────────────────────────────────────────────────────────
# PARSE
# ─────────────────────────────────────────────────────────────────────────────

def parse_workbook(filepath: str) -> dict:
    """Read all sheets into a dict of {sheet_name: [row_dict, ...]}."""
    wb  = load_workbook(filepath, read_only=True, data_only=True)
    out = {}
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws      = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        rows    = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            row_dict = {h: v for h, v in zip(headers, row) if h is not None}
            # Skip rows that start with a note marker
            name_val = row_dict.get("Name") or row_dict.get("MovieName") or row_dict.get("Constant")
            if name_val and str(name_val).startswith("-"):
                continue
            if name_val and str(name_val).startswith("NOTES"):
                break
            rows.append(row_dict)
        out[sheet_name] = rows
    wb.close()
    return out


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATE
# ─────────────────────────────────────────────────────────────────────────────

def validate(raw_data: dict) -> list[str]:
    """Validate all-or-nothing. Returns list of error strings (empty = valid)."""
    errors = []

    for sheet in REQUIRED_SHEETS:
        if sheet not in raw_data:
            errors.append(f"Missing required sheet: '{sheet}'")
    if errors:
        return errors

    _validate_classes(raw_data.get("Classes", []), errors)
    _validate_bosses(raw_data.get("Bosses", []), errors)
    _validate_minions(raw_data.get("Minions", []), errors)
    _validate_weapons(raw_data.get("Weapons", []), errors)
    _validate_armor(raw_data.get("Armor", []), errors)
    _validate_special_items(raw_data.get("SpecialItems", []), errors)
    _validate_random_events(raw_data.get("RandomEvents", []), errors)
    _validate_master(raw_data, errors)
    return errors


def _require(row: dict, fields: list, sheet: str, errors: list, row_name: str = ""):
    """Provide the internal require operation used by this module."""
    for f in fields:
        if row.get(f) is None or str(row.get(f, "")).strip() == "":
            errors.append(f"[{sheet}] Row '{row_name}': missing required field '{f}'")


def _validate_classes(rows: list, errors: list):
    """Validate classes worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Classes] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name"], "Classes", errors, name)


def _validate_bosses(rows: list, errors: list):
    """Validate bosses worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Bosses] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP",
                      "SpecialAttack_Name", "SpecialAttack_Die", "SpecialAttack_DamageType",
                      "SpecialBuff_Name", "SpecialBuff_Type", "SpecialBuff_Value"], "Bosses", errors, name)
        level = r.get("Level")
        if level is not None and (int(level) < 1 or int(level) > 15):
            errors.append(f"[Bosses] '{name}': Level must be 1-15, got {level}")
        buff_type = r.get("SpecialBuff_Type", "")
        if buff_type and buff_type not in VALID_BUFF_TYPES:
            errors.append(f"[Bosses] '{name}': Invalid SpecialBuff_Type '{buff_type}'")
        atk_type = r.get("SpecialAttack_DamageType", "")
        if atk_type and atk_type not in DAMAGE_TYPES:
            errors.append(f"[Bosses] '{name}': Invalid SpecialAttack_DamageType '{atk_type}'")
        # Resistance and weakness should not overlap
        for dtype in DAMAGE_TYPES:
            d = dtype.lower()
            if r.get(f"Res_{dtype}") and r.get(f"Weak_{dtype}"):
                errors.append(f"[Bosses] '{name}': {dtype} cannot be both resistant AND weak")


def _validate_minions(rows: list, errors: list):
    """Validate minions worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Minions] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP"], "Minions", errors, name)
        level = r.get("Level")
        if level is not None and (int(level) < 1 or int(level) > 15):
            errors.append(f"[Minions] '{name}': Level must be 1-15")


def _validate_weapons(rows: list, errors: list):
    """Validate weapons worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Weapons] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "Type", "DamageDie", "DamageType", "CreditCost"], "Weapons", errors, name)
        if r.get("Type") not in ("Melee", "Ranged", None):
            errors.append(f"[Weapons] '{name}': Type must be 'Melee' or 'Ranged'")
        if r.get("DamageType") and r["DamageType"] not in DAMAGE_TYPES:
            errors.append(f"[Weapons] '{name}': Invalid DamageType '{r['DamageType']}'")


def _validate_armor(rows: list, errors: list):
    """Validate armor worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Armor] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "CreditCost"], "Armor", errors, name)


def _validate_special_items(rows: list, errors: list):
    """Validate special items worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[SpecialItems] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "AssociatedTo", "AssociationType", "CreditCost"], "SpecialItems", errors, name)
        if r.get("AssociationType") not in ("Boss", "Minion", "Protagonist", None):
            errors.append(f"[SpecialItems] '{name}': invalid AssociationType")


def _validate_random_events(rows: list, errors: list):
    """Validate random events worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[RandomEvents] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Type", "Rarity", "EffectType", "EffectAmount", "Duration"], "RandomEvents", errors, name)
        if r.get("Type") not in ("Good", "Bad", None):
            errors.append(f"[RandomEvents] '{name}': Type must be 'Good' or 'Bad'")
        if r.get("Rarity") not in ("Common", "Uncommon", "Rare", None):
            errors.append(f"[RandomEvents] '{name}': Rarity must be Common/Uncommon/Rare")
        if r.get("EffectType") and r["EffectType"] not in VALID_EFFECT_TYPES:
            errors.append(f"[RandomEvents] '{name}': Invalid EffectType '{r['EffectType']}'")


def _validate_master(raw_data: dict, errors: list):
    """Validate master worksheet rows and report actionable errors."""
    boss_names    = {r.get("Name") for r in raw_data.get("Bosses", [])}
    minion_names  = {r.get("Name") for r in raw_data.get("Minions", [])}
    weapon_names  = {r.get("Name") for r in raw_data.get("Weapons", [])}
    armor_names   = {r.get("Name") for r in raw_data.get("Armor", [])}
    special_names = {r.get("Name") for r in raw_data.get("SpecialItems", [])}

    for r in raw_data.get("Master", []):
        movie = r.get("MovieName", "?")
        _require(r, ["MovieName", "BossName", "BossWeapon", "BossArmor", "BossSpecialItem",
                      "MinionName", "MinionWeapon", "MinionArmor", "MinionSpecialItem"],
                 "Master", errors, movie)
        if r.get("BossName")          and r["BossName"]          not in boss_names:
            errors.append(f"[Master] '{movie}': BossName '{r['BossName']}' not found in Bosses sheet")
        if r.get("MinionName")        and r["MinionName"]        not in minion_names:
            errors.append(f"[Master] '{movie}': MinionName '{r['MinionName']}' not found in Minions sheet")
        if r.get("BossWeapon")        and r["BossWeapon"]        not in weapon_names:
            errors.append(f"[Master] '{movie}': BossWeapon '{r['BossWeapon']}' not found in Weapons sheet")
        if r.get("BossArmor")         and r["BossArmor"]         not in armor_names:
            errors.append(f"[Master] '{movie}': BossArmor '{r['BossArmor']}' not found in Armor sheet")
        if r.get("BossSpecialItem")   and r["BossSpecialItem"]   not in special_names:
            errors.append(f"[Master] '{movie}': BossSpecialItem '{r['BossSpecialItem']}' not found in SpecialItems sheet")
        if r.get("MinionWeapon")      and r["MinionWeapon"]      not in weapon_names:
            errors.append(f"[Master] '{movie}': MinionWeapon '{r['MinionWeapon']}' not found in Weapons sheet")
        if r.get("MinionArmor")       and r["MinionArmor"]       not in armor_names:
            errors.append(f"[Master] '{movie}': MinionArmor '{r['MinionArmor']}' not found in Armor sheet")
        if r.get("MinionSpecialItem") and r["MinionSpecialItem"] not in special_names:
            errors.append(f"[Master] '{movie}': MinionSpecialItem '{r['MinionSpecialItem']}' not found in SpecialItems sheet")
        for field, pool in [
            ("ProtagonistWeapon", weapon_names),
            ("ProtagonistArmor", armor_names),
            ("ProtagonistSpecialItem", special_names),
        ]:
            value = r.get(field)
            if value and value not in pool:
                errors.append(f"[Master] '{movie}': {field} '{value}' not found")


# ─────────────────────────────────────────────────────────────────────────────
# DIFF
# ─────────────────────────────────────────────────────────────────────────────

def diff_content(raw_data: dict, full_reset: bool = False) -> dict:
    """Compare raw_data against current DB. Returns changes dict.
    If full_reset=True, treat everything as INSERT (no existing rows)."""
    changes = {}

    changes["classes"]       = _diff_table(raw_data.get("Classes", []),       "classes",       "Name", _map_class)
    changes["bosses"]        = _diff_table(raw_data.get("Bosses", []),         "bosses",        "Name", _map_boss,  full_reset=full_reset)
    changes["minions"]       = _diff_table(raw_data.get("Minions", []),        "minions",       "Name", _map_minion, full_reset=full_reset)
    changes["weapons"]       = _diff_table(raw_data.get("Weapons", []),        "weapons",       "Name", _map_weapon)
    changes["armor"]         = _diff_table(raw_data.get("Armor", []),          "armor",         "Name", _map_armor)
    changes["special_items"] = _diff_table(raw_data.get("SpecialItems", []),   "special_items", "Name", _map_special_item)
    changes["random_events"] = _diff_table(raw_data.get("RandomEvents", []),   "random_events", "Name", _map_random_event)
    changes["settings"]      = _diff_settings(raw_data.get("Settings", []))
    changes["master_rows"]   = raw_data.get("Master", [])  # always reprocess
    return changes


def _diff_table(rows: list, table: str, name_col_excel: str,
                mapper_fn, full_reset: bool = False) -> dict:
    """Compare Excel rows against DB rows matched by name.
    Returns {insert: [...], update: [...]} — never deletes."""
    existing = {}
    if not full_reset:
        db_rows = execute(f"SELECT * FROM {table}")
        existing = {r["name"]: r for r in db_rows}

    inserts = []
    updates = []
    for row in rows:
        name = row.get(name_col_excel)
        if not name:
            continue
        mapped = mapper_fn(row)
        mapped["name"] = str(name).strip()
        if mapped["name"] in existing:
            updates.append({"db_row": existing[mapped["name"]], "new_data": mapped})
        else:
            inserts.append(mapped)
    return {"insert": inserts, "update": updates}


def _diff_settings(rows: list) -> list:
    """Settings are always upsert by constant_name."""
    result = []
    for r in rows:
        name  = r.get("Constant")
        value = r.get("Value")
        desc  = r.get("Description", "")
        if name and value is not None:
            result.append({"constant_name": str(name), "value": str(value),
                           "description": str(desc) if desc else ""})
    return result


# ─────────────────────────────────────────────────────────────────────────────
# MAPPERS  (Excel row dict → DB column dict)
# ─────────────────────────────────────────────────────────────────────────────

def _b(val) -> int:
    """Convert Excel boolean/string to 0 or 1."""
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, str):
        return 1 if val.upper() in ("TRUE", "YES", "1") else 0
    return 1 if val else 0


def _i(val, default=0) -> int:
    """Provide the internal i operation used by this module."""
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _f(val, default=0.0) -> float:
    """Provide the internal f operation used by this module."""
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _s(val, default="") -> str:
    """Provide the internal s operation used by this module."""
    return str(val).strip() if val is not None else default


def _map_class(r: dict) -> dict:
    """Map a normalized class worksheet row to database fields."""
    return {
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")), "description": _s(r.get("Description")),
    }


def _map_boss(r: dict) -> dict:
    """Map a normalized boss worksheet row to database fields."""
    return {
        "level": _i(r.get("Level")),
        "str_stat": _i(r.get("STR")), "end_stat": _i(r.get("END")),
        "agi_stat": _i(r.get("AGI")), "lck_stat": _i(r.get("LCK")),
        "per_stat": _i(r.get("PER")), "max_hp": _i(r.get("HP")),
        "phase2_hp_percent": _i(r.get("Phase2_HPPercent"), 50),
        "phase3_hp_percent": _i(r.get("Phase3_HPPercent"), 25),
        "special_attack_name":        _s(r.get("SpecialAttack_Name")),
        "special_attack_die":         _s(r.get("SpecialAttack_Die")),
        "special_attack_damage_type": _s(r.get("SpecialAttack_DamageType")),
        "special_attack_flavor":      _s(r.get("SpecialAttack_Flavor")),
        "special_buff_name":          _s(r.get("SpecialBuff_Name")),
        "special_buff_type":          _s(r.get("SpecialBuff_Type")),
        "special_buff_value":         _f(r.get("SpecialBuff_Value")),
        "special_buff_damage_type":   _s(r.get("SpecialBuff_DamageType")),
        "special_buff_flavor":        _s(r.get("SpecialBuff_Flavor")),
        **{f"res_{d.lower()}":  _b(r.get(f"Res_{d}"))  for d in DAMAGE_TYPES},
        **{f"weak_{d.lower()}": _b(r.get(f"Weak_{d}")) for d in DAMAGE_TYPES},
        "drop_weapon_chance":        _f(r.get("Drop_Weapon_Chance")),
        "drop_armor_chance":         _f(r.get("Drop_Armor_Chance")),
        "drop_special_item_chance":  _f(r.get("Drop_SpecialItem_Chance")),
        "drop_credit_min":           _i(r.get("Drop_Credit_Min")),
        "drop_credit_max":           _i(r.get("Drop_Credit_Max")),
        "flavor_text":               _s(r.get("FlavorText")),
    }


def _map_minion(r: dict) -> dict:
    """Map a normalized minion worksheet row to database fields."""
    return {
        "level": _i(r.get("Level")),
        "str_stat": _i(r.get("STR")), "end_stat": _i(r.get("END")),
        "agi_stat": _i(r.get("AGI")), "lck_stat": _i(r.get("LCK")),
        "per_stat": _i(r.get("PER")), "max_hp": _i(r.get("HP")),
        "drop_weapon_chance":       _f(r.get("Drop_Weapon_Chance")),
        "drop_armor_chance":        _f(r.get("Drop_Armor_Chance")),
        "drop_special_item_chance": _f(r.get("Drop_SpecialItem_Chance")),
        "drop_credit_min":          _i(r.get("Drop_Credit_Min")),
        "drop_credit_max":          _i(r.get("Drop_Credit_Max")),
        "flavor_text":              _s(r.get("FlavorText")),
    }


def _map_weapon(r: dict) -> dict:
    """Map a normalized weapon worksheet row to database fields."""
    return {
        "level":       _i(r.get("Level")),
        "weapon_type": _s(r.get("Type")),
        "damage_die":  _s(r.get("DamageDie")),
        "damage_type": _s(r.get("DamageType")),
        "str_bonus":   _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus":   _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus":   _i(r.get("PER")),
        "associated_to":       _s(r.get("AssociatedTo")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_armor(r: dict) -> dict:
    """Map a normalized armor worksheet row to database fields."""
    return {
        "level":    _i(r.get("Level")),
        "ac_bonus": _i(r.get("AC_Bonus")),
        **{f"res_{d.lower()}": _b(r.get(f"Res_{d}")) for d in DAMAGE_TYPES},
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")),
        "associated_to":       _s(r.get("AssociatedTo")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_special_item(r: dict) -> dict:
    """Map a normalized special item worksheet row to database fields."""
    return {
        "associated_to":   _s(r.get("AssociatedTo")),
        "association_type": _s(r.get("AssociationType")),
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")),
        "initiative_bonus":    _i(r.get("InitiativeBonus")),
        "extra_attack":        _b(r.get("ExtraAttack")),
        "crit_chance_bonus":   _f(r.get("CritChanceBonus")),
        "crit_dmg_multiplier": _f(r.get("CritDmgMultiplier")),
        "ac_bonus":            _i(r.get("ACBonus")),
        **{f"res_{d.lower()}": _b(r.get(f"Res_{d}")) for d in DAMAGE_TYPES},
        "bonus_damage_type":   _s(r.get("BonusDamageType")),
        "bonus_damage_amount": _i(r.get("BonusDamageAmount")),
        "xp_multiplier":       _f(r.get("XPMultiplier")),
        "credit_multiplier":   _f(r.get("CreditMultiplier")),
        "steal_bonus":         _f(r.get("StealBonus")),
        "bonus_ap":            _i(r.get("BonusAP")),
        "hp_regen_bonus":      _i(r.get("HPRegenBonus")),
        "durability_reduction": _f(r.get("DurabilityReduction")),
        "shop_discount":       _f(r.get("ShopDiscount")),
        "sell_bonus":          _f(r.get("SellBonus")),
        "encounter_bonus":     _f(r.get("EncounterBonus")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_random_event(r: dict) -> dict:
    """Map a normalized random event worksheet row to database fields."""
    return {
        "event_type":    _s(r.get("Type")),
        "rarity":        _s(r.get("Rarity")),
        "flavor_text":   _s(r.get("FlavorText")),
        "effect_type":   _s(r.get("EffectType")),
        "effect_amount": _i(r.get("EffectAmount")),
        "duration":      _s(r.get("Duration")),
    }


# ─────────────────────────────────────────────────────────────────────────────
# APPLY CHANGES
# ─────────────────────────────────────────────────────────────────────────────

def apply_changes(changes: dict, full_reset: bool = False) -> dict:
    """Apply all inserts and updates. Must be called inside exclusive_transaction().
    Returns summary dict of counts."""
    summary = {}
    order   = ["classes", "bosses", "minions", "weapons", "armor", "special_items", "random_events"]

    for key in order:
        if key not in changes:
            continue
        tbl    = key
        data   = changes[key]
        inserts = data.get("insert", [])
        updates = data.get("update", [])
        for row in inserts:
            _upsert_row(tbl, row, None)
            # Create special_item_registry row for new special items
            if tbl == "special_items":
                new_id = execute_one("SELECT id FROM special_items WHERE name = ?", (row["name"],))
                if new_id:
                    execute_write(
                        """INSERT OR IGNORE INTO special_item_registry (special_item_id, status)
                           VALUES (?, 'IN_POOL')""",
                        (new_id["id"],)
                    )
        for item in updates:
            _upsert_row(tbl, item["new_data"], item["db_row"]["id"])
        summary[key] = {"insert": len(inserts), "update": len(updates)}

    # Settings: upsert by constant_name
    for s in changes.get("settings", []):
        execute_write(
            """INSERT OR REPLACE INTO settings (constant_name, value, description, imported_at)
               VALUES (?, ?, ?, ?)""",
            (s["constant_name"], s["value"], s["description"], datetime.utcnow().isoformat())
        )
    summary["settings"] = {"upsert": len(changes.get("settings", []))}

    # Master: process after all content tables are up to date
    _apply_master(changes.get("master_rows", []))
    summary["master"] = {"processed": len(changes.get("master_rows", []))}

    return summary


def _upsert_row(table: str, data: dict, existing_id: int | None):
    """Insert or update a content table row."""
    data["imported_at"] = datetime.utcnow().isoformat()
    if existing_id is None:
        cols   = ", ".join(data.keys())
        placeholders = ", ".join("?" * len(data))
        execute_write(
            f"INSERT OR IGNORE INTO {table} ({cols}) VALUES ({placeholders})",
            tuple(data.values())
        )
    else:
        sets = ", ".join(f"{k} = ?" for k in data)
        execute_write(
            f"UPDATE {table} SET {sets} WHERE id = ?",
            tuple(data.values()) + (existing_id,)
        )


def _apply_master(master_rows: list):
    """Process master sheet: upsert master rows, linking by name.
    Now includes protagonist FK columns."""
    for r in master_rows:
        movie = _s(r.get('MovieName'))
        if not movie:
            continue

        def get_id(table, name):
            """Handle the get id workflow."""
            if not name:
                return None
            row = execute_one(f"SELECT id FROM {table} WHERE name = ?", (_s(name),))
            return row['id'] if row else None

        boss_id          = get_id('bosses',        r.get('BossName'))
        minion_id        = get_id('minions',       r.get('MinionName'))
        boss_weapon_id   = get_id('weapons',       r.get('BossWeapon'))
        boss_armor_id    = get_id('armor',         r.get('BossArmor'))
        boss_special_id  = get_id('special_items', r.get('BossSpecialItem'))
        min_weapon_id    = get_id('weapons',       r.get('MinionWeapon'))
        min_armor_id     = get_id('armor',         r.get('MinionArmor'))
        min_special_id   = get_id('special_items', r.get('MinionSpecialItem'))
        prot_name        = _s(r.get('ProtagonistName')) or None
        prot_weapon_id   = get_id('weapons',       r.get('ProtagonistWeapon'))
        prot_armor_id    = get_id('armor',         r.get('ProtagonistArmor'))
        prot_special_id  = get_id('special_items', r.get('ProtagonistSpecialItem'))

        if not all([boss_id, minion_id, boss_weapon_id, boss_armor_id,
                    boss_special_id, min_weapon_id, min_armor_id, min_special_id]):
            logger.warning("Master row '%s': could not resolve all FK references, skipping", movie)
            continue

        now = datetime.utcnow().isoformat()
        existing = execute_one("SELECT id FROM master WHERE movie_name = ?", (movie,))
        if existing:
            execute_write(
                """UPDATE master SET
                   boss_id=?, boss_weapon_id=?, boss_armor_id=?, boss_special_item_id=?,
                   minion_id=?, minion_weapon_id=?, minion_armor_id=?, minion_special_item_id=?,
                   protagonist_name=?, protagonist_weapon_id=?, protagonist_armor_id=?,
                   protagonist_special_item_id=?, imported_at=?
                   WHERE movie_name=?""",
                (boss_id, boss_weapon_id, boss_armor_id, boss_special_id,
                 minion_id, min_weapon_id, min_armor_id, min_special_id,
                 prot_name, prot_weapon_id, prot_armor_id, prot_special_id,
                 now, movie)
            )
        else:
            execute_write(
                """INSERT INTO master
                   (movie_name, boss_id, boss_weapon_id, boss_armor_id, boss_special_item_id,
                    minion_id, minion_weapon_id, minion_armor_id, minion_special_item_id,
                    protagonist_name, protagonist_weapon_id, protagonist_armor_id,
                    protagonist_special_item_id, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (movie, boss_id, boss_weapon_id, boss_armor_id, boss_special_id,
                 minion_id, min_weapon_id, min_armor_id, min_special_id,
                 prot_name, prot_weapon_id, prot_armor_id, prot_special_id, now)
            )

def clear_stale_intel(changes: dict):
    """Clear boss_intel rows for any boss whose intel-sensitive columns changed."""
    for item in changes.get("bosses", {}).get("update", []):
        old = item["db_row"]
        new = item["new_data"]
        changed = any(
            str(old.get(col, "")) != str(new.get(col, ""))
            for col in INTEL_SENSITIVE_COLS
            if col in new
        )
        if changed:
            boss_id = old["id"]
            deleted = execute_write(
                "DELETE FROM boss_intel WHERE boss_id = ?", (boss_id,)
            )
            if deleted:
                logger.info("Cleared boss_intel for boss_id=%d (intel-sensitive columns changed)", boss_id)


# ─────────────────────────────────────────────────────────────────────────────
# AUTO-POPULATE ASSOCIATED_TO
# ─────────────────────────────────────────────────────────────────────────────

def auto_populate_associated_to():
    """Update weapons/armor associated_to field from master table.
    Format: 'MovieName (Boss)' or 'MovieName (Minion)'."""
    master_rows = execute("SELECT * FROM master")
    for m in master_rows:
        movie = execute_one("SELECT movie_name FROM master WHERE id = ?", (m["id"],))["movie_name"]
        for col, table in [
            ("boss_weapon_id",   "weapons"),
            ("boss_armor_id",    "armor"),
            ("minion_weapon_id", "weapons"),
            ("minion_armor_id",  "armor"),
        ]:
            item_id = m.get(col)
            if not item_id:
                continue
            side = "Boss" if "boss" in col else "Minion"
            execute_write(
                f"UPDATE {table} SET associated_to = ? WHERE id = ?",
                (f"{movie} ({side})", item_id)
            )


# ─────────────────────────────────────────────────────────────────────────────
# FILE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def _reject_import(filepath: str, reason: str):
    """Move the invalid file to the rejected folder and log the error."""
    os.makedirs(cfg.REJECTED_IMPORT_PATH, exist_ok=True)
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(cfg.REJECTED_IMPORT_PATH, f"rejected_{ts}.xlsx")
    try:
        shutil.move(filepath, dest)
    except Exception:
        pass
    os.makedirs(os.path.dirname(cfg.IMPORT_ERROR_LOG), exist_ok=True)
    with open(cfg.IMPORT_ERROR_LOG, "a") as f:
        f.write(f"{datetime.utcnow().isoformat()} | REJECTED | {reason}\n")
    logger.error("Import rejected: %s", reason)


def _archive_import(filepath: str):
    """Move successfully processed file to a timestamped archive."""
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(cfg.REJECTED_IMPORT_PATH, f"../applied_{ts}.xlsx")
    try:
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.move(filepath, dest)
    except Exception:
        pass


################################################################################

# FILE: npc.py
"""Automated player-character decisions using the normal queued action handlers."""

import logging
import math
import random
import hashlib
from datetime import datetime, timedelta

import config_defaults as cfg
from database import (execute, execute_one, execute_write, exclusive_transaction,
                      get_all_settings, get_player)
from queue_handler import enqueue_and_process

logger = logging.getLogger(__name__)


def _ensure_handlers_loaded():
    # Importing these modules registers their queue handlers.
    """Provide the internal ensure handlers loaded operation used by this module."""
    import routes.actions  # noqa: F401
    import routes.auth  # noqa: F401
    import routes.blacksmith  # noqa: F401
    import routes.combat  # noqa: F401
    import routes.shop  # noqa: F401


def run_due_npc_turns(now: datetime | None = None) -> dict:
    """Run staggered AP-driven NPC decisions.

    There is one deterministic, daily-random wake time in each three-hour
    window. During 23:00-23:55 UTC, NPCs receive repeated opportunities to use
    remaining AP. AP is never erased; only real player actions spend it.
    """
    now = now or datetime.utcnow()
    profiles = execute(
        """SELECT np.* FROM npc_profiles np JOIN players p ON p.id=np.player_id
           WHERE np.enabled=1 AND np.retired=0 AND p.is_banned=0
           ORDER BY COALESCE(np.last_action_at, '') ASC"""
    )
    results = []
    for profile in profiles:
        if not _npc_is_due(profile, now):
            continue
        try:
            attempts = 12 if now.hour == 23 else 1
            no_ap_progress = 0
            for _ in range(attempts):
                player = get_player(profile["player_id"])
                if not player or player["current_ap"] <= 0:
                    break
                ap_before = player["current_ap"]
                result = run_npc_turn(profile["player_id"])
                results.append(result)
                if now.hour != 23 or result["decision"] in ("WAIT", "SKIP"):
                    break
                player_after = get_player(profile["player_id"])
                if player_after and player_after["current_ap"] >= ap_before:
                    no_ap_progress += 1
                    if no_ap_progress >= 2:
                        break
                else:
                    no_ap_progress = 0
        except Exception as exc:
            logger.exception("NPC turn failed for player %d", profile["player_id"])
            _log(profile["player_id"], "ERROR", "Turn raised an exception", str(exc))
    return {"processed": len(results), "results": results}


def _npc_is_due(profile: dict, now: datetime) -> bool:
    """Provide the internal npc is due operation used by this module."""
    last = None
    if profile.get("last_action_at"):
        try:
            last = datetime.fromisoformat(profile["last_action_at"])
        except ValueError:
            pass
    if now.hour == 23:
        # One catch-up attempt per five-minute scheduler slot.
        slot = now.replace(minute=(now.minute // 5) * 5, second=0, microsecond=0)
        return last is None or last < slot
    window_start = now.replace(hour=(now.hour // 3) * 3, minute=0, second=0, microsecond=0)
    identity = f"{now.date().isoformat()}:{profile['player_id']}:{now.hour // 3}".encode()
    offset = int.from_bytes(hashlib.sha256(identity).digest()[:4], "big") % 165
    scheduled = window_start + timedelta(minutes=offset)
    return now >= scheduled and (last is None or last < scheduled)


def run_npc_turn(player_id: int) -> dict:
    """Handle the run npc turn workflow."""
    _ensure_handlers_loaded()
    profile = execute_one("SELECT * FROM npc_profiles WHERE player_id=?", (player_id,))
    player = get_player(player_id)
    if not profile or not player or not profile["enabled"] or profile["retired"]:
        return {"player_id": player_id, "decision": "SKIP", "result": "NPC inactive"}

    _assign_pending_levelup(player_id, profile)
    player = get_player(player_id)

    if player["in_combat"]:
        active = execute_one(
            "SELECT combat_type FROM combat_sessions WHERE attacker_player_id=? AND status='ACTIVE' ORDER BY id DESC LIMIT 1",
            (player_id,)
        )
        result = (_finish_thief_combat(player_id) if profile["thief"] > 0 and active
                  and active["combat_type"] == "PVP" else _finish_active_combat(player_id, profile=profile))
        return _finish_turn(profile, "COMBAT", "An active fight must be resolved first", result)

    _equip_best_core_items(player_id)
    player = get_player(player_id)
    settings = get_all_settings()

    # NPCs live under the same rules as human players, including random
    # events. An event modifies the turn but does not replace the chosen action.
    from routes.actions import check_random_event
    check_random_event(player, settings)
    player = get_player(player_id)

    repair_result = _maybe_repair(player, profile)
    if repair_result:
        return _finish_turn(profile, "REPAIR", repair_result[0], repair_result[1])

    heal_result = _maybe_heal(player, profile, settings)
    if heal_result:
        return _finish_turn(profile, "HEAL", heal_result[0], heal_result[1])

    hoard_result = _maybe_hoard(player, profile)
    if hoard_result:
        return _finish_turn(profile, "HOARD", hoard_result[0], hoard_result[1])

    pvp_targets = _eligible_pvp_targets(player)
    if profile["thief"] >= max(profile["player_hunter"], profile["boss_killer"],
                                profile["hoarder"]):
        last_mode = execute_one(
            """SELECT decision FROM npc_action_log WHERE player_id=?
               AND decision IN ('PVP_STEAL','BOSS') ORDER BY id DESC LIMIT 1""",
            (player_id,)
        )
        steal_turn = not last_mode or last_mode["decision"] == "BOSS"
        if steal_turn and pvp_targets:
            target = random.choice(pvp_targets)
            result = enqueue_and_process(
                player_id, "start_pvp_fight",
                {"target_id": target["id"], "cost_ap": settings.get("AP_COST_PVP", cfg.AP_COST_PVP)}
            )
            if result.get("error"):
                return _finish_turn(profile, "PVP_STEAL", "Random eligible target selected", result["error"])
            combat = _finish_thief_combat(player_id, result["session_id"])
            return _finish_turn(profile, "PVP_STEAL", f"Attempted theft from {target['character_name']}", combat)
        # Alternate with a boss fight for XP. If PvP was unavailable, boss is
        # also the normal fallback under the existing eligibility rules.
        opponent = _choose_boss(player)
        if opponent:
            result = enqueue_and_process(
                player_id, "start_boss_fight",
                {"opponent_id": opponent["id"], "encounter_type": "BOSS",
                 "cost_ap": settings.get("AP_COST_BOSS", cfg.AP_COST_BOSS)}
            )
            if not result.get("error"):
                combat = _finish_active_combat(player_id, result["session_id"], profile)
                return _finish_turn(profile, "BOSS", f"Alternated to boss {opponent['name']} for XP", combat)
            return _finish_turn(profile, "BOSS", "Thief's XP-building turn", result["error"])

    pvp_score = profile["player_hunter"] + random.randint(-10, 10)
    boss_score = profile["boss_killer"] + random.randint(-10, 10)
    if not pvp_targets:
        boss_score += profile["player_hunter"]  # hunter fallback

    if pvp_targets and pvp_score >= boss_score:
        target = _choose_pvp_target(player, pvp_targets, profile["aggression"])
        result = enqueue_and_process(
            player_id, "start_pvp_fight",
            {"target_id": target["id"], "cost_ap": settings.get("AP_COST_PVP", cfg.AP_COST_PVP)}
        )
        if result.get("error"):
            return _finish_turn(profile, "PVP", "Highest motivation was player hunting", result["error"])
        combat = _finish_active_combat(player_id, result["session_id"], profile)
        return _finish_turn(profile, "PVP", f"Targeted {target['character_name']}", combat)

    opponent = _choose_boss(player)
    if opponent:
        result = enqueue_and_process(
            player_id, "start_boss_fight",
            {"opponent_id": opponent["id"], "encounter_type": "BOSS",
             "cost_ap": settings.get("AP_COST_BOSS", cfg.AP_COST_BOSS)}
        )
        if not result.get("error"):
            combat = _finish_active_combat(player_id, result["session_id"], profile)
            return _finish_turn(profile, "BOSS", f"Hunted {opponent['name']}", combat)
        return _finish_turn(profile, "BOSS", "Boss hunt selected", result["error"])

    return _finish_turn(profile, "WAIT", "No legal useful action was available", "No action")


def spend_npc_ap_now(player_id: int, max_decisions: int = 24) -> dict:
    """Run normal NPC decisions immediately until AP or useful progress stops."""
    starting = get_player(player_id)
    if not starting:
        raise ValueError("NPC player not found.")
    if not execute_one("SELECT 1 FROM npc_profiles WHERE player_id=? AND enabled=1 AND retired=0",
                       (player_id,)):
        raise ValueError("NPC is paused or retired.")
    results = []
    no_ap_progress = 0
    for _ in range(max(1, min(50, max_decisions))):
        before = get_player(player_id)
        if before["current_ap"] <= 0:
            break
        result = run_npc_turn(player_id)
        results.append(result)
        after = get_player(player_id)
        if result["decision"] in ("WAIT", "SKIP"):
            break
        if after["current_ap"] >= before["current_ap"]:
            no_ap_progress += 1
            if no_ap_progress >= 2:
                break
        else:
            no_ap_progress = 0
    ending = get_player(player_id)
    return {
        "player_id": player_id, "decisions": len(results),
        "ap_spent": max(0, starting["current_ap"] - ending["current_ap"]),
        "ap_remaining": ending["current_ap"], "results": results,
    }


def retire_npc(player_id: int):
    """Safely retire an NPC and return its unique specials to the pool."""
    now = datetime.utcnow().isoformat()
    with exclusive_transaction():
        active_sessions = execute(
            """SELECT id,attacker_player_id,defender_player_id FROM combat_sessions
               WHERE status='ACTIVE' AND (attacker_player_id=? OR defender_player_id=?)""",
            (player_id, player_id)
        )
        for session in active_sessions:
            execute_write(
                "UPDATE combat_sessions SET status='ABANDONED',result='NPC_RETIRED',resolved_at=? WHERE id=?",
                (now, session["id"])
            )
            other_id = (session["defender_player_id"] if session["attacker_player_id"] == player_id
                        else session["attacker_player_id"])
            if other_id:
                execute_write("UPDATE players SET in_combat=0 WHERE id=?", (other_id,))
        specials = execute(
            "SELECT id,item_id FROM inventory_items WHERE player_id=? AND item_type='SPECIAL'",
            (player_id,)
        )
        for item in specials:
            execute_write(
                """UPDATE special_item_registry SET status='IN_POOL',current_owner_player_id=NULL,
                   inventory_item_id=NULL,last_released_method='NPC_RETIRED',updated_at=?
                   WHERE special_item_id=?""", (now, item["item_id"])
            )
        execute_write("DELETE FROM inventory_items WHERE player_id=? AND item_type='SPECIAL'", (player_id,))
        execute_write("UPDATE npc_profiles SET enabled=0,retired=1 WHERE player_id=?", (player_id,))
        execute_write("UPDATE players SET in_combat=0,is_banned=1 WHERE id=?", (player_id,))


def _finish_active_combat(player_id: int, session_id: int | None = None,
                          profile: dict | None = None) -> str:
    """Provide the internal finish active combat operation used by this module."""
    if session_id is None:
        row = execute_one(
            "SELECT id FROM combat_sessions WHERE attacker_player_id=? AND status='ACTIVE' ORDER BY id DESC LIMIT 1",
            (player_id,)
        )
        if not row:
            return "No active attacker combat found"
        session_id = row["id"]
    final = None
    for _ in range(20):
        action = _choose_combat_action(player_id, session_id, profile)
        final = enqueue_and_process(
            player_id, "combat_action", {"session_id": session_id, "action_type": action}
        )
        if any(entry.get("escaped") for entry in final.get("round_log", [])):
            return f"Combat {session_id} ended by escape"
        if final.get("combat_ended"):
            return f"Combat {session_id} completed ({final.get('winner_side')})"
        if final.get("at_round_limit"):
            enqueue_and_process(player_id, "combat_resolve", {"session_id": session_id})
            return f"Combat {session_id} resolved at round limit"
    return f"Combat {session_id} remains active after safety limit"


def _choose_combat_action(player_id: int, session_id: int, profile: dict | None) -> str:
    """Choose among the same Attack, Brace, Observe, and Escape actions as a player.

    Player ID supplies a stable temperament offset while the per-round roll
    prevents two otherwise identical NPCs from following the same script.
    """
    if not profile:
        return "attack"
    player = get_player(player_id)
    session_row = execute_one("SELECT * FROM combat_sessions WHERE id=?", (session_id,))
    if not player or not session_row:
        return "attack"

    temperament = random.Random(player_id * 104729)
    boldness = temperament.randint(-12, 12)
    curiosity = temperament.randint(-10, 10)
    caution = temperament.randint(-12, 12)
    hp_ratio = player["current_hp"] / max(1, player["max_hp"])
    aggression = max(0, min(100, profile["aggression"] + boldness))
    safety = max(0, min(100, profile["self_preservation"] + caution))

    weights = {"attack": 60.0, "brace": 5.0, "observe": 0.0, "escape": 0.0}
    if profile["player_hunter"] >= max(profile["boss_killer"], profile["hoarder"]):
        weights["attack"] += 20 + aggression * 0.2
        weights["brace"] += safety * 0.05
    if profile["boss_killer"] >= max(profile["player_hunter"], profile["hoarder"]):
        weights["observe"] += max(0, 35 + curiosity)
        weights["brace"] += 10 + safety * 0.08
    if profile["hoarder"] >= max(profile["player_hunter"], profile["boss_killer"]):
        weights["brace"] += 20 + safety * 0.15
        weights["escape"] += 10 + safety * 0.12

    # Observation is useful once per encounter; afterward its weight becomes 0.
    if session_row["attacker_observed"]:
        weights["observe"] = 0
    elif weights["observe"] == 0 and random.random() < 0.08:
        weights["observe"] = 8 + curiosity

    # Survival decisions intensify as HP falls, but aggression can keep a bold
    # character fighting longer. Escape remains subject to its normal AP cost.
    if hp_ratio < 0.55:
        weights["brace"] += (1 - hp_ratio) * safety
    if hp_ratio < 0.30:
        weights["escape"] += (0.30 - hp_ratio) * safety * 5
        weights["attack"] = max(5, weights["attack"] - safety * 0.35)
    settings = get_all_settings()
    if player["current_ap"] < settings.get("AP_COST_ESCAPE", cfg.AP_COST_ESCAPE):
        weights["escape"] = 0

    choices = [(action, max(0, weight + random.uniform(-8, 8)))
               for action, weight in weights.items()]
    total = sum(weight for _, weight in choices)
    if total <= 0:
        return "attack"
    pick = random.uniform(0, total)
    for action, weight in choices:
        pick -= weight
        if pick <= 0:
            return action
    return "attack"


def _finish_thief_combat(player_id: int, session_id: int | None = None) -> str:
    """Use only the normal Steal and Escape actions available to players."""
    if session_id is None:
        row = execute_one(
            "SELECT id FROM combat_sessions WHERE attacker_player_id=? AND status='ACTIVE' ORDER BY id DESC LIMIT 1",
            (player_id,)
        )
        if not row:
            return "No active thief combat found"
        session_id = row["id"]

    stole = False
    for _ in range(20):
        session_row = execute_one("SELECT status FROM combat_sessions WHERE id=?", (session_id,))
        if not session_row or session_row["status"] != "ACTIVE":
            return f"Combat {session_id} ended after {'a successful steal' if stole else 'the steal attempt'}"
        if not stole:
            result = enqueue_and_process(player_id, "combat_steal", {"session_id": session_id})
            steal_action = next((a for a in result.get("round_log", []) if a.get("action") == "STEAL"), {})
            stole = bool(steal_action.get("success"))
            if result.get("combat_ended"):
                return f"Combat {session_id} ended before escape"
        else:
            result = enqueue_and_process(
                player_id, "combat_action", {"session_id": session_id, "action_type": "escape"}
            )
            escape_action = next((a for a in result.get("round_log", []) if a.get("action") == "ESCAPE"), {})
            if escape_action.get("escaped"):
                return f"Stole successfully and escaped combat {session_id}"
            if result.get("combat_ended"):
                return f"Stole successfully but combat {session_id} ended before escape"
    return f"Combat {session_id} remains active after thief safety limit"


def _assign_pending_levelup(player_id: int, profile: dict):
    """Provide the internal assign pending levelup operation used by this module."""
    player = get_player(player_id)
    if not player or not player["pending_levelup"]:
        return
    if profile["thief"] >= max(profile["player_hunter"], profile["boss_killer"], profile["hoarder"]):
        priorities = ("agi", "lck", "per")
    elif profile["boss_killer"] >= max(profile["player_hunter"], profile["hoarder"]):
        priorities = ("str", "end", "agi")
    elif profile["hoarder"] >= profile["player_hunter"]:
        priorities = ("lck", "per", "end")
    else:
        priorities = ("agi", "per", "str")
    # Choose the currently lowest preferred stat, randomizing exact ties.
    lowest = min(player[f"{stat}_stat"] for stat in priorities)
    choices = [stat for stat in priorities if player[f"{stat}_stat"] == lowest]
    enqueue_and_process(player_id, "assign_levelup", {"stat": random.choice(choices)})


def _maybe_repair(player: dict, profile: dict):
    """Provide the internal maybe repair operation used by this module."""
    if random.randint(1, 100) > profile["repair_tendency"]:
        return None
    threshold = 40 + profile["repair_tendency"] // 2
    equipped = [player.get("equipped_weapon_id"), player.get("equipped_armor_id")]
    damaged = execute(
        "SELECT id,current_durability FROM inventory_items WHERE player_id=? AND id IN (?,?)",
        (player["id"], equipped[0] or -1, equipped[1] or -1)
    )
    if not any(item["current_durability"] < threshold for item in damaged):
        return None
    try:
        result = enqueue_and_process(player["id"], "blacksmith_repair", {"mode": "equipped", "inv_ids": []})
        return (f"Equipped gear fell below {threshold}% durability", str(result))
    except RuntimeError as exc:
        return ("Repair was desirable but unaffordable", str(exc))


def _maybe_heal(player: dict, profile: dict, settings: dict):
    """Provide the internal maybe heal operation used by this module."""
    hp_ratio = player["current_hp"] / max(1, player["max_hp"])
    threshold = 0.25 + (profile["self_preservation"] / 200)
    if hp_ratio >= threshold:
        return None
    cost_ap = settings.get("AP_COST_TAVERN", cfg.AP_COST_TAVERN)
    cost_cr = settings.get("TAVERN_HEAL_COST", cfg.TAVERN_HEAL_COST)
    if player["current_ap"] < cost_ap or player["credits"] < cost_cr:
        return None
    try:
        result = enqueue_and_process(player["id"], "tavern_heal", {
            "cost_ap": cost_ap, "cost_cr": cost_cr,
        })
        return (f"HP was below {int(threshold * 100)}% safety threshold", str(result))
    except RuntimeError as exc:
        return ("Healing was desirable but unavailable", str(exc))


def _maybe_hoard(player: dict, profile: dict):
    """Provide the internal maybe hoard operation used by this module."""
    if profile["hoarder"] <= 0:
        return None
    inv_count = execute_one("SELECT COUNT(*) cnt FROM inventory_items WHERE player_id=?", (player["id"],))["cnt"]
    if inv_count >= player["inventory_limit"]:
        cheapest = execute_one(
            """SELECT ii.id FROM inventory_items ii JOIN special_items si ON si.id=ii.item_id
               WHERE ii.player_id=? AND ii.item_type='SPECIAL' AND ii.id != COALESCE(?, -1)
               ORDER BY si.credit_cost ASC LIMIT 1""", (player["id"], player.get("equipped_special_id"))
        )
        if cheapest:
            result = enqueue_and_process(player["id"], "shop_sell", {"inv_id": cheapest["id"]})
            return ("Inventory was full; sold the cheapest unequipped special", str(result))
    listing = execute_one(
        """SELECT sl.id FROM shop_listings sl WHERE sl.item_type='SPECIAL' AND sl.price<=?
           ORDER BY sl.price ASC LIMIT 1""", (player["credits"],)
    )
    if listing and random.randint(1, 100) <= profile["hoarder"]:
        result = enqueue_and_process(player["id"], "shop_buy", {"listing_id": listing["id"]})
        return ("An affordable special item was available", str(result))
    return None


def _eligible_pvp_targets(player: dict) -> list[dict]:
    """Provide the internal eligible pvp targets operation used by this module."""
    return execute(
        """SELECT p.* FROM players p WHERE p.id != ? AND p.is_banned=0 AND p.in_combat=0
           AND p.level>=3 AND p.current_hp>1 AND (? - p.level)<=2""",
        (player["id"], player["level"])
    )


def _choose_pvp_target(player: dict, targets: list[dict], aggression: int) -> dict:
    """Choose using only eligibility and visible level information.

    NPCs must not exploit private credits, inventory, or exact HP values that a
    human challenger could not use when selecting an opponent.
    """
    if aggression >= 70:
        highest = max(p["level"] for p in targets)
        return random.choice([p for p in targets if p["level"] == highest])
    if aggression <= 30:
        lowest = min(p["level"] for p in targets)
        return random.choice([p for p in targets if p["level"] == lowest])
    nearest = min(abs(p["level"] - player["level"]) for p in targets)
    return random.choice([p for p in targets if abs(p["level"] - player["level"]) == nearest])


def _choose_boss(player: dict):
    """Provide the internal choose boss operation used by this module."""
    return execute_one(
        """SELECT b.* FROM bosses b LEFT JOIN boss_instances bi
           ON bi.boss_id=b.id AND bi.player_id=? WHERE b.is_active=1
           ORDER BY CASE WHEN COALESCE(bi.kill_count,0)=0 THEN 0 ELSE 1 END,
                    ABS(b.level-?) ASC, RANDOM() LIMIT 1""", (player["id"], player["level"])
    )


def _equip_best_core_items(player_id: int):
    """Provide the internal equip best core items operation used by this module."""
    updates = {}
    for item_type, table, field in (("WEAPON", "weapons", "equipped_weapon_id"),
                                    ("ARMOR", "armor", "equipped_armor_id")):
        item = execute_one(
            f"""SELECT ii.id FROM inventory_items ii JOIN {table} c ON c.id=ii.item_id
                WHERE ii.player_id=? AND ii.item_type=?
                ORDER BY c.level DESC,c.credit_cost DESC,ii.current_durability DESC LIMIT 1""",
            (player_id, item_type)
        )
        if item:
            updates[field] = item["id"]
    if updates:
        with exclusive_transaction():
            execute_write(
                "UPDATE players SET equipped_weapon_id=COALESCE(?,equipped_weapon_id),"
                "equipped_armor_id=COALESCE(?,equipped_armor_id) WHERE id=?",
                (updates.get("equipped_weapon_id"), updates.get("equipped_armor_id"), player_id)
            )


def _finish_turn(profile: dict, decision: str, reason: str, result) -> dict:
    """Provide the internal finish turn operation used by this module."""
    _assign_pending_levelup(profile["player_id"], profile)
    now = datetime.utcnow().isoformat()
    with exclusive_transaction():
        execute_write("UPDATE npc_profiles SET last_action_at=? WHERE player_id=?",
                      (now, profile["player_id"]))
    _log(profile["player_id"], decision, reason, str(result))
    return {"player_id": profile["player_id"], "decision": decision, "result": str(result)}


def _log(player_id: int, decision: str, reason: str, result: str):
    """Provide the internal log operation used by this module."""
    with exclusive_transaction():
        execute_write(
            "INSERT INTO npc_action_log(player_id,decision,reason,result) VALUES(?,?,?,?)",
            (player_id, decision, reason[:500], result[:1000])
        )

# FILE: queue_handler.py
"""Process auditable player actions synchronously through registered handlers."""
# queue_handler.py
# Synchronous action queue: writes a receipt to action_queue, processes inline
# inside an exclusive DB transaction, marks done or failed.
# On server restart, startup_cleanup() handles any orphaned PROCESSING rows.

import json
import logging
from datetime import datetime, timedelta

from database import execute, execute_one, execute_write, exclusive_transaction
import config_defaults as cfg

logger = logging.getLogger(__name__)

ACTION_HANDLERS: dict = {}


def register_handler(action_type: str):
    """Decorator to register an action handler function.

    Usage:
        @register_handler('tavern_heal')
        def handle_tavern_heal(player_id, payload):
            ...
    """
    def decorator(fn):
        """Handle the decorator workflow."""
        ACTION_HANDLERS[action_type] = fn
        return fn
    return decorator


def enqueue_and_process(player_id: int, action_type: str, payload: dict) -> dict:
    """Main entry point for all player write actions.
    Writes receipt, processes inline, marks done or failed."""
    if action_type not in ACTION_HANDLERS:
        raise ValueError(f"Unknown action_type: '{action_type}'")

    with exclusive_transaction():
        queue_id = execute_write(
            "INSERT INTO action_queue (player_id, action_type, payload, status) VALUES (?, ?, ?, 'PROCESSING')",
            (player_id, action_type, json.dumps(payload))
        )

    try:
        with exclusive_transaction():
            result = ACTION_HANDLERS[action_type](player_id, payload)

        with exclusive_transaction():
            execute_write(
                "UPDATE action_queue SET status = 'DONE', processed_at = ? WHERE id = ?",
                (datetime.utcnow().isoformat(), queue_id)
            )
            execute_write(
                """INSERT INTO player_activity_log
                   (player_id,category,action,status,message,details_json,queue_id,source)
                   VALUES(?, 'ACTION', ?, 'SUCCESS', ?, ?, ?, 'GAME')""",
                (player_id, action_type, f"{action_type} completed",
                 json.dumps(result, default=str)[:8000], queue_id)
            )
        return result

    except Exception as exc:
        try:
            with exclusive_transaction():
                execute_write(
                    "UPDATE action_queue SET status = 'FAILED', processed_at = ? WHERE id = ?",
                    (datetime.utcnow().isoformat(), queue_id)
                )
                execute_write(
                    """INSERT INTO player_activity_log
                       (player_id,category,action,status,message,details_json,queue_id,source)
                       VALUES(?, 'ERROR', ?, 'FAILED', ?, ?, ?, 'GAME')""",
                    (player_id, action_type, str(exc)[:1000],
                     json.dumps({"exception_type": type(exc).__name__}), queue_id)
                )
        except Exception:
            pass
        logger.exception("Action '%s' FAILED for player %d (queue_id=%d)", action_type, player_id, queue_id)
        raise RuntimeError(f"Action '{action_type}' failed: {exc}") from exc


def startup_cleanup():
    """Called once at app startup. Cleans up any PROCESSING rows from a prior crash.
    Refunds AP, clears in_combat, marks FAILED, logs to orphan log."""
    import sqlite3, os

    conn = sqlite3.connect(cfg.DB_PATH)
    conn.row_factory = lambda c, r: {col[0]: val for col, val in zip(c.description, r)}
    conn.execute("PRAGMA foreign_keys = ON")

    orphans = conn.execute("SELECT * FROM action_queue WHERE status = 'PROCESSING'").fetchall()
    if not orphans:
        conn.close()
        return

    logger.warning("startup_cleanup: %d orphaned actions found", len(orphans))
    os.makedirs(os.path.dirname(cfg.ORPHAN_LOG), exist_ok=True)

    with open(cfg.ORPHAN_LOG, "a") as log_file:
        for orphan in orphans:
            pid = orphan["player_id"]
            log_file.write(
                f"{datetime.utcnow().isoformat()} | ORPHAN | player={pid} "
                f"action={orphan['action_type']} queue_id={orphan['id']}\n"
            )
            ap_refund = _ap_cost_for_action(orphan["action_type"])
            conn.execute("BEGIN EXCLUSIVE")
            try:
                if ap_refund > 0:
                    conn.execute(
                        "UPDATE players SET current_ap = MIN(current_ap + ?, ?) WHERE id = ?",
                        (ap_refund, cfg.AP_CARRYOVER_CAP, pid)
                    )
                session = conn.execute(
                    """SELECT id, defender_player_id FROM combat_sessions
                       WHERE (attacker_player_id = ? OR defender_player_id = ?) AND status = 'ACTIVE'""",
                    (pid, pid)
                ).fetchone()
                if session:
                    conn.execute(
                        "UPDATE players SET in_combat = 0 WHERE id IN (?, ?)",
                        (pid, session["defender_player_id"] or pid)
                    )
                    conn.execute(
                        "UPDATE combat_sessions SET status = 'CANCELLED', result = 'CANCELLED' WHERE id = ?",
                        (session["id"],)
                    )
                conn.execute(
                    "UPDATE action_queue SET status = 'FAILED', processed_at = ? WHERE id = ?",
                    (datetime.utcnow().isoformat(), orphan["id"])
                )
                conn.execute("COMMIT")
            except Exception:
                conn.execute("ROLLBACK")
                logger.exception("startup_cleanup failed on queue_id=%d", orphan["id"])

    conn.close()
    logger.info("startup_cleanup: cleaned %d orphaned actions", len(orphans))


def _ap_cost_for_action(action_type: str) -> int:
    """Provide the internal ap cost for action operation used by this module."""
    costs = {
        "boss_fight": cfg.AP_COST_BOSS, "boss_confirm": cfg.AP_COST_BOSS,
        "pvp_start": cfg.AP_COST_PVP, "pvp_fight": cfg.AP_COST_PVP,
        "tavern_heal": cfg.AP_COST_TAVERN,
        "shop_buy": cfg.AP_COST_SHOP, "shop_sell": cfg.AP_COST_SHOP,
        "blacksmith_repair": cfg.AP_COST_BLACKSMITH,
    }
    return costs.get(action_type, 0)


def purge_old_done_rows():
    """Delete DONE rows older than 7 days. Called during midnight reset."""
    cutoff = (datetime.utcnow() - timedelta(days=7)).isoformat()
    with exclusive_transaction():
        deleted = execute_write(
            "DELETE FROM action_queue WHERE status = 'DONE' AND created_at < ?", (cutoff,)
        )
    logger.info("purge_old_done_rows: deleted %d rows", deleted)


################################################################################

# FILE: scheduler.py
"""Run UTC AP awards, feed archives, imports, recovery, and midnight maintenance."""
# scheduler.py  (Phase 7 — full implementation)
# Replaces the Phase 1 stub with complete midnight_reset and ap_trickle.

import math
import random
import logging
import os
from datetime import datetime

from database import execute, execute_one, execute_write, exclusive_transaction, get_all_settings
from queue_handler import purge_old_done_rows
import config_defaults as cfg

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# AP TRICKLE  (unchanged from Phase 1 — full from the start)
# ─────────────────────────────────────────────────────────────────────────────

def ap_trickle():
    """Award TRICKLE_AP_AMOUNT to all non-banned players, capped at AP_CARRYOVER_CAP.
    Runs at 03:00, 09:00, 15:00, 21:00 UTC daily."""
    settings = get_all_settings()
    trickle  = settings.get("TRICKLE_AP_AMOUNT", cfg.TRICKLE_AP_AMOUNT)
    cap      = settings.get("AP_CARRYOVER_CAP",  cfg.AP_CARRYOVER_CAP)

    with exclusive_transaction():
        updated = execute_write(
            "UPDATE players SET current_ap = MIN(current_ap + ?, ?) WHERE is_banned = 0",
            (trickle, cap)
        )
    logger.info("ap_trickle: +%d AP to %d players at %s",
                trickle, updated, datetime.utcnow().isoformat())


# ─────────────────────────────────────────────────────────────────────────────
# MIDNIGHT RESET  (full 12-step implementation)
# ─────────────────────────────────────────────────────────────────────────────

def midnight_reset():
    """Full UTC midnight reset sequence."""
    logger.info("=== MIDNIGHT RESET START %s ===", datetime.utcnow().isoformat())

    _step0_clear_status_effects()
    purge_old_done_rows()                # step 1
    _step2_apply_import()                # step 2
    _step3_archive_and_clear_feeds()     # step 3
    _step4_5_award_daily_ap()            # steps 4+5
    _step6_restore_midnight_hp()         # step 6
    _step7_midnight_encounters()         # step 7
    _step8_9_10_shop_rotation()          # steps 8-10
    _step11_pending_feed_entries()       # step 11

    logger.info("=== MIDNIGHT RESET COMPLETE %s ===", datetime.utcnow().isoformat())


# ─────────────────────────────────────────────────────────────────────────────
# STEP 0 — Clear all timed status effects
# ─────────────────────────────────────────────────────────────────────────────

def _step0_clear_status_effects():
    """Run the step0 clear status effects portion of scheduled maintenance."""
    with exclusive_transaction():
        deleted = execute_write("DELETE FROM status_effects")
    logger.info("step 0: cleared %d status effects", deleted)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Apply staged Excel import
# ─────────────────────────────────────────────────────────────────────────────

def _step2_apply_import():
    """Run the step2 apply import portion of scheduled maintenance."""
    if not os.path.exists(cfg.PENDING_IMPORT_PATH):
        logger.info("step 2: no pending import")
        return
    logger.info("step 2: applying staged import from %s", cfg.PENDING_IMPORT_PATH)
    from importer import run_import
    result = run_import(cfg.PENDING_IMPORT_PATH)
    if result["success"]:
        logger.info("step 2: import successful — %s", result["summary"])
    else:
        logger.error("step 2: import REJECTED — %s", result["errors"])


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Archive and clear daily feed
# ─────────────────────────────────────────────────────────────────────────────

def _step3_archive_and_clear_feeds():
    """Run the step3 archive and clear feeds portion of scheduled maintenance."""
    settings = get_all_settings()
    if settings.get("LOG_DAILY_ARCHIVE", cfg.LOG_DAILY_ARCHIVE):
        archive_feeds()

    with exclusive_transaction():
        deleted = execute_write("DELETE FROM daily_feed")
    logger.info("step 3: cleared %d daily feed entries", deleted)


def archive_feeds():
    """Export today's daily_feed to a timestamped text file."""
    os.makedirs(cfg.LOG_ARCHIVE_PATH, exist_ok=True)
    date_str  = datetime.utcnow().strftime("%Y_%m_%d")
    filepath  = os.path.join(cfg.LOG_ARCHIVE_PATH, f"game_log_{date_str}.txt")
    rows      = execute(
        "SELECT feed_scope, player_id, flavor_text, event_category, occurred_at "
        "FROM daily_feed ORDER BY occurred_at ASC"
    )
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"=== Daily Feed Archive — {date_str} UTC ===\n\n")
            for row in rows:
                scope = f"[{row['feed_scope']}]"
                pid   = f" player={row['player_id']}" if row["player_id"] else ""
                f.write(f"{row['occurred_at']} {scope}{pid} {row['flavor_text']}\n")
        logger.info("step 3: archived %d feed entries to %s", len(rows), filepath)
    except Exception as e:
        logger.exception("step 3: failed to archive feed: %s", e)


# ─────────────────────────────────────────────────────────────────────────────
# STEPS 4+5 — AP carryover + award daily AP
# ─────────────────────────────────────────────────────────────────────────────

def _step4_5_award_daily_ap():
    """Carryover is implicit (current_ap already holds remaining AP).
    Cap it at AP_CARRYOVER_CAP, then award new daily AP on top."""
    settings = get_all_settings()
    base_ap   = settings.get("BASE_DAILY_AP",      cfg.BASE_DAILY_AP)
    cap       = settings.get("AP_CARRYOVER_CAP",   cfg.AP_CARRYOVER_CAP)
    curse_red = settings.get("CURSE_AP_REDUCTION", cfg.CURSE_AP_REDUCTION)

    players = execute("SELECT id, end_stat FROM players WHERE is_banned = 0")
    cursed_ids = {
        r["player_id"] for r in execute(
            "SELECT player_id FROM status_effects WHERE effect_type = 'CURSED'"
        )
    }

    with exclusive_transaction():
        for p in players:
            daily_ap = base_ap + math.floor(p["end_stat"] / 2)
            if p["id"] in cursed_ids:
                daily_ap = int(daily_ap * (1 - curse_red))
            # Carryover cap first, then add daily AP, then cap again
            execute_write(
                "UPDATE players SET current_ap = MIN(MIN(current_ap, ?) + ?, ?) WHERE id = ?",
                (cap, daily_ap, cap, p["id"])
            )
    logger.info("steps 4+5: awarded daily AP to %d players (base=%d, cap=%d)",
                len(players), base_ap, cap)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Restore midnight HP
# ─────────────────────────────────────────────────────────────────────────────

def _step6_restore_midnight_hp():
    """Run the step6 restore midnight hp portion of scheduled maintenance."""
    settings  = get_all_settings()
    heal_pct  = settings.get("MIDNIGHT_HEAL_PERCENT", cfg.MIDNIGHT_HEAL_PERCENT)
    players   = execute(
        "SELECT id, current_hp, end_stat, level FROM players WHERE is_banned = 0"
    )
    with exclusive_transaction():
        for p in players:
            max_hp  = 10 + p["end_stat"] + (5 * p["level"])
            missing = max_hp - p["current_hp"]
            if missing > 0:
                restore = max(1, int(missing * heal_pct))
                execute_write(
                    "UPDATE players SET current_hp = MIN(current_hp + ?, ?) WHERE id = ?",
                    (restore, max_hp, p["id"])
                )
    logger.info("step 6: restored midnight HP for %d players", len(players))


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — Midnight random encounters
# ─────────────────────────────────────────────────────────────────────────────

def _step7_midnight_encounters():
    """Run a random event check for every active non-banned player."""
    from routes.actions import check_random_event
    settings = get_all_settings()
    players  = execute(
        """SELECT p.*, 0 as is_overencumbered FROM players p
           WHERE p.is_banned = 0 AND p.in_combat = 0"""
    )
    triggered = 0
    for p in players:
        # Temporarily set max_hp for the helper
        p["max_hp"]  = 10 + p["end_stat"] + (5 * p["level"])
        p["max_ap"]  = settings.get("BASE_DAILY_AP", cfg.BASE_DAILY_AP) + math.floor(p["end_stat"] / 2)
        event = check_random_event(p, settings)
        if event:
            triggered += 1
    logger.info("step 7: midnight encounters triggered for %d/%d players", triggered, len(players))


# ─────────────────────────────────────────────────────────────────────────────
# STEPS 8-10 — Shop rotation
# ─────────────────────────────────────────────────────────────────────────────

def _step8_9_10_shop_rotation():
    """Run the step8 9 10 shop rotation portion of scheduled maintenance."""
    settings      = get_all_settings()
    weapons_count = settings.get("SHOP_WEAPONS_COUNT", cfg.SHOP_WEAPONS_COUNT)
    armor_count   = settings.get("SHOP_ARMOR_COUNT",   cfg.SHOP_ARMOR_COUNT)

    with exclusive_transaction():
        # Step 8: Clear daily rotation listings
        execute_write("DELETE FROM shop_listings WHERE listing_source = 'DAILY_ROTATION'")

        # Step 9: Clear unsold special items from shop, return to loot pool
        unsold_specials = execute(
            "SELECT * FROM shop_listings WHERE item_type = 'SPECIAL'"
        )
        for s in unsold_specials:
            execute_write(
                """UPDATE special_item_registry
                   SET status = 'IN_POOL', current_owner_player_id = NULL,
                       inventory_item_id = NULL, shop_listing_price = NULL,
                       last_released_method = 'UNSOLD', updated_at = ?
                   WHERE special_item_id = ?""",
                (datetime.utcnow().isoformat(), s["item_id"])
            )
            execute_write("DELETE FROM shop_listings WHERE id = ?", (s["id"],))
            logger.info("step 9: returned special item id=%d to pool (unsold)", s["item_id"])

        # Populate new daily rotation — random selection weighted by drop_chance
        _populate_shop_rotation("weapons", weapons_count)
        _populate_shop_rotation("armor",   armor_count)

        # Step 10: Populate special item shop slots = floor(player_count / 2)
        player_count  = execute_one("SELECT COUNT(*) as cnt FROM players WHERE is_banned = 0")["cnt"]
        special_slots = max(0, player_count // 2)
        if special_slots > 0:
            _populate_special_slots(special_slots)

    logger.info("steps 8-10: shop rotated (%d weapons, %d armor, %d special slots)",
                weapons_count, armor_count, special_slots if player_count else 0)


def _populate_shop_rotation(table: str, count: int):
    """Select 'count' unique items from the content table and list them in the shop."""
    items = execute(
        f"SELECT * FROM {table} WHERE is_active = 1 ORDER BY RANDOM() LIMIT ?", (count * 3,)
    )
    # Weight by drop_chance
    weighted = []
    for item in items:
        w = max(1, int(item.get("drop_chance", 0.1) * 100))
        weighted.extend([item] * w)
    random.shuffle(weighted)
    seen   = set()
    chosen = []
    for item in weighted:
        if item["id"] not in seen:
            seen.add(item["id"])
            chosen.append(item)
        if len(chosen) >= count:
            break

    item_type = "WEAPON" if table == "weapons" else "ARMOR"
    for item in chosen:
        execute_write(
            """INSERT INTO shop_listings
               (item_type, item_id, listing_source, price)
               VALUES (?, ?, 'DAILY_ROTATION', ?)""",
            (item_type, item["id"], item["credit_cost"])
        )


def _populate_special_slots(slots: int):
    """Add up to 'slots' IN_POOL special items to the shop."""
    available = execute(
        """SELECT si.id, si.credit_cost
           FROM special_items si
           JOIN special_item_registry sir ON sir.special_item_id = si.id
           WHERE sir.status = 'IN_POOL' AND si.is_active = 1
           ORDER BY RANDOM()
           LIMIT ?""",
        (slots,)
    )
    for item in available:
        # Price special items significantly higher than their base cost
        price = int(item["credit_cost"] * 2.5)
        execute_write(
            """INSERT INTO shop_listings
               (item_type, item_id, listing_source, price)
               VALUES ('SPECIAL', ?, 'DAILY_ROTATION', ?)""",
            (item["id"], price)
        )
        execute_write(
            """UPDATE special_item_registry
               SET status = 'IN_SHOP', shop_listing_price = ?, updated_at = ?
               WHERE special_item_id = ?""",
            (price, datetime.utcnow().isoformat(), item["id"])
        )


# ─────────────────────────────────────────────────────────────────────────────
# STEP 11 — Process any pending feed entries
# ─────────────────────────────────────────────────────────────────────────────

def _step11_pending_feed_entries():
    """No deferred feed entries in current design — placeholder for future use."""
    logger.info("step 11: no pending feed entries")

# FILE: combat/actions.py
"""Apply combat actions and post-combat rewards to persistent game state."""
# combat/actions.py
# Per-action handlers for all 6 combat actions plus opponent automation.
# Each handler resolves the action, writes to DB, and returns a result dict.
# All DB writes happen inside exclusive_transaction() via queue_handler.

import math
import random
import logging
from datetime import datetime

from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_all_settings)
from combat import engine
from combat import flavour
import config_defaults as cfg

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# COMBAT STATE LOADER
# ─────────────────────────────────────────────────────────────────────────────

def get_combat_state(session_id: int) -> dict:
    """Load the full combat state for a given session.
    Returns dict with session, attacker, defender/boss/minion, equipped gear,
    active combat_buffs, and derived values."""
    session = execute_one(
        "SELECT * FROM combat_sessions WHERE id = ?", (session_id,)
    )
    if session is None:
        raise ValueError(f"Combat session {session_id} not found.")

    attacker = execute_one(
        "SELECT * FROM players WHERE id = ?", (session["attacker_player_id"],)
    )
    attacker_equipped = _load_equipped(attacker)

    # Load defender (player, boss, or minion)
    defender = defender_equipped = boss = minion = None

    if session["combat_type"] == "PVP":
        defender = execute_one(
            "SELECT * FROM players WHERE id = ?", (session["defender_player_id"],)
        )
        defender_equipped = _load_equipped(defender)

    elif session["combat_type"] == "BOSS":
        instance = execute_one(
            "SELECT * FROM boss_instances WHERE id = ?", (session["boss_instance_id"],)
        )
        boss = execute_one("SELECT * FROM bosses WHERE id = ?", (instance["boss_id"],))
        boss = {**boss,
                "current_hp":          instance["current_hp"],
                "special_attack_used": instance["special_attack_used"],
                "special_buff_used":   instance["special_buff_used"],
                "current_phase":       instance["current_phase"],
                "instance_id":         instance["id"]}

    elif session["combat_type"] == "MINION":
        instance = execute_one(
            "SELECT * FROM minion_instances WHERE id = ?", (session["minion_instance_id"],)
        )
        minion = execute_one("SELECT * FROM minions WHERE id = ?", (instance["minion_id"],))
        minion = {**minion,
                  "current_hp": instance["current_hp"],
                  "instance_id": instance["id"]}

    # Active combat buffs
    buffs = execute(
        "SELECT * FROM combat_buffs WHERE combat_session_id = ?", (session_id,)
    )
    attacker_buffs = [b for b in buffs if b["side"] == "ATTACKER"]
    defender_buffs = [b for b in buffs if b["side"] == "DEFENDER"]

    return {
        "session":            session,
        "attacker":           attacker,
        "attacker_equipped":  attacker_equipped,
        "defender":           defender,
        "defender_equipped":  defender_equipped,
        "boss":               boss,
        "minion":             minion,
        "attacker_buffs":     attacker_buffs,
        "defender_buffs":     defender_buffs,
    }


def _load_equipped(player: dict) -> dict:
    """Load weapon, armor, and special item rows for a player."""
    result = {"weapon": None, "armor": None, "special": None}
    if player is None:
        return result
    for slot, col, table in [
        ("weapon",  "equipped_weapon_id",  "weapons"),
        ("armor",   "equipped_armor_id",   "armor"),
        ("special", "equipped_special_id", "special_items"),
    ]:
        inv_id = player.get(col)
        if inv_id:
            inv = execute_one("SELECT * FROM inventory_items WHERE id = ?", (inv_id,))
            if inv:
                item = execute_one(f"SELECT * FROM {table} WHERE id = ?", (inv["item_id"],))
                if item:
                    result[slot] = {**item, "inv_id": inv_id,
                                    "current_durability": inv["current_durability"]}
    return result


def check_combat_end(state: dict) -> tuple[bool, str | None]:
    """Check if the fight should end. Returns (ended, winner_side).
    winner_side: 'ATTACKER', 'DEFENDER', or None."""
    session = state["session"]

    if session["combat_type"] == "PVP":
        # Check current HP from DB (may have changed mid-round)
        att = execute_one(
            "SELECT current_hp FROM players WHERE id = ?",
            (session["attacker_player_id"],)
        )
        dfn = execute_one(
            "SELECT current_hp FROM players WHERE id = ?",
            (session["defender_player_id"],)
        )
        if att["current_hp"] <= 1:
            return True, "DEFENDER"
        if dfn["current_hp"] <= 1:
            return True, "ATTACKER"

    elif session["combat_type"] in ("BOSS", "MINION"):
        table = "boss_instances" if session["combat_type"] == "BOSS" else "minion_instances"
        id_col = "boss_instance_id" if session["combat_type"] == "BOSS" else "minion_instance_id"
        inst = execute_one(
            f"SELECT current_hp FROM {table} WHERE id = ?", (session[id_col],)
        )
        if inst["current_hp"] <= 0:
            return True, "ATTACKER"
        # Player floor: 1 HP
        att = execute_one(
            "SELECT current_hp FROM players WHERE id = ?",
            (session["attacker_player_id"],)
        )
        if att["current_hp"] <= 1:
            return True, "DEFENDER"  # Boss/minion wins

    return False, None


# ─────────────────────────────────────────────────────────────────────────────
# ACTION: ATTACK
# ─────────────────────────────────────────────────────────────────────────────

def handle_attack(session_id: int, actor_side: str, state: dict) -> dict:
    """Resolve a full attack from one side.
    actor_side: 'ATTACKER' or 'DEFENDER'"""
    session = state["session"]
    is_attacker = actor_side == "ATTACKER"

    attacker    = state["attacker"] if is_attacker else state["defender"]
    defender    = state["defender"] if is_attacker else state["attacker"]
    att_eq      = state["attacker_equipped"] if is_attacker else state["defender_equipped"]
    def_eq      = state["defender_equipped"] if is_attacker else state["attacker_equipped"]
    att_buffs   = state["attacker_buffs"] if is_attacker else state["defender_buffs"]
    def_buffs   = state["defender_buffs"] if is_attacker else state["attacker_buffs"]
    boss        = state["boss"] if session["combat_type"] == "BOSS" else None
    minion      = state["minion"] if session["combat_type"] == "MINION" else None
    opponent    = boss or minion or defender

    weapon  = att_eq.get("weapon")
    armor   = def_eq.get("armor") if def_eq else None
    special = att_eq.get("special")
    def_special = def_eq.get("special") if def_eq else None

    if weapon is None:
        # Unarmed: d4 Blunt, no stat bonus weapon
        weapon = {"weapon_type": "Melee", "damage_die": "d4",
                  "damage_type": "Blunt", "name": "Fists",
                  "str_bonus": 0, "credit_cost": 0}

    # Brace dodge bonus for defender
    brace_dodge = sum(
        int(b["value"]) for b in def_buffs
        if b["buff_type"] == "BRACE_DODGE_BONUS"
    )

    result = engine.resolve_full_attack(
        attacker=attacker,
        defender=opponent,
        attacker_weapon=weapon,
        attacker_special=special,
        defender_armor=armor,
        defender_special=def_special,
        boss=boss,
        brace_dodge_bonus=brace_dodge,
        active_buffs=def_buffs,
        is_player_attacker=is_attacker,
    )

    # Extra attack (special item)
    extra_attack_result = None
    if special and special.get("extra_attack") and result["hit"]:
        extra_attack_result = engine.resolve_full_attack(
            attacker=attacker,
            defender=opponent,
            attacker_weapon=weapon,
            attacker_special=special,
            defender_armor=armor,
            defender_special=def_special,
            boss=boss,
            brace_dodge_bonus=brace_dodge,
            active_buffs=def_buffs,
            is_player_attacker=is_attacker,
        )

    # --- Write to DB ---
    with exclusive_transaction():
        damage_total = result["damage_total"]
        if extra_attack_result:
            damage_total += extra_attack_result["damage_total"]

        # Update HP
        if session["combat_type"] == "PVP":
            target_id = session["defender_player_id"] if is_attacker else session["attacker_player_id"]
            if is_attacker:
                current = execute_one("SELECT current_hp FROM players WHERE id = ?", (target_id,))
                new_hp  = max(1, current["current_hp"] - damage_total)  # PvP floor: 1 HP
            else:
                current = execute_one("SELECT current_hp FROM players WHERE id = ?", (target_id,))
                new_hp  = max(1, current["current_hp"] - damage_total)
            execute_write("UPDATE players SET current_hp = ? WHERE id = ?", (new_hp, target_id))
        else:
            # Boss or minion HP: floor 0
            inst_id  = (session["boss_instance_id"] if session["combat_type"] == "BOSS"
                        else session["minion_instance_id"])
            tbl      = "boss_instances" if session["combat_type"] == "BOSS" else "minion_instances"
            current  = execute_one(f"SELECT current_hp FROM {tbl} WHERE id = ?", (inst_id,))
            new_hp   = max(0, current["current_hp"] - damage_total)
            execute_write(f"UPDATE {tbl} SET current_hp = ? WHERE id = ?", (new_hp, inst_id))

        # Update damage totals on session
        if is_attacker:
            execute_write(
                "UPDATE combat_sessions SET attacker_total_damage_dealt = attacker_total_damage_dealt + ? WHERE id = ?",
                (damage_total, session_id)
            )
        else:
            execute_write(
                "UPDATE combat_sessions SET defender_total_damage_dealt = defender_total_damage_dealt + ? WHERE id = ?",
                (damage_total, session_id)
            )

        # Weapon durability (attacker)
        if result["hit"] and att_eq.get("weapon"):
            _apply_durability_loss(att_eq["weapon"]["inv_id"],
                                   result["weapon_durability_loss"],
                                   session["attacker_player_id"])
            if extra_attack_result and extra_attack_result["hit"]:
                _apply_durability_loss(att_eq["weapon"]["inv_id"],
                                       extra_attack_result["weapon_durability_loss"],
                                       session["attacker_player_id"])

        # Armor durability (defender — only for PvP)
        if result["hit"] and session["combat_type"] == "PVP" and def_eq and def_eq.get("armor"):
            def_player_id = (session["defender_player_id"] if is_attacker
                             else session["attacker_player_id"])
            _apply_durability_loss(def_eq["armor"]["inv_id"],
                                   result["armor_durability_loss"],
                                   def_player_id)

        # Expire BRACE buffs after defender is hit
        if result["hit"] or result["dodged"] is False:
            execute_write(
                """DELETE FROM combat_buffs
                   WHERE combat_session_id = ? AND side = ? AND expires_on = 'NEXT_HIT_RESOLVED'""",
                (session_id, "ATTACKER" if not is_attacker else "DEFENDER")
            )

        # Write combat log
        execute_write(
            """INSERT INTO combat_logs
               (combat_session_id, round_number, actor, action_type, roll_detail, outcome_detail,
                hp_after_attacker, hp_after_defender)
               VALUES (?, ?, ?, 'ATTACK', ?, ?, ?, ?)""",
            (session_id, session["current_round"], actor_side,
             result["roll_detail"], result["outcome_detail"],
             None, None)  # HP values filled in by caller
        )

    # Build flavor text
    weapon_name = weapon.get("name", "weapon")
    flavor = flavour.attack_flavor(
        attacker_name=attacker.get("character_name", "Attacker"),
        weapon_name=weapon_name,
        weapon_type=weapon.get("weapon_type", "Melee"),
        hit=result["hit"],
        dodged=result["dodged"],
        is_crit=result["is_crit"],
        damage=damage_total,
        damage_type=weapon.get("damage_type", "Blunt"),
        res_note=result["damage_breakdown"][0]["note"] if result["damage_breakdown"] else ""
    )

    return {
        "action":         "ATTACK",
        "hit":            result["hit"],
        "dodged":         result["dodged"],
        "damage_total":   damage_total,
        "is_crit":        result["is_crit"],
        "new_target_hp":  new_hp,
        "roll_detail":    result["roll_detail"],
        "flavor":         flavor,
        "extra_attack":   extra_attack_result is not None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# ACTION: STEAL
# ─────────────────────────────────────────────────────────────────────────────

def handle_steal(session_id: int, player_id: int, state: dict) -> dict:
    """Resolve a steal attempt by either PvP side using that actor's stats."""
    session  = state["session"]
    settings = get_all_settings()
    is_defender = (session["combat_type"] == "PVP" and
                   player_id == session["defender_player_id"])
    actor = state["defender"] if is_defender else state["attacker"]
    actor_side = "DEFENDER" if is_defender else "ATTACKER"
    actor_equipped = state["defender_equipped"] if is_defender else state["attacker_equipped"]

    special     = actor_equipped.get("special")
    steal_bonus = special.get("steal_bonus", 0.0) if special else 0.0

    if session["combat_type"] == "PVP":
        defender = state["attacker"] if is_defender else state["defender"]
        roll_result = engine.resolve_opposed_roll(
            actor_agi=actor["agi_stat"],
            actor_lck=actor["lck_stat"],
            defender_agi=defender["agi_stat"],
            defender_lck=defender["lck_stat"],
            steal_bonus_pct=steal_bonus,
            tie_goes_to="defender"
        )
        if not roll_result["success"]:
            _apply_steal_fail_penalty(session_id, actor_side)
            return {"action": "STEAL", "success": False,
                    "roll_detail": roll_result["detail"],
                    "flavor": flavour.steal_flavor(
                        actor["character_name"], defender["character_name"], False)}

        # Cascade: item → credits → XP
        result = _pvp_steal_cascade(player_id, defender["id"],
                                    steal_bonus, settings)
        _write_combat_log(session_id, session["current_round"], actor_side,
                          "STEAL", roll_result["detail"], str(result))
        return {"action": "STEAL", "success": True,
                "roll_detail": roll_result["detail"],
                "flavor": flavour.steal_flavor(
                    actor["character_name"], defender["character_name"], True,
                    item_name=result.get("item_name", ""),
                    credits=result.get("credits", 0),
                    xp_bonus=result.get("xp_bonus", 0))}

    else:
        # vs boss or minion
        opponent = state["boss"] or state["minion"]
        opp_agi  = opponent["agi_stat"]
        opp_lck  = opponent["lck_stat"]
        roll_result = engine.resolve_opposed_roll(
            actor_agi=actor["agi_stat"],
            actor_lck=actor["lck_stat"],
            defender_agi=opp_agi,
            defender_lck=opp_lck,
            steal_bonus_pct=steal_bonus,
            tie_goes_to="defender"
        )
        if not roll_result["success"]:
            _apply_steal_fail_penalty(session_id, "ATTACKER")
            return {"action": "STEAL", "success": False,
                    "roll_detail": roll_result["detail"],
                    "flavor": flavour.steal_flavor(
                        actor["character_name"], opponent["name"], False,
                        is_vs_boss=True)}

        result = _boss_steal_result(player_id, opponent, steal_bonus, settings,
                                    session["combat_type"])
        _write_combat_log(session_id, session["current_round"], "ATTACKER",
                          "STEAL", roll_result["detail"], str(result))
        return {"action": "STEAL", "success": True,
                "roll_detail": roll_result["detail"],
                "flavor": flavour.steal_flavor(
                    actor["character_name"], opponent["name"], True,
                    item_name=result.get("item_name", ""),
                    credits=result.get("credits", 0),
                    is_vs_boss=True)}


def _pvp_steal_cascade(attacker_id: int, defender_id: int,
                       steal_bonus: float, settings: dict) -> dict:
    """Provide the internal pvp steal cascade operation used by this module."""
    steal_cr_pct  = settings.get("STEAL_ACTION_CREDIT_PERCENT", cfg.STEAL_ACTION_CREDIT_PERCENT)
    zero_xp_bonus = settings.get("ZERO_CREDIT_XP_BONUS",        cfg.ZERO_CREDIT_XP_BONUS)

    # Step 1: try to steal a random unequipped item
    defender = execute_one("SELECT * FROM players WHERE id = ?", (defender_id,))
    equipped  = {defender.get("equipped_weapon_id"),
                 defender.get("equipped_armor_id"),
                 defender.get("equipped_special_id")} - {None}
    inv_items = execute(
        "SELECT * FROM inventory_items WHERE player_id = ?", (defender_id,)
    )
    stealable = [i for i in inv_items if i["id"] not in equipped]

    if stealable:
        target_inv = random.choice(stealable)
        # Transfer to attacker
        with exclusive_transaction():
            execute_write(
                "UPDATE inventory_items SET player_id = ?, acquired_method = 'PVP_STEAL' WHERE id = ?",
                (attacker_id, target_inv["id"])
            )
            item_detail = execute_one(
                f"SELECT name FROM {'weapons' if target_inv['item_type']=='WEAPON' else 'armor' if target_inv['item_type']=='ARMOR' else 'special_items'} WHERE id = ?",
                (target_inv["item_id"],)
            )
            item_name = item_detail["name"] if item_detail else "item"
            execute_write(
                """INSERT INTO item_history (player_id, item_type, item_id, item_name, event_type, related_player_id)
                   VALUES (?, ?, ?, ?, 'STOLEN_BY_ME', ?)""",
                (attacker_id, target_inv["item_type"], target_inv["item_id"], item_name, defender_id)
            )
            execute_write(
                """INSERT INTO item_history (player_id, item_type, item_id, item_name, event_type, related_player_id)
                   VALUES (?, ?, ?, ?, 'STOLEN_FROM_ME', ?)""",
                (defender_id, target_inv["item_type"], target_inv["item_id"], item_name, attacker_id)
            )
        return {"item_name": item_name, "credits": 0, "xp_bonus": 0}

    # Step 2: try to steal credits
    if defender["credits"] > 0:
        steal_pct = steal_cr_pct + steal_bonus
        amount    = max(0, int(defender["credits"] * steal_pct))
        if amount > 0:
            with exclusive_transaction():
                execute_write("UPDATE players SET credits = credits - ? WHERE id = ?",
                              (amount, defender_id))
                execute_write("UPDATE players SET credits = credits + ? WHERE id = ?",
                              (amount, attacker_id))
            return {"credits": amount, "xp_bonus": 0}

    # Step 3: nothing left — XP consolation
    with exclusive_transaction():
        execute_write("UPDATE players SET xp = xp + ? WHERE id = ?",
                      (zero_xp_bonus, attacker_id))
    return {"xp_bonus": zero_xp_bonus}


def _boss_steal_result(player_id, opponent, steal_bonus, settings, combat_type):
    """Provide the internal boss steal result operation used by this module."""
    import random, math
    from datetime import datetime

    base_chance   = settings.get("STEAL_SPECIAL_BASE_CHANCE", cfg.STEAL_SPECIAL_BASE_CHANCE)
    cr_multiplier = settings.get("STEAL_BOSS_CREDIT_MULTIPLIER", cfg.STEAL_BOSS_CREDIT_MULTIPLIER)
    player        = execute_one("SELECT lck_stat FROM players WHERE id = ?", (player_id,))
    lck_bonus     = math.floor(player["lck_stat"] / 2) / 100

    # Try special item first
    if random.random() < (base_chance + lck_bonus):
        association_type = "Boss" if combat_type == "BOSS" else "Minion"
        special_def = execute_one(
            """SELECT si.id, si.name, si.starting_durability
               FROM special_items si
               JOIN special_item_registry sir ON sir.special_item_id = si.id
               WHERE si.associated_to = ? AND si.association_type = ?
                 AND sir.status = 'IN_POOL'""",
            (opponent["name"], association_type)
        )
        if special_def:
            with exclusive_transaction():
                inv_id = execute_write(
                    """INSERT INTO inventory_items
                       (player_id, item_type, item_id, current_durability, acquired_method)
                       VALUES (?, 'SPECIAL', ?, ?, 'COMBAT_STEAL')""",
                    (player_id, special_def["id"], special_def.get("starting_durability", 100))
                )
                execute_write(
                    """UPDATE special_item_registry
                       SET status='IN_INVENTORY', current_owner_player_id=?,
                           inventory_item_id=?, last_acquired_method='COMBAT_STEAL', updated_at=?
                       WHERE special_item_id=?""",
                    (player_id, inv_id, datetime.utcnow().isoformat(), special_def["id"])
                )
                execute_write(
                    """INSERT INTO item_history
                       (player_id, item_type, item_id, item_name, event_type)
                       VALUES (?, 'SPECIAL', ?, ?, 'RECEIVED_COMBAT_STEAL')""",
                    (player_id, special_def["id"], special_def["name"])
                )
            return {"item_name": special_def["name"]}

    # Minion only: try to steal the minion weapon
    if combat_type == "MINION":
        master = execute_one(
            """SELECT m.minion_weapon_id, w.name as weapon_name, w.starting_durability
               FROM master m
               JOIN minions mn ON mn.id = m.minion_id
               JOIN weapons w  ON w.id  = m.minion_weapon_id
               WHERE mn.name = ?""",
            (opponent["name"],)
        )
        if master and master["minion_weapon_id"]:
            already_owned = execute_one(
                "SELECT id FROM inventory_items WHERE player_id = ? AND item_type = 'WEAPON' AND item_id = ?",
                (player_id, master["minion_weapon_id"])
            )
            if not already_owned:
                with exclusive_transaction():
                    execute_write(
                        """INSERT INTO inventory_items
                           (player_id, item_type, item_id, current_durability, acquired_method)
                           VALUES (?, 'WEAPON', ?, ?, 'COMBAT_STEAL')""",
                        (player_id, master["minion_weapon_id"], master.get("starting_durability", 100))
                    )
                    execute_write(
                        """INSERT INTO item_history
                           (player_id, item_type, item_id, item_name, event_type)
                           VALUES (?, 'WEAPON', ?, ?, 'RECEIVED_COMBAT_STEAL')""",
                        (player_id, master["minion_weapon_id"], master["weapon_name"])
                    )
                return {"item_name": master["weapon_name"]}

    # Credits fallback
    credits_stolen = int(opponent["level"] * (cr_multiplier + steal_bonus * cr_multiplier))
    with exclusive_transaction():
        execute_write("UPDATE players SET credits = credits + ? WHERE id = ?",
                      (credits_stolen, player_id))
    return {"credits": credits_stolen}

def handle_brace(session_id: int, player_id: int, state: dict) -> dict:
    """Process the queued brace action against validated game state."""
    session  = state["session"]
    attacker = state["attacker"]
    settings = get_all_settings()
    heal_pct    = settings.get("BRACE_HEAL_PERCENT",      cfg.BRACE_HEAL_PERCENT)
    ac_pct      = settings.get("BRACE_AC_BONUS_PERCENT",  cfg.BRACE_AC_BONUS_PERCENT)
    dodge_bonus = settings.get("BRACE_DODGE_BONUS",       cfg.BRACE_DODGE_BONUS)

    armor    = state["attacker_equipped"].get("armor")
    current_ac = engine.calc_ac(attacker, armor)
    ac_bonus   = int(current_ac * ac_pct)

    max_hp   = engine.calc_max_hp(attacker)
    missing  = max_hp - attacker["current_hp"]
    heal     = max(0, int(missing * heal_pct))
    new_hp   = min(attacker["current_hp"] + heal, max_hp)

    with exclusive_transaction():
        execute_write("UPDATE players SET current_hp = ? WHERE id = ?", (new_hp, player_id))
        execute_write(
            """INSERT INTO combat_buffs
               (combat_session_id, side, buff_type, value, expires_on)
               VALUES (?, 'ATTACKER', 'BRACE_AC_BONUS', ?, 'NEXT_HIT_RESOLVED')""",
            (session_id, ac_bonus)
        )
        execute_write(
            """INSERT INTO combat_buffs
               (combat_session_id, side, buff_type, value, expires_on)
               VALUES (?, 'ATTACKER', 'BRACE_DODGE_BONUS', ?, 'NEXT_HIT_RESOLVED')""",
            (session_id, dodge_bonus)
        )
        # Armor durability loss on Brace
        if armor:
            _apply_durability_loss(armor["inv_id"], 1, player_id)
        _write_combat_log(session_id, session["current_round"], "ATTACKER",
                          "BRACE", "Brace action",
                          f"Healed {heal} HP, AC+{ac_bonus}, Dodge+{dodge_bonus}")

    flv = flavour.brace_flavor(attacker["character_name"], heal, ac_bonus, dodge_bonus)
    return {"action": "BRACE", "new_hp": new_hp, "ac_bonus": ac_bonus,
            "dodge_bonus": dodge_bonus, "heal": heal, "flavor": flv}


# ─────────────────────────────────────────────────────────────────────────────
# ACTION: ESCAPE
# ─────────────────────────────────────────────────────────────────────────────

def handle_escape(session_id: int, player_id: int, state: dict) -> dict:
    """Process the queued escape action against validated game state."""
    session  = state["session"]
    attacker = state["attacker"]
    settings = get_all_settings()
    ap_cost  = settings.get("AP_COST_ESCAPE", cfg.AP_COST_ESCAPE)
    cr_drop  = settings.get("ESCAPE_CREDIT_DROP_CHANCE", cfg.ESCAPE_CREDIT_DROP_CHANCE)

    if attacker["current_ap"] < ap_cost:
        raise ValueError(f"Not enough AP to attempt escape (need {ap_cost}).")

    # Determine opponent stats for roll
    if session["combat_type"] == "PVP":
        opp = state["defender"]
        opp_agi, opp_lck = opp["agi_stat"], opp["lck_stat"]
    else:
        opp = state["boss"] or state["minion"]
        opp_agi, opp_lck = opp["agi_stat"], opp["lck_stat"]

    roll_result = engine.resolve_opposed_roll(
        actor_agi=attacker["agi_stat"],
        actor_lck=attacker["lck_stat"],
        defender_agi=opp_agi,
        defender_lck=opp_lck,
        tie_goes_to="defender"
    )

    credits_lost = 0
    with exclusive_transaction():
        execute_write(
            "UPDATE players SET current_ap = current_ap - ? WHERE id = ?",
            (ap_cost, player_id)
        )
        if roll_result["success"]:
            # Escape: cancel session
            execute_write(
                "UPDATE combat_sessions SET status='RESOLVED', result='ESCAPE', resolved_at=? WHERE id=?",
                (datetime.utcnow().isoformat(), session_id)
            )
            execute_write(
                "UPDATE players SET in_combat = 0 WHERE id = ?", (player_id,)
            )
            if session["combat_type"] == "PVP":
                execute_write(
                    "UPDATE players SET in_combat = 0 WHERE id = ?",
                    (session["defender_player_id"],)
                )
            # Reset boss/minion HP on escape
            if session["combat_type"] == "BOSS":
                boss = state["boss"]
                execute_write(
                    "UPDATE boss_instances SET current_hp=?, special_attack_used=0, special_buff_used=0, current_phase=1 WHERE id=?",
                    (boss["max_hp"], boss["instance_id"])
                )
            elif session["combat_type"] == "MINION":
                minion = state["minion"]
                execute_write(
                    "UPDATE minion_instances SET current_hp=? WHERE id=?",
                    (minion["max_hp"], minion["instance_id"])
                )
            # PvP credit drop
            if session["combat_type"] == "PVP" and random.random() < cr_drop:
                player_row = execute_one("SELECT credits FROM players WHERE id=?", (player_id,))
                if player_row["credits"] > 0:
                    credits_lost = max(1, int(player_row["credits"] * 0.05))
                    execute_write(
                        "UPDATE players SET credits = credits - ? WHERE id = ?",
                        (credits_lost, player_id)
                    )
                    execute_write(
                        "UPDATE players SET credits = credits + ? WHERE id = ?",
                        (credits_lost, session["defender_player_id"])
                    )
            # Delete combat buffs
            execute_write("DELETE FROM combat_buffs WHERE combat_session_id = ?", (session_id,))
        else:
            # Fail: AC penalty
            execute_write(
                """INSERT INTO combat_buffs
                   (combat_session_id, side, buff_type, value, expires_on)
                   VALUES (?, 'ATTACKER', 'ESCAPE_FAIL_AC_PENALTY', 3, 'NEXT_HIT_RESOLVED')""",
                (session_id,)
            )
        _write_combat_log(session_id, session["current_round"], "ATTACKER",
                          "ESCAPE", roll_result["detail"],
                          f"{'Escaped' if roll_result['success'] else 'Failed'}, credits lost: {credits_lost}")

    flv = flavour.escape_flavor(attacker["character_name"],
                                roll_result["success"], credits_lost)
    return {"action": "ESCAPE", "success": roll_result["success"],
            "credits_lost": credits_lost, "roll_detail": roll_result["detail"],
            "flavor": flv, "escaped": roll_result["success"]}


# ─────────────────────────────────────────────────────────────────────────────
# ACTION: OBSERVE
# ─────────────────────────────────────────────────────────────────────────────

def handle_observe(session_id: int, player_id: int, state: dict) -> dict:
    """Process the queued observe action against validated game state."""
    session  = state["session"]
    attacker = state["attacker"]

    if session["combat_type"] == "PVP":
        opp = state["defender"]
        opp_per, opp_agi, opp_lck = opp["per_stat"], opp["agi_stat"], opp["lck_stat"]
    else:
        opp = state["boss"] or state["minion"]
        opp_per, opp_agi, opp_lck = opp.get("per_stat", 0), opp["agi_stat"], opp["lck_stat"]

    roll_result = engine.resolve_opposed_roll(
        actor_agi=attacker["agi_stat"], actor_lck=attacker["lck_stat"],
        defender_agi=opp_agi, defender_lck=opp_lck,
        actor_per=attacker["per_stat"], defender_per=opp_per,
        tie_goes_to="defender"
    )

    revealed = {}
    with exclusive_transaction():
        if roll_result["success"]:
            execute_write(
                "UPDATE combat_sessions SET attacker_observed = 1 WHERE id = ?",
                (session_id,)
            )
            if session["combat_type"] == "BOSS":
                boss = state["boss"]
                resistances = [t for t in engine.DAMAGE_TYPES if boss.get(f"res_{t}")]
                weaknesses  = [t for t in engine.DAMAGE_TYPES if boss.get(f"weak_{t}")]
                revealed = {
                    "resistances": resistances,
                    "weaknesses":  weaknesses,
                    "exact_hp":    boss["current_hp"],
                }
                # Save to boss_intel permanently
                execute_write(
                    """INSERT OR IGNORE INTO boss_intel (player_id, boss_id)
                       SELECT ?, boss_id FROM boss_instances WHERE id = ?""",
                    (player_id, session["boss_instance_id"])
                )
            elif session["combat_type"] == "MINION":
                revealed = {"exact_hp": (state["minion"] or {}).get("current_hp")}
            else:
                # PvP: reveal equipped gear
                def_eq = state["defender_equipped"]
                revealed = {
                    "weapon": def_eq["weapon"]["name"] if def_eq.get("weapon") else None,
                    "armor":  def_eq["armor"]["name"]  if def_eq.get("armor")  else None,
                }
        _write_combat_log(session_id, session["current_round"], "ATTACKER",
                          "OBSERVE", roll_result["detail"], str(revealed))

    flv = flavour.observe_flavor(
        attacker["character_name"], roll_result["success"],
        opp.get("character_name") or opp.get("name", "opponent"), revealed
    )
    return {"action": "OBSERVE", "success": roll_result["success"],
            "revealed": revealed, "roll_detail": roll_result["detail"], "flavor": flv}


# ─────────────────────────────────────────────────────────────────────────────
# ACTION: SWAP GEAR
# ─────────────────────────────────────────────────────────────────────────────

def handle_swap_gear(session_id: int, player_id: int, state: dict,
                     new_weapon_inv_id: int | None = None,
                     new_armor_inv_id:  int | None = None,
                     new_special_inv_id: int | None = None) -> dict:
    """Process the queued swap gear action against validated game state."""
    session  = state["session"]
    attacker = state["attacker"]
    settings = get_all_settings()
    acc_pen  = settings.get("SWAP_GEAR_ACCURACY_PENALTY", cfg.SWAP_GEAR_ACCURACY_PENALTY)
    ac_pen   = settings.get("SWAP_GEAR_AC_PENALTY",       cfg.SWAP_GEAR_AC_PENALTY)

    current_ac = engine.calc_ac(attacker, state["attacker_equipped"].get("armor"))
    ac_penalty = int(current_ac * ac_pen)

    with exclusive_transaction():
        if new_weapon_inv_id:
            execute_write(
                "UPDATE players SET equipped_weapon_id = ? WHERE id = ?",
                (new_weapon_inv_id, player_id)
            )
        if new_armor_inv_id:
            execute_write(
                "UPDATE players SET equipped_armor_id = ? WHERE id = ?",
                (new_armor_inv_id, player_id)
            )
        if new_special_inv_id:
            execute_write(
                "UPDATE players SET equipped_special_id = ? WHERE id = ?",
                (new_special_inv_id, player_id)
            )
        execute_write(
            """INSERT INTO combat_buffs
               (combat_session_id, side, buff_type, value, expires_on)
               VALUES (?, 'ATTACKER', 'SWAP_GEAR_ACCURACY_PENALTY', ?, 'END_OF_ROUND')""",
            (session_id, int(acc_pen * 100))
        )
        execute_write(
            """INSERT INTO combat_buffs
               (combat_session_id, side, buff_type, value, expires_on)
               VALUES (?, 'ATTACKER', 'SWAP_GEAR_AC_PENALTY', ?, 'END_OF_ROUND')""",
            (session_id, ac_penalty)
        )
        new_item = execute_one(
            """SELECT name FROM weapons WHERE id = (
               SELECT item_id FROM inventory_items WHERE id = ?)""",
            (new_weapon_inv_id,)
        ) if new_weapon_inv_id else None
        new_item_name = new_item["name"] if new_item else "new gear"
        _write_combat_log(session_id, session["current_round"], "ATTACKER",
                          "SWAP_GEAR", "Swap gear action",
                          f"Swapped to {new_item_name}, penalties applied this round")

    flv = flavour.swap_gear_flavor(attacker["character_name"], new_item_name)
    return {"action": "SWAP_GEAR", "flavor": flv,
            "accuracy_penalty_pct": int(acc_pen * 100), "ac_penalty": ac_penalty}


# ─────────────────────────────────────────────────────────────────────────────
# OPPONENT AUTOMATION
# ─────────────────────────────────────────────────────────────────────────────

def handle_opponent_action(session_id: int, state: dict) -> dict:
    """Automated opponent action for PvP defender (offline) or boss."""
    session = state["session"]

    if session["combat_type"] == "PVP":
        return _pvp_defender_action(session_id, state)
    else:
        return _boss_action(session_id, state)


def _pvp_defender_action(session_id: int, state: dict) -> dict:
    """PvP defending player uses their combat preference."""
    settings  = get_all_settings()
    defender  = state["defender"]
    pref      = defender.get("combat_preference", "Balanced")
    balanced  = settings.get("COMBAT_PREF_BALANCED_SPLIT",   cfg.COMBAT_PREF_BALANCED_SPLIT)
    opportunist = settings.get("COMBAT_PREF_OPPORTUNIST_SPLIT", cfg.COMBAT_PREF_OPPORTUNIST_SPLIT)

    if pref == "Aggressive":
        action = "ATTACK"
    elif pref == "Defensive":
        action = "BRACE"
    elif pref == "Opportunist":
        action = "STEAL" if random.random() < opportunist else "ATTACK"
    else:  # Balanced
        action = "BRACE" if random.random() < balanced else "ATTACK"

    # Resolve the chosen action from DEFENDER perspective
    if action == "ATTACK":
        return handle_attack(session_id, "DEFENDER", state)
    elif action == "BRACE":
        return handle_brace(session_id, state["session"]["defender_player_id"], state)
    else:
        return handle_steal(session_id, state["session"]["defender_player_id"], state)


def _boss_action(session_id: int, state: dict) -> dict:
    """Boss chooses and executes its action for this round.
    Checks phase thresholds first, updates phase if needed,
    then scales behavior to current phase."""
    boss    = state["boss"]
    session = state["session"]

    # ── Phase check ───────────────────────────────────────────────────────────
    max_hp        = boss["max_hp"]
    current_hp    = boss["current_hp"]
    hp_pct        = (current_hp / max_hp * 100) if max_hp else 100
    phase2_thresh = boss.get("phase2_hp_percent", 50)
    phase3_thresh = boss.get("phase3_hp_percent", 25)
    current_phase = boss.get("current_phase", 1)

    new_phase = current_phase
    if hp_pct <= phase3_thresh:
        new_phase = 3
    elif hp_pct <= phase2_thresh:
        new_phase = 2

    # Persist phase change and handle transition effects
    if new_phase != current_phase:
        with exclusive_transaction():
            # On entering phase 3: reset special move flags (can use again)
            if new_phase == 3:
                execute_write(
                    """UPDATE boss_instances
                       SET current_phase = 3,
                           special_attack_used = 0,
                           special_buff_used = 0
                       WHERE id = ?""",
                    (boss["instance_id"],)
                )
                flavor_text = (
                    f"{boss['name'].upper()} ENTERS PHASE 3 — "
                    f"desperate, enraged, and more dangerous than ever!"
                )
            else:
                execute_write(
                    "UPDATE boss_instances SET current_phase = ? WHERE id = ?",
                    (new_phase, boss["instance_id"])
                )
                flavor_text = (
                    f"{boss['name'].upper()} ENTERS PHASE 2 — "
                    f"wounded and furious, its attacks grow more deliberate."
                )
            # Personal feed entry for the phase transition
            execute_write(
                """INSERT INTO daily_feed
                   (feed_scope, player_id, flavor_text, event_category)
                   VALUES ('PERSONAL', ?, ?, 'COMBAT')""",
                (session["attacker_player_id"], flavor_text)
            )
        # Refresh boss state after update
        boss["current_phase"]       = new_phase
        boss["special_attack_used"] = 0 if new_phase == 3 else boss["special_attack_used"]
        boss["special_buff_used"]   = 0 if new_phase == 3 else boss["special_buff_used"]
        current_phase = new_phase

    # ── Choose action based on phase ──────────────────────────────────────────
    s_atk_used = boss["special_attack_used"]
    s_buf_used = boss["special_buff_used"]

    if current_phase == 1:
        # Phase 1: 33/33/33 split
        r = random.random()
        if   r < 0.333 and not s_atk_used: chosen = "SPECIAL_ATTACK"
        elif r < 0.666 and not s_buf_used:  chosen = "SPECIAL_BUFF"
        else:                                chosen = "ATTACK"

    elif current_phase == 2:
        # Phase 2: specials first if available, else attack
        if not s_atk_used and not s_buf_used:
            chosen = "SPECIAL_ATTACK" if random.random() < 0.5 else "SPECIAL_BUFF"
        elif not s_atk_used:
            chosen = "SPECIAL_ATTACK"
        elif not s_buf_used:
            chosen = "SPECIAL_BUFF"
        else:
            chosen = "ATTACK"

    else:
        # Phase 3: specials reset — same as phase 2 but with extra attack
        # (extra attack handled in combat round handler, see routes/combat.py patch)
        if not s_atk_used and not s_buf_used:
            chosen = "SPECIAL_ATTACK" if random.random() < 0.5 else "SPECIAL_BUFF"
        elif not s_atk_used:
            chosen = "SPECIAL_ATTACK"
        elif not s_buf_used:
            chosen = "SPECIAL_BUFF"
        else:
            chosen = "ATTACK"

    # ── Execute chosen action ─────────────────────────────────────────────────
    if chosen == "SPECIAL_ATTACK" and not s_atk_used:
        primary = _boss_special_attack(session_id, state)
    elif chosen == "SPECIAL_BUFF" and not s_buf_used:
        primary = _boss_special_buff(session_id, state)
    else:
        primary = _boss_regular_attack(session_id, state, current_phase)

    primary["boss_phase"] = current_phase
    primary["phase_changed"] = new_phase != (boss.get("current_phase", 1) if new_phase == current_phase else current_phase - 1)

    # Phase 3 extra attack — always attacks again after any action
    if current_phase == 3:
        # Reload state (HP may have changed)
        state2 = get_combat_state(session_id)
        if state2["session"]["status"] == "ACTIVE":
            extra = _boss_regular_attack(session_id, state2, current_phase)
            extra["is_extra_attack"] = True
            primary["extra_attack_result"] = extra

    return primary

def _boss_special_attack(session_id: int, state: dict) -> dict:
    """Provide the internal boss special attack operation used by this module."""
    boss    = state["boss"]
    session = state["session"]
    player  = state["attacker"]
    att_eq  = state["attacker_equipped"]

    die_sides = int(boss["special_attack_die"].lstrip("d"))
    raw_dmg   = sum(engine.roll(die_sides) for _ in range(2))  # 2 dice for special

    final_dmg, res_note = engine.resolve_resistance(
        raw_dmg, boss["special_attack_damage_type"],
        att_eq.get("armor"), att_eq.get("special")
    )
    new_hp = max(1, player["current_hp"] - final_dmg)

    with exclusive_transaction():
        execute_write("UPDATE players SET current_hp = ? WHERE id = ?",
                      (new_hp, session["attacker_player_id"]))
        execute_write(
            "UPDATE combat_sessions SET defender_total_damage_dealt = defender_total_damage_dealt + ? WHERE id = ?",
            (final_dmg, session_id)
        )
        instance_id = boss["instance_id"]
        execute_write(
            "UPDATE boss_instances SET special_attack_used = 1 WHERE id = ?", (instance_id,)
        )
        _write_combat_log(session_id, session["current_round"], "DEFENDER",
                          "SPECIAL_ATTACK", f"Special: {boss['special_attack_name']}",
                          f"{final_dmg} {boss['special_attack_damage_type']} damage")

    flv = flavour.boss_special_attack_flavor(
        boss["name"], boss["special_attack_name"], final_dmg, boss["special_attack_flavor"]
    )
    return {"action": "SPECIAL_ATTACK", "damage_total": final_dmg,
            "new_player_hp": new_hp, "flavor": flv}


def _boss_special_buff(session_id: int, state: dict) -> dict:
    """Provide the internal boss special buff operation used by this module."""
    boss    = state["boss"]
    session = state["session"]
    buff_type  = boss["special_buff_type"]
    buff_value = boss["special_buff_value"]

    expires_on = "END_OF_COMBAT"

    with exclusive_transaction():
        if buff_type == "HP_RESTORE":
            restore = int(boss["max_hp"] * buff_value)
            inst_id = boss["instance_id"]
            execute_write(
                "UPDATE boss_instances SET current_hp = MIN(current_hp + ?, ?) WHERE id = ?",
                (restore, boss["max_hp"], inst_id)
            )
        else:
            execute_write(
                """INSERT INTO combat_buffs
                   (combat_session_id, side, buff_type, damage_type, value, expires_on)
                   VALUES (?, 'DEFENDER', ?, ?, ?, ?)""",
                (session_id, f"BOSS_{buff_type}", boss.get("special_buff_damage_type"),
                 buff_value, expires_on)
            )
        instance_id = boss["instance_id"]
        execute_write(
            "UPDATE boss_instances SET special_buff_used = 1 WHERE id = ?", (instance_id,)
        )
        _write_combat_log(session_id, session["current_round"], "DEFENDER",
                          "SPECIAL_BUFF", f"Special buff: {boss['special_buff_name']}",
                          f"Type: {buff_type}, Value: {buff_value}")

    flv = flavour.boss_special_buff_flavor(
        boss["name"], boss["special_buff_name"], boss["special_buff_flavor"]
    )
    return {"action": "SPECIAL_BUFF", "buff_type": buff_type,
            "buff_value": buff_value, "flavor": flv}


def _get_boss_weapon(boss: dict) -> dict:
    """Load the boss's weapon from master table."""
    master = execute_one("SELECT boss_weapon_id FROM master WHERE boss_id = ?", (boss["id"],))
    if master:
        weapon = execute_one("SELECT * FROM weapons WHERE id = ?", (master["boss_weapon_id"],))
        if weapon:
            return weapon
    return {"weapon_type": "Melee", "damage_die": "d8", "damage_type": "Blunt",
            "name": "Attack", "str_bonus": 0}


# ─────────────────────────────────────────────────────────────────────────────
# POST-COMBAT RESOLUTION
# ─────────────────────────────────────────────────────────────────────────────

def finalize_combat(session_id: int, winner_side: str, result_type: str,
                    state: dict) -> dict:
    """Run full post-combat resolution sequence.
    Steps: XP → credits stolen → durability hits → item steal → over-encumbered check
           → feed entries → boss intel → clear in_combat → clear combat buffs."""
    session  = state["session"]
    attacker = state["attacker"]
    settings = get_all_settings()

    winner_is_attacker = winner_side == "ATTACKER"
    winner  = attacker if winner_is_attacker else state.get("defender")
    loser   = state.get("defender") if winner_is_attacker else attacker

    xp_earned     = 0
    credits_stolen = 0
    item_stolen    = None
    drops           = None

    # Step 1: XP award
    if session["combat_type"] in ("BOSS", "MINION") and winner_is_attacker:
        opp = state.get("boss") or state.get("minion")
        base_xp = 100 * opp["level"]
        special = state["attacker_equipped"].get("special")
        xp_mult = special.get("xp_multiplier", 0.0) if special else 0.0
        xp_earned = engine.calc_xp_reward(
            base_xp, attacker["level"], opp["level"], xp_mult
        )
        with exclusive_transaction():
            execute_write("UPDATE players SET xp = xp + ? WHERE id = ?",
                          (xp_earned, attacker["id"]))
            leveled = engine.check_level_up(
                attacker["id"], attacker["xp"] + xp_earned, attacker["level"]
            )
        # Update kill count and reset persistent encounter state atomically.
        with exclusive_transaction():
            if session["combat_type"] == "BOSS":
                execute_write(
                    "UPDATE boss_instances SET kill_count = kill_count + 1 WHERE id = ?",
                    (session["boss_instance_id"],)
                )
                execute_write(
                    """UPDATE boss_instances
                       SET current_hp = (SELECT max_hp FROM bosses WHERE id = boss_id),
                           special_attack_used=0, special_buff_used=0, current_phase=1
                       WHERE id = ?""",
                    (session["boss_instance_id"],)
                )
            else:
                execute_write(
                    "UPDATE minion_instances SET kill_count = kill_count + 1 WHERE id = ?",
                    (session["minion_instance_id"],)
                )

        drops = _award_drops(
            player_id=attacker["id"], player=attacker, opponent=opp,
            combat_type=session["combat_type"],
            master_row=_get_master_for_opponent(opp, session["combat_type"]),
            settings=settings,
            equipped_special=state["attacker_equipped"].get("special"),
        )

    elif session["combat_type"] == "PVP":
        zero_xp_bonus = settings.get("ZERO_CREDIT_XP_BONUS", cfg.ZERO_CREDIT_XP_BONUS)
        xp_loss_div   = settings.get("XP_LOSS_DIVISOR",       cfg.XP_LOSS_DIVISOR)
        special       = state["attacker_equipped"].get("special")
        xp_mult       = special.get("xp_multiplier", 0.0) if special else 0.0

        if winner_is_attacker:
            # Winner XP
            base_xp = 80 * (state["defender"]["level"] if state.get("defender") else 1)
            xp_earned = engine.calc_xp_reward(
                base_xp, attacker["level"],
                state["defender"]["level"] if state.get("defender") else attacker["level"],
                xp_mult
            )
            with exclusive_transaction():
                execute_write("UPDATE players SET xp = xp + ? WHERE id = ?",
                              (xp_earned, attacker["id"]))
                execute_write("UPDATE player_stats SET pvp_kills = pvp_kills + 1 WHERE player_id = ?",
                              (attacker["id"],))
                execute_write(
                    "UPDATE player_stats SET times_reduced_to_1hp = times_reduced_to_1hp + 1 WHERE player_id = ?",
                    (state["defender"]["id"],)
                )
        else:
            # Initiator lost — XP penalty
            base_xp = 80 * (state["defender"]["level"] if state.get("defender") else 1)
            potential_win_xp = engine.calc_xp_reward(base_xp, attacker["level"],
                                                       attacker["level"], xp_mult)
            xp_penalty = max(0, potential_win_xp // xp_loss_div)
            with exclusive_transaction():
                execute_write(
                    "UPDATE players SET xp = MAX(0, xp - ?) WHERE id = ?",
                    (xp_penalty, attacker["id"])
                )
                execute_write("UPDATE player_stats SET pvp_kills = pvp_kills + 1 WHERE player_id = ?",
                              (state["defender"]["id"],))
                execute_write(
                    "UPDATE player_stats SET times_reduced_to_1hp = times_reduced_to_1hp + 1 WHERE player_id = ?",
                    (attacker["id"],)
                )

        # Step 2: Credits stolen
        if winner and loser:
            cr_pct    = settings.get("CREDIT_STEAL_PERCENT",       cfg.CREDIT_STEAL_PERCENT)
            cr_lck_mult = settings.get("CREDIT_STEAL_LUCK_MULTIPLIER", cfg.CREDIT_STEAL_LUCK_MULTIPLIER)
            steal_bonus = (special.get("steal_bonus", 0.0) if special else 0.0)
            final_pct   = cr_pct + steal_bonus
            loser_player = execute_one("SELECT credits FROM players WHERE id = ?", (loser["id"],))
            credits_stolen = max(0, int(loser_player["credits"] * final_pct))
            # LCK double roll
            if random.random() < (engine.stat_mod(winner["lck_stat"]) * 0.05):
                credits_stolen *= cr_lck_mult
            credits_stolen = min(credits_stolen, loser_player["credits"])

            if credits_stolen == 0:
                # Zero credit bonus
                with exclusive_transaction():
                    execute_write("UPDATE players SET xp = xp + ? WHERE id = ?",
                                  (zero_xp_bonus, winner["id"]))
            else:
                with exclusive_transaction():
                    execute_write(
                        "UPDATE players SET credits = credits - ? WHERE id = ?",
                        (credits_stolen, loser["id"])
                    )
                    execute_write(
                        "UPDATE players SET credits = credits + ? WHERE id = ?",
                        (credits_stolen, winner["id"])
                    )

        # Step 3: Durability hits on loser's gear (before item steal)
        if loser:
            loser_eq = (state["attacker_equipped"] if not winner_is_attacker
                        else state["defender_equipped"])
            if loser_eq:
                engine.apply_pvp_loss_durability_hits(loser["id"], loser_eq)

        # Step 4: Item steal roll
        if winner and loser and result_type == "1HP_WIN":
            loser_eq  = (state["attacker_equipped"] if not winner_is_attacker
                         else state["defender_equipped"])
            steal_bonus = (special.get("steal_bonus", 0.0) if special else 0.0)
            loser_player = (state["defender"] if winner_is_attacker else attacker)
            roll_r = engine.resolve_opposed_roll(
                actor_agi=winner["agi_stat"], actor_lck=winner["lck_stat"],
                defender_agi=loser["agi_stat"], defender_lck=loser["lck_stat"],
                steal_bonus_pct=steal_bonus, tie_goes_to="defender"
            )
            if roll_r["success"]:
                loser_unequipped = [
                    i for i in execute(
                        "SELECT * FROM inventory_items WHERE player_id = ?", (loser["id"],)
                    )
                    if i["id"] not in {
                        loser_player.get("equipped_weapon_id"),
                        loser_player.get("equipped_armor_id"),
                        loser_player.get("equipped_special_id"),
                    }
                ]
                if loser_unequipped:
                    target = random.choice(loser_unequipped)
                    with exclusive_transaction():
                        execute_write(
                            "UPDATE inventory_items SET player_id = ?, acquired_method = 'PVP_STEAL' WHERE id = ?",
                            (winner["id"], target["id"])
                        )
                    item_detail = execute_one(
                        f"SELECT name FROM {'weapons' if target['item_type']=='WEAPON' else 'armor' if target['item_type']=='ARMOR' else 'special_items'} WHERE id = ?",
                        (target["item_id"],)
                    )
                    item_stolen = item_detail["name"] if item_detail else "item"

    # Finalize session
    with exclusive_transaction():
        execute_write(
            "UPDATE combat_sessions SET status='RESOLVED', result=?, resolved_at=? WHERE id=?",
            (result_type, datetime.utcnow().isoformat(), session_id)
        )
        execute_write(
            "UPDATE players SET in_combat = 0 WHERE id = ?",
            (session["attacker_player_id"],)
        )
        if session["combat_type"] == "PVP" and session.get("defender_player_id"):
            execute_write(
                "UPDATE players SET in_combat = 0 WHERE id = ?",
                (session["defender_player_id"],)
            )
        execute_write("DELETE FROM combat_buffs WHERE combat_session_id = ?", (session_id,))

    # Feed entries
    winner_name = winner.get("character_name") if winner else "Unknown"
    loser_name  = (loser.get("character_name") if loser
                   else (state.get("boss") or state.get("minion") or {}).get("name", "opponent"))
    global_text = flavour.combat_result_flavor(
        winner_name, loser_name, session["combat_type"],
        credits_stolen, item_stolen, result_type
    )
    with exclusive_transaction():
        execute_write(
            """INSERT INTO daily_feed (feed_scope, player_id, flavor_text, event_category, combat_session_id)
               VALUES ('GLOBAL', NULL, ?, 'COMBAT', ?)""",
            (global_text, session_id)
        )
        execute_write(
            """INSERT INTO daily_feed (feed_scope, player_id, flavor_text, event_category, combat_session_id)
               VALUES ('PERSONAL', ?, ?, 'COMBAT', ?)""",
            (attacker["id"], global_text, session_id)
        )

    return {
        "winner_side":     winner_side,
        "result_type":     result_type,
        "xp_earned":       xp_earned,
        "credits_stolen":  credits_stolen,
        "item_stolen":     item_stolen,
        "drops":           drops,
        "flavor":          global_text,
    }


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _apply_durability_loss(inv_id: int, loss: int, player_id: int):
    """Apply durability loss to an inventory item. Destroys item if it hits 0.
    Must be called inside exclusive_transaction()."""
    row = execute_one("SELECT current_durability, item_type, item_id, player_id FROM inventory_items WHERE id = ?", (inv_id,))
    if row is None:
        return
    new_dur = max(0, row["current_durability"] - loss)
    if new_dur == 0:
        _destroy_item(inv_id, row, player_id)
    else:
        execute_write(
            "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
            (new_dur, inv_id)
        )


def _destroy_item(inv_id: int, row: dict, player_id: int):
    """Delete an item at 0 durability, null out equipped slot, return special to pool."""
    execute_write("DELETE FROM inventory_items WHERE id = ?", (inv_id,))
    # Null out equipped slot if this was equipped
    for col in ("equipped_weapon_id", "equipped_armor_id", "equipped_special_id"):
        execute_write(
            f"UPDATE players SET {col} = NULL WHERE id = ? AND {col} = ?",
            (player_id, inv_id)
        )
    if row["item_type"] == "SPECIAL":
        execute_write(
            """UPDATE special_item_registry
               SET status='IN_POOL', current_owner_player_id=NULL, inventory_item_id=NULL,
                   last_released_method='DESTROYED', updated_at=?
               WHERE special_item_id=?""",
            (datetime.utcnow().isoformat(), row["item_id"])
        )
    # Log destruction
    item_detail = execute_one(
        f"SELECT name FROM {'weapons' if row['item_type']=='WEAPON' else 'armor' if row['item_type']=='ARMOR' else 'special_items'} WHERE id = ?",
        (row["item_id"],)
    )
    item_name = item_detail["name"] if item_detail else "Unknown Item"
    execute_write(
        """INSERT INTO item_history (player_id, item_type, item_id, item_name, event_type)
           VALUES (?, ?, ?, ?, 'DESTROYED')""",
        (player_id, row["item_type"], row["item_id"], item_name)
    )


def _apply_steal_fail_penalty(session_id: int, side: str):
    """Insert steal fail AC penalty buff. Inside exclusive_transaction()."""
    with exclusive_transaction():
        execute_write(
            """INSERT INTO combat_buffs
               (combat_session_id, side, buff_type, value, expires_on)
               VALUES (?, ?, 'STEAL_FAIL_AC_PENALTY', 3, 'NEXT_HIT_RESOLVED')""",
            (session_id, side)
        )


def _write_combat_log(session_id: int, round_num: int, actor: str,
                      action_type: str, roll_detail: str, outcome_detail: str):
    """Provide the internal write combat log operation used by this module."""
    execute_write(
        """INSERT INTO combat_logs
           (combat_session_id, round_number, actor, action_type, roll_detail, outcome_detail)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (session_id, round_num, actor, action_type, roll_detail, outcome_detail)
    )


def _get_master_for_opponent(opponent: dict, combat_type: str) -> dict | None:
    """Load the master table row for a boss or minion by name."""
    col = "boss_id" if combat_type == "BOSS" else "minion_id"
    # opponent dict has 'id' which is the bosses/minions table id
    return execute_one(
        f"SELECT * FROM master WHERE {col} = (SELECT id FROM "
        f"{'bosses' if combat_type == 'BOSS' else 'minions'} WHERE name = ?)",
        (opponent["name"],)
    )


def _award_drops(player_id: int, player: dict, opponent: dict,
                 combat_type: str, master_row: dict | None,
                 settings: dict, equipped_special: dict | None) -> dict:
    """Roll all drop table entries for a defeated boss or minion.
    Returns dict with keys: credits, weapon, armor, special (each None or item name)."""
    import random
    from datetime import datetime

    result = {"credits": 0, "weapon": None, "armor": None, "special": None}

    if not master_row:
        return result

    # ── Credits ───────────────────────────────────────────────────────────────
    cr_min  = opponent.get("drop_credit_min", 0)
    cr_max  = opponent.get("drop_credit_max", 0)
    if cr_max > 0:
        credits = random.randint(cr_min, cr_max)
        # Apply credit multiplier from equipped special item
        if equipped_special and equipped_special.get("credit_multiplier"):
            credits = int(credits * (1 + equipped_special["credit_multiplier"]))
        if credits > 0:
            with exclusive_transaction():
                execute_write(
                    "UPDATE players SET credits = credits + ? WHERE id = ?",
                    (credits, player_id)
                )
            result["credits"] = credits

    # ── Determine which item IDs to roll for ──────────────────────────────────
    if combat_type == "BOSS":
        weapon_id  = master_row.get("boss_weapon_id")
        armor_id   = master_row.get("boss_armor_id")
        special_id = master_row.get("boss_special_item_id")
    else:
        weapon_id  = master_row.get("minion_weapon_id")
        armor_id   = master_row.get("minion_armor_id")
        special_id = master_row.get("minion_special_item_id")

    # ── Weapon drop ───────────────────────────────────────────────────────────
    weapon_chance = opponent.get("drop_weapon_chance", 0.0)
    if weapon_id and random.random() < weapon_chance:
        # Skip if player already owns this weapon
        already_owned = execute_one(
            "SELECT id FROM inventory_items WHERE player_id = ? AND item_type = 'WEAPON' AND item_id = ?",
            (player_id, weapon_id)
        )
        if not already_owned:
            weapon_detail = execute_one("SELECT * FROM weapons WHERE id = ?", (weapon_id,))
            if weapon_detail:
                with exclusive_transaction():
                    execute_write(
                        """INSERT INTO inventory_items
                           (player_id, item_type, item_id, current_durability, acquired_method)
                           VALUES (?, 'WEAPON', ?, ?, ?)""",
                        (player_id, weapon_id,
                         weapon_detail.get("starting_durability", 100),
                         "BOSS_DROP" if combat_type == "BOSS" else "MINION_DROP")
                    )
                    execute_write(
                        """INSERT INTO item_history
                           (player_id, item_type, item_id, item_name, event_type)
                           VALUES (?, 'WEAPON', ?, ?, ?)""",
                        (player_id, weapon_id, weapon_detail["name"],
                         "RECEIVED_BOSS_DROP" if combat_type == "BOSS" else "RECEIVED_MINION_DROP")
                    )
                result["weapon"] = weapon_detail["name"]

    # ── Armor drop ────────────────────────────────────────────────────────────
    armor_chance = opponent.get("drop_armor_chance", 0.0)
    if armor_id and random.random() < armor_chance:
        already_owned = execute_one(
            "SELECT id FROM inventory_items WHERE player_id = ? AND item_type = 'ARMOR' AND item_id = ?",
            (player_id, armor_id)
        )
        if not already_owned:
            armor_detail = execute_one("SELECT * FROM armor WHERE id = ?", (armor_id,))
            if armor_detail:
                with exclusive_transaction():
                    execute_write(
                        """INSERT INTO inventory_items
                           (player_id, item_type, item_id, current_durability, acquired_method)
                           VALUES (?, 'ARMOR', ?, ?, ?)""",
                        (player_id, armor_id,
                         armor_detail.get("starting_durability", 100),
                         "BOSS_DROP" if combat_type == "BOSS" else "MINION_DROP")
                    )
                    execute_write(
                        """INSERT INTO item_history
                           (player_id, item_type, item_id, item_name, event_type)
                           VALUES (?, 'ARMOR', ?, ?, ?)""",
                        (player_id, armor_id, armor_detail["name"],
                         "RECEIVED_BOSS_DROP" if combat_type == "BOSS" else "RECEIVED_MINION_DROP")
                    )
                result["armor"] = armor_detail["name"]

    # ── Special item drop ─────────────────────────────────────────────────────
    special_chance = opponent.get("drop_special_item_chance", 0.0)
    if special_id and random.random() < special_chance:
        # Check registry — must be IN_POOL
        reg = execute_one(
            "SELECT * FROM special_item_registry WHERE special_item_id = ?",
            (special_id,)
        )
        if reg and reg["status"] == "IN_POOL":
            special_detail = execute_one("SELECT * FROM special_items WHERE id = ?", (special_id,))
            if special_detail:
                with exclusive_transaction():
                    inv_id = execute_write(
                        """INSERT INTO inventory_items
                           (player_id, item_type, item_id, current_durability, acquired_method)
                           VALUES (?, 'SPECIAL', ?, ?, ?)""",
                        (player_id, special_id,
                         special_detail.get("starting_durability", 100),
                         "BOSS_DROP" if combat_type == "BOSS" else "MINION_DROP")
                    )
                    execute_write(
                        """UPDATE special_item_registry
                           SET status = 'IN_INVENTORY',
                               current_owner_player_id = ?,
                               inventory_item_id = ?,
                               last_acquired_method = ?,
                               updated_at = ?
                           WHERE special_item_id = ?""",
                        (player_id, inv_id,
                         "BOSS_DROP" if combat_type == "BOSS" else "MINION_DROP",
                         datetime.utcnow().isoformat(), special_id)
                    )
                    execute_write(
                        """INSERT INTO item_history
                           (player_id, item_type, item_id, item_name, event_type)
                           VALUES (?, 'SPECIAL', ?, ?, ?)""",
                        (player_id, special_id, special_detail["name"],
                         "RECEIVED_BOSS_DROP" if combat_type == "BOSS" else "RECEIVED_MINION_DROP")
                    )
                    # Global feed: special item enters world
                    execute_write(
                        """INSERT INTO daily_feed
                           (feed_scope, player_id, flavor_text, event_category)
                           VALUES ('GLOBAL', NULL, ?, 'ITEM')""",
                        (f"The {special_detail['name']} has been claimed from {opponent['name']}.",)
                    )
                result["special"] = special_detail["name"]

    # ── Personal feed entry summarising all drops ─────────────────────────────
    drop_lines = []
    if result["credits"]: drop_lines.append(f"+{result['credits']} credits")
    if result["weapon"]:  drop_lines.append(f"Found: {result['weapon']}")
    if result["armor"]:   drop_lines.append(f"Found: {result['armor']}")
    if result["special"]: drop_lines.append(f"★ Seized: {result['special']}")

    if drop_lines:
        with exclusive_transaction():
            execute_write(
                """INSERT INTO daily_feed
                   (feed_scope, player_id, flavor_text, event_category)
                   VALUES ('PERSONAL', ?, ?, 'ITEM')""",
                (player_id, " | ".join(drop_lines))
            )

    return result


def _boss_regular_attack(session_id: int, state: dict, phase: int) -> dict:
    """Execute a regular boss attack, with phase-based attack bonuses.
    Phase 2: +2 to attack roll. Phase 3: +4 to attack roll."""
    boss    = state["boss"]
    session = state["session"]

    # Phase attack bonus
    phase_attack_bonus = {1: 0, 2: 2, 3: 4}.get(phase, 0)

    # Inject phase bonus as a temporary combat buff for this round
    if phase_attack_bonus > 0:
        with exclusive_transaction():
            execute_write(
                """INSERT INTO combat_buffs
                   (combat_session_id, side, buff_type, value, expires_on)
                   VALUES (?, 'DEFENDER', 'BOSS_ATTACK_BONUS', ?, 'END_OF_ROUND')""",
                (session_id, phase_attack_bonus)
            )
        # Reload state to pick up the new buff
        state = get_combat_state(session_id)
        boss  = state["boss"]

    boss_as_attacker = {**boss}
    boss_weapon      = _get_boss_weapon(boss)
    attacker_player  = state["attacker"]
    att_armor        = state["attacker_equipped"].get("armor")
    att_special      = state["attacker_equipped"].get("special")
    att_buffs        = state["attacker_buffs"]
    brace_dodge      = sum(
        int(b["value"]) for b in att_buffs if b["buff_type"] == "BRACE_DODGE_BONUS"
    )

    result = engine.resolve_full_attack(
        attacker=boss_as_attacker,
        defender=attacker_player,
        attacker_weapon=boss_weapon,
        attacker_special=None,
        defender_armor=att_armor,
        defender_special=att_special,
        boss=None,
        brace_dodge_bonus=brace_dodge,
        active_buffs=att_buffs,
        is_player_attacker=False,
    )

    with exclusive_transaction():
        if result["hit"]:
            new_hp = max(1, attacker_player["current_hp"] - result["damage_total"])
            execute_write(
                "UPDATE players SET current_hp = ? WHERE id = ?",
                (new_hp, session["attacker_player_id"])
            )
            execute_write(
                """UPDATE combat_sessions
                   SET defender_total_damage_dealt = defender_total_damage_dealt + ?
                   WHERE id = ?""",
                (result["damage_total"], session_id)
            )
            if att_armor:
                _apply_durability_loss(att_armor["inv_id"], 1,
                                       session["attacker_player_id"])
            execute_write(
                """DELETE FROM combat_buffs
                   WHERE combat_session_id = ? AND side = 'ATTACKER'
                   AND expires_on = 'NEXT_HIT_RESOLVED'""",
                (session_id,)
            )
        _write_combat_log(session_id, session["current_round"], "DEFENDER",
                          "ATTACK", result["roll_detail"], result["outcome_detail"])

    flv = flavour.attack_flavor(
        attacker_name=boss["name"],
        weapon_name=boss_weapon.get("name", "attack"),
        weapon_type=boss_weapon.get("weapon_type", "Melee"),
        hit=result["hit"], dodged=result["dodged"],
        is_crit=result["is_crit"],
        damage=result["damage_total"],
        damage_type=boss_weapon.get("damage_type", "Blunt"),
    )
    return {
        "action": "ATTACK",
        "hit": result["hit"],
        "dodged": result["dodged"],
        "damage_total": result["damage_total"],
        "roll_detail": result["roll_detail"],
        "flavor": flv,
    }

# FILE: combat/engine.py
"""Pure dice, damage, initiative, progression, and opposed-roll calculations."""
# combat/engine.py
# Core combat math. All dice rolls, stat modifiers, damage resolution,
# resistance/weakness checks, dodge, crit, and durability.
# Stateless pure functions — takes input dicts, returns result dicts.
# Never writes to the DB directly.

import math
import random
import logging

import config_defaults as cfg
from database import get_all_settings

logger = logging.getLogger(__name__)

# The 7 damage types
DAMAGE_TYPES = {"blade", "blunt", "ballistic", "energy", "arcane", "explosive", "venom"}

# Resistance column names on armor/special_items/bosses
RES_COLS  = [f"res_{t}"  for t in DAMAGE_TYPES]
WEAK_COLS = [f"weak_{t}" for t in DAMAGE_TYPES]


# ─────────────────────────────────────────────────────────────────────────────
# DICE & STAT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def roll(sides: int) -> int:
    """Roll a single die with the given number of sides."""
    return random.randint(1, sides)


def roll_damage_die(die_str: str) -> int:
    """Parse and roll a damage die string like 'd8', 'd12', etc."""
    sides = int(die_str.lstrip("d"))
    return roll(sides)


def stat_mod(stat: int) -> int:
    """Standard stat modifier: floor(stat / 2)."""
    return math.floor(stat / 2)


# ─────────────────────────────────────────────────────────────────────────────
# DERIVED STAT CALCULATIONS
# ─────────────────────────────────────────────────────────────────────────────

def calc_max_hp(player: dict) -> int:
    """10 + END + (5 * level)"""
    return 10 + player["end_stat"] + (5 * player["level"])


def calc_ac(combatant: dict, armor: dict | None) -> int:
    """10 + floor(AGI/2) + armor.ac_bonus (if equipped)"""
    ac = 10 + stat_mod(combatant["agi_stat"])
    if armor:
        ac += armor.get("ac_bonus", 0)
    return ac


def calc_max_ap(player: dict, is_cursed: bool = False) -> int:
    """BASE_DAILY_AP + floor(END/2), reduced by CURSE_AP_REDUCTION if cursed."""
    settings  = get_all_settings()
    base      = settings.get("BASE_DAILY_AP",       cfg.BASE_DAILY_AP)
    cap       = settings.get("AP_CARRYOVER_CAP",    cfg.AP_CARRYOVER_CAP)
    curse_red = settings.get("CURSE_AP_REDUCTION",  cfg.CURSE_AP_REDUCTION)
    raw       = base + stat_mod(player["end_stat"])
    if is_cursed:
        raw = int(raw * (1 - curse_red))
    return min(raw, cap)


def calc_passive_regen(player: dict, special: dict | None = None) -> int:
    """AP_PASSIVE_HP_REGEN + floor(END/END_HP_REGEN_DIVISOR) + HP_REGEN_BONUS (special)"""
    settings  = get_all_settings()
    base_regen = settings.get("AP_PASSIVE_HP_REGEN",   cfg.AP_PASSIVE_HP_REGEN)
    divisor    = settings.get("END_HP_REGEN_DIVISOR",  cfg.END_HP_REGEN_DIVISOR)
    bonus      = special.get("hp_regen_bonus", 0) if special else 0
    return base_regen + math.floor(player["end_stat"] / divisor) + bonus


def calc_initiative(combatant: dict, initiative_bonus: int = 0) -> tuple[int, int]:
    """Roll initiative: d20 + floor(AGI/2) + initiative_bonus + initiative_modifier.
    initiative_modifier comes from status_effects (STAT_BOOST/PENALTY_INITIATIVE).
    Returns (total, raw_agi) — raw AGI used for tie-breaking."""
    raw_roll = roll(20)
    status_init_mod = combatant.get("initiative_modifier", 0)
    total = raw_roll + stat_mod(combatant["agi_stat"]) + initiative_bonus + status_init_mod
    return total, combatant["agi_stat"]

def calc_crit_threshold(combatant: dict, special: dict | None = None) -> int:
    """max(CRIT_MIN_THRESHOLD, 20 - floor(LCK / CRIT_LCK_DIVISOR))
    Further reduced by special.crit_chance_bonus if equipped."""
    settings  = get_all_settings()
    base      = settings.get("CRIT_BASE_THRESHOLD", cfg.CRIT_BASE_THRESHOLD)
    divisor   = settings.get("CRIT_LCK_DIVISOR",   cfg.CRIT_LCK_DIVISOR)
    min_thr   = settings.get("CRIT_MIN_THRESHOLD",  cfg.CRIT_MIN_THRESHOLD)
    threshold = base - math.floor(combatant["lck_stat"] / divisor)
    if special:
        threshold -= int(special.get("crit_chance_bonus", 0))
    return max(min_thr, threshold)


# ─────────────────────────────────────────────────────────────────────────────
# ATTACK ROLL
# ─────────────────────────────────────────────────────────────────────────────

def calc_attack_roll(attacker: dict, weapon: dict) -> tuple[int, int, bool]:
    """Roll an attack.
    Melee: d20 + floor(STR/2). Ranged: d20 + floor(AGI/2).
    Returns (total, raw_d20, is_crit_range_roll)."""
    raw_d20 = roll(20)
    if weapon["weapon_type"] == "Melee":
        modifier = stat_mod(attacker["str_stat"])
    else:
        modifier = stat_mod(attacker["agi_stat"])
    total = raw_d20 + modifier
    return total, raw_d20, modifier


def hits_ac(attack_total: int, target_ac: int) -> bool:
    """Handle the hits ac workflow."""
    return attack_total >= target_ac


# ─────────────────────────────────────────────────────────────────────────────
# DAMAGE CALCULATION
# ─────────────────────────────────────────────────────────────────────────────

def calc_weapon_damage(attacker: dict, weapon: dict, is_crit: bool) -> tuple[int, str]:
    """Roll weapon damage + stat modifier.
    Doubles on crit. Returns (damage, detail_str)."""
    die_roll = roll_damage_die(weapon["damage_die"])
    if weapon["weapon_type"] == "Melee":
        modifier = stat_mod(attacker["str_stat"])
    else:
        modifier = stat_mod(attacker["agi_stat"])
    base = die_roll + modifier
    if is_crit:
        base *= 2
    detail = f"{weapon['damage_die']}({die_roll})+{modifier}={'CRIT:' if is_crit else ''}{base}"
    return base, detail


def calc_bonus_damage(special: dict, is_crit: bool) -> tuple[int, str]:
    """Calculate bonus damage from an equipped special item.
    Doubles on crit. Returns (damage, damage_type)."""
    if not special or not special.get("bonus_damage_amount"):
        return 0, ""
    amount = special["bonus_damage_amount"]
    if is_crit:
        amount *= 2
    return amount, special.get("bonus_damage_type", "")


# ─────────────────────────────────────────────────────────────────────────────
# RESISTANCE & WEAKNESS
# ─────────────────────────────────────────────────────────────────────────────

def resolve_resistance(damage: int, damage_type: str,
                       armor: dict | None,
                       special: dict | None,
                       boss_buff_resistance: str | None = None) -> tuple[int, str]:
    """Apply resistance stacking rule.
    0 sources → full damage
    1 source  → half damage
    2+ sources → floor at RESISTANCE_STACK_MIN_DAMAGE_PERCENT
    Returns (final_damage, note_str)."""
    settings   = get_all_settings()
    floor_pct  = settings.get("RESISTANCE_STACK_MIN_DAMAGE_PERCENT",
                               cfg.RESISTANCE_STACK_MIN_DAMAGE_PERCENT)
    dtype_col  = f"res_{damage_type.lower()}"

    sources = 0
    if armor  and armor.get(dtype_col):
        sources += 1
    if special and special.get(dtype_col):
        sources += 1
    if boss_buff_resistance and boss_buff_resistance.lower() == damage_type.lower():
        sources += 1

    if sources == 0:
        return damage, ""
    elif sources == 1:
        final = max(1, damage // 2)
        return final, f"Resisted ({damage}→{final})"
    else:
        floor_dmg = max(1, int(damage * floor_pct))
        return floor_dmg, f"Stacked resistance ({damage}→{floor_dmg})"


def resolve_weakness(damage: int, damage_type: str, boss: dict) -> tuple[int, str]:
    """If boss has weakness to damage_type, double damage. Players never have weaknesses."""
    dtype_col = f"weak_{damage_type.lower()}"
    if boss and boss.get(dtype_col):
        doubled = damage * 2
        return doubled, f"Weakness! ({damage}→{doubled})"
    return damage, ""


# ─────────────────────────────────────────────────────────────────────────────
# DODGE
# ─────────────────────────────────────────────────────────────────────────────

def resolve_dodge(defender: dict, attacker: dict,
                  brace_dodge_bonus: int = 0) -> tuple[bool, str]:
    """Player-only dodge check. Bosses/minions do not dodge.
    Defender: d20 + floor(AGI/2) + floor(LCK/2) + BRACE_DODGE_BONUS
    Attacker: d20 + floor(AGI/2)  (Initiative Bonus does NOT apply here)
    Ties go to attacker (harder to dodge).
    Returns (dodged: bool, detail_str)."""
    def_roll = roll(20)
    def_mod  = stat_mod(defender["agi_stat"]) + stat_mod(defender["lck_stat"]) + brace_dodge_bonus
    def_total = def_roll + def_mod

    att_roll  = roll(20)
    att_mod   = stat_mod(attacker["agi_stat"])
    att_total = att_roll + att_mod

    dodged = def_total > att_total  # ties go to attacker
    detail = (f"Dodge: {def_roll}+{def_mod}={def_total} vs "
              f"{att_roll}+{att_mod}={att_total} → {'DODGE!' if dodged else 'Hit'}")
    return dodged, detail


# ─────────────────────────────────────────────────────────────────────────────
# FULL ATTACK RESOLUTION
# ─────────────────────────────────────────────────────────────────────────────

def resolve_full_attack(attacker: dict, defender: dict,
                        attacker_weapon: dict,
                        attacker_special: dict | None,
                        defender_armor: dict | None,
                        defender_special: dict | None,
                        boss: dict | None = None,
                        brace_dodge_bonus: int = 0,
                        active_buffs: list | None = None,
                        is_player_attacker: bool = True) -> dict:
    """Run the full attack sequence for one attack action.
    Returns a result dict with all details for combat log rendering.

    Sequence:
    1. Dodge check (player defenders only)
    2. Attack roll vs AC
    3. Crit check
    4. Weapon damage roll + stat mod
    5. Resistance + weakness resolution
    6. Special item bonus damage (separate resistance check)
    7. Durability effects
    Returns:
        hit, dodged, damage_total, is_crit,
        weapon_durability_loss, armor_durability_loss,
        roll_detail, outcome_detail
    """
    settings = get_all_settings()

    # --- Get active boss resistance buff if any ---
    boss_resistance_type = None
    if active_buffs:
        for buff in active_buffs:
            if buff.get("buff_type") == "BOSS_RESISTANCE_TYPE":
                boss_resistance_type = buff.get("damage_type")

    # --- Get active combat modifiers from buffs ---
    attack_bonus = 0
    if active_buffs:
        for buff in active_buffs:
            if buff.get("buff_type") == "BOSS_ATTACK_BONUS":
                attack_bonus += int(buff.get("value", 0))

    # --- Step 1: Dodge check (only when defender is a player) ---
    defender_is_player = boss is None
    dodged      = False
    dodge_detail = ""
    if defender_is_player:
        dodged, dodge_detail = resolve_dodge(defender, attacker, brace_dodge_bonus)
        if dodged:
            return {
                "hit": False, "dodged": True, "damage_total": 0, "is_crit": False,
                "weapon_durability_loss": 0, "armor_durability_loss": 0,
                "roll_detail": dodge_detail, "outcome_detail": "Attack dodged — no damage.",
                "damage_breakdown": []
            }

    # --- Step 2: Attack roll vs AC ---
    attack_total, raw_d20, attack_mod = calc_attack_roll(attacker, attacker_weapon)
    attack_total += attack_bonus
    defender_ac   = calc_ac(defender, defender_armor)

    # Apply over-encumbered attack penalty if attacker is a player
    if is_player_attacker and attacker.get("is_overencumbered"):
        over_penalty = settings.get("OVERENCUMBERED_ATTACK_PENALTY", cfg.OVERENCUMBERED_ATTACK_PENALTY)
        attack_total -= over_penalty

    # Apply swap gear penalty if active
    swap_penalty = 0
    if active_buffs and is_player_attacker:
        for buff in active_buffs:
            if buff.get("buff_type") == "SWAP_GEAR_ACCURACY_PENALTY":
                swap_penalty = int(buff.get("value", 0))
    attack_total -= swap_penalty

    hit = hits_ac(attack_total, defender_ac)
    attack_detail = (f"Attack: d20({raw_d20})+{attack_mod}"
                     f"{'−'+str(swap_penalty) if swap_penalty else ''}"
                     f"={attack_total} vs AC {defender_ac} → {'HIT' if hit else 'MISS'}")

    if not hit:
        return {
            "hit": False, "dodged": False, "damage_total": 0, "is_crit": False,
            "weapon_durability_loss": 0, "armor_durability_loss": 0,
            "roll_detail": attack_detail, "outcome_detail": "Miss — no damage.",
            "damage_breakdown": []
        }

    # --- Step 3: Crit check ---
    crit_threshold = calc_crit_threshold(attacker, attacker_special)
    # Extra crit bonus from boss buff
    if active_buffs:
        for buff in active_buffs:
            if buff.get("buff_type") == "BOSS_CRIT_BONUS":
                crit_threshold = max(
                    settings.get("CRIT_MIN_THRESHOLD", cfg.CRIT_MIN_THRESHOLD),
                    crit_threshold - int(buff.get("value", 0))
                )
    is_crit = raw_d20 >= crit_threshold
    if is_crit:
        attack_detail += f" CRITICAL HIT (rolled {raw_d20} ≥ {crit_threshold})!"

    # --- Step 4: Weapon damage ---
    weapon_dmg, weapon_detail = calc_weapon_damage(attacker, attacker_weapon, is_crit)

    # --- Step 5: Resistance + weakness ---
    weapon_dmg, res_note = resolve_resistance(
        weapon_dmg, attacker_weapon["damage_type"],
        defender_armor, defender_special, boss_resistance_type
    )
    if boss:
        weapon_dmg, weak_note = resolve_weakness(
            weapon_dmg, attacker_weapon["damage_type"], boss
        )
    else:
        weak_note = ""

    damage_breakdown = [{
        "type": attacker_weapon["damage_type"],
        "raw": weapon_dmg,
        "note": " ".join(filter(None, [res_note, weak_note]))
    }]

    # --- Step 6: Special item bonus damage ---
    bonus_dmg = 0
    if attacker_special and attacker_special.get("bonus_damage_amount"):
        raw_bonus, bonus_type = calc_bonus_damage(attacker_special, is_crit)
        if raw_bonus and bonus_type:
            final_bonus, bonus_res_note = resolve_resistance(
                raw_bonus, bonus_type, defender_armor, defender_special, boss_resistance_type
            )
            if boss:
                final_bonus, bonus_weak_note = resolve_weakness(final_bonus, bonus_type, boss)
            else:
                bonus_weak_note = ""
            bonus_dmg = final_bonus
            damage_breakdown.append({
                "type": bonus_type,
                "raw": final_bonus,
                "note": " ".join(filter(None, [bonus_res_note, bonus_weak_note]))
            })

    # Crit DMG multiplier from special item (applies on top)
    if is_crit and attacker_special and attacker_special.get("crit_dmg_multiplier"):
        mult = attacker_special["crit_dmg_multiplier"]
        weapon_dmg  = int(weapon_dmg  * (1 + mult))
        bonus_dmg   = int(bonus_dmg   * (1 + mult))

    damage_total = weapon_dmg + bonus_dmg

    # Boss DMG_REDUCTION buff
    if active_buffs:
        for buff in active_buffs:
            if buff.get("buff_type") == "BOSS_DMG_REDUCTION":
                reduction = int(buff.get("value", 0))
                damage_total = max(0, damage_total - reduction)

    # --- Step 7: Durability ---
    weapon_dur_loss = _calc_durability_loss(1, attacker_special)
    armor_dur_loss  = _calc_durability_loss(1, defender_special)

    outcome = (f"{weapon_detail} → {damage_total} damage"
               f"{' (' + damage_breakdown[0]['note'] + ')' if damage_breakdown[0]['note'] else ''}")

    return {
        "hit":                   True,
        "dodged":                False,
        "damage_total":          max(0, damage_total),
        "is_crit":               is_crit,
        "weapon_durability_loss": weapon_dur_loss,
        "armor_durability_loss":  armor_dur_loss,
        "roll_detail":           attack_detail,
        "outcome_detail":        outcome,
        "damage_breakdown":      damage_breakdown,
        "weapon_damage_type":    attacker_weapon["damage_type"],
    }


# ─────────────────────────────────────────────────────────────────────────────
# OPPOSED ROLLS (steal, escape, observe, minion PER)
# ─────────────────────────────────────────────────────────────────────────────

def resolve_opposed_roll(actor_agi: int, actor_lck: int,
                         defender_agi: int, defender_lck: int,
                         actor_per: int = 0, defender_per: int = 0,
                         steal_bonus_pct: float = 0.0,
                         tie_goes_to: str = "defender") -> dict:
    """Generic opposed roll for steal, escape, observe, minion PER check.
    Actor:    d20 + floor(AGI/2) + floor(LCK/2) [+ floor(PER/2) if observe]
    Defender: d20 + floor(AGI/2) + floor(LCK/2) [+ floor(PER/2) if observe]
    Steal bonus adds a flat roll bonus on top.

    tie_goes_to: 'defender' (steal, observe, minion PER) or 'opponent' (escape)
    Returns: {actor_roll, defender_roll, success, detail}"""
    actor_roll    = roll(20) + stat_mod(actor_agi) + stat_mod(actor_lck) + stat_mod(actor_per)
    actor_roll   += int(steal_bonus_pct * 20)  # steal bonus as flat roll bonus
    defender_roll = roll(20) + stat_mod(defender_agi) + stat_mod(defender_lck) + stat_mod(defender_per)

    if tie_goes_to == "defender":
        success = actor_roll > defender_roll
    else:  # tie goes to opponent/defender — actor needs strict win
        success = actor_roll > defender_roll

    detail = f"Roll: {actor_roll} vs {defender_roll} → {'Success' if success else 'Fail'}"
    return {
        "actor_roll":    actor_roll,
        "defender_roll": defender_roll,
        "success":       success,
        "detail":        detail,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PVP SCORE FORMULA
# ─────────────────────────────────────────────────────────────────────────────

def calc_pvp_score(session: dict, attacker_max_hp: int,
                   defender_max_hp: int) -> tuple[float, float]:
    """Tiebreak score formula:
    (HP% * COMBAT_WIN_HP_WEIGHT) + (Damage Dealt% * COMBAT_WIN_DMG_WEIGHT)
    Always produces a winner."""
    settings   = get_all_settings()
    hp_weight  = settings.get("COMBAT_WIN_HP_WEIGHT",  cfg.COMBAT_WIN_HP_WEIGHT)
    dmg_weight = settings.get("COMBAT_WIN_DMG_WEIGHT", cfg.COMBAT_WIN_DMG_WEIGHT)

    att_hp_pct  = session["attacker_hp_start"] / attacker_max_hp if attacker_max_hp else 0
    def_hp_pct  = session["defender_hp_start"] / defender_max_hp if defender_max_hp else 0

    total_dmg = (session["attacker_total_damage_dealt"] +
                 session["defender_total_damage_dealt"])

    att_dmg_pct = session["attacker_total_damage_dealt"] / total_dmg if total_dmg else 0
    def_dmg_pct = session["defender_total_damage_dealt"] / total_dmg if total_dmg else 0

    att_score = (att_hp_pct * hp_weight) + (att_dmg_pct * dmg_weight)
    def_score = (def_hp_pct * hp_weight) + (def_dmg_pct * dmg_weight)

    # Guarantee a winner — nudge attacker score if truly tied
    if att_score == def_score:
        att_score += 0.001

    return att_score, def_score


# ─────────────────────────────────────────────────────────────────────────────
# DURABILITY HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _calc_durability_loss(base_loss: int, special: dict | None) -> int:
    """Apply durability_reduction modifier from special item.
    base_loss * (1 - durability_reduction), minimum 1."""
    if not special or not special.get("durability_reduction"):
        return base_loss
    reduction = special["durability_reduction"]
    return max(1, int(base_loss * (1 - reduction)))


def calc_special_item_round_loss(special: dict | None) -> int:
    """SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND % of 100 per round, both sides."""
    if not special:
        return 0
    settings = get_all_settings()
    pct = settings.get("SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND",
                       cfg.SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND)
    return max(1, int(100 * pct))


def apply_pvp_loss_durability_hits(player_id: int, equipped: dict):
    """Additional durability hits on all equipped gear when losing PvP.
    Called from post-combat resolution BEFORE item steal roll.
    Imported and called by combat/actions.py."""
    from database import execute_write, exclusive_transaction
    for slot, item in equipped.items():
        if item is None:
            continue
        inv_id   = item["inv_id"]
        new_dur  = max(0, item["current_durability"] - 10)  # flat -10 on PvP loss
        with exclusive_transaction():
            execute_write(
                "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
                (new_dur, inv_id)
            )


# ─────────────────────────────────────────────────────────────────────────────
# XP SCALING
# ─────────────────────────────────────────────────────────────────────────────

def calc_xp_reward(base_xp: int, winner_level: int, opponent_level: int,
                   xp_multiplier: float = 0.0) -> int:
    """Scale XP by level difference. Higher opponent = bonus, lower = penalty."""
    level_diff = opponent_level - winner_level
    if level_diff > 0:
        scale = 1.0 + (level_diff * 0.1)    # +10% per level above
    elif level_diff < 0:
        scale = max(0.1, 1.0 + (level_diff * 0.15))  # -15% per level below, min 10%
    else:
        scale = 1.0

    raw = int(base_xp * scale)
    if xp_multiplier:
        raw = int(raw * (1 + xp_multiplier))
    return max(0, raw)


def check_level_up(player_id: int, current_xp: int, current_level: int) -> bool:
    """Check if player has enough XP to level up. Sets pending_levelup if so.
    Returns True if a level-up occurred."""
    settings  = get_all_settings()
    xp_curve  = cfg.XP_CURVE
    next_level = current_level + 1
    if next_level > 15:
        return False  # Max level — XP accumulates but no more level-ups
    threshold = xp_curve.get(next_level)
    if threshold is None or current_xp < threshold:
        return False

    from database import execute_write, exclusive_transaction
    with exclusive_transaction():
        execute_write(
            "UPDATE players SET level = level + 1, pending_levelup = 1 WHERE id = ?",
            (player_id,)
        )
    return True


################################################################################

# FILE: combat/flavour.py
"""Build concise player-facing narrative text for combat outcomes."""
# combat/flavour.py
# Generates all flavor text strings for combat logs, feed entries,
# and event results. Keeps narrative text out of logic files.
# All functions return plain strings ready for template rendering.

import random


# ─────────────────────────────────────────────────────────────────────────────
# COMBAT INTRO
# ─────────────────────────────────────────────────────────────────────────────

def combat_intro(combat_type: str, opponent_name: str,
                 boss_flavor: str = "", boss_phase: int = 1) -> str:
    """Build player-facing narrative text for combat intro."""
    if combat_type == "BOSS":
        lines = [
            f"═══ BOSS FIGHT: {opponent_name.upper()} ════════════════════════════════",
        ]
        if boss_flavor:
            lines.append(boss_flavor)
        if boss_phase > 1:
            lines.append(f"[Phase {boss_phase} active]")
        return "\n".join(lines)
    elif combat_type == "MINION":
        return (f"═══ MINION ENCOUNTER: {opponent_name.upper()} ══════════════════════\n"
                f"A lesser foe blocks your path!")
    else:  # PVP
        return f"═══ PVP: You challenge {opponent_name} ═══════════════════════════════"


def round_header(round_num: int) -> str:
    """Build player-facing narrative text for round header."""
    return f"─── Round {round_num} ─────────────────────────────────────────────────────"


def combat_warning(warning_type: str, opponent_name: str = "",
                   level_diff: int = 0) -> str:
    """Build player-facing narrative text for combat warning."""
    if warning_type == "empty_weapon":
        return "⚠ WARNING: You are unarmed. Fists deal minimal damage."
    elif warning_type == "empty_armor":
        return "⚠ WARNING: You have no armor equipped."
    elif warning_type == "level_mismatch":
        return (f"⚠ WARNING: {opponent_name} is {level_diff} levels above you. "
                f"This fight may be extremely dangerous. Proceed?")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# ATTACK FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

ATTACK_VERBS_MELEE  = ["swings", "strikes at", "slashes at", "lunges at", "hammers"]
ATTACK_VERBS_RANGED = ["fires at", "shoots at", "takes aim at", "blasts", "unleashes on"]
DODGE_VERBS         = ["sidesteps", "ducks under", "narrowly evades", "deflects", "rolls away from"]
HIT_VERBS           = ["connects with", "lands a hit on", "strikes", "hits"]


def attack_flavor(attacker_name: str, weapon_name: str,
                  weapon_type: str,
                  hit: bool, dodged: bool, is_crit: bool,
                  damage: int, damage_type: str,
                  res_note: str = "") -> str:
    """Build player-facing narrative text for attack flavor."""
    verb = random.choice(ATTACK_VERBS_MELEE if weapon_type == "Melee" else ATTACK_VERBS_RANGED)

    if dodged:
        evade = random.choice(DODGE_VERBS)
        return f"{attacker_name} {verb} with the {weapon_name} — opponent {evade}! (Dodged)"

    if not hit:
        return f"{attacker_name} {verb} with the {weapon_name} — Miss!"

    if is_crit:
        line = f"★ CRITICAL HIT! {attacker_name} {verb} with the {weapon_name}!"
    else:
        hit_v = random.choice(HIT_VERBS)
        line  = f"{attacker_name} {verb} with the {weapon_name} and {hit_v}!"

    line += f" {damage} {damage_type} damage"
    if res_note:
        line += f" ({res_note})"
    line += "."
    return line


def bonus_damage_flavor(damage: int, damage_type: str, res_note: str = "") -> str:
    """Build player-facing narrative text for bonus damage flavor."""
    line = f"  → Bonus {damage_type} damage: {damage}"
    if res_note:
        line += f" ({res_note})"
    return line + "."


# ─────────────────────────────────────────────────────────────────────────────
# STEAL FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

def steal_flavor(attacker_name: str, target_name: str, success: bool,
                 item_name: str = "", credits: int = 0,
                 xp_bonus: int = 0, is_vs_boss: bool = False) -> str:
    """Build player-facing narrative text for steal flavor."""
    if not success:
        return (f"{attacker_name} attempts to steal from {target_name} — "
                f"caught! AC penalty incoming.")
    if is_vs_boss:
        if item_name:
            return (f"{attacker_name} makes a daring grab and seizes "
                    f"the {item_name} from {target_name}!")
        else:
            return (f"{attacker_name} pilfers {credits} credits worth of valuables "
                    f"from {target_name}.")
    # vs player cascade
    parts = []
    if item_name:
        parts.append(f"snatched the {item_name}")
    if credits:
        parts.append(f"took {credits} credits")
    if xp_bonus:
        parts.append(f"+{xp_bonus} XP consolation")
    result = " → ".join(parts) if parts else "nothing left to steal"
    return f"{attacker_name} steals from {target_name}: {result}."


# ─────────────────────────────────────────────────────────────────────────────
# BRACE FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

def brace_flavor(player_name: str, hp_restored: int,
                 ac_bonus: int, dodge_bonus: int) -> str:
    """Build player-facing narrative text for brace flavor."""
    line = f"{player_name} takes a defensive stance, bracing for impact."
    if hp_restored:
        line += f" +{hp_restored} HP."
    line += f" AC+{ac_bonus}, Dodge+{dodge_bonus} until next hit."
    return line


# ─────────────────────────────────────────────────────────────────────────────
# ESCAPE FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

def escape_flavor(player_name: str, success: bool,
                  credits_lost: int = 0) -> str:
    """Build player-facing narrative text for escape flavor."""
    if success:
        line = f"{player_name} breaks away and flees the fight!"
        if credits_lost:
            line += f" {credits_lost} credits spilled in the retreat."
        return line
    return (f"{player_name} tries to escape but is cut off! "
            f"AC reduced for the next incoming attack.")


# ─────────────────────────────────────────────────────────────────────────────
# OBSERVE FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

def observe_flavor(player_name: str, success: bool,
                   opponent_name: str, revealed: dict | None = None) -> str:
    """Build player-facing narrative text for observe flavor."""
    if not success:
        return (f"{player_name} tries to read {opponent_name} but reveals nothing.")
    line = f"{player_name} studies {opponent_name} carefully."
    if revealed:
        parts = []
        if revealed.get("resistances"):
            parts.append(f"Resistant: {', '.join(revealed['resistances'])}")
        if revealed.get("weaknesses"):
            parts.append(f"Weak to: {', '.join(revealed['weaknesses'])}")
        if revealed.get("exact_hp") is not None:
            parts.append(f"HP: {revealed['exact_hp']}")
        if parts:
            line += " — " + " | ".join(parts)
    return line


# ─────────────────────────────────────────────────────────────────────────────
# SWAP GEAR FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

def swap_gear_flavor(player_name: str, new_item_name: str) -> str:
    """Build player-facing narrative text for swap gear flavor."""
    return (f"{player_name} quickly swaps to {new_item_name}. "
            f"Attack and AC reduced this round.")


# ─────────────────────────────────────────────────────────────────────────────
# BOSS SPECIAL MOVES
# ─────────────────────────────────────────────────────────────────────────────

def boss_special_attack_flavor(boss_name: str, attack_name: str,
                                damage: int, attack_flavor_text: str = "") -> str:
    """Build player-facing narrative text for boss special attack flavor."""
    line = f"★ {boss_name.upper()} uses {attack_name}!"
    if attack_flavor_text:
        line += f" {attack_flavor_text}"
    line += f" {damage} damage!"
    return line


def boss_special_buff_flavor(boss_name: str, buff_name: str,
                              buff_flavor_text: str = "") -> str:
    """Build player-facing narrative text for boss special buff flavor."""
    line = f"★ {boss_name.upper()} activates {buff_name}!"
    if buff_flavor_text:
        line += f" {buff_flavor_text}"
    line += " (Effect lasts rest of fight.)"
    return line


# ─────────────────────────────────────────────────────────────────────────────
# POST-COMBAT RESULT (for feeds)
# ─────────────────────────────────────────────────────────────────────────────

def combat_result_flavor(winner_name: str, loser_name: str,
                         combat_type: str,
                         credits_stolen: int = 0,
                         item_stolen: str = "",
                         result_type: str = "1HP_WIN") -> str:
    """Global feed entry for a completed fight."""
    if combat_type == "BOSS":
        line = f"{winner_name} has defeated {loser_name}!"
    elif combat_type == "MINION":
        line = f"{winner_name} dispatched a {loser_name}."
    else:  # PVP
        if result_type == "ESCAPE":
            return f"{loser_name} fled a fight with {winner_name}."
        elif result_type == "SCORE_WIN":
            line = f"{winner_name} outlasted {loser_name} in a grinding battle!"
        else:
            line = f"{winner_name} defeated {loser_name} in combat!"

    if credits_stolen:
        line += f" Claimed {credits_stolen} credits."
    if item_stolen:
        line += f" Seized the {item_stolen}!"
    return line


def level_up_flavor(player_name: str, new_level: int) -> str:
    """Build player-facing narrative text for level up flavor."""
    return f"⬆ {player_name} has reached Level {new_level}!"


def special_item_pool_flavor(item_name: str, reason: str = "released") -> str:
    """Global feed: special item returned to/entered the loot pool."""
    if reason == "destroyed":
        return f"The {item_name} has been destroyed and returned to the loot pool."
    elif reason == "sold":
        return f"The {item_name} is now available in the shop."
    else:
        return f"The {item_name} has returned to the loot pool."


# ─────────────────────────────────────────────────────────────────────────────
# RANDOM EVENT FLAVOR
# ─────────────────────────────────────────────────────────────────────────────

def random_event_flavor(event: dict, player_name: str) -> str:
    """Render event flavor text with player name substituted."""
    text = event.get("flavor_text", "Something happens...")
    return text.replace("{player}", player_name).replace("{name}", player_name)


# ─────────────────────────────────────────────────────────────────────────────
# HP DESCRIPTOR
# ─────────────────────────────────────────────────────────────────────────────

def hp_descriptor(current_hp: int, max_hp: int, observed: bool = False) -> str:
    """Return HP description. Exact value if observed, tier name otherwise."""
    if observed:
        return f"{current_hp}/{max_hp} HP"
    pct = (current_hp / max_hp * 100) if max_hp else 0
    if pct >= 76: return "Healthy"
    if pct >= 51: return "Wounded"
    if pct >= 26: return "Hurt"
    if pct >= 2:  return "Critical"
    return "Near Death"


################################################################################

# FILE: combat/__init__.py

"""Combat calculation, state transition, and narrative modules."""

# FILE: routes/actions.py
"""HTTP and queued handlers for Tavern, boss, minion, PvP, and random events."""
# routes/actions.py  (Phase 5 — full implementation)
# Terminal-fragment POST routes for all AP actions.
# Replaces the Phase 3 stub with full boss/minion/random event logic.

import math
import random
import logging
from datetime import datetime

from flask import Blueprint, render_template, request, session, g
from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_all_settings, get_player)
from queue_handler import enqueue_and_process, register_handler
from combat import actions as combat_actions
from combat import flavour
import config_defaults as cfg

bp = Blueprint("actions", __name__)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _error_fragment(message: str) -> str:
    """Provide the internal error fragment operation used by this module."""
    return render_template("fragments/error.html", message=message,
                           player=g.get("player"))


def _with_random_event(event: dict | None, player: dict, content: str) -> str:
    """Prepend a triggered event without replacing the requested action."""
    if not event:
        return content
    event_html = render_template("fragments/event_result.html", event=event, player=player)
    return event_html + content


def _check_ap(player: dict, cost: int) -> str | None:
    """Provide the internal check ap operation used by this module."""
    if player["current_ap"] < cost:
        return _error_fragment(f"Not enough AP. Need {cost}, have {player['current_ap']}.")
    return None


def _deduct_ap_and_regen(player_id: int, player: dict, cost: int, settings: dict):
    """Deduct AP cost and apply passive HP regen. Must be inside exclusive_transaction."""
    ap_regen    = settings.get("AP_PASSIVE_HP_REGEN",   cfg.AP_PASSIVE_HP_REGEN)
    end_divisor = settings.get("END_HP_REGEN_DIVISOR",  cfg.END_HP_REGEN_DIVISOR)
    hp_regen    = ap_regen + math.floor(player["end_stat"] / end_divisor)

    if player.get("equipped_special_id"):
        inv = execute_one("SELECT item_id FROM inventory_items WHERE id = ?",
                          (player["equipped_special_id"],))
        if inv:
            s = execute_one("SELECT hp_regen_bonus FROM special_items WHERE id = ?",
                            (inv["item_id"],))
            if s:
                hp_regen += s["hp_regen_bonus"]

    max_hp = 10 + player["end_stat"] + (5 * player["level"])
    new_ap = player["current_ap"] - cost
    new_hp = min(player["current_hp"] + hp_regen, max_hp)
    execute_write(
        "UPDATE players SET current_ap = ?, current_hp = ? WHERE id = ?",
        (new_ap, new_hp, player_id)
    )
    return new_ap, new_hp


# ─────────────────────────────────────────────────────────────────────────────
# RANDOM EVENT CHECK
# ─────────────────────────────────────────────────────────────────────────────

def _describe_random_event_effect(event: dict) -> str:
    """Return a short player-facing description of an event's mechanics."""
    effect = event["effect_type"]
    amount = int(event["effect_amount"])
    stat_names = {
        "STAT_BOOST_STR": "STR", "STAT_BOOST_END": "END",
        "STAT_BOOST_AGI": "AGI", "STAT_BOOST_LCK": "LCK",
        "STAT_BOOST_PER": "PER", "STAT_BOOST_INITIATIVE": "Initiative",
        "STAT_PENALTY_STR": "STR", "STAT_PENALTY_END": "END",
        "STAT_PENALTY_AGI": "AGI", "STAT_PENALTY_LCK": "LCK",
        "STAT_PENALTY_PER": "PER", "STAT_PENALTY_INITIATIVE": "Initiative",
    }
    if effect in stat_names:
        return f"{stat_names[effect]} {amount:+d}"
    if effect == "CREDITS": return f"Credits {amount:+d}"
    if effect == "BONUS_AP": return f"AP +{amount} (up to your AP cap)"
    if effect == "HP_LOSS": return f"HP {amount:+d} (cannot reduce you below 1 HP)"
    if effect == "XP_LOSS": return f"XP {amount:+d} (cannot reduce XP below 0)"
    if effect == "AP_REDUCTION_PERCENT": return f"Daily AP award -{abs(amount)}%"
    if effect == "DURABILITY_RESTORE_RANDOM": return f"Restore up to {abs(amount)} durability on one random item"
    if effect == "DURABILITY_LOSS_RANDOM": return f"One random item's durability {amount:+d}"
    if effect == "ITEM_AT_LEVEL": return "Receive one level-appropriate weapon or armor"
    if effect == "SPECIAL_ITEM_FROM_POOL": return "Receive one available unique special item"
    if effect == "PROTAGONIST_ENCOUNTER": return "Receive a reward from a protagonist encounter"
    return effect.replace("_", " ").title()


def check_random_event(player: dict, settings: dict) -> dict | None:
    """Roll for a random event. Returns event dict if triggered, else None.
    Fires BEFORE AP is deducted. If triggered: apply effect, write feeds."""
    base_chance   = settings.get("RANDOM_EVENT_BASE_CHANCE", cfg.RANDOM_EVENT_BASE_CHANCE)
    max_chance    = settings.get("RANDOM_EVENT_MAX_CHANCE",  cfg.RANDOM_EVENT_MAX_CHANCE)
    good_base     = settings.get("RANDOM_EVENT_GOOD_BASE",   cfg.RANDOM_EVENT_GOOD_BASE)
    good_max      = settings.get("RANDOM_EVENT_GOOD_MAX",    cfg.RANDOM_EVENT_GOOD_MAX)
    bad_min       = settings.get("RANDOM_EVENT_BAD_MIN",     cfg.RANDOM_EVENT_BAD_MIN)
    lck_bonus_per = settings.get("RANDOM_EVENT_LCK_BONUS",   cfg.RANDOM_EVENT_LCK_BONUS)

    lck_steps    = math.floor(player["lck_stat"] / 2)
    # Check for encounter_bonus from equipped special
    encounter_bonus = 0
    if player.get("equipped_special_id"):
        inv = execute_one("SELECT item_id FROM inventory_items WHERE id = ?",
                          (player["equipped_special_id"],))
        if inv:
            s = execute_one("SELECT encounter_bonus FROM special_items WHERE id = ?",
                            (inv["item_id"],))
            if s:
                encounter_bonus = math.floor(s["encounter_bonus"] * 20)  # treat as bonus LCK steps

    effective_lck_steps = lck_steps + encounter_bonus
    trigger_chance = min(base_chance + effective_lck_steps * lck_bonus_per, max_chance)
    if random.random() >= trigger_chance:
        return None

    # Determine good vs bad
    good_chance = min(good_base + effective_lck_steps * lck_bonus_per, good_max)
    bad_chance  = max(1.0 - good_chance, bad_min)
    is_good     = random.random() < good_chance

    # Pull matching events from DB
    events = execute(
        "SELECT * FROM random_events WHERE is_active = 1 AND event_type = ?",
        ("Good" if is_good else "Bad",)
    )
    if not events:
        return None

    # Avoid showing the same event twice in a row to the same player. The
    # previous feed text is enough to identify it without adding schema state.
    last_event = execute_one(
        """SELECT flavor_text FROM daily_feed
           WHERE player_id = ? AND event_category = 'RANDOM_EVENT'
           ORDER BY id DESC LIMIT 1""",
        (player["id"],)
    )
    if last_event and len(events) > 1:
        previous_text = last_event["flavor_text"]
        alternatives = [
            event for event in events
            if flavour.random_event_flavor(
                event, player.get("character_name", "Player")
            ) != previous_text
        ]
        if alternatives:
            events = alternatives

    # Weight by rarity (LCK improves chance of Rare/Uncommon in good pool)
    if is_good:
        weights = {"Common": 60, "Uncommon": 30 + effective_lck_steps,
                   "Rare": 10 + effective_lck_steps * 2}
    else:
        weights = {"Common": 60, "Uncommon": 30, "Rare": 10}

    weighted_pool = []
    for e in events:
        w = weights.get(e["rarity"], 30)
        weighted_pool.extend([e] * w)
    if not weighted_pool:
        return None

    event = random.choice(weighted_pool)
    event["effect_summary"] = _describe_random_event_effect(event)
    _apply_random_event(player["id"], player, event, settings)
    return event


def _apply_random_event(player_id: int, player: dict, event: dict, settings: dict):
    """Apply event effect to the player. Writes to DB and daily_feed."""
    effect    = event["effect_type"]
    amount    = event["effect_amount"]
    ap_cap    = settings.get("AP_CARRYOVER_CAP", cfg.AP_CARRYOVER_CAP)
    player_name = player.get("character_name", "Player")

    with exclusive_transaction():
        if effect == "CREDITS":
            if amount >= 0:
                execute_write("UPDATE players SET credits = credits + ? WHERE id = ?",
                              (amount, player_id))
            else:
                execute_write(
                    "UPDATE players SET credits = MAX(0, credits + ?) WHERE id = ?",
                    (amount, player_id)
                )

        elif effect == "BONUS_AP":
            execute_write(
                "UPDATE players SET current_ap = MIN(current_ap + ?, ?) WHERE id = ?",
                (amount, ap_cap, player_id)
            )

        elif effect == "HP_LOSS":
            max_hp = 10 + player["end_stat"] + (5 * player["level"])
            new_hp = max(1, player["current_hp"] + amount)  # amount is negative
            execute_write("UPDATE players SET current_hp = ? WHERE id = ?", (new_hp, player_id))

        elif effect == "XP_LOSS":
            execute_write(
                "UPDATE players SET xp = MAX(0, xp + ?) WHERE id = ?",
                (amount, player_id)  # amount is negative
            )

        elif effect == "AP_REDUCTION_PERCENT":
            # Cursed: insert status_effect
            execute_write(
                """INSERT INTO status_effects (player_id, effect_type, value)
                   VALUES (?, 'CURSED', ?)""",
                (player_id, abs(amount) / 100)
            )

        elif effect == "ITEM_AT_LEVEL":
            # Random weapon or armor at or above player level
            table = random.choice(["weapons", "armor"])
            item  = execute_one(
                f"SELECT * FROM {table} WHERE level >= ? AND is_active = 1 ORDER BY RANDOM() LIMIT 1",
                (player["level"],)
            )
            if item is None:
                item = execute_one(
                    f"SELECT * FROM {table} WHERE is_active = 1 ORDER BY level DESC LIMIT 1"
                )
            if item:
                item_type = "WEAPON" if table == "weapons" else "ARMOR"
                inv_id = execute_write(
                    """INSERT INTO inventory_items
                       (player_id, item_type, item_id, current_durability, acquired_method)
                       VALUES (?, ?, ?, ?, 'RANDOM_EVENT')""",
                    (player_id, item_type, item["id"], item["starting_durability"])
                )
                execute_write(
                    """INSERT INTO item_history
                       (player_id, item_type, item_id, item_name, event_type)
                       VALUES (?, ?, ?, ?, 'RECEIVED_RANDOM_EVENT')""",
                    (player_id, item_type, item["id"], item["name"])
                )

        elif effect == "DURABILITY_RESTORE_RANDOM":
            # Random item including equipped — use repair formula, no cost
            all_inv = execute(
                "SELECT * FROM inventory_items WHERE player_id = ?", (player_id,)
            )
            if all_inv:
                target = random.choice(all_inv)
                if target["current_durability"] < 100:
                    missing  = 100 - target["current_durability"]
                    base_pct = settings.get("REPAIR_BASE_PERCENT", cfg.REPAIR_BASE_PERCENT)
                    restore  = max(1, int(missing * base_pct))
                    new_dur  = min(100, target["current_durability"] + restore)
                    execute_write(
                        "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
                        (new_dur, target["id"])
                    )

        elif effect == "DURABILITY_LOSS_RANDOM":
            all_inv = execute(
                "SELECT * FROM inventory_items WHERE player_id = ?", (player_id,)
            )
            if all_inv:
                target  = random.choice(all_inv)
                new_dur = max(0, target["current_durability"] + amount)  # amount negative
                if new_dur == 0:
                    execute_write("DELETE FROM inventory_items WHERE id = ?", (target["id"],))
                else:
                    execute_write(
                        "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
                        (new_dur, target["id"])
                    )

        elif effect in (
            "STAT_BOOST_STR", "STAT_BOOST_END", "STAT_BOOST_AGI",
            "STAT_BOOST_LCK", "STAT_BOOST_PER", "STAT_BOOST_INITIATIVE",
            "STAT_PENALTY_STR", "STAT_PENALTY_END", "STAT_PENALTY_AGI",
            "STAT_PENALTY_LCK", "STAT_PENALTY_PER", "STAT_PENALTY_INITIATIVE",
        ):
            execute_write(
                "INSERT INTO status_effects (player_id, effect_type, value) VALUES (?, ?, ?)",
                (player_id, effect, float(amount))
            )

        elif effect == "PROTAGONIST_ENCOUNTER":
            # Applied after this transaction because the handler owns its transaction.
            pass

        elif effect == "SPECIAL_ITEM_FROM_POOL":
            # Rare: give player a random special item from the pool
            available = execute(
                """SELECT si.id FROM special_items si
                   JOIN special_item_registry sir ON sir.special_item_id = si.id
                   WHERE sir.status = 'IN_POOL' AND si.is_active = 1 ORDER BY RANDOM() LIMIT 1"""
            )
            if available:
                special_id = available[0]["id"]
                inv_id = execute_write(
                    """INSERT INTO inventory_items
                       (player_id, item_type, item_id, current_durability, acquired_method)
                       VALUES (?, 'SPECIAL', ?, 100, 'RANDOM_EVENT')""",
                    (player_id, special_id)
                )
                execute_write(
                    """UPDATE special_item_registry
                       SET status='IN_INVENTORY', current_owner_player_id=?,
                           inventory_item_id=?, last_acquired_method='RANDOM_EVENT', updated_at=?
                       WHERE special_item_id=?""",
                    (player_id, inv_id, datetime.utcnow().isoformat(), special_id)
                )

        # Personal feed entry
        feed_text = flavour.random_event_flavor(event, player_name)
        execute_write(
            """INSERT INTO daily_feed (feed_scope, player_id, flavor_text, event_category)
               VALUES ('PERSONAL', ?, ?, 'RANDOM_EVENT')""",
            (player_id, feed_text)
        )

    if effect == "PROTAGONIST_ENCOUNTER":
        _handle_protagonist_encounter(player_id, player, settings)


# ─────────────────────────────────────────────────────────────────────────────
# TAVERN  (carried forward from Phase 3 — already complete)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/action/tavern", methods=["POST"])
def action_tavern():
    """Handle the action tavern workflow."""
    player   = g.player
    settings = get_all_settings()
    cost_ap  = settings.get("AP_COST_TAVERN",   cfg.AP_COST_TAVERN)
    cost_cr  = settings.get("TAVERN_HEAL_COST", cfg.TAVERN_HEAL_COST)

    if player["current_hp"] >= player["max_hp"]:
        return _error_fragment("You are already at full health.")
    if err := _check_ap(player, cost_ap):
        return err
    if player["credits"] < cost_cr:
        return _error_fragment(f"Not enough credits. Need {cost_cr}.")

    result = enqueue_and_process(
        player["id"], "tavern_heal", {"cost_ap": cost_ap, "cost_cr": cost_cr}
    )
    return render_template("fragments/tavern_result.html", **result, player=player)


@register_handler("tavern_heal")
def handle_tavern_heal(player_id: int, payload: dict) -> dict:
    """Process the queued tavern heal action against validated game state."""
    settings  = get_all_settings()
    cost_ap   = payload["cost_ap"]
    cost_cr   = payload["cost_cr"]
    heal_pct  = settings.get("TAVERN_HEAL_PERCENT", cfg.TAVERN_HEAL_PERCENT)
    player    = get_player(player_id)
    max_hp    = 10 + player["end_stat"] + (5 * player["level"])
    missing   = max_hp - player["current_hp"]
    if missing <= 0:
        raise ValueError("Already at full health.")
    heal_amount = max(1, int(missing * heal_pct))
    with exclusive_transaction():
        new_ap, new_hp_regen = _deduct_ap_and_regen(player_id, player, cost_ap, settings)
        final_hp = min(new_hp_regen + heal_amount, max_hp)
        execute_write(
            "UPDATE players SET current_hp = ?, credits = credits - ? WHERE id = ?",
            (final_hp, cost_cr, player_id)
        )
    return {"heal_amount": heal_amount, "new_hp": final_hp, "max_hp": max_hp,
            "new_ap": new_ap, "max_ap": player["max_ap"],
            "new_credits": player["credits"] - cost_cr, "cost_cr": cost_cr}


# ─────────────────────────────────────────────────────────────────────────────
# BOSS FIGHT
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/action/boss", methods=["POST"])
def action_boss():
    """Handle the action boss workflow."""
    player   = g.player
    settings = get_all_settings()
    cost_ap  = settings.get("AP_COST_BOSS", cfg.AP_COST_BOSS)

    if player["in_combat"]:
        return _error_fragment("You are already in combat.")
    if g.get("blackout"):
        return _error_fragment("Combat unavailable — midnight reset approaching.")
    if err := _check_ap(player, cost_ap):
        return err

    # Random event check (fires before AP deducted)
    event = check_random_event(player, settings)
    if event:
        player = get_player(player["id"]) or player

    # 50/50 boss vs minion roll
    minion_chance = settings.get("MINION_ENCOUNTER_CHANCE", cfg.MINION_ENCOUNTER_CHANCE)
    encounter_type = "MINION" if random.random() < minion_chance else "BOSS"

    # Determine which boss/minion
    existing_instances = execute(
        f"SELECT * FROM {'boss' if encounter_type == 'BOSS' else 'minion'}_instances "
        f"WHERE player_id = ? ORDER BY discovered_at DESC",
        (player["id"],)
    )
    discovered_ids = [i[f"{'boss' if encounter_type == 'BOSS' else 'minion'}_id"]
                      for i in existing_instances]

    # Get a random undiscovered one, or random from all if all discovered
    tbl = "bosses" if encounter_type == "BOSS" else "minions"
    if discovered_ids:
        undiscovered = execute(
            f"SELECT * FROM {tbl} WHERE is_active = 1 AND id NOT IN ({','.join('?' * len(discovered_ids))}) ORDER BY RANDOM() LIMIT 1",
            tuple(discovered_ids)
        )
        opponent = undiscovered[0] if undiscovered else random.choice(
            execute(f"SELECT * FROM {tbl} WHERE is_active = 1")
        )
    else:
        result = execute(f"SELECT * FROM {tbl} WHERE is_active = 1 ORDER BY RANDOM() LIMIT 1")
        if not result:
            return _with_random_event(
                event, player,
                _error_fragment("No content available yet. Ask the admin to import game content.")
            )
        opponent = result[0]

    # Level warning check
    warn_threshold = settings.get("BOSS_LEVEL_WARNING_THRESHOLD", cfg.BOSS_LEVEL_WARNING_THRESHOLD)
    level_diff     = opponent["level"] - player["level"]
    if level_diff >= warn_threshold:
        content = render_template("fragments/boss_confirm.html",
                                  opponent=opponent,
                                  encounter_type=encounter_type,
                                  level_diff=level_diff,
                                  player=player)
        return _with_random_event(event, player, content)

    # Minion: PER check
    if encounter_type == "MINION":
        per_result = _minion_per_check(player, opponent)
        if per_result["spotted"]:
            content = render_template("fragments/minion_spotted.html",
                                      minion=opponent,
                                      per_result=per_result,
                                      player=player)
            return _with_random_event(event, player, content)

    # Start the fight
    content = _start_boss_fight(player, opponent, encounter_type, cost_ap, settings)
    return _with_random_event(event, player, content)


@bp.route("/action/boss/confirm", methods=["POST"])
def action_boss_confirm():
    """Player confirmed they want to fight despite level warning or spotted minion."""
    player       = g.player
    settings     = get_all_settings()
    cost_ap      = settings.get("AP_COST_BOSS",   cfg.AP_COST_BOSS)
    opponent_id  = request.form.get("opponent_id", type=int)
    encounter_type = request.form.get("encounter_type", "BOSS")
    action       = request.form.get("action", "fight")  # fight or avoid

    if action == "avoid":
        # AP refunded — just return a message
        return render_template("fragments/event_result.html",
                               event={"flavor_text": "You slip past undetected.",
                                      "event_type": "Good"},
                               player=player)

    tbl = "bosses" if encounter_type == "BOSS" else "minions"
    opponent = execute_one(f"SELECT * FROM {tbl} WHERE id = ?", (opponent_id,))
    if not opponent:
        return _error_fragment("Opponent not found.")

    return _start_boss_fight(player, opponent, encounter_type, cost_ap, settings)


def _minion_per_check(player: dict, minion: dict) -> dict:
    """Provide the internal minion per check operation used by this module."""
    from combat.engine import resolve_opposed_roll
    result = resolve_opposed_roll(
        actor_agi=player["agi_stat"], actor_lck=player["lck_stat"],
        defender_agi=minion["agi_stat"], defender_lck=minion["lck_stat"],
        actor_per=player["per_stat"], defender_per=minion.get("per_stat", 0),
        tie_goes_to="defender"
    )
    return {"spotted": result["success"], "detail": result["detail"]}


def _start_boss_fight(player: dict, opponent: dict, encounter_type: str,
                      cost_ap: int, settings: dict):
    """Initiate a boss or minion fight. Deducts AP, sets in_combat, creates session."""
    result = enqueue_and_process(
        player["id"], "start_boss_fight",
        {"opponent_id": opponent["id"], "encounter_type": encounter_type, "cost_ap": cost_ap}
    )
    if result.get("error"):
        return _error_fragment(result["error"])

    session["combat_session_id"] = result["session_id"]

    opponent_full = execute_one(
        f"SELECT * FROM {'bosses' if encounter_type == 'BOSS' else 'minions'} WHERE id = ?",
        (opponent["id"],)
    )
    # Show boss intel if previously observed
    intel = None
    intel_detail = None
    if encounter_type == "BOSS":
        intel = execute_one(
            "SELECT * FROM boss_intel WHERE player_id = ? AND boss_id = ?",
            (player["id"], opponent["id"])
        )
        if intel:
            damage_types = ("blade", "blunt", "ballistic", "energy", "arcane", "explosive", "venom")
            intel_detail = {
                "resistances": [t.upper() for t in damage_types if opponent_full.get(f"res_{t}")],
                "weaknesses": [t.upper() for t in damage_types if opponent_full.get(f"weak_{t}")],
                "special_attack_name": opponent_full.get("special_attack_name"),
                "special_attack_type": opponent_full.get("special_attack_damage_type"),
                "special_buff_name": opponent_full.get("special_buff_name"),
                "special_buff_type": opponent_full.get("special_buff_type"),
                "current_hp": opponent_full.get("current_hp"),
                "max_hp": opponent_full.get("max_hp"),
            }

    return render_template("fragments/combat_open.html",
                           opponent=opponent_full,
                           encounter_type=encounter_type,
                           session_id=result["session_id"],
                           intel=intel,
                           intel_detail=intel_detail,
                           player=player,
                           boss_flavor=opponent_full.get("flavor_text", ""))


@register_handler("start_boss_fight")
def handle_start_boss_fight(player_id: int, payload: dict) -> dict:
    """Process the queued start boss fight action against validated game state."""
    opponent_id    = payload["opponent_id"]
    encounter_type = payload["encounter_type"]
    cost_ap        = payload["cost_ap"]

    player   = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    settings = get_all_settings()

    if player["in_combat"]:
        return {"error": "Already in combat."}
    if player["current_ap"] < cost_ap:
        return {"error": "Not enough AP."}

    tbl = "bosses" if encounter_type == "BOSS" else "minions"
    opponent = execute_one(f"SELECT * FROM {tbl} WHERE id = ?", (opponent_id,))
    if not opponent:
        return {"error": "Opponent not found."}

    with exclusive_transaction():
        new_ap, new_hp = _deduct_ap_and_regen(player_id, player, cost_ap, settings)
        execute_write("UPDATE players SET in_combat = 1 WHERE id = ?", (player_id,))

        # Get or create boss/minion instance
        inst_tbl  = f"{'boss' if encounter_type == 'BOSS' else 'minion'}_instances"
        id_col    = f"{'boss' if encounter_type == 'BOSS' else 'minion'}_id"
        existing  = execute_one(
            f"SELECT * FROM {inst_tbl} WHERE player_id = ? AND {id_col} = ?",
            (player_id, opponent_id)
        )
        if existing:
            # Reset for new fight
            execute_write(
                f"UPDATE {inst_tbl} SET current_hp = ?, special_attack_used = 0, "
                f"special_buff_used = 0, current_phase = 1 WHERE id = ?",
                (opponent["max_hp"], existing["id"])
            )
            inst_id = existing["id"]
        else:
            inst_id = execute_write(
                f"INSERT INTO {inst_tbl} (player_id, {id_col}, current_hp) VALUES (?, ?, ?)",
                (player_id, opponent_id, opponent["max_hp"])
            )

        # Create combat session
        if encounter_type == "BOSS":
            session_id = execute_write(
                """INSERT INTO combat_sessions
                   (combat_type, attacker_player_id, boss_instance_id, status,
                    attacker_hp_start)
                   VALUES ('BOSS', ?, ?, 'ACTIVE', ?)""",
                (player_id, inst_id, new_hp)
            )
        else:
            session_id = execute_write(
                """INSERT INTO combat_sessions
                   (combat_type, attacker_player_id, minion_instance_id, status,
                    attacker_hp_start)
                   VALUES ('MINION', ?, ?, 'ACTIVE', ?)""",
                (player_id, inst_id, new_hp)
            )

    return {"session_id": session_id, "new_ap": new_ap, "new_hp": new_hp}


# ─────────────────────────────────────────────────────────────────────────────
# PVP
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/action/pvp", methods=["POST"])
def action_pvp():
    """Handle the action pvp workflow."""
    player   = g.player
    settings = get_all_settings()
    cost_ap  = settings.get("AP_COST_PVP", cfg.AP_COST_PVP)

    if player["in_combat"]:
        return _error_fragment("You are already in combat.")
    if g.get("blackout"):
        return _error_fragment("Combat unavailable — midnight reset approaching.")
    if err := _check_ap(player, cost_ap):
        return err

    # Random event check
    event = check_random_event(player, settings)
    if event:
        player = get_player(player["id"]) or player

    # Build eligible opponent list
    opponents = _get_eligible_opponents(player, settings)
    content = render_template("fragments/opponent_list.html",
                              opponents=opponents, player=player)
    return _with_random_event(event, player, content)


def _get_eligible_opponents(player: dict, settings: dict) -> list[dict]:
    """Return list of PvP-eligible opponents with HP tier and wealth tier."""
    all_players = execute(
        """SELECT p.*, ps.pvp_kills FROM players p
           LEFT JOIN player_stats ps ON ps.player_id = p.id
           WHERE p.id != ? AND p.is_banned = 0""",
        (player["id"],)
    )
    # Filter eligibility
    eligible = []
    for p in all_players:
        if p["level"] < 3:
            continue  # Level 1-2 protected
        if p["current_hp"] <= 1:
            continue
        level_diff = player["level"] - p["level"]
        if level_diff > 2:
            continue  # Can't attack more than 2 levels below
        if p["level"] < player["level"] - 2:
            continue
        max_hp = 10 + p["end_stat"] + (5 * p["level"])
        hp_pct = p["current_hp"] / max_hp * 100 if max_hp else 0
        if hp_pct >= 76:   hp_tier = "Healthy"
        elif hp_pct >= 51: hp_tier = "Wounded"
        elif hp_pct >= 26: hp_tier = "Hurt"
        else:              hp_tier = "Critical"
        eligible.append({**p, "hp_tier": hp_tier, "max_hp": max_hp})

    # Calculate wealth tiers based on credit percentile
    if eligible:
        sorted_by_credits = sorted(eligible, key=lambda x: x["credits"])
        n = len(sorted_by_credits)
        for i, p in enumerate(sorted_by_credits):
            percentile = i / n
            poor_max   = settings.get("WEALTH_TIER_POOR_MAX",   cfg.WEALTH_TIER_POOR_MAX)
            middle_max = settings.get("WEALTH_TIER_MIDDLE_MAX", cfg.WEALTH_TIER_MIDDLE_MAX)
            if percentile <= poor_max:
                p["wealth_tier"] = "Poor"
            elif percentile <= middle_max:
                p["wealth_tier"] = "Middle"
            else:
                p["wealth_tier"] = "Rich"

    return eligible


@bp.route("/action/pvp/fight", methods=["POST"])
def action_pvp_fight():
    """Handle the action pvp fight workflow."""
    player     = g.player
    settings   = get_all_settings()
    cost_ap    = settings.get("AP_COST_PVP", cfg.AP_COST_PVP)
    target_id  = request.form.get("target_id", type=int)

    if not target_id:
        return _error_fragment("No opponent selected.")

    # Re-validate server-side
    target = execute_one("SELECT * FROM players WHERE id = ?", (target_id,))
    if target is None:
        return _error_fragment("Opponent not found.")
    if target["level"] < 3:
        return _error_fragment("That player cannot be attacked.")
    if target["current_hp"] <= 1:
        return _error_fragment("That player is at 1 HP and cannot be attacked.")
    if target["in_combat"]:
        return _error_fragment("That player is already in combat.")
    level_diff = player["level"] - target["level"]
    if level_diff > 2:
        return _error_fragment("That player is too far below your level.")

    result = enqueue_and_process(
        player["id"], "start_pvp_fight",
        {"target_id": target_id, "cost_ap": cost_ap}
    )
    if result.get("error"):
        return _error_fragment(result["error"])

    session["combat_session_id"] = result["session_id"]

    return render_template("fragments/combat_open.html",
                           opponent=target,
                           encounter_type="PVP",
                           session_id=result["session_id"],
                           intel=None,
                           player=player,
                           boss_flavor="")


@register_handler("start_pvp_fight")
def handle_start_pvp_fight(player_id: int, payload: dict) -> dict:
    """Process the queued start pvp fight action against validated game state."""
    target_id = payload["target_id"]
    cost_ap   = payload["cost_ap"]
    settings  = get_all_settings()

    player  = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    target  = execute_one("SELECT * FROM players WHERE id = ?", (target_id,))

    if player["in_combat"] or target["in_combat"]:
        return {"error": "A player is already in combat."}
    if player["current_ap"] < cost_ap:
        return {"error": "Not enough AP."}

    with exclusive_transaction():
        new_ap, new_hp = _deduct_ap_and_regen(player_id, player, cost_ap, settings)
        execute_write("UPDATE players SET in_combat = 1 WHERE id = ?", (player_id,))
        execute_write("UPDATE players SET in_combat = 1 WHERE id = ?", (target_id,))
        target_max_hp = 10 + target["end_stat"] + (5 * target["level"])
        session_id = execute_write(
            """INSERT INTO combat_sessions
               (combat_type, attacker_player_id, defender_player_id, status,
                attacker_hp_start, defender_hp_start)
               VALUES ('PVP', ?, ?, 'ACTIVE', ?, ?)""",
            (player_id, target_id, new_hp, target["current_hp"])
        )

    return {"session_id": session_id, "new_ap": new_ap, "new_hp": new_hp}


################################################################################


def _handle_protagonist_encounter(player_id: int, player: dict, settings: dict):
    """Find a level-appropriate protagonist, roll 40/40/20 for weapon/armor/special,
    award the item, write feed entry. If item already taken, fall back to credits."""
    from database import execute, execute_one, execute_write, exclusive_transaction
    from datetime import datetime
    import random, math

    # Find all movies with a protagonist defined, ordered by how close
    # their boss level is to the player's current level
    movies = execute(
        """SELECT m.id, m.movie_name,
                  m.protagonist_name,
                  m.protagonist_weapon_id,
                  m.protagonist_armor_id,
                  m.protagonist_special_item_id,
                  b.level as boss_level
           FROM master m
           JOIN bosses b ON b.id = m.boss_id
           WHERE m.protagonist_name IS NOT NULL
             AND m.is_active = 1
           ORDER BY ABS(b.level - ?) ASC""",
        (player["level"],)
    )

    if not movies:
        # Fallback: award credits
        with exclusive_transaction():
            execute_write(
            "UPDATE players SET credits = credits + 50 WHERE id = ?", (player_id,)
            )
            execute_write(
            """INSERT INTO daily_feed (feed_scope, player_id, flavor_text, event_category)
               VALUES ('PERSONAL', ?, ?, 'RANDOM_EVENT')""",
            (player_id,
             "A familiar figure passes in the crowd — but vanishes before you can speak. +50 credits left behind.")
            )
        return

    # Pick from the 3 closest level matches (weighted toward closest)
    candidates = movies[:3]
    weights    = [3, 2, 1][:len(candidates)]
    movie      = random.choices(candidates, weights=weights, k=1)[0]

    protagonist = movie["protagonist_name"]
    roll        = random.random()

    if roll < 0.40:
        # Weapon
        item_type = "WEAPON"
        item_id   = movie["protagonist_weapon_id"]
        table     = "weapons"
    elif roll < 0.80:
        # Armor
        item_type = "ARMOR"
        item_id   = movie["protagonist_armor_id"]
        table     = "armor"
    else:
        # Special — check registry first
        item_type = "SPECIAL"
        item_id   = movie["protagonist_special_item_id"]
        table     = "special_items"

    if not item_id:
        # Protagonist item not defined — fallback credits
        with exclusive_transaction():
            execute_write(
                "UPDATE players SET credits = credits + 50 WHERE id = ?", (player_id,)
            )
        return

    # For specials: check if already in world
    if item_type == "SPECIAL":
        reg = execute_one(
            "SELECT status FROM special_item_registry WHERE special_item_id = ?",
            (item_id,)
        )
        if reg and reg["status"] != "IN_POOL":
            # Already taken — fall back to weapon instead
            item_type = "WEAPON"
            item_id   = movie["protagonist_weapon_id"]
            table     = "weapons"

    # Get item detail for name
    item_detail = execute_one(f"SELECT name, starting_durability FROM {table} WHERE id = ?", (item_id,))
    if not item_detail:
        return

    item_name = item_detail["name"]
    durability = item_detail.get("starting_durability", 100) or 100

    with exclusive_transaction():
        inv_id = execute_write(
            """INSERT INTO inventory_items
               (player_id, item_type, item_id, current_durability, acquired_method)
               VALUES (?, ?, ?, ?, 'RANDOM_EVENT')""",
            (player_id, item_type, item_id, durability)
        )
        execute_write(
            """INSERT INTO item_history
               (player_id, item_type, item_id, item_name, event_type)
               VALUES (?, ?, ?, ?, 'RECEIVED_RANDOM_EVENT')""",
            (player_id, item_type, item_id, item_name)
        )
        if item_type == "SPECIAL":
            execute_write(
                """UPDATE special_item_registry
                   SET status = 'IN_INVENTORY', current_owner_player_id = ?,
                       inventory_item_id = ?, last_acquired_method = 'RANDOM_EVENT',
                       updated_at = ?
                   WHERE special_item_id = ?""",
                (player_id, inv_id, datetime.utcnow().isoformat(), item_id)
            )
        # Personal feed entry
        execute_write(
            """INSERT INTO daily_feed (feed_scope, player_id, flavor_text, event_category)
               VALUES ('PERSONAL', ?, ?, 'RANDOM_EVENT')""",
            (player_id,
             f"{protagonist} looks you over and hands you the {item_name}. No words. Just a nod.")
        )
        # Global feed for special items
        if item_type == "SPECIAL":
            execute_write(
                """INSERT INTO daily_feed (feed_scope, player_id, flavor_text, event_category)
                   VALUES ('GLOBAL', NULL, ?, 'ITEM')""",
                (f"The {item_name} has entered the world.",)
            )

# FILE: routes/auth.py
"""Account registration, login, character creation, and level-up routes."""
# routes/auth.py
# Handles: login, logout, register, character creation, level-up prompt.
# All write operations go through enqueue_and_process except auth itself
# (login/register are not game actions, no AP involved).

import math
import logging
from datetime import datetime

from flask import (Blueprint, render_template, request, session,
                   redirect, url_for, flash, g)
from werkzeug.security import generate_password_hash, check_password_hash

from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_player, get_all_settings)
from queue_handler import enqueue_and_process, register_handler
import config_defaults as cfg

bp = Blueprint("auth", __name__)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/login", methods=["GET"])
def login():
    """Handle the login workflow."""
    if session.get("player_id"):
        return redirect(url_for("dashboard.index"))
    return render_template("auth/login.html")


@bp.route("/login", methods=["POST"])
def login_post():
    """Handle the login post workflow."""
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    if not username or not password:
        return render_template("auth/login.html", error="Username and password required.")

    player = execute_one(
        "SELECT id, password_hash, is_banned, retired_at FROM players WHERE username = ?",
        (username,)
    )

    if player is None or not check_password_hash(player["password_hash"], password):
        return render_template("auth/login.html", error="Invalid username or password.")

    if player.get("retired_at"):
        return render_template("auth/login.html", error="This character has been retired.")
    if player["is_banned"]:
        return render_template("auth/login.html", error="This account has been banned.")

    # Update last_login_at
    with exclusive_transaction():
        execute_write(
            "UPDATE players SET last_login_at = ? WHERE id = ?",
            (datetime.utcnow().isoformat(), player["id"])
        )

    session.clear()
    session["player_id"] = player["id"]
    return redirect(url_for("dashboard.index"))


# ─────────────────────────────────────────────────────────────────────────────
# LOGOUT
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/logout", methods=["POST"])
def logout():
    """Handle the logout workflow."""
    session.clear()
    return redirect(url_for("auth.login"))


# ─────────────────────────────────────────────────────────────────────────────
# REGISTER
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/register", methods=["GET"])
def register():
    """Handle the register workflow."""
    if session.get("player_id"):
        return redirect(url_for("dashboard.index"))
    return render_template("auth/register.html")


@bp.route("/register", methods=["POST"])
def register_post():
    """Handle the register post workflow."""
    username   = request.form.get("username", "").strip()
    password   = request.form.get("password", "")
    email      = request.form.get("email", "").strip().lower()

    # Validation
    errors = []
    if not username:
        errors.append("Username is required.")
    if len(password) < 8:
        errors.append("Password must be at least 8 characters.")
    if not email or "@" not in email:
        errors.append("A valid email address is required.")

    if not errors:
        if execute_one("SELECT id FROM players WHERE username = ?", (username,)):
            errors.append("That username is already taken.")
        if execute_one("SELECT id FROM players WHERE email = ?", (email,)):
            errors.append("That email address is already registered.")

    if errors:
        return render_template("auth/register.html", errors=errors,
                               username=username, email=email)

    # Create account with placeholder stats — character creation completes them
    password_hash = generate_password_hash(password)
    with exclusive_transaction():
        # Players row needs class_id — set to 0 as sentinel until character creation
        # current_hp and current_ap set to 0; character creation sets real values
        player_id = execute_write(
            """INSERT INTO players
               (username, password_hash, email, character_name, sex, class_id,
                str_stat, end_stat, agi_stat, lck_stat, per_stat,
                level, xp, current_hp, current_ap, credits, last_login_at)
               VALUES (?, ?, ?, '', '', NULL, 1, 1, 1, 1, 1, 1, 0, 0, 0, ?, ?)""",
            (username, password_hash, email,
             cfg.STARTING_CREDITS, datetime.utcnow().isoformat())
        )
        # Create player_stats row
        execute_write(
            "INSERT INTO player_stats (player_id) VALUES (?)", (player_id,)
        )

    session.clear()
    session["player_id"] = player_id
    return redirect(url_for("auth.character_create"))


# ─────────────────────────────────────────────────────────────────────────────
# CHARACTER CREATION
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character-create", methods=["GET"])
def character_create():
    """Handle the character create workflow."""
    classes = execute("SELECT * FROM classes WHERE is_active = 1 ORDER BY name")
    if not classes:
        return render_template("auth/character_create.html",
                               classes=[],
                               error="No classes available yet. Ask the admin to import content first.")
    settings = get_all_settings()
    stat_points = settings.get("STARTING_STAT_POINTS", cfg.STARTING_STAT_POINTS)
    return render_template("auth/character_create.html",
                           classes=classes,
                           stat_points=stat_points)


@bp.route("/character-create", methods=["POST"])
def character_create_post():
    """Handle the character create post workflow."""
    player_id      = session["player_id"]
    character_name = request.form.get("character_name", "").strip()
    sex            = request.form.get("sex", "").strip()
    class_id       = request.form.get("class_id", type=int)

    # Stat allocations from form
    try:
        alloc = {
            "str": int(request.form.get("str_alloc", 0)),
            "end": int(request.form.get("end_alloc", 0)),
            "agi": int(request.form.get("agi_alloc", 0)),
            "lck": int(request.form.get("lck_alloc", 0)),
            "per": int(request.form.get("per_alloc", 0)),
        }
    except (ValueError, TypeError):
        alloc = {"str": 0, "end": 0, "agi": 0, "lck": 0, "per": 0}

    settings    = get_all_settings()
    stat_points = settings.get("STARTING_STAT_POINTS", cfg.STARTING_STAT_POINTS)
    classes     = execute("SELECT * FROM classes WHERE is_active = 1 ORDER BY name")

    # Validate
    errors = []
    if not character_name:
        errors.append("Character name is required.")
    if not sex:
        errors.append("Please select a sex.")
    if not class_id:
        errors.append("Please select a class.")

    selected_class = execute_one("SELECT * FROM classes WHERE id = ? AND is_active = 1", (class_id,))
    if not selected_class and not errors:
        errors.append("Invalid class selected.")

    total_alloc = sum(alloc.values())
    if total_alloc != stat_points:
        errors.append(f"You must allocate exactly {stat_points} stat points (you allocated {total_alloc}).")
    if any(v < 0 for v in alloc.values()):
        errors.append("Stat allocations cannot be negative.")

    if errors:
        return render_template("auth/character_create.html",
                               classes=classes, errors=errors,
                               stat_points=stat_points,
                               character_name=character_name, sex=sex,
                               class_id=class_id, alloc=alloc)

    # Apply class bonuses on top of base 1 per stat + player allocation
    final_stats = {
        "str": 1 + selected_class["str_bonus"] + alloc["str"],
        "end": 1 + selected_class["end_bonus"] + alloc["end"],
        "agi": 1 + selected_class["agi_bonus"] + alloc["agi"],
        "lck": 1 + selected_class["lck_bonus"] + alloc["lck"],
        "per": 1 + selected_class["per_bonus"] + alloc["per"],
    }

    # Derive starting HP and AP
    base_daily_ap = settings.get("BASE_DAILY_AP", cfg.BASE_DAILY_AP)
    starting_hp   = 10 + final_stats["end"] + 5   # level 1
    starting_ap   = base_daily_ap + math.floor(final_stats["end"] / 2)

    with exclusive_transaction():
        execute_write(
            """UPDATE players SET
               character_name = ?, sex = ?, class_id = ?,
               str_stat = ?, end_stat = ?, agi_stat = ?, lck_stat = ?, per_stat = ?,
               current_hp = ?, current_ap = ?
               WHERE id = ?""",
            (character_name, sex, class_id,
             final_stats["str"], final_stats["end"], final_stats["agi"],
             final_stats["lck"], final_stats["per"],
             starting_hp, starting_ap, player_id)
        )

    # Award starter gear (random level 1 weapon + armor)
    _award_starter_gear(player_id)
    _write_tutorial_feed(player_id)

    return redirect(url_for("dashboard.index"))


def _award_starter_gear(player_id: int):
    """Select a random level-1 weapon and armor, add to inventory."""
    for item_type, table in [("WEAPON", "weapons"), ("ARMOR", "armor")]:
        item = execute_one(
            f"SELECT * FROM {table} WHERE level = 1 AND is_active = 1 ORDER BY RANDOM() LIMIT 1"
        )
        if item is None:
            # Fallback: any active item
            item = execute_one(
                f"SELECT * FROM {table} WHERE is_active = 1 ORDER BY level ASC, RANDOM() LIMIT 1"
            )
        if item is None:
            continue  # No items imported yet — admin will need to import content

        with exclusive_transaction():
            inv_id = execute_write(
                """INSERT INTO inventory_items
                   (player_id, item_type, item_id, current_durability, acquired_method)
                   VALUES (?, ?, ?, ?, 'STARTER')""",
                (player_id, item_type, item["id"], item["starting_durability"])
            )
            execute_write(
                """INSERT INTO item_history
                   (player_id, item_type, item_id, item_name, event_type)
                   VALUES (?, ?, ?, ?, 'RECEIVED_STARTER')""",
                (player_id, item_type, item["id"], item["name"])
            )


# ─────────────────────────────────────────────────────────────────────────────
# LEVEL UP
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/levelup", methods=["GET"])
def levelup():
    """Show stat point assignment page.
    Enforced by before_request — only reachable when pending_levelup = True."""
    player = g.player
    return render_template("auth/levelup.html", player=player)


@bp.route("/levelup", methods=["POST"])
def levelup_post():
    """Handle the levelup post workflow."""
    stat = request.form.get("stat", "").strip().upper()
    if stat not in ("STR", "END", "AGI", "LCK", "PER"):
        return render_template("auth/levelup.html", player=g.player,
                               error="Please choose a valid stat.")

    result = enqueue_and_process(
        session["player_id"], "assign_levelup", {"stat": stat}
    )
    return redirect(url_for("dashboard.index"))


@register_handler("assign_levelup")
def handle_assign_levelup(player_id: int, payload: dict) -> dict:
    """Assign one stat point, recalculate and fully restore HP, clear pending flag."""
    stat = payload["stat"].lower()  # str / end / agi / lck / per
    col  = f"{stat}_stat"

    player = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    if not player or not player["pending_levelup"]:
        raise ValueError("No pending level-up for this player.")

    new_stat_val = player[col] + 1
    new_level    = player["level"]

    # Recalculate max HP with new stat (end may have just increased)
    new_end = new_stat_val if stat == "end" else player["end_stat"]
    new_max_hp = 10 + new_end + (5 * new_level)

    with exclusive_transaction():
        execute_write(f"UPDATE players SET {col} = ? WHERE id = ?", (new_stat_val, player_id))
        # Always fully restore HP on level up
        execute_write(
            "UPDATE players SET current_hp = ?, pending_levelup = 0 WHERE id = ?",
            (new_max_hp, player_id)
        )
        execute_write(
            """INSERT INTO level_up_history (player_id, level_reached, stat_increased)
               VALUES (?, ?, ?)""",
            (player_id, new_level, stat.upper())
        )

    logger.info("Player %d assigned level-up stat point to %s (now %d)",
                player_id, stat.upper(), new_stat_val)
    return {"stat": stat.upper(), "new_value": new_stat_val,
            "level": new_level}


################################################################################


def _write_tutorial_feed(player_id: int):
    """Write onboarding feed entries so the terminal has context on first login."""
    from datetime import datetime, timedelta
    messages = [
        ("SYSTEM",       "Welcome. The world is dangerous. Here is what you need to know."),
        ("SYSTEM",       "AP (Action Points) fuel everything. You earn a daily allotment at midnight plus trickle bonuses every 6 hours. Spend them wisely."),
        ("SYSTEM",       "BOSS — Challenge a movie villain. Defeat them for XP, credits, and gear. Watch for phase transitions as their HP drops."),
        ("SYSTEM",       "PVP — Fight another player. Win to steal credits and items. Lose and you drop to 1 HP. Choose your targets carefully."),
        ("SYSTEM",       "TAVERN — Spend credits to restore HP. No AP cost once inside."),
        ("SYSTEM",       "BLACKSMITH — Repair damaged gear. Durability matters — broken weapons deal less damage."),
        ("SYSTEM",       "SHOP — Buy and sell weapons, armor, and special items. Special items are unique. Only one copy exists in the world at a time."),
        ("SYSTEM",       "OBSERVE in combat to learn an enemy's resistances and weaknesses. That intel is stored permanently for future fights."),
        ("SYSTEM",       "Level up by earning XP. Each level grants one permanent stat point. Choose carefully — there is no going back."),
        ("SYSTEM",       "You have been given starter gear. Visit your Character Sheet to equip it before your first fight."),
        ("RANDOM_EVENT", "Good luck out there. You will need it."),
    ]
    base_time = datetime.utcnow()
    with exclusive_transaction():
        for i, (category, text) in enumerate(messages):
            ts = (base_time + timedelta(seconds=i)).isoformat()
            execute_write(
                """INSERT INTO daily_feed
                   (feed_scope, player_id, flavor_text, event_category, occurred_at)
                   VALUES ('PERSONAL', ?, ?, ?, ?)""",
                (player_id, text, category, ts)
            )

# FILE: routes/blacksmith.py
"""Blacksmith display and queued durability-repair operations."""
# routes/blacksmith.py
# Full-page repair interface. Players select damaged items to repair,
# pay credits per item, with a LCK bonus roll for enhanced restoration.

import math
import logging
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, session, g
from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_all_settings)
from queue_handler import enqueue_and_process, register_handler
import config_defaults as cfg
import random

bp = Blueprint("blacksmith", __name__)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# GET /blacksmith
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/blacksmith")
def index():
    """Handle the index workflow."""
    player   = g.player
    settings = get_all_settings()

    if player["credits"] == 0:
        return render_template("blacksmith/blacksmith.html",
                               items=[], blocked=True,
                               blocked_reason="You have no credits.",
                               feedback=request.args.get("feedback"),
                               error=request.args.get("error"))

    items = _get_repairable_items(player, settings)

    all_full = all(i["current_durability"] >= 100 for i in items)
    if all_full:
        return render_template("blacksmith/blacksmith.html",
                               items=[], blocked=True,
                               blocked_reason="All your items are at full durability.",
                               feedback=request.args.get("feedback"),
                               error=request.args.get("error"))

    return render_template("blacksmith/blacksmith.html",
                           items=items,
                           blocked=False,
                           feedback=request.args.get("feedback"),
                           error=request.args.get("error"))


def _get_repairable_items(player: dict, settings: dict) -> list[dict]:
    """Load player inventory with repair cost calculated per item."""
    repair_cost_pct = settings.get("REPAIR_COST_PERCENT", cfg.REPAIR_COST_PERCENT)
    items = execute(
        "SELECT * FROM inventory_items WHERE player_id = ?", (player["id"],)
    )
    result = []
    for inv in items:
        if inv["current_durability"] >= 100:
            continue  # skip fully repaired items
        detail = _get_item_detail(inv["item_type"], inv["item_id"])
        if detail is None:
            continue
        missing_dur = 100 - inv["current_durability"]
        # Cost = credit_cost * REPAIR_COST_PERCENT * (missing_dur / 100)
        # Free if credit_cost is 0
        repair_cost = max(0, int(
            detail["credit_cost"] * repair_cost_pct * (missing_dur / 100)
        ))
        equipped = inv["id"] in {
            player.get("equipped_weapon_id"),
            player.get("equipped_armor_id"),
            player.get("equipped_special_id"),
        }
        result.append({
            **inv, **detail,
            "inv_id":      inv["id"],
            "repair_cost": repair_cost,
            "missing_dur": missing_dur,
            "is_equipped": equipped,
        })
    return result


def _get_item_detail(item_type: str, item_id: int) -> dict | None:
    """Load item detail from current database state."""
    table = {"WEAPON": "weapons", "ARMOR": "armor", "SPECIAL": "special_items"}.get(item_type)
    if not table:
        return None
    return execute_one(f"SELECT * FROM {table} WHERE id = ?", (item_id,))


# ─────────────────────────────────────────────────────────────────────────────
# POST /blacksmith/repair
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/blacksmith/repair", methods=["POST"])
def repair():
    # inv_ids is a list of inventory item IDs the player wants to repair
    """Handle the repair workflow."""
    inv_ids = request.form.getlist("inv_ids", type=int)
    repair_mode = request.form.get("mode", "selected")  # selected / equipped / all

    try:
        result = enqueue_and_process(
            session["player_id"], "blacksmith_repair",
            {"inv_ids": inv_ids, "mode": repair_mode}
        )
        feedback = f"Repaired {result['items_repaired']} item(s). Spent {result['total_cost']} credits."
        return redirect(url_for("blacksmith.index", feedback=feedback))
    except RuntimeError as e:
        return redirect(url_for("blacksmith.index", error=str(e)))


@register_handler("blacksmith_repair")
def handle_blacksmith_repair(player_id: int, payload: dict) -> dict:
    """Process the queued blacksmith repair action against validated game state."""
    settings    = get_all_settings()
    repair_base = settings.get("REPAIR_BASE_PERCENT",    cfg.REPAIR_BASE_PERCENT)
    lck_mult    = settings.get("REPAIR_LCK_MULTIPLIER",  cfg.REPAIR_LCK_MULTIPLIER)
    lck_cap     = settings.get("REPAIR_LCK_CAP",         cfg.REPAIR_LCK_CAP)
    cost_pct    = settings.get("REPAIR_COST_PERCENT",    cfg.REPAIR_COST_PERCENT)
    ap_cost     = settings.get("AP_COST_BLACKSMITH",     cfg.AP_COST_BLACKSMITH)

    player  = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    if player["credits"] == 0:
        raise ValueError("You have no credits.")
    if player["current_ap"] < ap_cost:
        raise ValueError(f"Not enough AP. Need {ap_cost}.")

    inv_ids = payload.get("inv_ids", [])
    mode    = payload.get("mode", "selected")

    # Build list of items to repair based on mode
    all_items = execute(
        "SELECT * FROM inventory_items WHERE player_id = ?", (player_id,)
    )
    equipped_ids = {
        player.get("equipped_weapon_id"),
        player.get("equipped_armor_id"),
        player.get("equipped_special_id"),
    } - {None}

    to_repair = []
    for inv in all_items:
        if inv["current_durability"] >= 100:
            continue
        if mode == "equipped" and inv["id"] not in equipped_ids:
            continue
        if mode == "selected" and inv["id"] not in inv_ids:
            continue
        detail = _get_item_detail(inv["item_type"], inv["item_id"])
        if detail is None:
            continue
        missing = 100 - inv["current_durability"]
        cost    = max(0, int(detail["credit_cost"] * cost_pct * (missing / 100)))
        to_repair.append({**inv, "detail": detail, "repair_cost": cost, "missing": missing})

    if not to_repair:
        raise ValueError("No items selected for repair.")

    total_cost = sum(i["repair_cost"] for i in to_repair)
    if player["credits"] < total_cost:
        raise ValueError(f"Not enough credits. Need {total_cost}, have {player['credits']}.")

    lck = player["lck_stat"]
    lck_roll_chance = math.floor(lck / 2) * 0.05  # 5% per floor(LCK/2)
    results = []

    with exclusive_transaction():
        # Deduct AP (no passive regen on blacksmith entry)
        execute_write(
            "UPDATE players SET current_ap = current_ap - ? WHERE id = ?",
            (ap_cost, player_id)
        )
        # Deduct total credit cost
        execute_write(
            "UPDATE players SET credits = credits - ? WHERE id = ?",
            (total_cost, player_id)
        )

        for item in to_repair:
            # Base repair
            base_restore = int(item["missing"] * repair_base)

            # LCK bonus roll
            lck_bonus_applied = False
            if random.random() < lck_roll_chance:
                lck_cap_restore = int(item["missing"] * min(0.50 + (lck * lck_mult / 100), lck_cap))
                final_restore   = max(base_restore, lck_cap_restore)
                lck_bonus_applied = True
            else:
                final_restore = base_restore

            new_durability = min(item["current_durability"] + final_restore, 100)
            execute_write(
                "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
                (new_durability, item["id"])
            )
            results.append({
                "name":             item["detail"]["name"],
                "restored":         final_restore,
                "new_durability":   new_durability,
                "lck_bonus":        lck_bonus_applied,
            })

    logger.info("Player %d repaired %d items for %d credits",
                player_id, len(results), total_cost)
    return {
        "items_repaired": len(results),
        "total_cost":     total_cost,
        "results":        results,
    }


################################################################################

# FILE: routes/character.py
"""Character sheet, inventory equipment, item dropping, and combat preferences."""
# routes/character.py
# Full-page character sheet with inventory management.
# Equip, unequip, drop items. Update combat preference.
# Live stat preview via lightweight JSON endpoint.

import math
import logging
from datetime import datetime

from flask import (Blueprint, render_template, request, redirect,
                   url_for, session, g, jsonify)
from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_all_settings)
from queue_handler import enqueue_and_process, register_handler
import config_defaults as cfg

bp = Blueprint("character", __name__)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# GET /character
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character")
def index():
    """Handle the index workflow."""
    player   = g.player
    settings = get_all_settings()

    inventory = _get_full_inventory(player)
    equipped  = {
        "weapon":  _find_equipped(inventory, player.get("equipped_weapon_id")),
        "armor":   _find_equipped(inventory, player.get("equipped_armor_id")),
        "special": _find_equipped(inventory, player.get("equipped_special_id")),
    }
    derived = _calc_derived_stats(player, equipped, settings)
    active_effects = _get_active_effects(player["id"])

    return render_template(
        "character/character.html",
        inventory=inventory,
        equipped=equipped,
        derived=derived,
        active_effects=active_effects,
        preferences=["Aggressive", "Defensive", "Opportunist", "Balanced"],
        feedback=request.args.get("feedback"),
        error=request.args.get("error"),
    )


def _get_active_effects(player_id: int) -> list[dict]:
    """Build readable character-sheet entries for midnight status effects."""
    rows = execute(
        "SELECT effect_type, value FROM status_effects WHERE player_id = ? ORDER BY id",
        (player_id,)
    )
    stat_names = {
        "STAT_BOOST_STR": "Strength", "STAT_BOOST_END": "Endurance",
        "STAT_BOOST_AGI": "Agility", "STAT_BOOST_LCK": "Luck",
        "STAT_BOOST_PER": "Perception", "STAT_BOOST_INITIATIVE": "Initiative",
        "STAT_PENALTY_STR": "Strength", "STAT_PENALTY_END": "Endurance",
        "STAT_PENALTY_AGI": "Agility", "STAT_PENALTY_LCK": "Luck",
        "STAT_PENALTY_PER": "Perception", "STAT_PENALTY_INITIATIVE": "Initiative",
    }
    effects = []
    for row in rows:
        effect_type = row["effect_type"]
        value = row["value"]
        if effect_type == "CURSED":
            description = f"Daily AP award -{int(round(value * 100))}%"
            is_good = False
        elif effect_type in stat_names:
            description = f"{stat_names[effect_type]} {int(value):+d}"
            is_good = value > 0
        else:
            description = effect_type.replace("_", " ").title()
            is_good = value >= 0
        effects.append({
            "description": description,
            "is_good": is_good,
            "expires": "Midnight reset",
        })
    return effects


def _get_full_inventory(player: dict) -> list[dict]:
    """Load full inventory from current database state."""
    equipped_ids = {
        player.get("equipped_weapon_id"),
        player.get("equipped_armor_id"),
        player.get("equipped_special_id"),
    } - {None}

    rows = execute(
        "SELECT * FROM inventory_items WHERE player_id = ? ORDER BY item_type, acquired_at",
        (player["id"],)
    )
    result = []
    for inv in rows:
        detail = _get_item_detail(inv["item_type"], inv["item_id"])
        if detail is None:
            continue
        result.append({
            **inv, **detail,
            "inv_id":     inv["id"],
            "is_equipped": inv["id"] in equipped_ids,
        })
    return result


def _find_equipped(inventory: list, inv_id) -> dict | None:
    """Provide the internal find equipped operation used by this module."""
    if inv_id is None:
        return None
    return next((i for i in inventory if i["inv_id"] == inv_id), None)


def _get_item_detail(item_type: str, item_id: int) -> dict | None:
    """Load item detail from current database state."""
    table = {"WEAPON": "weapons", "ARMOR": "armor", "SPECIAL": "special_items"}.get(item_type)
    if not table:
        return None
    return execute_one(f"SELECT * FROM {table} WHERE id = ?", (item_id,))


def _calc_derived_stats(player: dict, equipped: dict, settings: dict) -> dict:
    """Compute all derived stats for display on character sheet."""
    w = equipped.get("weapon")
    a = equipped.get("armor")
    s = equipped.get("special")

    str_total = player["str_stat"] + (w.get("str_bonus", 0) if w else 0) + \
                (a.get("str_bonus", 0) if a else 0) + (s.get("str_bonus", 0) if s else 0)
    end_total = player["end_stat"] + (w.get("end_bonus", 0) if w else 0) + \
                (a.get("end_bonus", 0) if a else 0) + (s.get("end_bonus", 0) if s else 0)
    agi_total = player["agi_stat"] + (w.get("agi_bonus", 0) if w else 0) + \
                (a.get("agi_bonus", 0) if a else 0) + (s.get("agi_bonus", 0) if s else 0)
    lck_total = player["lck_stat"] + (w.get("lck_bonus", 0) if w else 0) + \
                (a.get("lck_bonus", 0) if a else 0) + (s.get("lck_bonus", 0) if s else 0)
    per_total = player["per_stat"] + (w.get("per_bonus", 0) if w else 0) + \
                (a.get("per_bonus", 0) if a else 0) + (s.get("per_bonus", 0) if s else 0)

    ac_bonus     = (a.get("ac_bonus", 0) if a else 0) + (s.get("ac_bonus", 0) if s else 0)
    ac           = 10 + math.floor(agi_total / 2) + ac_bonus
    max_hp       = 10 + end_total + (5 * player["level"])
    inv_limit    = settings.get("INVENTORY_LIMIT", cfg.INVENTORY_LIMIT) + \
                   math.floor(str_total / 2)
    crit_thresh  = max(
        settings.get("CRIT_MIN_THRESHOLD", cfg.CRIT_MIN_THRESHOLD),
        settings.get("CRIT_BASE_THRESHOLD", cfg.CRIT_BASE_THRESHOLD) -
        math.floor(lck_total / settings.get("CRIT_LCK_DIVISOR", cfg.CRIT_LCK_DIVISOR))
    )
    if s:
        crit_thresh = max(
            settings.get("CRIT_MIN_THRESHOLD", cfg.CRIT_MIN_THRESHOLD),
            crit_thresh - int(s.get("crit_chance_bonus", 0))
        )
    shop_discount= min(
        math.floor(per_total / 2) + int((s.get("shop_discount", 0) if s else 0) * 100),
        int(settings.get("SHOP_DISCOUNT_MAX", cfg.SHOP_DISCOUNT_MAX) * 100)
    )
    daily_ap     = settings.get("BASE_DAILY_AP", cfg.BASE_DAILY_AP) + \
                   math.floor(end_total / 2) + (s.get("bonus_ap", 0) if s else 0)
    passive_regen= settings.get("AP_PASSIVE_HP_REGEN", cfg.AP_PASSIVE_HP_REGEN) + \
                   math.floor(end_total / settings.get("END_HP_REGEN_DIVISOR", cfg.END_HP_REGEN_DIVISOR)) + \
                   (s.get("hp_regen_bonus", 0) if s else 0)

    return {
        "str": str_total, "end": end_total, "agi": agi_total,
        "lck": lck_total, "per": per_total,
        "ac": ac, "max_hp": max_hp, "inv_limit": inv_limit,
        "crit_threshold": crit_thresh, "shop_discount_pct": shop_discount,
        "daily_ap": daily_ap, "passive_regen": passive_regen,
        "extra_attack": bool(s.get("extra_attack")) if s else False,
        "xp_multiplier_pct": int((s.get("xp_multiplier", 0) if s else 0) * 100),
        "credit_multiplier_pct": int((s.get("credit_multiplier", 0) if s else 0) * 100),
    }


# ─────────────────────────────────────────────────────────────────────────────
# STAT PREVIEW (live AJAX — the third JS feature)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character/preview")
def preview():
    """Return JSON of derived stats for a hypothetical equipment loadout.
    Used by character.html JS for live stat preview on equip/unequip changes."""
    player   = g.player
    settings = get_all_settings()

    def load_equipped(inv_id_str):
        """Handle the load equipped workflow."""
        if not inv_id_str or inv_id_str == "none":
            return None
        try:
            inv_id = int(inv_id_str)
        except ValueError:
            return None
        inv = execute_one(
            "SELECT * FROM inventory_items WHERE id = ? AND player_id = ?",
            (inv_id, player["id"])
        )
        if not inv:
            return None
        detail = _get_item_detail(inv["item_type"], inv["item_id"])
        return {**(detail or {}), "current_durability": inv["current_durability"]}

    equipped = {
        "weapon":  load_equipped(request.args.get("weapon")),
        "armor":   load_equipped(request.args.get("armor")),
        "special": load_equipped(request.args.get("special")),
    }
    derived = _calc_derived_stats(player, equipped, settings)
    return jsonify(derived)


# ─────────────────────────────────────────────────────────────────────────────
# POST /character/equip
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character/equip", methods=["POST"])
def equip():
    """Handle the equip workflow."""
    inv_id = request.form.get("inv_id", type=int)
    try:
        enqueue_and_process(session["player_id"], "equip", {"inv_id": inv_id})
        return redirect(url_for("character.index", feedback="Item equipped."))
    except RuntimeError as e:
        return redirect(url_for("character.index", error=str(e)))


@register_handler("equip")
def handle_equip(player_id: int, payload: dict) -> dict:
    """Process the queued equip action against validated game state."""
    inv_id = payload["inv_id"]
    player = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    inv    = execute_one(
        "SELECT * FROM inventory_items WHERE id = ? AND player_id = ?",
        (inv_id, player_id)
    )
    if inv is None:
        raise ValueError("Item not found in your inventory.")

    # Determine correct slot column
    slot_col = {
        "WEAPON":  "equipped_weapon_id",
        "ARMOR":   "equipped_armor_id",
        "SPECIAL": "equipped_special_id",
    }.get(inv["item_type"])
    if slot_col is None:
        raise ValueError("Unknown item type.")

    with exclusive_transaction():
        execute_write(
            f"UPDATE players SET {slot_col} = ? WHERE id = ?",
            (inv_id, player_id)
        )
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# POST /character/unequip
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character/unequip", methods=["POST"])
def unequip():
    """Handle the unequip workflow."""
    slot = request.form.get("slot")  # weapon / armor / special
    try:
        enqueue_and_process(session["player_id"], "unequip", {"slot": slot})
        return redirect(url_for("character.index", feedback="Item unequipped."))
    except RuntimeError as e:
        return redirect(url_for("character.index", error=str(e)))


@register_handler("unequip")
def handle_unequip(player_id: int, payload: dict) -> dict:
    """Process the queued unequip action against validated game state."""
    slot = payload.get("slot", "").lower()
    slot_col = {
        "weapon":  "equipped_weapon_id",
        "armor":   "equipped_armor_id",
        "special": "equipped_special_id",
    }.get(slot)
    if not slot_col:
        raise ValueError("Invalid slot.")

    with exclusive_transaction():
        execute_write(
            f"UPDATE players SET {slot_col} = NULL WHERE id = ?",
            (player_id,)
        )
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# POST /character/drop
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character/drop", methods=["POST"])
def drop():
    """Handle the drop workflow."""
    inv_id = request.form.get("inv_id", type=int)
    try:
        enqueue_and_process(session["player_id"], "drop_item", {"inv_id": inv_id})
        return redirect(url_for("character.index", feedback="Item dropped."))
    except RuntimeError as e:
        return redirect(url_for("character.index", error=str(e)))


@register_handler("drop_item")
def handle_drop_item(player_id: int, payload: dict) -> dict:
    """Process the queued drop item action against validated game state."""
    inv_id = payload["inv_id"]
    player = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    inv    = execute_one(
        "SELECT * FROM inventory_items WHERE id = ? AND player_id = ?",
        (inv_id, player_id)
    )
    if inv is None:
        raise ValueError("Item not found.")

    # Cannot drop equipped items
    equipped = {player.get("equipped_weapon_id"),
                player.get("equipped_armor_id"),
                player.get("equipped_special_id")}
    if inv_id in equipped:
        raise ValueError("Unequip the item before dropping it.")

    detail = _get_item_detail(inv["item_type"], inv["item_id"])
    item_name = detail["name"] if detail else "Unknown Item"

    with exclusive_transaction():
        execute_write("DELETE FROM inventory_items WHERE id = ?", (inv_id,))
        execute_write(
            """INSERT INTO item_history
               (player_id, item_type, item_id, item_name, event_type)
               VALUES (?, ?, ?, ?, 'DROPPED')""",
            (player_id, inv["item_type"], inv["item_id"], item_name)
        )
        # If special: return to pool
        if inv["item_type"] == "SPECIAL":
            execute_write(
                """UPDATE special_item_registry
                   SET status = 'IN_POOL', current_owner_player_id = NULL,
                       inventory_item_id = NULL, last_released_method = 'DROPPED',
                       updated_at = ?
                   WHERE special_item_id = ?""",
                (datetime.utcnow().isoformat(), inv["item_id"])
            )
            # Global feed: special item returned to pool
            execute_write(
                """INSERT INTO daily_feed
                   (feed_scope, player_id, flavor_text, event_category)
                   VALUES ('GLOBAL', NULL, ?, 'ITEM')""",
                (f"{item_name} has returned to the loot pool.",)
            )
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# POST /character/preference
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/character/preference", methods=["POST"])
def preference():
    """Handle the preference workflow."""
    pref = request.form.get("preference", "")
    try:
        enqueue_and_process(session["player_id"], "set_preference", {"preference": pref})
        return redirect(url_for("character.index", feedback=f"Combat preference set to {pref}."))
    except RuntimeError as e:
        return redirect(url_for("character.index", error=str(e)))


@register_handler("set_preference")
def handle_set_preference(player_id: int, payload: dict) -> dict:
    """Process the queued set preference action against validated game state."""
    pref = payload.get("preference", "")
    valid = {"Aggressive", "Defensive", "Opportunist", "Balanced"}
    if pref not in valid:
        raise ValueError(f"Invalid preference. Choose from: {', '.join(valid)}")
    with exclusive_transaction():
        execute_write(
            "UPDATE players SET combat_preference = ? WHERE id = ?",
            (pref, player_id)
        )
    return {"success": True}


################################################################################

# FILE: routes/combat.py
"""Combat request routes that execute complete player/opponent rounds."""
# routes/combat.py  (Phase 5 — full implementation)
# All in-combat terminal-fragment POST routes.
# Each route loads combat state, resolves the action, checks for combat end,
# and returns an HTML fragment appended to the terminal by terminal.js.

import logging
from datetime import datetime

from flask import Blueprint, render_template, request, session, g, has_request_context
from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_all_settings)
from queue_handler import enqueue_and_process, register_handler
from combat import actions as combat_actions
from combat import engine, flavour
import config_defaults as cfg

bp = Blueprint("combat", __name__)
logger = logging.getLogger(__name__)


def _clear_browser_combat_session():
    """Clear browser-only state when a human request context exists."""
    if has_request_context():
        session.pop("combat_session_id", None)


def _error_fragment(message: str) -> str:
    """Provide the internal error fragment operation used by this module."""
    return render_template("fragments/error.html", message=message,
                           player=g.get("player"))


def _get_session_id() -> int | None:
    """Load session id from current database state."""
    return session.get("combat_session_id")


# ─────────────────────────────────────────────────────────────────────────────
# POST /combat/action
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/combat/action", methods=["POST"])
def action():
    """Handle the action workflow."""
    session_id = _get_session_id()
    if not session_id:
        return _error_fragment("No active combat session.")

    action_type = request.form.get("action_type", "attack").lower()
    player      = g.player

    result = enqueue_and_process(
        player["id"], "combat_action",
        {"session_id": session_id, "action_type": action_type}
    )
    return _render_combat_round(result, session_id, player)


@register_handler("combat_action")
def handle_combat_action(player_id: int, payload: dict) -> dict:
    """Resolve one full combat round:
    1. Roll initiative — determine who acts first
    2. First actor resolves chosen action
    3. Check for combat end
    4. Second actor resolves automated action (or round ends)
    5. Check for combat end again
    6. Apply end-of-round effects (special item durability, expire END_OF_ROUND buffs)
    7. Advance round counter"""
    session_id  = payload["session_id"]
    action_type = payload["action_type"]
    settings    = get_all_settings()

    state = combat_actions.get_combat_state(session_id)
    sess  = state["session"]

    if sess["status"] != "ACTIVE":
        raise ValueError("Combat session is not active.")

    attacker = state["attacker"]
    att_eq   = state["attacker_equipped"]
    att_special = att_eq.get("special")

    # Determine opponent for initiative
    if sess["combat_type"] == "PVP":
        defender = state["defender"]
        def_special = state["defender_equipped"].get("special")
    else:
        defender = state["boss"] or state["minion"]
        def_special = None

    # --- Initiative ---
    att_init_bonus = att_special.get("initiative_bonus", 0) if att_special else 0
    att_init, att_agi = engine.calc_initiative(attacker, att_init_bonus)
    def_init, def_agi = engine.calc_initiative(defender)

    if att_init > def_init:
        attacker_first = True
    elif def_init > att_init:
        attacker_first = False
    else:
        # Tie: higher raw AGI wins; if still tied, reroll
        if att_agi != def_agi:
            attacker_first = att_agi > def_agi
        else:
            while att_init == def_init:
                att_init, _ = engine.calc_initiative(attacker, att_init_bonus)
                def_init, _ = engine.calc_initiative(defender)
            attacker_first = att_init > def_init

    round_log = []
    first_result  = None
    second_result = None
    combat_ended  = False
    winner_side   = None
    result_type   = None

    # --- First actor ---
    if attacker_first:
        first_result = _resolve_player_action(
            session_id, player_id, action_type, state, settings
        )
        round_log.append(first_result)
        if first_result.get("escaped"):
            return _escaped_round_result(round_log, session_id, attacker_first, att_init, def_init)
        ended, winner_side = combat_actions.check_combat_end(state)
        if ended:
            combat_ended = True
            result_type  = "1HP_WIN"
    else:
        first_result = combat_actions.handle_opponent_action(session_id, state)
        round_log.append(first_result)
        ended, winner_side = combat_actions.check_combat_end(state)
        if ended:
            combat_ended = True
            winner_side  = "DEFENDER"  # boss/opponent won by hitting player to 1 HP
            result_type  = "1HP_WIN"

    # --- Second actor (if fight not over) ---
    if not combat_ended:
        # Reload state after first action (HP may have changed)
        state = combat_actions.get_combat_state(session_id)
        if attacker_first:
            second_result = combat_actions.handle_opponent_action(session_id, state)
        else:
            second_result = _resolve_player_action(
                session_id, player_id, action_type, state, settings
            )
        round_log.append(second_result)
        if second_result.get("escaped"):
            return _escaped_round_result(round_log, session_id, attacker_first, att_init, def_init)
        ended, winner_side = combat_actions.check_combat_end(state)
        if ended:
            combat_ended = True
            result_type  = "1HP_WIN"

    # --- End of round effects ---
    if not combat_ended:
        _apply_end_of_round(session_id, state, settings)
        with exclusive_transaction():
            execute_write(
                "UPDATE combat_sessions SET current_round = current_round + 1 WHERE id = ?",
                (session_id,)
            )

    # --- PvP round limit check ---
    reload_sess = execute_one("SELECT * FROM combat_sessions WHERE id = ?", (session_id,))
    pvp_rounds  = settings.get("COMBAT_ROUNDS_DEFAULT", cfg.COMBAT_ROUNDS_DEFAULT)
    rounds_extended = reload_sess["rounds_extended"]
    max_rounds  = pvp_rounds + (rounds_extended * settings.get("COMBAT_ROUNDS_EXTENSION",
                                                                cfg.COMBAT_ROUNDS_EXTENSION))

    at_round_limit = (not combat_ended and
                      sess["combat_type"] == "PVP" and
                      reload_sess["current_round"] > max_rounds)

    # --- Post-combat resolution ---
    final_result = None
    if combat_ended and winner_side:
        state = combat_actions.get_combat_state(session_id)
        final_result = combat_actions.finalize_combat(
            session_id, winner_side, result_type, state
        )
        _clear_browser_combat_session()

    return {
        "round_log":       round_log,
        "combat_ended":    combat_ended,
        "at_round_limit":  at_round_limit,
        "winner_side":     winner_side,
        "final_result":    final_result,
        "session_id":      session_id,
        "attacker_first":  attacker_first,
        "att_init":        att_init,
        "def_init":        def_init,
    }


def _escaped_round_result(round_log, session_id, attacker_first, att_init, def_init):
    """Return immediately after a successful normal Escape action."""
    return {
        "round_log": round_log, "combat_ended": True, "at_round_limit": False,
        "winner_side": None, "final_result": {"result_type": "ESCAPE"},
        "session_id": session_id, "attacker_first": attacker_first,
        "att_init": att_init, "def_init": def_init,
    }


def _resolve_player_action(session_id: int, player_id: int,
                            action_type: str, state: dict,
                            settings: dict) -> dict:
    """Route the player's chosen action to the correct handler."""
    if action_type == "attack":
        return combat_actions.handle_attack(session_id, "ATTACKER", state)
    elif action_type == "brace":
        return combat_actions.handle_brace(session_id, player_id, state)
    elif action_type == "escape":
        return combat_actions.handle_escape(session_id, player_id, state)
    elif action_type == "observe":
        return combat_actions.handle_observe(session_id, player_id, state)
    elif action_type == "swap_gear":
        new_weapon  = request.form.get("new_weapon_id",  type=int)
        new_armor   = request.form.get("new_armor_id",   type=int)
        new_special = request.form.get("new_special_id", type=int)
        return combat_actions.handle_swap_gear(
            session_id, player_id, state, new_weapon, new_armor, new_special
        )
    else:
        raise ValueError(f"Unknown action type: {action_type}")


def _apply_end_of_round(session_id: int, state: dict, settings: dict):
    """Apply end-of-round effects:
    - Special item durability loss (both sides)
    - Expire END_OF_ROUND combat buffs"""
    loss_pct = settings.get("SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND",
                             cfg.SPECIAL_ITEM_DURABILITY_LOSS_PER_ROUND)
    loss_pts = max(1, int(100 * loss_pct))

    with exclusive_transaction():
        # Attacker special item
        att_special = state["attacker_equipped"].get("special")
        if att_special:
            new_dur = max(0, att_special["current_durability"] - loss_pts)
            execute_write(
                "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
                (new_dur, att_special["inv_id"])
            )
            if new_dur == 0:
                execute_write(
                    "DELETE FROM inventory_items WHERE id = ?", (att_special["inv_id"],)
                )
                execute_write(
                    "UPDATE players SET equipped_special_id = NULL WHERE id = ?",
                    (state["session"]["attacker_player_id"],)
                )

        # Defender special item (PvP only)
        if state["session"]["combat_type"] == "PVP" and state.get("defender_equipped"):
            def_special = state["defender_equipped"].get("special")
            if def_special:
                new_dur = max(0, def_special["current_durability"] - loss_pts)
                execute_write(
                    "UPDATE inventory_items SET current_durability = ? WHERE id = ?",
                    (new_dur, def_special["inv_id"])
                )
                if new_dur == 0:
                    execute_write(
                        "DELETE FROM inventory_items WHERE id = ?", (def_special["inv_id"],)
                    )
                    execute_write(
                        "UPDATE players SET equipped_special_id = NULL WHERE id = ?",
                        (state["session"]["defender_player_id"],)
                    )

        # Expire END_OF_ROUND buffs
        execute_write(
            "DELETE FROM combat_buffs WHERE combat_session_id = ? AND expires_on = 'END_OF_ROUND'",
            (session_id,)
        )


# ─────────────────────────────────────────────────────────────────────────────
# POST /combat/steal  (confirmation step)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/combat/steal", methods=["POST"])
def steal():
    """Return a confirmation fragment showing risk/reward before committing."""
    session_id = _get_session_id()
    if not session_id:
        return _error_fragment("No active combat session.")
    player = g.player
    state  = combat_actions.get_combat_state(session_id)
    opponent = state.get("defender") or state.get("boss") or state.get("minion")
    return render_template("fragments/combat_steal_confirm.html",
                           opponent=opponent,
                           session_id=session_id,
                           player=player)


@bp.route("/combat/steal/confirm", methods=["POST"])
def steal_confirm():
    """Execute the steal action after player confirms."""
    session_id = _get_session_id()
    if not session_id:
        return _error_fragment("No active combat session.")
    player = g.player
    result = enqueue_and_process(
        player["id"], "combat_steal", {"session_id": session_id}
    )
    return _render_combat_round(result, session_id, player)


@register_handler("combat_steal")
def handle_combat_steal(player_id: int, payload: dict) -> dict:
    """Steal resolves as a full combat round (player steals, then opponent acts)."""
    session_id = payload["session_id"]
    state = combat_actions.get_combat_state(session_id)
    sess  = state["session"]

    if sess["status"] != "ACTIVE":
        raise ValueError("Combat session is not active.")

    round_log = []
    # Player steal attempt
    steal_result = combat_actions.handle_steal(session_id, player_id, state)
    round_log.append(steal_result)

    # Check if escape happened (steal on escaped opponent?)
    ended, winner_side = combat_actions.check_combat_end(state)
    final_result = None
    if ended:
        state = combat_actions.get_combat_state(session_id)
        final_result = combat_actions.finalize_combat(session_id, winner_side, "1HP_WIN", state)
        _clear_browser_combat_session()
    else:
        # Opponent counter-acts
        state = combat_actions.get_combat_state(session_id)
        opp_result = combat_actions.handle_opponent_action(session_id, state)
        round_log.append(opp_result)
        ended, winner_side = combat_actions.check_combat_end(state)
        if ended:
            state = combat_actions.get_combat_state(session_id)
            final_result = combat_actions.finalize_combat(
                session_id, winner_side, "1HP_WIN", state
            )
            _clear_browser_combat_session()
        else:
            _apply_end_of_round(session_id, state, get_all_settings())
            with exclusive_transaction():
                execute_write(
                    "UPDATE combat_sessions SET current_round = current_round + 1 WHERE id = ?",
                    (session_id,)
                )

    return {
        "round_log":      round_log,
        "combat_ended":   ended,
        "at_round_limit": False,
        "winner_side":    winner_side,
        "final_result":   final_result,
        "session_id":     session_id,
        "attacker_first": True,
        "att_init": 0, "def_init": 0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /combat/extend  (PvP round extension)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/combat/extend", methods=["POST"])
def extend():
    """Handle the extend workflow."""
    session_id = _get_session_id()
    if not session_id:
        return _error_fragment("No active combat session.")
    player = g.player
    result = enqueue_and_process(
        player["id"], "combat_extend", {"session_id": session_id}
    )
    if result.get("error"):
        return _error_fragment(result["error"])
    settings = get_all_settings()
    timeout  = settings.get("COMBAT_EXTENSION_TIMEOUT", cfg.COMBAT_EXTENSION_TIMEOUT)
    return render_template("fragments/combat_extend.html",
                           session_id=session_id,
                           timeout=timeout,
                           player=player)


@register_handler("combat_extend")
def handle_combat_extend(player_id: int, payload: dict) -> dict:
    """Process the queued combat extend action against validated game state."""
    session_id = payload["session_id"]
    settings   = get_all_settings()
    ap_cost    = settings.get("AP_COST_ESCAPE", 1)  # extension costs 1 AP

    player = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))
    if player["current_ap"] < ap_cost:
        return {"error": f"Not enough AP to extend. Need {ap_cost}."}

    sess = execute_one("SELECT * FROM combat_sessions WHERE id = ?", (session_id,))
    if sess["combat_type"] != "PVP":
        return {"error": "Round extension is only available in PvP."}

    with exclusive_transaction():
        execute_write(
            "UPDATE players SET current_ap = current_ap - ? WHERE id = ?",
            (ap_cost, player_id)
        )
        execute_write(
            "UPDATE combat_sessions SET rounds_extended = rounds_extended + 1 WHERE id = ?",
            (session_id,)
        )
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# POST /combat/resolve  (score formula or auto-triggered by timer)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/combat/resolve", methods=["POST"])
def resolve():
    """Handle the resolve workflow."""
    session_id = _get_session_id()
    if not session_id:
        return _error_fragment("No active combat session.")
    player = g.player
    result = enqueue_and_process(
        player["id"], "combat_resolve", {"session_id": session_id}
    )
    return _render_combat_round(result, session_id, player)


@register_handler("combat_resolve")
def handle_combat_resolve(player_id: int, payload: dict) -> dict:
    """Resolve combat via the PvP score formula."""
    session_id = payload["session_id"]
    state = combat_actions.get_combat_state(session_id)
    sess  = state["session"]

    if sess["status"] != "ACTIVE":
        raise ValueError("Session not active.")
    if sess["combat_type"] != "PVP":
        raise ValueError("Score resolution only applies to PvP.")

    attacker = state["attacker"]
    defender = state["defender"]
    att_max_hp = engine.calc_max_hp(attacker)
    def_max_hp = engine.calc_max_hp(defender)

    att_score, def_score = engine.calc_pvp_score(sess, att_max_hp, def_max_hp)
    winner_side  = "ATTACKER" if att_score > def_score else "DEFENDER"
    final_result = combat_actions.finalize_combat(
        session_id, winner_side, "SCORE_WIN", state
    )
    _clear_browser_combat_session()

    return {
        "round_log":      [],
        "combat_ended":   True,
        "at_round_limit": False,
        "winner_side":    winner_side,
        "final_result":   final_result,
        "session_id":     session_id,
        "att_score":      round(att_score, 3),
        "def_score":      round(def_score, 3),
        "attacker_first": True,
        "att_init": 0, "def_init": 0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# FRAGMENT RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def _render_combat_round(result: dict, session_id: int, player) -> str:
    """Choose which fragment to render based on combat state."""
    if result.get("combat_ended"):
        return render_template("fragments/combat_result.html",
                               result=result, player=player)
    if result.get("at_round_limit"):
        settings = get_all_settings()
        timeout  = settings.get("COMBAT_EXTENSION_TIMEOUT", cfg.COMBAT_EXTENSION_TIMEOUT)
        return render_template("fragments/combat_extend.html",
                               session_id=session_id, timeout=timeout, player=player)
    return render_template("fragments/combat_round.html",
                           result=result, session_id=session_id, player=player)


################################################################################

# FILE: routes/dashboard.py
"""Main player dashboard and context-sensitive action availability."""
# routes/dashboard.py  (Phase 9 — adds now_iso injection for JS feed polling)
import logging
from datetime import datetime
from flask import Blueprint, render_template, g
from database import execute, get_all_settings
import config_defaults as cfg

bp = Blueprint('dashboard', __name__)
logger = logging.getLogger(__name__)


@bp.route('/')
def index():
    """Handle the index workflow."""
    player   = g.player
    settings = get_all_settings()
    history_count = settings.get('TERMINAL_HISTORY_ENTRIES', cfg.TERMINAL_HISTORY_ENTRIES)

    terminal_history = execute(
        '''SELECT flavor_text, event_category, occurred_at, combat_session_id
           FROM daily_feed
           WHERE player_id = ? OR feed_scope = 'GLOBAL'
           ORDER BY occurred_at DESC
           LIMIT ?''',
        (player['id'], history_count)
    )
    terminal_history = list(reversed(terminal_history))

    button_states = _get_button_states(player, settings)

    # Inject current UTC timestamp for JS feed polling start point
    now_iso = datetime.utcnow().isoformat()

    active_effects = execute(
        "SELECT effect_type, value FROM status_effects WHERE player_id = ?", (player["id"],)
    )
    label_map = {
        "STAT_BOOST_STR": "+STR", "STAT_BOOST_END": "+END", "STAT_BOOST_AGI": "+AGI",
        "STAT_BOOST_LCK": "+LCK", "STAT_BOOST_PER": "+PER", "STAT_BOOST_INITIATIVE": "+INIT",
        "STAT_PENALTY_STR": "-STR", "STAT_PENALTY_END": "-END", "STAT_PENALTY_AGI": "-AGI",
        "STAT_PENALTY_LCK": "-LCK", "STAT_PENALTY_PER": "-PER", "STAT_PENALTY_INITIATIVE": "-INIT",
        "CURSED": "CURSED",
    }
    effect_labels = [
        f"{label_map.get(e['effect_type'], e['effect_type'])} {int(abs(e['value']))}"
        for e in active_effects
    ]

    return render_template(
        'dashboard.html',
        terminal_history=terminal_history,
        button_states=button_states,
        blackout=g.get('blackout', False),
        now_iso=now_iso,
        effect_labels=effect_labels,
    )


def _get_button_states(player: dict, settings: dict) -> dict:
    """Load button states from current database state."""
    in_combat  = player['in_combat']
    current_ap = player['current_ap']
    credits    = player['credits']
    current_hp = player['current_hp']
    max_hp     = player['max_hp']
    blackout   = g.get('blackout', False)

    ap_boss       = settings.get('AP_COST_BOSS',       cfg.AP_COST_BOSS)
    ap_pvp        = settings.get('AP_COST_PVP',        cfg.AP_COST_PVP)
    ap_tavern     = settings.get('AP_COST_TAVERN',     cfg.AP_COST_TAVERN)
    ap_blacksmith = settings.get('AP_COST_BLACKSMITH', cfg.AP_COST_BLACKSMITH)
    ap_shop       = settings.get('AP_COST_SHOP',       cfg.AP_COST_SHOP)
    tavern_cost   = settings.get('TAVERN_HEAL_COST',   cfg.TAVERN_HEAL_COST)

    # Boss/PvP: blocked by in_combat, blackout, or insufficient AP
    if in_combat:
        boss_ok, boss_reason = False, 'In combat'
        pvp_ok,  pvp_reason  = False, 'In combat'
    elif blackout:
        boss_ok, boss_reason = False, 'Approaching midnight reset'
        pvp_ok,  pvp_reason  = False, 'Approaching midnight reset'
    elif current_ap < ap_boss:
        boss_ok, boss_reason = False, f'Need {ap_boss} AP'
        pvp_ok,  pvp_reason  = False, f'Need {ap_pvp} AP'
    else:
        boss_ok, boss_reason = True, None
        pvp_ok,  pvp_reason  = current_ap >= ap_pvp, (None if current_ap >= ap_pvp else f'Need {ap_pvp} AP')

    # Tavern: no blackout restriction
    if in_combat:
        tavern_ok, tavern_reason = False, 'In combat'
    elif current_ap < ap_tavern:
        tavern_ok, tavern_reason = False, f'Need {ap_tavern} AP'
    elif credits < tavern_cost:
        tavern_ok, tavern_reason = False, f'Need {tavern_cost} credits'
    elif current_hp >= max_hp:
        tavern_ok, tavern_reason = False, 'Already at full health'
    else:
        tavern_ok, tavern_reason = True, None

    # Blacksmith: no blackout restriction; blocked at 0 credits
    if in_combat:
        bs_ok, bs_reason = False, 'In combat'
    elif current_ap < ap_blacksmith:
        bs_ok, bs_reason = False, f'Need {ap_blacksmith} AP'
    elif credits == 0:
        bs_ok, bs_reason = False, 'No credits'
    else:
        bs_ok, bs_reason = True, None

    # Shop: no blackout restriction
    if in_combat:
        shop_ok, shop_reason = False, 'In combat'
    elif current_ap < ap_shop:
        shop_ok, shop_reason = False, f'Need {ap_shop} AP'
    else:
        shop_ok, shop_reason = True, None

    return {
        'boss':       {'enabled': boss_ok,   'reason': boss_reason,   'ap_cost': ap_boss},
        'pvp':        {'enabled': pvp_ok,    'reason': pvp_reason,    'ap_cost': ap_pvp},
        'tavern':     {'enabled': tavern_ok, 'reason': tavern_reason, 'ap_cost': ap_tavern},
        'blacksmith': {'enabled': bs_ok,     'reason': bs_reason,     'ap_cost': ap_blacksmith},
        'shop':       {'enabled': shop_ok,   'reason': shop_reason,   'ap_cost': ap_shop},
    }

# FILE: routes/feeds.py
"""Personal and global feed endpoints used by the terminal interface."""
# routes/feeds.py
# Lightweight JSON polling endpoints for the live terminal feed.
# Called every 5 seconds by terminal.js.
# These are the only two JSON-returning routes in the main app.

from flask import Blueprint, jsonify, request, session
from database import execute

bp = Blueprint("feeds", __name__)


@bp.route("/feed/personal/latest")
def personal_latest():
    """Return new personal feed entries since a given timestamp.
    Query param: since=<ISO datetime string>"""
    player_id = session.get("player_id")
    since = request.args.get("since", "1970-01-01T00:00:00")

    rows = execute(
        """SELECT flavor_text, event_category, occurred_at, combat_session_id
           FROM daily_feed
           WHERE player_id = ? AND occurred_at > ?
           ORDER BY occurred_at ASC""",
        (player_id, since)
    )
    return jsonify(rows)


@bp.route("/feed/global/latest")
def global_latest():
    """Return new global feed entries since a given timestamp.
    Query param: since=<ISO datetime string>"""
    since = request.args.get("since", "1970-01-01T00:00:00")

    rows = execute(
        """SELECT flavor_text, event_category, occurred_at
           FROM daily_feed
           WHERE feed_scope = 'GLOBAL' AND occurred_at > ?
           ORDER BY occurred_at ASC""",
        (since,)
    )
    return jsonify(rows)


################################################################################

# FILE: routes/scoreboards.py
"""Read-only public progression, combat, economy, and shame rankings."""
# routes/scoreboards.py  (Phase 6 — full implementation)
# Full-page leaderboards. All data computed via live DB queries.
# Inactive players (7+ days without login) excluded from all boards.

import logging
from flask import Blueprint, render_template, g
from database import execute, execute_one, get_all_settings
import config_defaults as cfg

bp = Blueprint("scoreboards", __name__)
logger = logging.getLogger(__name__)


@bp.route("/scoreboards")
def index():
    """Handle the index workflow."""
    settings      = get_all_settings()
    inactive_days = settings.get("INACTIVE_DAYS_THRESHOLD", cfg.INACTIVE_DAYS_THRESHOLD)
    cutoff_expr   = f"datetime('now', '-{inactive_days} days')"

    top_level      = _top_level_xp(cutoff_expr)
    top_pvp_kills  = _top_pvp_kills(cutoff_expr)
    top_boss_global= _top_boss_kills_global(cutoff_expr)
    top_boss_each  = _top_boss_kills_per_boss(cutoff_expr)
    top_minion_gl  = _top_minion_kills_global(cutoff_expr)
    top_minion_each= _top_minion_kills_per_minion(cutoff_expr)
    top_credits    = _top_credits(cutoff_expr)
    shame_board    = _shame_board(cutoff_expr)

    return render_template(
        "scoreboards/scoreboards.html",
        top_level=top_level,
        top_pvp_kills=top_pvp_kills,
        top_boss_global=top_boss_global,
        top_boss_each=top_boss_each,
        top_minion_global=top_minion_gl,
        top_minion_each=top_minion_each,
        top_credits=top_credits,
        shame_board=shame_board,
    )


def _top_level_xp(cutoff_expr: str, limit: int = 20) -> list:
    """Query and rank players for the level xp scoreboard."""
    return execute(
        f"""SELECT character_name, level, xp, class_id
            FROM players
            WHERE is_banned = 0
              AND (last_login_at IS NULL OR last_login_at >= {cutoff_expr})
            ORDER BY level DESC, xp DESC
            LIMIT ?""",
        (limit,)
    )


def _top_pvp_kills(cutoff_expr: str, limit: int = 20) -> list:
    """Query and rank players for the pvp kills scoreboard."""
    return execute(
        f"""SELECT p.character_name, p.level, ps.pvp_kills
            FROM players p
            JOIN player_stats ps ON ps.player_id = p.id
            WHERE p.is_banned = 0
              AND (p.last_login_at IS NULL OR p.last_login_at >= {cutoff_expr})
            ORDER BY ps.pvp_kills DESC
            LIMIT ?""",
        (limit,)
    )


def _top_boss_kills_global(cutoff_expr: str, limit: int = 20) -> list:
    """Query and rank players for the boss kills global scoreboard."""
    return execute(
        f"""SELECT p.character_name, p.level, SUM(bi.kill_count) as total_kills
            FROM players p
            JOIN boss_instances bi ON bi.player_id = p.id
            WHERE p.is_banned = 0
              AND (p.last_login_at IS NULL OR p.last_login_at >= {cutoff_expr})
            GROUP BY p.id
            ORDER BY total_kills DESC
            LIMIT ?""",
        (limit,)
    )


def _top_boss_kills_per_boss(cutoff_expr: str, limit_per: int = 5) -> dict:
    """Returns dict of {boss_name: [top players]}."""
    bosses = execute("SELECT id, name FROM bosses WHERE is_active = 1 ORDER BY name")
    result = {}
    for boss in bosses:
        rows = execute(
            f"""SELECT p.character_name, bi.kill_count
                FROM boss_instances bi
                JOIN players p ON p.id = bi.player_id
                WHERE bi.boss_id = ?
                  AND p.is_banned = 0
                  AND (p.last_login_at IS NULL OR p.last_login_at >= {cutoff_expr})
                  AND bi.kill_count > 0
                ORDER BY bi.kill_count DESC
                LIMIT ?""",
            (boss["id"], limit_per)
        )
        if rows:
            result[boss["name"]] = rows
    return result


def _top_minion_kills_global(cutoff_expr: str, limit: int = 20) -> list:
    """Query and rank players for the minion kills global scoreboard."""
    return execute(
        f"""SELECT p.character_name, p.level, SUM(mi.kill_count) as total_kills
            FROM players p
            JOIN minion_instances mi ON mi.player_id = p.id
            WHERE p.is_banned = 0
              AND (p.last_login_at IS NULL OR p.last_login_at >= {cutoff_expr})
            GROUP BY p.id
            ORDER BY total_kills DESC
            LIMIT ?""",
        (limit,)
    )


def _top_minion_kills_per_minion(cutoff_expr: str, limit_per: int = 5) -> dict:
    """Query and rank players for the minion kills per minion scoreboard."""
    minions = execute("SELECT id, name FROM minions WHERE is_active = 1 ORDER BY name")
    result  = {}
    for minion in minions:
        rows = execute(
            f"""SELECT p.character_name, mi.kill_count
                FROM minion_instances mi
                JOIN players p ON p.id = mi.player_id
                WHERE mi.minion_id = ?
                  AND p.is_banned = 0
                  AND (p.last_login_at IS NULL OR p.last_login_at >= {cutoff_expr})
                  AND mi.kill_count > 0
                ORDER BY mi.kill_count DESC
                LIMIT ?""",
            (minion["id"], limit_per)
        )
        if rows:
            result[minion["name"]] = rows
    return result


def _top_credits(cutoff_expr: str, limit: int = 20) -> list:
    """Query and rank players for the credits scoreboard."""
    return execute(
        f"""SELECT character_name, level, credits
            FROM players
            WHERE is_banned = 0
              AND (last_login_at IS NULL OR last_login_at >= {cutoff_expr})
            ORDER BY credits DESC
            LIMIT ?""",
        (limit,)
    )


def _shame_board(cutoff_expr: str, limit: int = 20) -> list:
    """Provide the internal shame board operation used by this module."""
    return execute(
        f"""SELECT p.character_name, p.level, ps.times_reduced_to_1hp
            FROM players p
            JOIN player_stats ps ON ps.player_id = p.id
            WHERE p.is_banned = 0
              AND (p.last_login_at IS NULL OR p.last_login_at >= {cutoff_expr})
            ORDER BY ps.times_reduced_to_1hp DESC
            LIMIT ?""",
        (limit,)
    )


################################################################################

# FILE: routes/shop.py
"""Shop listings plus queued purchases, sales, pricing, and unique-item transfers."""
# routes/shop.py
# Full-page shop. Players buy from daily rotation and player-sold listings,
# and sell unequipped gear back. Every transaction redirects back to GET /shop.

import math
import logging
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, session, g
from database import (execute, execute_one, execute_write,
                      exclusive_transaction, get_all_settings)
from queue_handler import enqueue_and_process, register_handler
import config_defaults as cfg

bp = Blueprint("shop", __name__)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# GET /shop
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/shop")
def index():
    """Handle the index workflow."""
    player   = g.player
    settings = get_all_settings()

    # Calculate player's effective discount
    per_discount = math.floor(player["per_stat"] / 2) / 100
    special_discount = _get_special_shop_discount(player)
    max_discount = settings.get("SHOP_DISCOUNT_MAX", cfg.SHOP_DISCOUNT_MAX)
    discount = min(per_discount + special_discount, max_discount)

    # Load all current shop listings with content detail
    listings = _get_listings_with_detail(discount)

    # Load player's unequipped inventory for the sell panel
    sellable = _get_sellable_items(player)

    # Check if player spent AP to enter (AP deducted on first visit each session)
    # For simplicity: AP is deducted when the player clicks Shop from the dashboard.
    # The dashboard action_shop POST (handled here via redirect) deducts AP.

    return render_template(
        "shop/shop.html",
        listings=listings,
        sellable=sellable,
        discount_pct=int(discount * 100),
        feedback=request.args.get("feedback"),
        error=request.args.get("error"),
    )


def _get_listings_with_detail(discount: float) -> list[dict]:
    """Load shop_listings joined with content table for display."""
    rows = execute(
        """SELECT sl.*, sl.id as listing_id
           FROM shop_listings sl
           ORDER BY sl.item_type, sl.listed_at ASC"""
    )
    result = []
    for row in rows:
        detail = _get_item_detail(row["item_type"], row["item_id"])
        if detail is None:
            continue
        discounted_price = max(0, int(row["price"] * (1 - discount)))
        result.append({**row, **detail,
                       "discounted_price": discounted_price,
                       "listing_id": row["id"]})
    return result


def _get_item_detail(item_type: str, item_id: int) -> dict | None:
    """Load item detail from current database state."""
    table = {"WEAPON": "weapons", "ARMOR": "armor", "SPECIAL": "special_items"}.get(item_type)
    if not table:
        return None
    return execute_one(f"SELECT * FROM {table} WHERE id = ?", (item_id,))


def _get_sellable_items(player: dict) -> list[dict]:
    """Return inventory items that can be sold (unequipped, active content)."""
    equipped_ids = {
        player.get("equipped_weapon_id"),
        player.get("equipped_armor_id"),
        player.get("equipped_special_id"),
    } - {None}

    items = execute(
        "SELECT * FROM inventory_items WHERE player_id = ?", (player["id"],)
    )
    result = []
    for inv in items:
        if inv["id"] in equipped_ids:
            continue
        detail = _get_item_detail(inv["item_type"], inv["item_id"])
        if detail is None:
            continue
        sell_price = _calc_sell_price(detail, player)
        result.append({**inv, **detail,
                       "inv_id": inv["id"],
                       "sell_price": sell_price})
    return result


def _calc_sell_price(item_detail: dict, player: dict) -> int:
    """Calculate sell price: credit_cost * SELL_PRICE_PERCENT, boosted by Sell Bonus."""
    settings     = get_all_settings()
    sell_pct     = settings.get("SELL_PRICE_PERCENT", cfg.SELL_PRICE_PERCENT)
    sell_bonus   = _get_special_sell_bonus(player)
    final_pct    = min(sell_pct + sell_bonus, 1.0)
    return max(0, int(item_detail["credit_cost"] * final_pct))


def _get_special_shop_discount(player: dict) -> float:
    """Load special shop discount from current database state."""
    if not player.get("equipped_special_id"):
        return 0.0
    inv = execute_one(
        "SELECT item_id FROM inventory_items WHERE id = ?",
        (player["equipped_special_id"],)
    )
    if not inv:
        return 0.0
    s = execute_one("SELECT shop_discount FROM special_items WHERE id = ?", (inv["item_id"],))
    return s["shop_discount"] if s else 0.0


def _get_special_sell_bonus(player: dict) -> float:
    """Load special sell bonus from current database state."""
    if not player.get("equipped_special_id"):
        return 0.0
    inv = execute_one(
        "SELECT item_id FROM inventory_items WHERE id = ?",
        (player["equipped_special_id"],)
    )
    if not inv:
        return 0.0
    s = execute_one("SELECT sell_bonus FROM special_items WHERE id = ?", (inv["item_id"],))
    return s["sell_bonus"] if s else 0.0


# ─────────────────────────────────────────────────────────────────────────────
# POST /shop/buy
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/shop/buy", methods=["POST"])
def buy():
    """Handle the buy workflow."""
    listing_id = request.form.get("listing_id", type=int)
    if not listing_id:
        return redirect(url_for("shop.index", error="Invalid listing."))

    try:
        enqueue_and_process(session["player_id"], "shop_buy", {"listing_id": listing_id})
        return redirect(url_for("shop.index", feedback="Purchase successful."))
    except RuntimeError as e:
        return redirect(url_for("shop.index", error=str(e)))


@register_handler("shop_buy")
def handle_shop_buy(player_id: int, payload: dict) -> dict:
    """Process the queued shop buy action against validated game state."""
    listing_id = payload["listing_id"]
    settings   = get_all_settings()
    player     = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))

    listing = execute_one("SELECT * FROM shop_listings WHERE id = ?", (listing_id,))
    if listing is None:
        raise ValueError("Item is no longer available.")

    # Calculate discounted price
    per_discount     = math.floor(player["per_stat"] / 2) / 100
    special_discount = 0.0
    if player.get("equipped_special_id"):
        inv = execute_one(
            "SELECT item_id FROM inventory_items WHERE id = ?",
            (player["equipped_special_id"],)
        )
        if inv:
            s = execute_one("SELECT shop_discount FROM special_items WHERE id = ?", (inv["item_id"],))
            if s:
                special_discount = s["shop_discount"]
    max_discount     = settings.get("SHOP_DISCOUNT_MAX", cfg.SHOP_DISCOUNT_MAX)
    discount         = min(per_discount + special_discount, max_discount)
    final_price      = max(0, int(listing["price"] * (1 - discount)))

    if player["credits"] < final_price:
        raise ValueError(f"Not enough credits. Need {final_price}.")

    # Check inventory limit — buying always allowed but check for over-encumbered flag
    inv_limit = settings.get("INVENTORY_LIMIT", cfg.INVENTORY_LIMIT) + \
                math.floor(player["str_stat"] / 2)

    with exclusive_transaction():
        # Re-check listing still exists (race condition guard)
        listing = execute_one("SELECT * FROM shop_listings WHERE id = ?", (listing_id,))
        if listing is None:
            raise ValueError("Item was purchased by another player.")

        durability = listing["durability_at_listing"] or 100
        inv_id = execute_write(
            """INSERT INTO inventory_items
               (player_id, item_type, item_id, current_durability, acquired_method)
               VALUES (?, ?, ?, ?, 'SHOP_PURCHASE')""",
            (player_id, listing["item_type"], listing["item_id"], durability)
        )
        execute_write("DELETE FROM shop_listings WHERE id = ?", (listing_id,))
        execute_write(
            "UPDATE players SET credits = credits - ? WHERE id = ?",
            (final_price, player_id)
        )
        execute_write(
            """INSERT INTO item_history
               (player_id, item_type, item_id, item_name, event_type, credit_amount)
               VALUES (?, ?, ?, ?, 'PURCHASED', ?)""",
            (player_id, listing["item_type"], listing["item_id"],
             _get_item_name(listing["item_type"], listing["item_id"]), final_price)
        )
        # If special item: update registry
        if listing["item_type"] == "SPECIAL":
            execute_write(
                """UPDATE special_item_registry
                   SET status = 'IN_INVENTORY', current_owner_player_id = ?,
                       inventory_item_id = ?, last_acquired_method = 'SHOP_PURCHASE',
                       updated_at = ?
                   WHERE special_item_id = ?""",
                (player_id, inv_id, datetime.utcnow().isoformat(), listing["item_id"])
            )

    logger.info("Player %d bought item %s/%d for %d credits",
                player_id, listing["item_type"], listing["item_id"], final_price)
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# POST /shop/sell
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/shop/sell", methods=["POST"])
def sell():
    """Handle the sell workflow."""
    inv_id = request.form.get("inv_id", type=int)
    if not inv_id:
        return redirect(url_for("shop.index", error="Invalid item."))

    try:
        enqueue_and_process(session["player_id"], "shop_sell", {"inv_id": inv_id})
        return redirect(url_for("shop.index", feedback="Item listed for sale."))
    except RuntimeError as e:
        return redirect(url_for("shop.index", error=str(e)))


@register_handler("shop_sell")
def handle_shop_sell(player_id: int, payload: dict) -> dict:
    """Process the queued shop sell action against validated game state."""
    inv_id   = payload["inv_id"]
    settings = get_all_settings()
    sell_pct = settings.get("SELL_PRICE_PERCENT", cfg.SELL_PRICE_PERCENT)
    player   = execute_one("SELECT * FROM players WHERE id = ?", (player_id,))

    inv = execute_one(
        "SELECT * FROM inventory_items WHERE id = ? AND player_id = ?",
        (inv_id, player_id)
    )
    if inv is None:
        raise ValueError("Item not found in your inventory.")

    # Cannot sell equipped items
    equipped = {player.get("equipped_weapon_id"),
                player.get("equipped_armor_id"),
                player.get("equipped_special_id")}
    if inv_id in equipped:
        raise ValueError("Unequip the item before selling it.")

    detail    = _get_item_detail(inv["item_type"], inv["item_id"])
    if detail is None:
        raise ValueError("Item content not found.")

    # Apply sell bonus from special item
    sell_bonus = _get_special_sell_bonus(player)
    final_pct  = min(sell_pct + sell_bonus, 1.0)
    sell_price = max(0, int(detail["credit_cost"] * final_pct))

    with exclusive_transaction():
        # Delete from inventory
        execute_write("DELETE FROM inventory_items WHERE id = ?", (inv_id,))
        # Credit player
        execute_write(
            "UPDATE players SET credits = credits + ? WHERE id = ?",
            (sell_price, player_id)
        )
        # Create shop listing
        listing_id = execute_write(
            """INSERT INTO shop_listings
               (item_type, item_id, listing_source, seller_player_id,
                durability_at_listing, price)
               VALUES (?, ?, 'PLAYER_SOLD', ?, ?, ?)""",
            (inv["item_type"], inv["item_id"], player_id,
             inv["current_durability"], detail["credit_cost"])
        )
        # Log to item_history
        execute_write(
            """INSERT INTO item_history
               (player_id, item_type, item_id, item_name, event_type, credit_amount)
               VALUES (?, ?, ?, ?, 'SOLD', ?)""",
            (player_id, inv["item_type"], inv["item_id"],
             detail["name"], sell_price)
        )
        # If special item: update registry to IN_SHOP
        if inv["item_type"] == "SPECIAL":
            execute_write(
                """UPDATE special_item_registry
                   SET status = 'IN_SHOP', current_owner_player_id = NULL,
                       inventory_item_id = NULL, shop_listing_price = ?,
                       last_released_method = 'SOLD', updated_at = ?
                   WHERE special_item_id = ?""",
                (sell_price, datetime.utcnow().isoformat(), inv["item_id"])
            )

    logger.info("Player %d sold item %s/%d for %d credits",
                player_id, inv["item_type"], inv["item_id"], sell_price)
    return {"success": True, "sell_price": sell_price}


def _get_item_name(item_type: str, item_id: int) -> str:
    """Load item name from current database state."""
    detail = _get_item_detail(item_type, item_id)
    return detail["name"] if detail else "Unknown Item"


################################################################################

# FILE: routes/__init__.py

"""Flask route blueprints for player-facing game features."""

# FILE: schema.sql
-- schema.sql
-- Full database schema. Executed by database.init_db().
-- All tables use IF NOT EXISTS — safe to re-run on an existing DB.

-- ─────────────────────────────────────────────────────────────────────────────
-- PLAYERS & IDENTITY
-- ─────────────────────────────────────────────────────────────────────────────

-- Account identity, character statistics, resources, equipment pointers, and active-state flags.
CREATE TABLE IF NOT EXISTS players (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    username            TEXT    UNIQUE NOT NULL,
    password_hash       TEXT    NOT NULL,
    email               TEXT    UNIQUE NOT NULL,
    character_name      TEXT    NOT NULL,
    sex                 TEXT    NOT NULL,
    class_id            INTEGER REFERENCES classes(id),
    str_stat            INTEGER NOT NULL DEFAULT 1,
    end_stat            INTEGER NOT NULL DEFAULT 1,
    agi_stat            INTEGER NOT NULL DEFAULT 1,
    lck_stat            INTEGER NOT NULL DEFAULT 1,
    per_stat            INTEGER NOT NULL DEFAULT 1,
    level               INTEGER NOT NULL DEFAULT 1,
    xp                  INTEGER NOT NULL DEFAULT 0,
    current_hp          INTEGER NOT NULL,
    current_ap          INTEGER NOT NULL,
    credits             INTEGER NOT NULL DEFAULT 25,
    equipped_weapon_id  INTEGER REFERENCES inventory_items(id),
    equipped_armor_id   INTEGER REFERENCES inventory_items(id),
    equipped_special_id INTEGER REFERENCES inventory_items(id),
    in_combat           INTEGER NOT NULL DEFAULT 0,
    pending_levelup     INTEGER NOT NULL DEFAULT 0,
    combat_preference   TEXT    NOT NULL DEFAULT "Balanced",
    is_banned           INTEGER NOT NULL DEFAULT 0,
    retired_at          TEXT,
    last_login_at       TEXT,
    created_at          TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Long-lived aggregate player records that do not belong on the core character row.
CREATE TABLE IF NOT EXISTS player_stats (
    player_id            INTEGER PRIMARY KEY REFERENCES players(id),
    pvp_kills            INTEGER NOT NULL DEFAULT 0,
    times_reduced_to_1hp INTEGER NOT NULL DEFAULT 0,
    updated_at           TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Automated player characters. A profile row marks a normal player as an NPC.
-- Automation motivations and scheduling state attached to otherwise normal player characters.
CREATE TABLE IF NOT EXISTS npc_profiles (
    player_id          INTEGER PRIMARY KEY REFERENCES players(id),
    enabled            INTEGER NOT NULL DEFAULT 1,
    retired            INTEGER NOT NULL DEFAULT 0,
    player_hunter      INTEGER NOT NULL DEFAULT 0 CHECK(player_hunter BETWEEN 0 AND 100),
    boss_killer        INTEGER NOT NULL DEFAULT 0 CHECK(boss_killer BETWEEN 0 AND 100),
    hoarder            INTEGER NOT NULL DEFAULT 0 CHECK(hoarder BETWEEN 0 AND 100),
    thief              INTEGER NOT NULL DEFAULT 0 CHECK(thief BETWEEN 0 AND 100),
    aggression         INTEGER NOT NULL DEFAULT 50 CHECK(aggression BETWEEN 0 AND 100),
    self_preservation  INTEGER NOT NULL DEFAULT 50 CHECK(self_preservation BETWEEN 0 AND 100),
    repair_tendency    INTEGER NOT NULL DEFAULT 50 CHECK(repair_tendency BETWEEN 0 AND 100),
    actions_per_day    INTEGER NOT NULL DEFAULT 4 CHECK(actions_per_day BETWEEN 1 AND 24),
    actions_today      INTEGER NOT NULL DEFAULT 0,
    last_action_at     TEXT,
    created_at         TEXT NOT NULL DEFAULT (datetime('now'))
);

-- Administrator-readable explanations of automated NPC decisions and outcomes.
CREATE TABLE IF NOT EXISTS npc_action_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id   INTEGER NOT NULL REFERENCES players(id),
    decision    TEXT NOT NULL,
    reason      TEXT NOT NULL,
    result      TEXT NOT NULL,
    occurred_at TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE INDEX IF NOT EXISTS idx_npc_profiles_active ON npc_profiles(enabled, retired);
CREATE INDEX IF NOT EXISTS idx_npc_action_log_player ON npc_action_log(player_id, occurred_at);

-- Permanent record of each level-up stat point assigned to a character.
CREATE TABLE IF NOT EXISTS level_up_history (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id      INTEGER NOT NULL REFERENCES players(id),
    level_reached  INTEGER NOT NULL,
    stat_increased TEXT    NOT NULL,
    timestamp      TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Temporary character modifiers, normally cleared by the UTC reset.
CREATE TABLE IF NOT EXISTS status_effects (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id   INTEGER NOT NULL REFERENCES players(id),
    effect_type TEXT    NOT NULL,
    value       REAL    NOT NULL,
    created_at  TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Round-scoped or combat-scoped modifiers applied to one combat side.
CREATE TABLE IF NOT EXISTS combat_buffs (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    combat_session_id INTEGER NOT NULL REFERENCES combat_sessions(id),
    side              TEXT    NOT NULL,
    buff_type         TEXT    NOT NULL,
    damage_type       TEXT,
    value             REAL    NOT NULL,
    expires_on        TEXT    NOT NULL,
    created_at        TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- ─────────────────────────────────────────────────────────────────────────────
-- COMBAT
-- ─────────────────────────────────────────────────────────────────────────────

-- Authoritative lifecycle and summary state for PvP, boss, and minion fights.
CREATE TABLE IF NOT EXISTS combat_sessions (
    id                          INTEGER PRIMARY KEY AUTOINCREMENT,
    combat_type                 TEXT    NOT NULL,
    attacker_player_id          INTEGER NOT NULL REFERENCES players(id),
    defender_player_id          INTEGER REFERENCES players(id),
    boss_instance_id            INTEGER REFERENCES boss_instances(id),
    minion_instance_id          INTEGER REFERENCES minion_instances(id),
    status                      TEXT    NOT NULL DEFAULT "ACTIVE",
    result                      TEXT,
    current_round               INTEGER NOT NULL DEFAULT 1,
    rounds_extended             INTEGER NOT NULL DEFAULT 0,
    attacker_hp_start           INTEGER NOT NULL,
    defender_hp_start           INTEGER,
    attacker_total_damage_dealt INTEGER NOT NULL DEFAULT 0,
    defender_total_damage_dealt INTEGER NOT NULL DEFAULT 0,
    attacker_observed           INTEGER NOT NULL DEFAULT 0,
    defender_observed           INTEGER NOT NULL DEFAULT 0,
    started_at                  TEXT    NOT NULL DEFAULT (datetime('now')),
    resolved_at                 TEXT
);

-- Per-player persistent discovery, HP, phase, and kill state for bosses.
CREATE TABLE IF NOT EXISTS boss_instances (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id           INTEGER NOT NULL REFERENCES players(id),
    boss_id             INTEGER NOT NULL REFERENCES bosses(id),
    current_hp          INTEGER NOT NULL,
    special_attack_used INTEGER NOT NULL DEFAULT 0,
    special_buff_used   INTEGER NOT NULL DEFAULT 0,
    current_phase       INTEGER NOT NULL DEFAULT 1,
    discovered_at       TEXT    NOT NULL DEFAULT (datetime('now')),
    kill_count          INTEGER NOT NULL DEFAULT 0,
    UNIQUE(player_id, boss_id)
);

-- Per-player persistent discovery, HP, and kill state for minions.
CREATE TABLE IF NOT EXISTS minion_instances (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id     INTEGER NOT NULL REFERENCES players(id),
    minion_id     INTEGER NOT NULL REFERENCES minions(id),
    current_hp    INTEGER NOT NULL,
    discovered_at TEXT    NOT NULL DEFAULT (datetime('now')),
    kill_count    INTEGER NOT NULL DEFAULT 0,
    UNIQUE(player_id, minion_id)
);

-- Permanent record that a player has learned a boss’s combat information.
CREATE TABLE IF NOT EXISTS boss_intel (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id  INTEGER NOT NULL REFERENCES players(id),
    boss_id    INTEGER NOT NULL REFERENCES bosses(id),
    learned_at TEXT    NOT NULL DEFAULT (datetime('now')),
    UNIQUE(player_id, boss_id)
);

-- Detailed round-by-round audit records for combat actions and outcomes.
CREATE TABLE IF NOT EXISTS combat_logs (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    combat_session_id INTEGER NOT NULL REFERENCES combat_sessions(id),
    round_number      INTEGER NOT NULL,
    actor             TEXT    NOT NULL,
    action_type       TEXT    NOT NULL,
    roll_detail       TEXT    NOT NULL,
    outcome_detail    TEXT    NOT NULL,
    hp_after_attacker INTEGER,
    hp_after_defender INTEGER,
    created_at        TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- ─────────────────────────────────────────────────────────────────────────────
-- INVENTORY & ITEMS
-- ─────────────────────────────────────────────────────────────────────────────

-- Physical item copies owned by players, including durability and acquisition method.
CREATE TABLE IF NOT EXISTS inventory_items (
    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id          INTEGER NOT NULL REFERENCES players(id),
    item_type          TEXT    NOT NULL,
    item_id            INTEGER NOT NULL,
    current_durability INTEGER NOT NULL DEFAULT 100,
    acquired_at        TEXT    NOT NULL DEFAULT (datetime('now')),
    acquired_method    TEXT    NOT NULL
);

-- Permanent acquisition, sale, theft, drop, grant, and loss history.
CREATE TABLE IF NOT EXISTS item_history (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id         INTEGER NOT NULL REFERENCES players(id),
    item_type         TEXT    NOT NULL,
    item_id           INTEGER NOT NULL,
    item_name         TEXT    NOT NULL,
    event_type        TEXT    NOT NULL,
    credit_amount     INTEGER,
    related_player_id INTEGER REFERENCES players(id),
    occurred_at       TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Single-source ownership and location state for globally unique special items.
CREATE TABLE IF NOT EXISTS special_item_registry (
    id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    special_item_id         INTEGER NOT NULL UNIQUE REFERENCES special_items(id),
    status                  TEXT    NOT NULL DEFAULT "IN_POOL",
    current_owner_player_id INTEGER REFERENCES players(id),
    inventory_item_id       INTEGER REFERENCES inventory_items(id),
    shop_listing_price      INTEGER,
    last_acquired_method    TEXT,
    last_released_method    TEXT,
    updated_at              TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- ─────────────────────────────────────────────────────────────────────────────
-- ECONOMY
-- ─────────────────────────────────────────────────────────────────────────────

-- Items currently offered by the system or a player through the Shop.
CREATE TABLE IF NOT EXISTS shop_listings (
    id                    INTEGER PRIMARY KEY AUTOINCREMENT,
    item_type             TEXT    NOT NULL,
    item_id               INTEGER NOT NULL,
    listing_source        TEXT    NOT NULL,
    seller_player_id      INTEGER REFERENCES players(id),
    durability_at_listing INTEGER,
    price                 INTEGER NOT NULL,
    listed_at             TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- ─────────────────────────────────────────────────────────────────────────────
-- FEEDS
-- ─────────────────────────────────────────────────────────────────────────────

-- Time-ordered personal and global messages displayed in the terminal interface.
CREATE TABLE IF NOT EXISTS daily_feed (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    feed_scope        TEXT    NOT NULL,
    player_id         INTEGER REFERENCES players(id),
    flavor_text       TEXT    NOT NULL,
    event_category    TEXT    NOT NULL,
    combat_session_id INTEGER REFERENCES combat_sessions(id),
    occurred_at       TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- ─────────────────────────────────────────────────────────────────────────────
-- QUEUE
-- ─────────────────────────────────────────────────────────────────────────────

-- Auditable receipts for every shared state-changing player or NPC action.
CREATE TABLE IF NOT EXISTS action_queue (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id    INTEGER NOT NULL REFERENCES players(id),
    action_type  TEXT    NOT NULL,
    payload      TEXT    NOT NULL,
    status       TEXT    NOT NULL DEFAULT "PROCESSING",
    created_at   TEXT    NOT NULL DEFAULT (datetime('now')),
    processed_at TEXT
);

-- Unified per-character success, failure, diagnostic, and action history.
CREATE TABLE IF NOT EXISTS player_activity_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id   INTEGER NOT NULL REFERENCES players(id),
    category    TEXT NOT NULL,
    action      TEXT NOT NULL,
    status      TEXT NOT NULL,
    message     TEXT NOT NULL,
    details_json TEXT,
    queue_id    INTEGER REFERENCES action_queue(id),
    source      TEXT NOT NULL DEFAULT 'GAME',
    occurred_at TEXT NOT NULL DEFAULT (datetime('now'))
);

-- Permanent reasons and before/after details for significant administrator changes.
CREATE TABLE IF NOT EXISTS admin_audit_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    action      TEXT NOT NULL,
    target_type TEXT NOT NULL,
    target_id   INTEGER,
    reason      TEXT,
    details_json TEXT,
    occurred_at TEXT NOT NULL DEFAULT (datetime('now'))
);

-- Start, completion, failure, and result summaries for background jobs.
CREATE TABLE IF NOT EXISTS scheduler_run_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    job_name    TEXT NOT NULL,
    status      TEXT NOT NULL,
    result_summary TEXT,
    started_at  TEXT NOT NULL,
    finished_at TEXT
);

-- ─────────────────────────────────────────────────────────────────────────────
-- CONTENT TABLES (Excel-imported)
-- ─────────────────────────────────────────────────────────────────────────────

-- Excel-imported class definitions and permanent creation bonuses.
CREATE TABLE IF NOT EXISTS classes (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT    UNIQUE NOT NULL,
    is_active   INTEGER NOT NULL DEFAULT 1,
    str_bonus   INTEGER NOT NULL DEFAULT 0,
    end_bonus   INTEGER NOT NULL DEFAULT 0,
    agi_bonus   INTEGER NOT NULL DEFAULT 0,
    lck_bonus   INTEGER NOT NULL DEFAULT 0,
    per_bonus   INTEGER NOT NULL DEFAULT 0,
    description TEXT,
    imported_at TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Excel-imported boss statistics, phases, attacks, rewards, and narrative definitions.
CREATE TABLE IF NOT EXISTS bosses (
    id                          INTEGER PRIMARY KEY AUTOINCREMENT,
    name                        TEXT    UNIQUE NOT NULL,
    is_active                   INTEGER NOT NULL DEFAULT 1,
    level                       INTEGER NOT NULL,
    str_stat                    INTEGER NOT NULL,
    end_stat                    INTEGER NOT NULL,
    agi_stat                    INTEGER NOT NULL,
    lck_stat                    INTEGER NOT NULL,
    per_stat                    INTEGER NOT NULL,
    max_hp                      INTEGER NOT NULL,
    phase2_hp_percent           INTEGER NOT NULL,
    phase3_hp_percent           INTEGER NOT NULL,
    special_attack_name         TEXT    NOT NULL,
    special_attack_die          TEXT    NOT NULL,
    special_attack_damage_type  TEXT    NOT NULL,
    special_attack_flavor       TEXT    NOT NULL,
    special_buff_name           TEXT    NOT NULL,
    special_buff_type           TEXT    NOT NULL,
    special_buff_value          REAL    NOT NULL,
    special_buff_damage_type    TEXT,
    special_buff_flavor         TEXT    NOT NULL,
    res_blade     INTEGER NOT NULL DEFAULT 0,
    res_blunt     INTEGER NOT NULL DEFAULT 0,
    res_ballistic INTEGER NOT NULL DEFAULT 0,
    res_energy    INTEGER NOT NULL DEFAULT 0,
    res_arcane    INTEGER NOT NULL DEFAULT 0,
    res_explosive INTEGER NOT NULL DEFAULT 0,
    res_venom     INTEGER NOT NULL DEFAULT 0,
    weak_blade    INTEGER NOT NULL DEFAULT 0,
    weak_blunt    INTEGER NOT NULL DEFAULT 0,
    weak_ballistic INTEGER NOT NULL DEFAULT 0,
    weak_energy   INTEGER NOT NULL DEFAULT 0,
    weak_arcane   INTEGER NOT NULL DEFAULT 0,
    weak_explosive INTEGER NOT NULL DEFAULT 0,
    weak_venom    INTEGER NOT NULL DEFAULT 0,
    drop_weapon_chance       REAL    NOT NULL,
    drop_armor_chance        REAL    NOT NULL,
    drop_special_item_chance REAL    NOT NULL,
    drop_credit_min          INTEGER NOT NULL,
    drop_credit_max          INTEGER NOT NULL,
    flavor_text              TEXT    NOT NULL,
    imported_at              TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Excel-imported minion statistics and combat definitions.
CREATE TABLE IF NOT EXISTS minions (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT    UNIQUE NOT NULL,
    is_active     INTEGER NOT NULL DEFAULT 1,
    level         INTEGER NOT NULL,
    str_stat      INTEGER NOT NULL,
    end_stat      INTEGER NOT NULL,
    agi_stat      INTEGER NOT NULL,
    lck_stat      INTEGER NOT NULL,
    per_stat      INTEGER NOT NULL,
    max_hp        INTEGER NOT NULL,
    drop_weapon_chance       REAL    NOT NULL,
    drop_armor_chance        REAL    NOT NULL,
    drop_special_item_chance REAL    NOT NULL,
    drop_credit_min          INTEGER NOT NULL,
    drop_credit_max          INTEGER NOT NULL,
    flavor_text   TEXT    NOT NULL,
    imported_at   TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Weapon balance definitions shared by all inventory copies.
CREATE TABLE IF NOT EXISTS weapons (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT    UNIQUE NOT NULL,
    is_active       INTEGER NOT NULL DEFAULT 1,
    level           INTEGER NOT NULL,
    weapon_type     TEXT    NOT NULL,
    damage_die      TEXT    NOT NULL,
    damage_type     TEXT    NOT NULL,
    str_bonus       INTEGER NOT NULL DEFAULT 0,
    end_bonus       INTEGER NOT NULL DEFAULT 0,
    agi_bonus       INTEGER NOT NULL DEFAULT 0,
    lck_bonus       INTEGER NOT NULL DEFAULT 0,
    per_bonus       INTEGER NOT NULL DEFAULT 0,
    associated_to   TEXT,
    credit_cost     INTEGER NOT NULL,
    drop_chance     REAL    NOT NULL,
    starting_durability INTEGER NOT NULL DEFAULT 100,
    imported_at     TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Armor class, resistance, stat, economy, and durability definitions.
CREATE TABLE IF NOT EXISTS armor (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT    UNIQUE NOT NULL,
    is_active       INTEGER NOT NULL DEFAULT 1,
    level           INTEGER NOT NULL,
    ac_bonus        INTEGER NOT NULL DEFAULT 0,
    res_blade       INTEGER NOT NULL DEFAULT 0,
    res_blunt       INTEGER NOT NULL DEFAULT 0,
    res_ballistic   INTEGER NOT NULL DEFAULT 0,
    res_energy      INTEGER NOT NULL DEFAULT 0,
    res_arcane      INTEGER NOT NULL DEFAULT 0,
    res_explosive   INTEGER NOT NULL DEFAULT 0,
    res_venom       INTEGER NOT NULL DEFAULT 0,
    str_bonus       INTEGER NOT NULL DEFAULT 0,
    end_bonus       INTEGER NOT NULL DEFAULT 0,
    agi_bonus       INTEGER NOT NULL DEFAULT 0,
    lck_bonus       INTEGER NOT NULL DEFAULT 0,
    per_bonus       INTEGER NOT NULL DEFAULT 0,
    associated_to   TEXT,
    credit_cost     INTEGER NOT NULL,
    drop_chance     REAL    NOT NULL,
    starting_durability INTEGER NOT NULL DEFAULT 100,
    imported_at     TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Unique special-item modifiers, associations, economy values, and durability definitions.
CREATE TABLE IF NOT EXISTS special_items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT    UNIQUE NOT NULL,
    is_active       INTEGER NOT NULL DEFAULT 1,
    associated_to   TEXT    NOT NULL,
    association_type TEXT   NOT NULL,
    str_bonus       INTEGER NOT NULL DEFAULT 0,
    end_bonus       INTEGER NOT NULL DEFAULT 0,
    agi_bonus       INTEGER NOT NULL DEFAULT 0,
    lck_bonus       INTEGER NOT NULL DEFAULT 0,
    per_bonus       INTEGER NOT NULL DEFAULT 0,
    initiative_bonus    INTEGER NOT NULL DEFAULT 0,
    extra_attack        INTEGER NOT NULL DEFAULT 0,
    crit_chance_bonus   REAL    NOT NULL DEFAULT 0,
    crit_dmg_multiplier REAL    NOT NULL DEFAULT 0,
    ac_bonus            INTEGER NOT NULL DEFAULT 0,
    res_blade       INTEGER NOT NULL DEFAULT 0,
    res_blunt       INTEGER NOT NULL DEFAULT 0,
    res_ballistic   INTEGER NOT NULL DEFAULT 0,
    res_energy      INTEGER NOT NULL DEFAULT 0,
    res_arcane      INTEGER NOT NULL DEFAULT 0,
    res_explosive   INTEGER NOT NULL DEFAULT 0,
    res_venom       INTEGER NOT NULL DEFAULT 0,
    bonus_damage_type   TEXT,
    bonus_damage_amount INTEGER NOT NULL DEFAULT 0,
    xp_multiplier       REAL    NOT NULL DEFAULT 0,
    credit_multiplier   REAL    NOT NULL DEFAULT 0,
    steal_bonus         REAL    NOT NULL DEFAULT 0,
    bonus_ap            INTEGER NOT NULL DEFAULT 0,
    hp_regen_bonus      INTEGER NOT NULL DEFAULT 0,
    durability_reduction REAL   NOT NULL DEFAULT 0,
    shop_discount       REAL    NOT NULL DEFAULT 0,
    sell_bonus          REAL    NOT NULL DEFAULT 0,
    encounter_bonus     REAL    NOT NULL DEFAULT 0,
    credit_cost         INTEGER NOT NULL,
    drop_chance         REAL    NOT NULL,
    starting_durability INTEGER NOT NULL DEFAULT 100,
    imported_at         TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Weighted good and bad encounters plus their mechanical effects.
CREATE TABLE IF NOT EXISTS random_events (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT    UNIQUE NOT NULL,
    is_active     INTEGER NOT NULL DEFAULT 1,
    event_type    TEXT    NOT NULL,
    rarity        TEXT    NOT NULL,
    flavor_text   TEXT    NOT NULL,
    effect_type   TEXT    NOT NULL,
    effect_amount INTEGER NOT NULL,
    duration      TEXT    NOT NULL,
    imported_at   TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Movie-level relationships connecting bosses, minions, protagonists, and their equipment.
CREATE TABLE IF NOT EXISTS master (
    id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    movie_name              TEXT    UNIQUE NOT NULL,
    is_active               INTEGER NOT NULL DEFAULT 1,
    boss_id                 INTEGER NOT NULL REFERENCES bosses(id),
    boss_weapon_id          INTEGER NOT NULL REFERENCES weapons(id),
    boss_armor_id           INTEGER NOT NULL REFERENCES armor(id),
    boss_special_item_id    INTEGER NOT NULL REFERENCES special_items(id),
    minion_id               INTEGER NOT NULL REFERENCES minions(id),
    minion_weapon_id        INTEGER NOT NULL REFERENCES weapons(id),
    minion_armor_id         INTEGER NOT NULL REFERENCES armor(id),
    minion_special_item_id  INTEGER NOT NULL REFERENCES special_items(id),
    protagonist_name        TEXT,
    protagonist_weapon_id   INTEGER REFERENCES weapons(id),
    protagonist_armor_id    INTEGER REFERENCES armor(id),
    protagonist_special_item_id INTEGER REFERENCES special_items(id),
    imported_at             TEXT    NOT NULL DEFAULT (datetime('now'))
);

-- Database overrides for typed gameplay defaults in config_defaults.py.
CREATE TABLE IF NOT EXISTS settings (
    constant_name TEXT PRIMARY KEY,
    value         TEXT NOT NULL,
    description   TEXT,
    imported_at   TEXT NOT NULL DEFAULT (datetime('now'))
);

-- ─────────────────────────────────────────────────────────────────────────────
-- INDEXES
-- ─────────────────────────────────────────────────────────────────────────────

CREATE INDEX IF NOT EXISTS idx_players_username        ON players(username);
CREATE INDEX IF NOT EXISTS idx_players_email           ON players(email);
CREATE INDEX IF NOT EXISTS idx_players_in_combat       ON players(in_combat);
CREATE INDEX IF NOT EXISTS idx_inventory_player        ON inventory_items(player_id);
CREATE INDEX IF NOT EXISTS idx_inventory_type          ON inventory_items(player_id, item_type);
CREATE INDEX IF NOT EXISTS idx_combat_sessions_status  ON combat_sessions(status);
CREATE INDEX IF NOT EXISTS idx_combat_sessions_attacker ON combat_sessions(attacker_player_id);
CREATE INDEX IF NOT EXISTS idx_combat_logs_session     ON combat_logs(combat_session_id);
CREATE INDEX IF NOT EXISTS idx_daily_feed_player       ON daily_feed(player_id, occurred_at);
CREATE INDEX IF NOT EXISTS idx_daily_feed_global       ON daily_feed(feed_scope, occurred_at);
CREATE INDEX IF NOT EXISTS idx_boss_instances_player   ON boss_instances(player_id);
CREATE INDEX IF NOT EXISTS idx_minion_instances_player ON minion_instances(player_id);
CREATE INDEX IF NOT EXISTS idx_action_queue_status     ON action_queue(status, created_at);
CREATE INDEX IF NOT EXISTS idx_player_activity_date    ON player_activity_log(player_id, occurred_at);
CREATE INDEX IF NOT EXISTS idx_player_activity_status  ON player_activity_log(status, occurred_at);
CREATE INDEX IF NOT EXISTS idx_admin_audit_date        ON admin_audit_log(occurred_at);
CREATE INDEX IF NOT EXISTS idx_scheduler_run_date      ON scheduler_run_log(job_name, started_at);
CREATE INDEX IF NOT EXISTS idx_item_history_player     ON item_history(player_id);
CREATE INDEX IF NOT EXISTS idx_special_registry_status ON special_item_registry(status);

# FILE: static/help.js
/*
 * Contextual help catalog for Movie Multiverse.
 *
 * Keys under FIELD_HELP match HTML form `name` attributes. TEXT_HELP keys
 * match visible labels, buttons, links, and table headings. Keeping definitions
 * here makes player/admin wording consistent and gives maintainers one concise
 * place to update a mechanic's explanation.
 */
(() => {
  'use strict';
  const FIELD_HELP = {
    username: 'Account login name. This is separate from the public character name.',
    email: 'Account recovery/contact address. It is not displayed to other players.',
    password: 'Account password. Passwords are never displayed or written to activity logs.',
    character_name: 'Public, permanent character name shown in combat, feeds, and scoreboards.',
    sex: 'Character identity selection. It does not change combat statistics.',
    class_id: 'Permanent class choice. Class bonuses are added to the character’s base statistics.',
    level: 'Progress tier. Level affects HP, eligible opponents, rewards, and content difficulty.',
    str_bonus: 'Strength added by this class or equipped item.',
    end_bonus: 'Endurance added by this class or equipped item.',
    agi_bonus: 'Agility added by this class or equipped item.',
    lck_bonus: 'Luck added by this class or equipped item.',
    per_bonus: 'Perception added by this class or equipped item.',
    xp: 'Experience earned through combat and certain events. Thresholds award a level and stat point.',
    current_hp: 'Current health. At 1 HP a player is defeated but not killed.',
    current_ap: 'Action Points available for activities. AP is awarded at reset and through scheduled trickles.',
    credits: 'Spendable currency used by the Tavern, Blacksmith, Shop, and other systems.',
    str_stat: 'Strength: improves melee damage and inventory capacity.',
    end_stat: 'Endurance: improves maximum HP, AP allowance, and regeneration.',
    agi_stat: 'Agility: improves ranged damage, armor class, dodge, initiative, stealing, and escape.',
    lck_stat: 'Luck: improves critical results, random events, stealing, escape, and some economy rolls.',
    per_stat: 'Perception: improves observation, detection, and selected economy checks.',
    str_alloc: 'Creation points assigned to Strength before class bonuses.',
    end_alloc: 'Creation points assigned to Endurance before class bonuses.',
    agi_alloc: 'Creation points assigned to Agility before class bonuses.',
    lck_alloc: 'Creation points assigned to Luck before class bonuses.',
    per_alloc: 'Creation points assigned to Perception before class bonuses.',
    preset: 'Fills the NPC behavior fields with a suggested archetype. You may customize the values afterward.',
    player_hunter: 'Motivation to seek legal PvP fights. Higher values make PvP more important than other goals.',
    boss_killer: 'Motivation to challenge bosses for progression, rewards, and completion.',
    hoarder: 'Motivation to acquire and retain unique special items.',
    thief: 'Motivation to alternate legal PvP stealing attempts with boss fights for progression.',
    aggression: 'Willingness to select stronger opponents and keep attacking instead of defending or escaping.',
    self_preservation: 'Willingness to heal, brace, avoid risk, and escape when health becomes dangerous.',
    repair_tendency: 'Likelihood of visiting the Blacksmith when equipped gear loses durability.',
    enabled: 'Paused NPCs remain in the world but do not receive scheduled automated decisions.',
    item_key: 'Content item to grant directly. Unique specials must currently be available in the global pool.',
    constant_name: 'Configuration key used by the game. Change only a setting whose effect you understand.',
    value: 'Stored configuration value. It is converted to the setting’s expected number, boolean, or text type.',
    reason: 'Required audit explanation describing why this administrative or balancing change was made.',
    name: 'Display name used anywhere this content appears in the game.',
    is_active: 'Disabled content remains in historical records but is excluded from new gameplay selection.',
    weapon_type: 'Weapon category used by combat flavor and related rules.',
    damage_die: 'Dice expression rolled for base weapon damage, such as d6 or 2d4.',
    damage_type: 'Damage category checked against armor resistance.',
    credit_cost: 'Base credit value used when purchasing, selling, rewarding, or comparing this item.',
    drop_chance: 'Probability weight used when this content is considered for a drop.',
    starting_durability: 'Durability assigned when a new copy enters an inventory.',
    ac_bonus: 'Armor Class added while equipped, making attacks less likely to hit.',
    res_blade: 'Damage reduction against blade attacks.', res_blunt: 'Damage reduction against blunt attacks.',
    res_ballistic: 'Damage reduction against ballistic attacks.', res_energy: 'Damage reduction against energy attacks.',
    res_arcane: 'Damage reduction against arcane attacks.', res_explosive: 'Damage reduction against explosive attacks.',
    res_venom: 'Damage reduction against venom attacks.',
    associated_to: 'Movie character or content record this unique item belongs to.',
    association_type: 'Whether the associated owner is a Boss, Minion, or Protagonist.',
    initiative_bonus: 'Bonus applied when determining who acts first in combat.',
    extra_attack: 'Chance or flag allowing an additional attack when the special’s rules trigger.',
    crit_chance_bonus: 'Additional critical-hit probability supplied by this special.',
    crit_dmg_multiplier: 'Additional multiplier applied to critical-hit damage.',
    xp_multiplier: 'Fractional bonus to XP rewards; 0.10 means ten percent.',
    credit_multiplier: 'Fractional bonus to credit rewards.',
    steal_bonus: 'Bonus applied to opposed stealing rolls and applicable steal rewards.',
    bonus_ap: 'Additional AP capacity supplied while this special is equipped.',
    hp_regen_bonus: 'Additional health restored by applicable regeneration effects.',
    durability_reduction: 'Fractional reduction to durability loss; 0.10 means ten percent less loss.',
    shop_discount: 'Fractional reduction to shop purchase prices.',
    sell_bonus: 'Fractional increase to eligible sale proceeds.',
    encounter_bonus: 'Bonus applied to the chance or quality of random encounters.',
    start: 'Include log entries on or after this date.', end: 'Include log entries through this date.',
    category: 'Restrict results to one kind of recorded activity.', errors_only: 'Show only failed actions and errors.'
  };
  const TEXT_HELP = {
    'dashboard': 'Summary of current game state and administrative warnings.',
    'import excel': 'Stage a game-content workbook for validated import at the next reset.',
    'players': 'Inspect characters, histories, inventory, statistics, and administrative state.',
    'npcs': 'Create and manage automated characters that follow ordinary player rules.',
    'items': 'Inspect and rebalance weapon, armor, and special-item definitions.',
    'analytics': 'Aggregate gameplay data used to evaluate progression and balance.',
    'health & audit': 'Inspect failed actions, scheduler runs, integrity warnings, and admin changes.',
    'config': 'Edit global gameplay constants. Changes can affect every player.',
    'logs': 'Inspect import errors, orphan recovery, and failed queued actions.',
    'midnight reset': 'Immediately run the normal UTC reset sequence. This affects the entire game.',
    'create npc': 'Create an ordinary player character controlled by the NPC decision scheduler.',
    'run one turn': 'Run one NPC decision immediately without waiting for its scheduled time.',
    'spend ap now': 'Run repeated ordinary NPC decisions until AP or useful progress stops.',
    'retire': 'Permanently remove a character from active play while retaining historical records.',
    'retire character': 'Permanently disable login, end combat, and return unique specials to circulation.',
    'ban player': 'Disciplinary removal that wipes credits and inventory. Retirement is a different operation.',
    'save item': 'Apply visible item changes and record the required reason in the admin audit log.',
    'grant item': 'Place this item in the selected NPC inventory under normal uniqueness constraints.',
    'activity log': 'Open this character’s chronological action and error history.',
    'boss': 'Spend AP to challenge an available movie villain.',
    'pvp': 'Spend AP to challenge a player allowed by the PvP eligibility rules.',
    'tavern': 'Spend AP and credits to restore missing health.',
    'blacksmith': 'Spend resources to restore durability on damaged equipment.',
    'shop': 'Browse available equipment and unique special items.',
    'character': 'Review statistics, effects, inventory, equipment, and combat preference.',
    'scoreboards': 'Compare public progression and combat records.',
    'attack': 'Use the equipped weapon to damage the opponent.',
    'brace': 'Take a defensive combat action that improves survival for the round.',
    'observe': 'Study an opponent to reveal useful combat information.',
    'steal': 'Attempt the normal opposed combat steal action; failure creates a defensive penalty.',
    'escape': 'Spend AP and make an opposed roll to leave combat.',
    'str': 'Strength improves melee damage and inventory capacity.',
    'end': 'Endurance improves HP, AP allowance, and regeneration.',
    'agi': 'Agility improves ranged combat, defense, initiative, stealing, and escape.',
    'lck': 'Luck influences critical results, events, stealing, escape, and economy rolls.',
    'per': 'Perception influences observation, detection, and selected economy checks.',
    'lvl': 'Current character level.', 'hp': 'Current health compared with maximum health.',
    'ap': 'Current Action Points compared with the character’s allowance.', 'cr': 'Current spendable credits.'
  };
  const normalize = value => (value || '').replace(/\s+/g, ' ').trim().toLowerCase().replace(/\s*\([^)]*\)\s*$/, '');
  function attach(element, text, showBadge) {
    if (!text || element.dataset.helpAttached) return;
    element.dataset.helpAttached = '1'; element.title = text; element.setAttribute('aria-description', text);
    if (!showBadge) return;
    const badge = document.createElement('span'); badge.className = 'context-help'; badge.tabIndex = 0;
    badge.setAttribute('role', 'note'); badge.setAttribute('aria-label', text); badge.dataset.help = text; badge.textContent = '?';
    element.insertAdjacentElement('afterend', badge);
  }
  function install() {
    document.querySelectorAll('input[name],select[name],textarea[name]').forEach(control => {
      const text = FIELD_HELP[control.name]; if (!text) return;
      attach(control, text, false);
      const label = control.closest('label') || (control.id && document.querySelector(`label[for="${control.id}"]`));
      if (label) attach(label, text, true);
    });
    document.querySelectorAll('button,a,th,.label,[data-help-key]').forEach(element => {
      const key = element.dataset.helpKey || normalize(element.textContent);
      const text = TEXT_HELP[key]; if (text) attach(element, text, element.matches('th,.label'));
    });
  }
  document.readyState === 'loading' ? document.addEventListener('DOMContentLoaded', install) : install();
})();

# FILE: static/style.css
/* style.css
   Dark terminal theme for the BBS-inspired dueling game.
   Monospace throughout, color-coded terminal output. */

/* ── CSS Variables ─────────────────────────────────────────── */
:root {
    --bg:         #0a0a0a;
    --bg-panel:   #111111;
    --bg-input:   #1a1a1a;
    --border:     #2a2a2a;
    --green:      #00cc66;
    --red:        #cc2222;
    --amber:      #ffaa00;
    --blue:       #4499ff;
    --grey:       #666666;
    --white:      #dddddd;
    --dim:        #444444;
    --font:       'Courier New', Courier, monospace;
    --left-width: 220px;
    --ticker-h:   32px;
}

:root[data-theme="light"] {
    --bg:       #f4f0e7;
    --bg-panel: #ffffff;
    --bg-input: #faf7f0;
    --border:   #c9bea9;
    --green:    #176b42;
    --red:      #a12b2b;
    --amber:    #8a5400;
    --blue:     #185f9d;
    --grey:     #655f55;
    --white:    #201d18;
    --dim:      #8a8174;
    color-scheme: light;
}

/* ── Reset & Base ──────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body {
    height: 100%;
    background: var(--bg);
    color: var(--white);
    font-family: var(--font);
    font-size: 14px;
    overflow: hidden;
}

a { color: var(--blue); text-decoration: none; }
a:hover { text-decoration: underline; }

#theme-toggle {
    position: fixed;
    top: 12px;
    right: 14px;
    z-index: 1000;
    min-width: 112px;
    padding: 7px 10px;
    border: 1px solid var(--amber);
    background: var(--bg-panel);
    color: var(--amber);
    font: bold 11px/1 var(--font);
    letter-spacing: 0.08em;
    cursor: pointer;
    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.18);
}

#theme-toggle:hover,
#theme-toggle:focus-visible {
    background: var(--amber);
    color: var(--bg);
    outline: none;
}

:root[data-theme="light"] .action-btn:hover:not(:disabled),
:root[data-theme="light"] .auth-btn { background: #e5f2e9; }
:root[data-theme="light"] .action-btn.disabled,
:root[data-theme="light"] .action-btn:disabled { border-color: var(--border); }
:root[data-theme="light"] #auth-box {
    border-color: #9b7a3f;
    box-shadow: 0 8px 30px rgba(54, 43, 25, 0.12);
}
:root[data-theme="light"] .auth-title {
    color: #176b42;
    text-shadow: none;
}
:root[data-theme="light"] .auth-btn:hover { background: #d5e9dc; }
:root[data-theme="light"] tr:nth-child(even) td { background: #eee8dc; }
:root[data-theme="light"] .effect-good { background: #e5f2e9; }
:root[data-theme="light"] .effect-bad { background: #f7e6e3; }

/* ── Layout ────────────────────────────────────────────────── */
#left-col {
    position: fixed;
    top: 0; left: 0;
    width: var(--left-width);
    height: calc(100vh - var(--ticker-h));
    background: var(--bg-panel);
    border-right: 1px solid var(--border);
    display: flex;
    flex-direction: column;
    padding: 12px 10px;
    overflow-y: auto;
    z-index: 10;
}

#main {
    margin-left: var(--left-width);
    height: calc(100vh - var(--ticker-h));
    overflow-y: auto;
    overflow-x: hidden;
    scrollbar-gutter: stable;
    display: flex;
    flex-direction: column;
}

body.auth-page #main {
    margin-left: 0;
    height: 100vh;
}

.sr-only {
    position: absolute;
    width: 1px;
    height: 1px;
    padding: 0;
    margin: -1px;
    overflow: hidden;
    clip: rect(0, 0, 0, 0);
    white-space: nowrap;
    border: 0;
}

body.poster-page #main {
    min-height: 100vh;
    height: auto;
    background-color: #05070c;
    align-items: center;
    justify-content: flex-start;
}

.poster-stage {
    position: relative;
    flex: 0 0 auto;
    width: min(100vw, 150vh);
    aspect-ratio: 3 / 2;
    margin: 0 auto;
    background: #05070c;
}

.poster-art {
    display: block;
    width: 100%;
    height: 100%;
    object-fit: contain;
}

body.poster-page #auth-box.poster-auth-card {
    position: absolute;
    margin: 0;
    border-color: rgba(190, 142, 54, 0.72);
    background: rgba(5, 8, 13, 0.92);
    box-shadow: 0 12px 36px rgba(0, 0, 0, 0.5);
    backdrop-filter: blur(3px);
}

body.poster-login #auth-box.poster-auth-card {
    left: 25.5%;
    top: 62.5%;
    width: 49%;
    max-width: none;
    padding: clamp(12px, 1.5vw, 24px) clamp(18px, 2.5vw, 42px);
}

body.poster-register #auth-box.poster-auth-card {
    left: 36%;
    top: 53.5%;
    width: 28%;
    max-width: none;
    padding: clamp(10px, 1.2vw, 20px) clamp(14px, 2vw, 32px);
}

:root[data-theme="light"] body.poster-page #auth-box.poster-auth-card {
    background: rgba(255, 252, 245, 0.96);
    border-color: #9b7a3f;
}

@media (max-width: 760px), (max-height: 620px) {
    body.poster-page #main {
        min-height: 100vh;
        padding-bottom: 28px;
        background: var(--bg);
    }

    .poster-stage {
        width: 100%;
        aspect-ratio: auto;
    }

    .poster-art {
        height: auto;
    }

    body.poster-login #auth-box.poster-auth-card,
    body.poster-register #auth-box.poster-auth-card {
        position: relative;
        inset: auto;
        width: min(100%, 520px);
        max-width: 520px;
        margin: 18px auto 0;
        padding: 24px;
    }
}

/* ── Status Block ──────────────────────────────────────────── */
#status-block {
    border: 1px solid var(--border);
    padding: 8px;
    margin-bottom: 14px;
}

.status-name {
    color: var(--amber);
    font-size: 13px;
    font-weight: bold;
    margin-bottom: 6px;
    text-transform: uppercase;
    letter-spacing: 1px;
}

.status-line {
    display: flex;
    justify-content: space-between;
    margin: 2px 0;
    font-size: 13px;
}

.status-line .label { color: var(--grey); }

.status-warning { color: var(--amber); font-size: 12px; margin-top: 4px; }
.status-combat  { color: var(--red);   font-size: 12px; margin-top: 4px; }

/* ── Action Buttons ────────────────────────────────────────── */
#action-buttons {
    display: flex;
    flex-direction: column;
    gap: 4px;
    margin-bottom: 14px;
}

#action-buttons form { margin: 0; }

.action-btn {
    width: 100%;
    background: var(--bg-input);
    color: var(--green);
    border: 1px solid var(--border);
    padding: 7px 8px;
    text-align: left;
    cursor: pointer;
    font-family: var(--font);
    font-size: 13px;
    letter-spacing: 0.5px;
    transition: background 0.1s, border-color 0.1s;
}

.action-btn:hover:not(:disabled) {
    background: #1a2a1a;
    border-color: var(--green);
}

.action-btn.disabled, .action-btn:disabled {
    color: var(--dim);
    cursor: not-allowed;
    border-color: #1a1a1a;
}

.ap-cost   { font-size: 11px; color: var(--grey); float: right; }
.btn-reason{ display: block; font-size: 10px; color: var(--dim); margin-top: 2px; }

/* ── Left Nav ──────────────────────────────────────────────── */
#left-nav {
    margin-top: auto;
    display: flex;
    flex-direction: column;
    gap: 4px;
    border-top: 1px solid var(--border);
    padding-top: 10px;
}

#left-nav a, .nav-link-btn {
    color: var(--grey);
    font-size: 12px;
    font-family: var(--font);
    background: none;
    border: none;
    cursor: pointer;
    text-align: left;
    padding: 0;
}
#left-nav a:hover, .nav-link-btn:hover { color: var(--white); }

/* ── Terminal Area ─────────────────────────────────────────── */
#terminal {
    flex: 1;
    overflow-y: auto;
    padding: 12px 16px;
    scroll-behavior: smooth;
}

/* Terminal color coding */
.term-line         { margin: 2px 0; line-height: 1.5; word-wrap: break-word; }
.term-combat       { color: var(--white); }
.term-item         { color: var(--blue); }
.term-level_up     { color: var(--amber); font-weight: bold; }
.term-random_event { color: var(--green); }
.term-system       { color: var(--amber); }
.term-error        { color: var(--red); }
.term-good         { color: var(--green); }
.term-bad          { color: var(--red); }
.term-opponent     { color: var(--grey); }

.term-ts { color: var(--dim); font-size: 12px; margin-right: 6px; }

/* Terminal output panel on full pages (shop, blacksmith, etc.) */
#terminal-output {
    background: var(--bg-panel);
    border: 1px solid var(--border);
    padding: 8px 12px;
    min-height: 40px;
    font-size: 13px;
    margin-bottom: 16px;
}

/* ── Bottom Ticker ─────────────────────────────────────────── */
#ticker-wrap {
    position: fixed;
    bottom: 0; left: 0; right: 0;
    height: var(--ticker-h);
    background: var(--bg-panel);
    border-top: 1px solid var(--border);
    overflow: hidden;
    z-index: 20;
    display: flex;
    align-items: center;
}

#ticker {
    white-space: nowrap;
    color: var(--dim);
    font-size: 12px;
    padding-left: 100%;
    animation: ticker-scroll 60s linear infinite;
}

#ticker:hover { animation-play-state: paused; }

@keyframes ticker-scroll {
    0%   { transform: translateX(0); }
    100% { transform: translateX(-100%); }
}

/* ── Auth Pages ────────────────────────────────────────────── */
#auth-box {
    max-width: 520px;
    margin: 60px auto;
    padding: 32px;
    background: var(--bg-panel);
    border: 1px solid #1d6b42;
    box-shadow: 0 0 24px rgba(0, 204, 102, 0.08), inset 0 0 24px rgba(0, 204, 102, 0.025);
}

.action-link {
    display: block;
    text-decoration: none;
}

.action-link:hover { text-decoration: none; }

.auth-banner {
    overflow: hidden;
    margin-bottom: 24px;
    text-align: center;
}

.auth-banner-rule {
    color: var(--dim);
    font-size: 12px;
    line-height: 1;
    white-space: nowrap;
}

.auth-title {
    margin: 12px 0 8px;
    color: #40ee86;
    font-size: clamp(32px, 10vw, 56px);
    font-weight: 700;
    line-height: 1;
    letter-spacing: 0.14em;
    text-indent: 0.14em;
    text-shadow: 0 0 9px rgba(0, 204, 102, 0.45);
}

.auth-subtitle  { color: var(--amber); margin-bottom: 12px; font-size: 12px; letter-spacing: 0.08em; }
.auth-heading   { color: var(--amber); margin-bottom: 16px; font-size: 16px; }
.auth-link      { color: var(--grey); font-size: 12px; margin-top: 16px; }

.form-row {
    display: flex;
    flex-direction: column;
    margin-bottom: 14px;
}

.form-row label { color: var(--grey); font-size: 12px; margin-bottom: 4px; }

.form-row input,
.form-row select {
    background: var(--bg-input);
    border: 1px solid var(--border);
    color: var(--white);
    padding: 7px 10px;
    font-family: var(--font);
    font-size: 13px;
    outline: none;
}

.form-row input:focus,
.form-row select:focus { border-color: var(--green); }

.field-hint { color: var(--dim); font-size: 11px; margin-top: 3px; }

.auth-btn {
    background: #0a2a0a;
    color: var(--green);
    border: 1px solid var(--green);
    padding: 9px 24px;
    font-family: var(--font);
    font-size: 14px;
    letter-spacing: 1px;
    cursor: pointer;
    margin-top: 8px;
    width: 100%;
}
.auth-btn:hover { background: #0f3a0f; }

.term-errors { margin-bottom: 12px; }
.term-error  { color: var(--red); font-size: 13px; margin: 3px 0; }

/* ── Character Creation ────────────────────────────────────── */
.form-section { margin-bottom: 24px; }
.form-section h3 { color: var(--amber); margin-bottom: 10px; font-size: 14px; }

.class-option {
    display: block;
    background: var(--bg-input);
    border: 1px solid var(--border);
    padding: 8px 10px;
    margin-bottom: 6px;
    cursor: pointer;
}
.class-option:hover { border-color: var(--green); }
.class-option.selected {
    border-color: var(--green);
    background: #10281b;
    box-shadow: inset 4px 0 0 var(--green), 0 0 10px rgba(0, 204, 102, 0.18);
}
.class-option input[type=radio] {
    position: absolute;
    opacity: 0;
    pointer-events: none;
}
.class-option:focus-within { outline: 2px solid var(--blue); outline-offset: 2px; }
.class-option.selected .class-name::after {
    content: " [SELECTED]";
    color: var(--green);
    font-size: 11px;
    letter-spacing: 0.06em;
}
:root[data-theme="light"] .class-option.selected { background: #e5f2e9; }
.class-name    { color: var(--amber); font-size: 13px; display: block; }
.class-bonuses { color: var(--green); font-size: 12px; display: block; }
.class-desc    { color: var(--grey);  font-size: 11px; display: block; margin-top: 3px; }

.stat-row {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 8px;
}
.stat-row label { color: var(--grey); font-size: 12px; flex: 1; }
.stat-row input {
    width: 60px;
    background: var(--bg-input);
    border: 1px solid var(--border);
    color: var(--white);
    padding: 4px 8px;
    font-family: var(--font);
    text-align: center;
}

#points-remaining {
    color: var(--amber);
    font-size: 13px;
    margin-top: 8px;
    text-align: right;
}

/* ── Level Up ──────────────────────────────────────────────── */
.stat-choice {
    display: block;
    background: var(--bg-input);
    border: 1px solid var(--border);
    padding: 8px 10px;
    margin-bottom: 6px;
    cursor: pointer;
}
.stat-choice:hover { border-color: var(--green); }
.stat-choice input[type=radio] { display: none; }
.stat-name    { color: var(--amber); display: block; font-size: 13px; }
.stat-current { color: var(--green); display: block; font-size: 12px; }
.stat-detail  { color: var(--grey);  display: block; font-size: 11px; margin-top: 2px; }

/* ── Fragments ─────────────────────────────────────────────── */
.fragment { padding: 6px 0; border-top: 1px solid var(--border); margin-top: 8px; }
.fragment-header { color: var(--amber); margin-bottom: 6px; }

/* ── Full Pages (shop, blacksmith, character, scoreboards) ─── */
#page-content {
    padding: 20px 24px;
    overflow-y: auto;
    height: 100%;
}

.page-title {
    color: var(--amber);
    font-size: 16px;
    margin-bottom: 16px;
    border-bottom: 1px solid var(--border);
    padding-bottom: 8px;
}

.back-link {
    color: var(--grey);
    font-size: 12px;
    display: inline-block;
    margin-bottom: 16px;
}

table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
    margin-bottom: 16px;
}
th {
    background: var(--bg-input);
    color: var(--amber);
    padding: 6px 10px;
    text-align: left;
    border: 1px solid var(--border);
}
td {
    padding: 5px 10px;
    border: 1px solid var(--border);
    color: var(--white);
    vertical-align: middle;
}
tr:nth-child(even) td { background: #0d0d0d; }

.btn-small {
    background: var(--bg-input);
    color: var(--green);
    border: 1px solid var(--border);
    padding: 3px 10px;
    font-family: var(--font);
    font-size: 12px;
    cursor: pointer;
}
.btn-small:hover { border-color: var(--green); }
.btn-small.danger { color: var(--red); }
.btn-small.danger:hover { border-color: var(--red); }

/* Durability bar */
.dur-bar {
    display: inline-block;
    height: 6px;
    background: var(--green);
    transition: width 0.2s;
}
.dur-bar.medium { background: var(--amber); }
.dur-bar.low    { background: var(--red); }

/* ── Scrollbar ─────────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--dim); }


.status-effects { display: flex; flex-wrap: wrap; gap: 3px; margin-top: 4px; }
.effect-tag { font-size: 10px; padding: 1px 5px; border-radius: 2px; }
.effect-good { background: #0a2a0a; color: var(--green); border: 1px solid var(--green); }
.effect-bad { background: #2a0a0a; color: var(--red); border: 1px solid var(--red); }

.random-event-effect { margin-top: 5px; }

.active-effects-panel { margin-bottom: 24px; }
.active-effects-panel h3 { color: var(--amber); margin-bottom: 10px; }
.active-effects-list { display: grid; gap: 7px; }
.active-effect {
    display: flex;
    justify-content: space-between;
    gap: 16px;
    padding: 8px 10px;
}
.active-effect-duration { color: var(--grey); font-size: 11px; white-space: nowrap; }
.active-effects-empty { color: var(--grey); font-size: 12px; }
/* Shared contextual-help badge. Hover with a mouse or focus with a keyboard. */
.context-help { position:relative;display:inline-flex;align-items:center;justify-content:center;
  width:15px;height:15px;margin-left:5px;border:1px solid var(--green);border-radius:50%;
  color:var(--green);font-size:10px;cursor:help;vertical-align:middle; }
.context-help::after { content:attr(data-help);display:none;position:absolute;z-index:10000;
  left:18px;top:-8px;width:280px;padding:8px 10px;background:var(--panel-bg,#111);color:var(--text,#eee);
  border:1px solid var(--green);box-shadow:0 4px 14px #000;text-align:left;white-space:normal;
  font-size:12px;line-height:1.35;font-weight:normal;text-transform:none;letter-spacing:0; }
.context-help:hover::after,.context-help:focus::after { display:block; }

# FILE: templates/base.html
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{% block title %}Movie Multiverse{% endblock %}</title>
    <script>
        try {
            document.documentElement.dataset.theme = localStorage.getItem('movie-multiverse-theme') || 'dark';
        } catch (error) {
            document.documentElement.dataset.theme = 'dark';
        }
    </script>
<link rel="stylesheet" href="{{ url_for('static', filename='style.css') }}">
    <script defer src="{{ url_for('static', filename='help.js') }}"></script>
</head>
<body class="{{ 'authenticated' if player else 'auth-page' }} {% block body_class %}{% endblock %}">

<button id="theme-toggle" type="button" aria-label="Switch color theme" aria-pressed="false">
    LIGHT MODE
</button>

{% if player %}
<!-- ═══════════════════════════════════════════════════════════════ -->
<!-- LEFT COLUMN — status block + action buttons + nav             -->
<!-- ═══════════════════════════════════════════════════════════════ -->
<div id="left-col">

    <div id="status-block">
        <div class="status-name">{{ player.character_name }}</div>
        <div class="status-line">
            <span class="label">LVL</span>
            <span id="status-level">{{ player.level }}</span>
        </div>
        <div class="status-line">
            <span class="label">HP</span>
            <span id="status-hp">{{ player.current_hp }}</span>/<span id="status-maxhp">{{ player.max_hp }}</span>
        </div>
        <div class="status-line">
            <span class="label">AP</span>
            <span id="status-ap">{{ player.current_ap }}</span>/<span id="status-maxap">{{ player.max_ap }}</span>
        </div>
        <div class="status-line">
            <span class="label">CR</span>
            <span id="status-credits">{{ player.credits }}</span>
        </div>
        {% if player.is_overencumbered %}
        <div class="status-warning">⚠ OVER ENCUMBERED</div>
        {% endif %}
        {% if player.is_cursed %}
        <div class="status-warning">☠ CURSED</div>
        {% endif %}
        {% if effect_labels is defined and effect_labels %}
        <div class="status-effects">
            {% for label in effect_labels %}
            <span class="effect-tag {{ 'effect-good' if label.startswith('+') else 'effect-bad' }}">{{ label }}</span>
            {% endfor %}
        </div>
        {% endif %}
        {% if player.in_combat %}
        <div class="status-combat">⚔ IN COMBAT</div>
        {% endif %}
    </div>

    <div id="action-buttons">
        {% if button_states is defined %}
            {% set action_endpoints = {
                'boss': 'actions.action_boss',
                'pvp': 'actions.action_pvp',
                'tavern': 'actions.action_tavern',
                'blacksmith': 'blacksmith.index',
                'shop': 'shop.index'
            } %}
            {% for action, state in button_states.items() %}
                {% if state.enabled %}
                    {% if action in ('blacksmith', 'shop') %}
                    <a class="action-btn action-link" href="{{ url_for(action_endpoints[action]) }}">
                        {{ action|upper }} <span class="ap-cost">({{ state.ap_cost }} AP)</span>
                    </a>
                    {% else %}
                    <form class="terminal-action" action="{{ url_for(action_endpoints[action]) }}" method="POST">
                    <button type="submit" class="action-btn">
                        {{ action|upper }} <span class="ap-cost">({{ state.ap_cost }} AP)</span>
                    </button>
                    </form>
                    {% endif %}
                {% else %}
                <button class="action-btn disabled" title="{{ state.reason }}" disabled>
                    {{ action|upper }} <span class="ap-cost">({{ state.ap_cost }} AP)</span>
                    <span class="btn-reason">{{ state.reason }}</span>
                </button>
                {% endif %}
            {% endfor %}
        {% endif %}
    </div>

    <nav id="left-nav">
        <a href="{{ url_for('character.index') }}">Character</a>
        <a href="{{ url_for('scoreboards.index') }}">Scoreboards</a>
        <form action="{{ url_for('auth.logout') }}" method="POST" style="margin:0">
            <button type="submit" class="nav-link-btn">Logout</button>
        </form>
    </nav>

</div><!-- /left-col -->
{% endif %}

<!-- ═══════════════════════════════════════════════════════════════ -->
<!-- MAIN CONTENT AREA                                             -->
<!-- ═══════════════════════════════════════════════════════════════ -->
<div id="main">
    {% block content %}{% endblock %}
</div>

{% if player %}
<!-- ═══════════════════════════════════════════════════════════════ -->
<!-- BOTTOM TICKER — global feed                                   -->
<!-- ═══════════════════════════════════════════════════════════════ -->
<div id="ticker-wrap">
    <div id="ticker">
        <span id="ticker-content">Loading global feed...</span>
    </div>
</div>
{% endif %}

{% block scripts %}{% endblock %}
<script src="{{ url_for('static', filename='terminal.js') }}"></script>

</body>
</html>

# FILE: templates/admin/base_admin.html
<!-- ============================================================ -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Admin — {% block title %}Dashboard{% endblock %}</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Courier New', monospace; background: #0a0a0a;
               color: #dddddd; font-size: 13px; }
        #admin-wrap { display: flex; min-height: 100vh; }
        #admin-nav  { width: 180px; background: #111; border-right: 1px solid #222;
                      padding: 16px 12px; flex-shrink: 0; }
        #admin-nav h2 { color: #ffaa00; font-size: 13px; margin-bottom: 16px;
                        text-transform: uppercase; letter-spacing: 1px; }
        #admin-nav a  { display: block; color: #666; text-decoration: none;
                        padding: 5px 0; font-size: 12px; }
        #admin-nav a:hover { color: #fff; }
        #admin-main { flex: 1; padding: 24px; overflow-y: auto; }
        h1 { color: #ffaa00; font-size: 16px; margin-bottom: 16px; }
        h2 { color: #4499ff; font-size: 14px; margin: 20px 0 10px; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 16px; }
        th { background: #1a1a1a; color: #ffaa00; padding: 6px 10px;
             text-align: left; border: 1px solid #222; }
        td { padding: 5px 10px; border: 1px solid #1a1a1a; }
        tr:nth-child(even) td { background: #0d0d0d; }
        .btn { background: #1a1a1a; color: #00cc66; border: 1px solid #222;
               padding: 5px 14px; cursor: pointer; font-family: inherit;
               font-size: 12px; text-decoration: none; display: inline-block; }
        .btn:hover { border-color: #00cc66; }
        .btn-danger { color: #cc2222; }
        .btn-danger:hover { border-color: #cc2222; }
        .feedback { color: #00cc66; margin-bottom: 12px; }
        .error    { color: #cc2222; margin-bottom: 12px; }
        input, select, textarea { background: #1a1a1a; border: 1px solid #333;
                                   color: #ddd; padding: 5px 8px; font-family: inherit;
                                   font-size: 12px; }
        input:focus, select:focus { border-color: #4499ff; outline: none; }
        .stat-grid { display: grid; grid-template-columns: repeat(5,1fr); gap: 8px; }
        .stat-box  { background: #111; border: 1px solid #222; padding: 8px;
                     text-align: center; }
        .stat-val  { color: #ffaa00; font-size: 18px; display: block; }
        .stat-lbl  { color: #666; font-size: 10px; }
        .context-help { position:relative;display:inline-flex;align-items:center;justify-content:center;
                        width:15px;height:15px;margin-left:5px;border:1px solid #4499ff;border-radius:50%;
                        color:#4499ff;font-size:10px;cursor:help;vertical-align:middle; }
        .context-help::after { content:attr(data-help);display:none;position:absolute;z-index:10000;
                        left:18px;top:-8px;width:280px;padding:8px 10px;background:#f3f0e8;color:#111;
                        border:1px solid #4499ff;box-shadow:0 4px 14px #000;text-align:left;
                        white-space:normal;font-size:12px;line-height:1.35;font-weight:normal; }
        .context-help:hover::after,.context-help:focus::after { display:block; }
    </style>
</head>
<body>
<div id="admin-wrap">
    <nav id="admin-nav">
        <h2>⚙ Admin</h2>
        <a href="/admin">Dashboard</a>
        <a href="/admin/import">Import Excel</a>
        <a href="/admin/players">Players</a>
        <a href="/admin/npcs">NPCs</a>
        <a href="/admin/items">Items</a>
        <a href="/admin/analytics">Analytics</a>
        <a href="/admin/health">Health & Audit</a>
        <a href="/admin/config">Config</a>
        <a href="/admin/logs">Logs</a>
        <br>
        <form method="POST" action="/admin/reset/midnight"
              onsubmit="return confirm('Trigger midnight reset now?');">
            <button type="submit" class="btn" style="width:100%;margin-top:8px;">
                ↺ Midnight Reset
            </button>
        </form>
    </nav>
    <main id="admin-main">
        {% if request.args.get('feedback') %}
        <div class="feedback">✓ {{ request.args.get('feedback') }}</div>
        {% endif %}
        {% if request.args.get('error') %}
        <div class="error">⚠ {{ request.args.get('error') }}</div>
        {% endif %}
        {% block content %}{% endblock %}
    </main>
</div>
<script src="/static/help.js"></script>
</body>
</html>


<!-- ============================================================ -->


