################################################################################
# PHASE 7 CODE — Importer + Full Midnight Reset
# BBS-Inspired Multiplayer Dueling Game
#
# Files included:
#   1. importer.py   — Full Excel parse/validate/diff/apply implementation
#   2. scheduler.py  — Full midnight_reset (replaces Phase 1 stub)
#                      ap_trickle unchanged from Phase 1
#
# IMPORTANT: scheduler.py REPLACES the Phase 1 version.
# importer.py is a new file — place at project root.
# Requires: pip install openpyxl
################################################################################

################################################################################
# FILE: importer.py (Phase 7 — full implementation)
################################################################################

# importer.py  (Phase 7 — full implementation)
# Reads the staged Excel file, validates it, diffs against current DB content,
# and applies changes atomically at midnight reset.
# Called by scheduler.py (midnight) and admin.py (manual trigger / full reset).

import os
import math
import logging
import shutil
from datetime import datetime

from openpyxl import load_workbook

from database import execute, execute_one, execute_write, exclusive_transaction
import config_defaults as cfg

logger = logging.getLogger(__name__)

REQUIRED_SHEETS = {
    "Master", "Bosses", "Minions", "Weapons",
    "Armor", "SpecialItems", "Classes", "RandomEvents", "Settings"
}

DAMAGE_TYPES = ("Blade", "Blunt", "Ballistic", "Energy", "Arcane", "Explosive", "Venom")

VALID_BUFF_TYPES = {
    "AC_BONUS", "DMG_REDUCTION", "ATTACK_BONUS",
    "CRIT_BONUS", "RESISTANCE_TYPE", "HP_RESTORE"
}

VALID_EFFECT_TYPES = {
    "CREDITS", "ITEM_AT_LEVEL", "BONUS_AP", "DURABILITY_RESTORE_RANDOM",
    "SPECIAL_ITEM_FROM_POOL", "HP_LOSS", "DURABILITY_LOSS_RANDOM",
    "XP_LOSS", "AP_REDUCTION_PERCENT"
}

# Intel-sensitive columns on the bosses table — changing these clears boss_intel
INTEL_SENSITIVE_COLS = {
    "res_blade", "res_blunt", "res_ballistic", "res_energy",
    "res_arcane", "res_explosive", "res_venom",
    "weak_blade", "weak_blunt", "weak_ballistic", "weak_energy",
    "weak_arcane", "weak_explosive", "weak_venom",
    "special_attack_damage_type", "special_buff_damage_type",
}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def run_import(filepath: str = None, full_reset: bool = False) -> dict:
    """Main entry point. Returns {'success': bool, 'errors': list, 'summary': dict}."""
    if filepath is None:
        filepath = cfg.PENDING_IMPORT_PATH

    if not os.path.exists(filepath):
        return {"success": False, "errors": ["No staged import file found."], "summary": {}}

    logger.info("Starting import from %s (full_reset=%s)", filepath, full_reset)

    try:
        raw_data = parse_workbook(filepath)
    except Exception as e:
        _reject_import(filepath, f"Failed to parse workbook: {e}")
        return {"success": False, "errors": [str(e)], "summary": {}}

    errors = validate(raw_data)
    if errors:
        _reject_import(filepath, "\n".join(errors))
        return {"success": False, "errors": errors, "summary": {}}

    try:
        changes = diff_content(raw_data, full_reset)
        with exclusive_transaction():
            summary = apply_changes(changes, full_reset)
            clear_stale_intel(changes)
            auto_populate_associated_to()
        # Move the processed file out of the way
        _archive_import(filepath)
        logger.info("Import complete: %s", summary)
        return {"success": True, "errors": [], "summary": summary}
    except Exception as e:
        logger.exception("Import failed during apply_changes")
        _reject_import(filepath, str(e))
        return {"success": False, "errors": [str(e)], "summary": {}}


# ─────────────────────────────────────────────────────────────────────────────
# PARSE
# ─────────────────────────────────────────────────────────────────────────────

def parse_workbook(filepath: str) -> dict:
    """Read all sheets into a dict of {sheet_name: [row_dict, ...]}."""
    wb  = load_workbook(filepath, read_only=True, data_only=True)
    out = {}
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws      = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        rows    = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            row_dict = {h: v for h, v in zip(headers, row) if h is not None}
            # Skip rows that start with a note marker
            name_val = row_dict.get("Name") or row_dict.get("MovieName") or row_dict.get("Constant")
            if name_val and str(name_val).startswith("-"):
                continue
            if name_val and str(name_val).startswith("NOTES"):
                break
            rows.append(row_dict)
        out[sheet_name] = rows
    wb.close()
    return out


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATE
# ─────────────────────────────────────────────────────────────────────────────

def validate(raw_data: dict) -> list[str]:
    """Validate all-or-nothing. Returns list of error strings (empty = valid)."""
    errors = []

    for sheet in REQUIRED_SHEETS:
        if sheet not in raw_data:
            errors.append(f"Missing required sheet: '{sheet}'")
    if errors:
        return errors

    _validate_classes(raw_data.get("Classes", []), errors)
    _validate_bosses(raw_data.get("Bosses", []), errors)
    _validate_minions(raw_data.get("Minions", []), errors)
    _validate_weapons(raw_data.get("Weapons", []), errors)
    _validate_armor(raw_data.get("Armor", []), errors)
    _validate_special_items(raw_data.get("SpecialItems", []), errors)
    _validate_random_events(raw_data.get("RandomEvents", []), errors)
    _validate_master(raw_data, errors)
    return errors


def _require(row: dict, fields: list, sheet: str, errors: list, row_name: str = ""):
    for f in fields:
        if row.get(f) is None or str(row.get(f, "")).strip() == "":
            errors.append(f"[{sheet}] Row '{row_name}': missing required field '{f}'")


def _validate_classes(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Classes] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name"], "Classes", errors, name)


def _validate_bosses(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Bosses] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP",
                      "SpecialAttack_Name", "SpecialAttack_Die", "SpecialAttack_DamageType",
                      "SpecialBuff_Name", "SpecialBuff_Type", "SpecialBuff_Value"], "Bosses", errors, name)
        level = r.get("Level")
        if level is not None and (int(level) < 1 or int(level) > 15):
            errors.append(f"[Bosses] '{name}': Level must be 1-15, got {level}")
        buff_type = r.get("SpecialBuff_Type", "")
        if buff_type and buff_type not in VALID_BUFF_TYPES:
            errors.append(f"[Bosses] '{name}': Invalid SpecialBuff_Type '{buff_type}'")
        atk_type = r.get("SpecialAttack_DamageType", "")
        if atk_type and atk_type not in DAMAGE_TYPES:
            errors.append(f"[Bosses] '{name}': Invalid SpecialAttack_DamageType '{atk_type}'")
        # Resistance and weakness should not overlap
        for dtype in DAMAGE_TYPES:
            d = dtype.lower()
            if r.get(f"Res_{dtype}") and r.get(f"Weak_{dtype}"):
                errors.append(f"[Bosses] '{name}': {dtype} cannot be both resistant AND weak")


def _validate_minions(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Minions] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP"], "Minions", errors, name)
        level = r.get("Level")
        if level is not None and (int(level) < 1 or int(level) > 15):
            errors.append(f"[Minions] '{name}': Level must be 1-15")


def _validate_weapons(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Weapons] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "Type", "DamageDie", "DamageType", "CreditCost"], "Weapons", errors, name)
        if r.get("Type") not in ("Melee", "Ranged", None):
            errors.append(f"[Weapons] '{name}': Type must be 'Melee' or 'Ranged'")
        if r.get("DamageType") and r["DamageType"] not in DAMAGE_TYPES:
            errors.append(f"[Weapons] '{name}': Invalid DamageType '{r['DamageType']}'")


def _validate_armor(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Armor] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "CreditCost"], "Armor", errors, name)


def _validate_special_items(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[SpecialItems] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "AssociatedTo", "AssociationType", "CreditCost"], "SpecialItems", errors, name)
        if r.get("AssociationType") not in ("Boss", "Minion", None):
            errors.append(f"[SpecialItems] '{name}': AssociationType must be 'Boss' or 'Minion'")


def _validate_random_events(rows: list, errors: list):
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[RandomEvents] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Type", "Rarity", "EffectType", "EffectAmount", "Duration"], "RandomEvents", errors, name)
        if r.get("Type") not in ("Good", "Bad", None):
            errors.append(f"[RandomEvents] '{name}': Type must be 'Good' or 'Bad'")
        if r.get("Rarity") not in ("Common", "Uncommon", "Rare", None):
            errors.append(f"[RandomEvents] '{name}': Rarity must be Common/Uncommon/Rare")
        if r.get("EffectType") and r["EffectType"] not in VALID_EFFECT_TYPES:
            errors.append(f"[RandomEvents] '{name}': Invalid EffectType '{r['EffectType']}'")


def _validate_master(raw_data: dict, errors: list):
    boss_names    = {r.get("Name") for r in raw_data.get("Bosses", [])}
    minion_names  = {r.get("Name") for r in raw_data.get("Minions", [])}
    weapon_names  = {r.get("Name") for r in raw_data.get("Weapons", [])}
    armor_names   = {r.get("Name") for r in raw_data.get("Armor", [])}
    special_names = {r.get("Name") for r in raw_data.get("SpecialItems", [])}

    for r in raw_data.get("Master", []):
        movie = r.get("MovieName", "?")
        _require(r, ["MovieName", "BossName", "BossWeapon", "BossArmor", "BossSpecialItem",
                      "MinionName", "MinionWeapon", "MinionArmor", "MinionSpecialItem"],
                 "Master", errors, movie)
        if r.get("BossName")          and r["BossName"]          not in boss_names:
            errors.append(f"[Master] '{movie}': BossName '{r['BossName']}' not found in Bosses sheet")
        if r.get("MinionName")        and r["MinionName"]        not in minion_names:
            errors.append(f"[Master] '{movie}': MinionName '{r['MinionName']}' not found in Minions sheet")
        if r.get("BossWeapon")        and r["BossWeapon"]        not in weapon_names:
            errors.append(f"[Master] '{movie}': BossWeapon '{r['BossWeapon']}' not found in Weapons sheet")
        if r.get("BossArmor")         and r["BossArmor"]         not in armor_names:
            errors.append(f"[Master] '{movie}': BossArmor '{r['BossArmor']}' not found in Armor sheet")
        if r.get("BossSpecialItem")   and r["BossSpecialItem"]   not in special_names:
            errors.append(f"[Master] '{movie}': BossSpecialItem '{r['BossSpecialItem']}' not found in SpecialItems sheet")
        if r.get("MinionWeapon")      and r["MinionWeapon"]      not in weapon_names:
            errors.append(f"[Master] '{movie}': MinionWeapon '{r['MinionWeapon']}' not found in Weapons sheet")
        if r.get("MinionArmor")       and r["MinionArmor"]       not in armor_names:
            errors.append(f"[Master] '{movie}': MinionArmor '{r['MinionArmor']}' not found in Armor sheet")
        if r.get("MinionSpecialItem") and r["MinionSpecialItem"] not in special_names:
            errors.append(f"[Master] '{movie}': MinionSpecialItem '{r['MinionSpecialItem']}' not found in SpecialItems sheet")


# ─────────────────────────────────────────────────────────────────────────────
# DIFF
# ─────────────────────────────────────────────────────────────────────────────

def diff_content(raw_data: dict, full_reset: bool = False) -> dict:
    """Compare raw_data against current DB. Returns changes dict.
    If full_reset=True, treat everything as INSERT (no existing rows)."""
    changes = {}

    changes["classes"]       = _diff_table(raw_data.get("Classes", []),       "classes",       "Name", _map_class)
    changes["bosses"]        = _diff_table(raw_data.get("Bosses", []),         "bosses",        "Name", _map_boss,  full_reset=full_reset)
    changes["minions"]       = _diff_table(raw_data.get("Minions", []),        "minions",       "Name", _map_minion, full_reset=full_reset)
    changes["weapons"]       = _diff_table(raw_data.get("Weapons", []),        "weapons",       "Name", _map_weapon)
    changes["armor"]         = _diff_table(raw_data.get("Armor", []),          "armor",         "Name", _map_armor)
    changes["special_items"] = _diff_table(raw_data.get("SpecialItems", []),   "special_items", "Name", _map_special_item)
    changes["random_events"] = _diff_table(raw_data.get("RandomEvents", []),   "random_events", "Name", _map_random_event)
    changes["settings"]      = _diff_settings(raw_data.get("Settings", []))
    changes["master_rows"]   = raw_data.get("Master", [])  # always reprocess
    return changes


def _diff_table(rows: list, table: str, name_col_excel: str,
                mapper_fn, full_reset: bool = False) -> dict:
    """Compare Excel rows against DB rows matched by name.
    Returns {insert: [...], update: [...]} — never deletes."""
    existing = {}
    if not full_reset:
        db_rows = execute(f"SELECT * FROM {table}")
        existing = {r["name"]: r for r in db_rows}

    inserts = []
    updates = []
    for row in rows:
        name = row.get(name_col_excel)
        if not name:
            continue
        mapped = mapper_fn(row)
        mapped["name"] = str(name).strip()
        if mapped["name"] in existing:
            updates.append({"db_row": existing[mapped["name"]], "new_data": mapped})
        else:
            inserts.append(mapped)
    return {"insert": inserts, "update": updates}


def _diff_settings(rows: list) -> list:
    """Settings are always upsert by constant_name."""
    result = []
    for r in rows:
        name  = r.get("Constant")
        value = r.get("Value")
        desc  = r.get("Description", "")
        if name and value is not None:
            result.append({"constant_name": str(name), "value": str(value),
                           "description": str(desc) if desc else ""})
    return result


# ─────────────────────────────────────────────────────────────────────────────
# MAPPERS  (Excel row dict → DB column dict)
# ─────────────────────────────────────────────────────────────────────────────

def _b(val) -> int:
    """Convert Excel boolean/string to 0 or 1."""
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, str):
        return 1 if val.upper() in ("TRUE", "YES", "1") else 0
    return 1 if val else 0


def _i(val, default=0) -> int:
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _f(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _s(val, default="") -> str:
    return str(val).strip() if val is not None else default


def _map_class(r: dict) -> dict:
    return {
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")), "description": _s(r.get("Description")),
    }


def _map_boss(r: dict) -> dict:
    return {
        "level": _i(r.get("Level")),
        "str_stat": _i(r.get("STR")), "end_stat": _i(r.get("END")),
        "agi_stat": _i(r.get("AGI")), "lck_stat": _i(r.get("LCK")),
        "per_stat": _i(r.get("PER")), "max_hp": _i(r.get("HP")),
        "phase2_hp_percent": _i(r.get("Phase2_HPPercent"), 50),
        "phase3_hp_percent": _i(r.get("Phase3_HPPercent"), 25),
        "special_attack_name":        _s(r.get("SpecialAttack_Name")),
        "special_attack_die":         _s(r.get("SpecialAttack_Die")),
        "special_attack_damage_type": _s(r.get("SpecialAttack_DamageType")),
        "special_attack_flavor":      _s(r.get("SpecialAttack_Flavor")),
        "special_buff_name":          _s(r.get("SpecialBuff_Name")),
        "special_buff_type":          _s(r.get("SpecialBuff_Type")),
        "special_buff_value":         _f(r.get("SpecialBuff_Value")),
        "special_buff_damage_type":   _s(r.get("SpecialBuff_DamageType")),
        "special_buff_flavor":        _s(r.get("SpecialBuff_Flavor")),
        **{f"res_{d.lower()}":  _b(r.get(f"Res_{d}"))  for d in DAMAGE_TYPES},
        **{f"weak_{d.lower()}": _b(r.get(f"Weak_{d}")) for d in DAMAGE_TYPES},
        "drop_weapon_chance":        _f(r.get("Drop_Weapon_Chance")),
        "drop_armor_chance":         _f(r.get("Drop_Armor_Chance")),
        "drop_special_item_chance":  _f(r.get("Drop_SpecialItem_Chance")),
        "drop_credit_min":           _i(r.get("Drop_Credit_Min")),
        "drop_credit_max":           _i(r.get("Drop_Credit_Max")),
        "flavor_text":               _s(r.get("FlavorText")),
    }


def _map_minion(r: dict) -> dict:
    return {
        "level": _i(r.get("Level")),
        "str_stat": _i(r.get("STR")), "end_stat": _i(r.get("END")),
        "agi_stat": _i(r.get("AGI")), "lck_stat": _i(r.get("LCK")),
        "per_stat": _i(r.get("PER")), "max_hp": _i(r.get("HP")),
        "drop_weapon_chance":       _f(r.get("Drop_Weapon_Chance")),
        "drop_armor_chance":        _f(r.get("Drop_Armor_Chance")),
        "drop_special_item_chance": _f(r.get("Drop_SpecialItem_Chance")),
        "drop_credit_min":          _i(r.get("Drop_Credit_Min")),
        "drop_credit_max":          _i(r.get("Drop_Credit_Max")),
        "flavor_text":              _s(r.get("FlavorText")),
    }


def _map_weapon(r: dict) -> dict:
    return {
        "level":       _i(r.get("Level")),
        "weapon_type": _s(r.get("Type")),
        "damage_die":  _s(r.get("DamageDie")),
        "damage_type": _s(r.get("DamageType")),
        "str_bonus":   _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus":   _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus":   _i(r.get("PER")),
        "associated_to":       _s(r.get("AssociatedTo")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_armor(r: dict) -> dict:
    return {
        "level":    _i(r.get("Level")),
        "ac_bonus": _i(r.get("AC_Bonus")),
        **{f"res_{d.lower()}": _b(r.get(f"Res_{d}")) for d in DAMAGE_TYPES},
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")),
        "associated_to":       _s(r.get("AssociatedTo")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_special_item(r: dict) -> dict:
    return {
        "associated_to":   _s(r.get("AssociatedTo")),
        "association_type": _s(r.get("AssociationType")),
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")),
        "initiative_bonus":    _i(r.get("InitiativeBonus")),
        "extra_attack":        _b(r.get("ExtraAttack")),
        "crit_chance_bonus":   _f(r.get("CritChanceBonus")),
        "crit_dmg_multiplier": _f(r.get("CritDmgMultiplier")),
        "ac_bonus":            _i(r.get("ACBonus")),
        **{f"res_{d.lower()}": _b(r.get(f"Res_{d}")) for d in DAMAGE_TYPES},
        "bonus_damage_type":   _s(r.get("BonusDamageType")),
        "bonus_damage_amount": _i(r.get("BonusDamageAmount")),
        "xp_multiplier":       _f(r.get("XPMultiplier")),
        "credit_multiplier":   _f(r.get("CreditMultiplier")),
        "steal_bonus":         _f(r.get("StealBonus")),
        "bonus_ap":            _i(r.get("BonusAP")),
        "hp_regen_bonus":      _i(r.get("HPRegenBonus")),
        "durability_reduction": _f(r.get("DurabilityReduction")),
        "shop_discount":       _f(r.get("ShopDiscount")),
        "sell_bonus":          _f(r.get("SellBonus")),
        "encounter_bonus":     _f(r.get("EncounterBonus")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_random_event(r: dict) -> dict:
    return {
        "event_type":    _s(r.get("Type")),
        "rarity":        _s(r.get("Rarity")),
        "flavor_text":   _s(r.get("FlavorText")),
        "effect_type":   _s(r.get("EffectType")),
        "effect_amount": _i(r.get("EffectAmount")),
        "duration":      _s(r.get("Duration")),
    }


# ─────────────────────────────────────────────────────────────────────────────
# APPLY CHANGES
# ─────────────────────────────────────────────────────────────────────────────

def apply_changes(changes: dict, full_reset: bool = False) -> dict:
    """Apply all inserts and updates. Must be called inside exclusive_transaction().
    Returns summary dict of counts."""
    summary = {}
    order   = ["classes", "bosses", "minions", "weapons", "armor", "special_items", "random_events"]

    for key in order:
        if key not in changes:
            continue
        tbl    = key
        data   = changes[key]
        inserts = data.get("insert", [])
        updates = data.get("update", [])
        for row in inserts:
            _upsert_row(tbl, row, None)
            # Create special_item_registry row for new special items
            if tbl == "special_items":
                new_id = execute_one("SELECT id FROM special_items WHERE name = ?", (row["name"],))
                if new_id:
                    execute_write(
                        """INSERT OR IGNORE INTO special_item_registry (special_item_id, status)
                           VALUES (?, 'IN_POOL')""",
                        (new_id["id"],)
                    )
        for item in updates:
            _upsert_row(tbl, item["new_data"], item["db_row"]["id"])
        summary[key] = {"insert": len(inserts), "update": len(updates)}

    # Settings: upsert by constant_name
    for s in changes.get("settings", []):
        execute_write(
            """INSERT OR REPLACE INTO settings (constant_name, value, description, imported_at)
               VALUES (?, ?, ?, ?)""",
            (s["constant_name"], s["value"], s["description"], datetime.utcnow().isoformat())
        )
    summary["settings"] = {"upsert": len(changes.get("settings", []))}

    # Master: process after all content tables are up to date
    _apply_master(changes.get("master_rows", []))
    summary["master"] = {"processed": len(changes.get("master_rows", []))}

    return summary


def _upsert_row(table: str, data: dict, existing_id: int | None):
    """Insert or update a content table row."""
    data["imported_at"] = datetime.utcnow().isoformat()
    if existing_id is None:
        cols   = ", ".join(data.keys())
        placeholders = ", ".join("?" * len(data))
        execute_write(
            f"INSERT OR IGNORE INTO {table} ({cols}) VALUES ({placeholders})",
            tuple(data.values())
        )
    else:
        sets = ", ".join(f"{k} = ?" for k in data)
        execute_write(
            f"UPDATE {table} SET {sets} WHERE id = ?",
            tuple(data.values()) + (existing_id,)
        )


def _apply_master(master_rows: list):
    """Process master sheet: upsert master rows, linking by name."""
    for r in master_rows:
        movie = _s(r.get("MovieName"))
        if not movie:
            continue

        def get_id(table, name):
            if not name:
                return None
            row = execute_one(f"SELECT id FROM {table} WHERE name = ?", (_s(name),))
            return row["id"] if row else None

        boss_id          = get_id("bosses",        r.get("BossName"))
        minion_id        = get_id("minions",       r.get("MinionName"))
        boss_weapon_id   = get_id("weapons",       r.get("BossWeapon"))
        boss_armor_id    = get_id("armor",         r.get("BossArmor"))
        boss_special_id  = get_id("special_items", r.get("BossSpecialItem"))
        min_weapon_id    = get_id("weapons",       r.get("MinionWeapon"))
        min_armor_id     = get_id("armor",         r.get("MinionArmor"))
        min_special_id   = get_id("special_items", r.get("MinionSpecialItem"))

        if not all([boss_id, minion_id, boss_weapon_id, boss_armor_id,
                    boss_special_id, min_weapon_id, min_armor_id, min_special_id]):
            logger.warning("Master row '%s': could not resolve all FK references, skipping", movie)
            continue

        existing = execute_one("SELECT id FROM master WHERE movie_name = ?", (movie,))
        if existing:
            execute_write(
                """UPDATE master SET boss_id=?, boss_weapon_id=?, boss_armor_id=?,
                   boss_special_item_id=?, minion_id=?, minion_weapon_id=?,
                   minion_armor_id=?, minion_special_item_id=?, imported_at=?
                   WHERE movie_name=?""",
                (boss_id, boss_weapon_id, boss_armor_id, boss_special_id,
                 minion_id, min_weapon_id, min_armor_id, min_special_id,
                 datetime.utcnow().isoformat(), movie)
            )
        else:
            execute_write(
                """INSERT INTO master
                   (movie_name, boss_id, boss_weapon_id, boss_armor_id, boss_special_item_id,
                    minion_id, minion_weapon_id, minion_armor_id, minion_special_item_id, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (movie, boss_id, boss_weapon_id, boss_armor_id, boss_special_id,
                 minion_id, min_weapon_id, min_armor_id, min_special_id,
                 datetime.utcnow().isoformat())
            )


# ─────────────────────────────────────────────────────────────────────────────
# INTEL CLEARING
# ─────────────────────────────────────────────────────────────────────────────

def clear_stale_intel(changes: dict):
    """Clear boss_intel rows for any boss whose intel-sensitive columns changed."""
    for item in changes.get("bosses", {}).get("update", []):
        old = item["db_row"]
        new = item["new_data"]
        changed = any(
            str(old.get(col, "")) != str(new.get(col, ""))
            for col in INTEL_SENSITIVE_COLS
            if col in new
        )
        if changed:
            boss_id = old["id"]
            deleted = execute_write(
                "DELETE FROM boss_intel WHERE boss_id = ?", (boss_id,)
            )
            if deleted:
                logger.info("Cleared boss_intel for boss_id=%d (intel-sensitive columns changed)", boss_id)


# ─────────────────────────────────────────────────────────────────────────────
# AUTO-POPULATE ASSOCIATED_TO
# ─────────────────────────────────────────────────────────────────────────────

def auto_populate_associated_to():
    """Update weapons/armor associated_to field from master table.
    Format: 'MovieName (Boss)' or 'MovieName (Minion)'."""
    master_rows = execute("SELECT * FROM master")
    for m in master_rows:
        movie = execute_one("SELECT movie_name FROM master WHERE id = ?", (m["id"],))["movie_name"]
        for col, table in [
            ("boss_weapon_id",   "weapons"),
            ("boss_armor_id",    "armor"),
            ("minion_weapon_id", "weapons"),
            ("minion_armor_id",  "armor"),
        ]:
            item_id = m.get(col)
            if not item_id:
                continue
            side = "Boss" if "boss" in col else "Minion"
            execute_write(
                f"UPDATE {table} SET associated_to = ? WHERE id = ?",
                (f"{movie} ({side})", item_id)
            )


# ─────────────────────────────────────────────────────────────────────────────
# FILE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def _reject_import(filepath: str, reason: str):
    """Move the invalid file to the rejected folder and log the error."""
    os.makedirs(cfg.REJECTED_IMPORT_PATH, exist_ok=True)
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(cfg.REJECTED_IMPORT_PATH, f"rejected_{ts}.xlsx")
    try:
        shutil.move(filepath, dest)
    except Exception:
        pass
    os.makedirs(os.path.dirname(cfg.IMPORT_ERROR_LOG), exist_ok=True)
    with open(cfg.IMPORT_ERROR_LOG, "a") as f:
        f.write(f"{datetime.utcnow().isoformat()} | REJECTED | {reason}\n")
    logger.error("Import rejected: %s", reason)


def _archive_import(filepath: str):
    """Move successfully processed file to a timestamped archive."""
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(cfg.REJECTED_IMPORT_PATH, f"../applied_{ts}.xlsx")
    try:
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.move(filepath, dest)
    except Exception:
        pass


################################################################################
# FILE: scheduler.py (Phase 7 — full midnight_reset, replaces Phase 1 stub)
################################################################################

# scheduler.py  (Phase 7 — full implementation)
# Replaces the Phase 1 stub with complete midnight_reset and ap_trickle.

import math
import random
import logging
import os
from datetime import datetime

from database import execute, execute_one, execute_write, exclusive_transaction, get_all_settings
from queue_handler import purge_old_done_rows
import config_defaults as cfg

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# AP TRICKLE  (unchanged from Phase 1 — full from the start)
# ─────────────────────────────────────────────────────────────────────────────

def ap_trickle():
    """Award TRICKLE_AP_AMOUNT to all non-banned players, capped at AP_CARRYOVER_CAP.
    Runs at 03:00, 09:00, 15:00, 21:00 UTC daily."""
    settings = get_all_settings()
    trickle  = settings.get("TRICKLE_AP_AMOUNT", cfg.TRICKLE_AP_AMOUNT)
    cap      = settings.get("AP_CARRYOVER_CAP",  cfg.AP_CARRYOVER_CAP)

    with exclusive_transaction():
        updated = execute_write(
            "UPDATE players SET current_ap = MIN(current_ap + ?, ?) WHERE is_banned = 0",
            (trickle, cap)
        )
    logger.info("ap_trickle: +%d AP to %d players at %s",
                trickle, updated, datetime.utcnow().isoformat())


# ─────────────────────────────────────────────────────────────────────────────
# MIDNIGHT RESET  (full 12-step implementation)
# ─────────────────────────────────────────────────────────────────────────────

def midnight_reset():
    """Full UTC midnight reset sequence."""
    logger.info("=== MIDNIGHT RESET START %s ===", datetime.utcnow().isoformat())

    _step0_clear_status_effects()
    purge_old_done_rows()                # step 1
    _step2_apply_import()                # step 2
    _step3_archive_and_clear_feeds()     # step 3
    _step4_5_award_daily_ap()            # steps 4+5
    _step6_restore_midnight_hp()         # step 6
    _step7_midnight_encounters()         # step 7
    _step8_9_10_shop_rotation()          # steps 8-10
    _step11_pending_feed_entries()       # step 11

    logger.info("=== MIDNIGHT RESET COMPLETE %s ===", datetime.utcnow().isoformat())


# ─────────────────────────────────────────────────────────────────────────────
# STEP 0 — Clear all timed status effects
# ─────────────────────────────────────────────────────────────────────────────

def _step0_clear_status_effects():
    with exclusive_transaction():
        deleted = execute_write("DELETE FROM status_effects")
    logger.info("step 0: cleared %d status effects", deleted)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Apply staged Excel import
# ─────────────────────────────────────────────────────────────────────────────

def _step2_apply_import():
    if not os.path.exists(cfg.PENDING_IMPORT_PATH):
        logger.info("step 2: no pending import")
        return
    logger.info("step 2: applying staged import from %s", cfg.PENDING_IMPORT_PATH)
    from importer import run_import
    result = run_import(cfg.PENDING_IMPORT_PATH)
    if result["success"]:
        logger.info("step 2: import successful — %s", result["summary"])
    else:
        logger.error("step 2: import REJECTED — %s", result["errors"])


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Archive and clear daily feed
# ─────────────────────────────────────────────────────────────────────────────

def _step3_archive_and_clear_feeds():
    settings = get_all_settings()
    if settings.get("LOG_DAILY_ARCHIVE", cfg.LOG_DAILY_ARCHIVE):
        archive_feeds()

    with exclusive_transaction():
        deleted = execute_write("DELETE FROM daily_feed")
    logger.info("step 3: cleared %d daily feed entries", deleted)


def archive_feeds():
    """Export today's daily_feed to a timestamped text file."""
    os.makedirs(cfg.LOG_ARCHIVE_PATH, exist_ok=True)
    date_str  = datetime.utcnow().strftime("%Y_%m_%d")
    filepath  = os.path.join(cfg.LOG_ARCHIVE_PATH, f"game_log_{date_str}.txt")
    rows      = execute(
        "SELECT feed_scope, player_id, flavor_text, event_category, occurred_at "
        "FROM daily_feed ORDER BY occurred_at ASC"
    )
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"=== Daily Feed Archive — {date_str} UTC ===\n\n")
            for row in rows:
                scope = f"[{row['feed_scope']}]"
                pid   = f" player={row['player_id']}" if row["player_id"] else ""
                f.write(f"{row['occurred_at']} {scope}{pid} {row['flavor_text']}\n")
        logger.info("step 3: archived %d feed entries to %s", len(rows), filepath)
    except Exception as e:
        logger.exception("step 3: failed to archive feed: %s", e)


# ─────────────────────────────────────────────────────────────────────────────
# STEPS 4+5 — AP carryover + award daily AP
# ─────────────────────────────────────────────────────────────────────────────

def _step4_5_award_daily_ap():
    """Carryover is implicit (current_ap already holds remaining AP).
    Cap it at AP_CARRYOVER_CAP, then award new daily AP on top."""
    settings = get_all_settings()
    base_ap   = settings.get("BASE_DAILY_AP",      cfg.BASE_DAILY_AP)
    cap       = settings.get("AP_CARRYOVER_CAP",   cfg.AP_CARRYOVER_CAP)
    curse_red = settings.get("CURSE_AP_REDUCTION", cfg.CURSE_AP_REDUCTION)

    players = execute("SELECT id, end_stat FROM players WHERE is_banned = 0")
    cursed_ids = {
        r["player_id"] for r in execute(
            "SELECT player_id FROM status_effects WHERE effect_type = 'CURSED'"
        )
    }

    with exclusive_transaction():
        for p in players:
            daily_ap = base_ap + math.floor(p["end_stat"] / 2)
            if p["id"] in cursed_ids:
                daily_ap = int(daily_ap * (1 - curse_red))
            # Carryover cap first, then add daily AP, then cap again
            execute_write(
                "UPDATE players SET current_ap = MIN(MIN(current_ap, ?) + ?, ?) WHERE id = ?",
                (cap, daily_ap, cap, p["id"])
            )
    logger.info("steps 4+5: awarded daily AP to %d players (base=%d, cap=%d)",
                len(players), base_ap, cap)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Restore midnight HP
# ─────────────────────────────────────────────────────────────────────────────

def _step6_restore_midnight_hp():
    settings  = get_all_settings()
    heal_pct  = settings.get("MIDNIGHT_HEAL_PERCENT", cfg.MIDNIGHT_HEAL_PERCENT)
    players   = execute(
        "SELECT id, current_hp, end_stat, level FROM players WHERE is_banned = 0"
    )
    with exclusive_transaction():
        for p in players:
            max_hp  = 10 + p["end_stat"] + (5 * p["level"])
            missing = max_hp - p["current_hp"]
            if missing > 0:
                restore = max(1, int(missing * heal_pct))
                execute_write(
                    "UPDATE players SET current_hp = MIN(current_hp + ?, ?) WHERE id = ?",
                    (restore, max_hp, p["id"])
                )
    logger.info("step 6: restored midnight HP for %d players", len(players))


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — Midnight random encounters
# ─────────────────────────────────────────────────────────────────────────────

def _step7_midnight_encounters():
    """Run a random event check for every active non-banned player."""
    from routes.actions import check_random_event
    settings = get_all_settings()
    players  = execute(
        """SELECT p.*, 0 as is_overencumbered FROM players p
           WHERE p.is_banned = 0 AND p.in_combat = 0"""
    )
    triggered = 0
    for p in players:
        # Temporarily set max_hp for the helper
        p["max_hp"]  = 10 + p["end_stat"] + (5 * p["level"])
        p["max_ap"]  = settings.get("BASE_DAILY_AP", cfg.BASE_DAILY_AP) + math.floor(p["end_stat"] / 2)
        event = check_random_event(p, settings)
        if event:
            triggered += 1
    logger.info("step 7: midnight encounters triggered for %d/%d players", triggered, len(players))


# ─────────────────────────────────────────────────────────────────────────────
# STEPS 8-10 — Shop rotation
# ─────────────────────────────────────────────────────────────────────────────

def _step8_9_10_shop_rotation():
    settings      = get_all_settings()
    weapons_count = settings.get("SHOP_WEAPONS_COUNT", cfg.SHOP_WEAPONS_COUNT)
    armor_count   = settings.get("SHOP_ARMOR_COUNT",   cfg.SHOP_ARMOR_COUNT)

    with exclusive_transaction():
        # Step 8: Clear daily rotation listings
        execute_write("DELETE FROM shop_listings WHERE listing_source = 'DAILY_ROTATION'")

        # Step 9: Clear unsold special items from shop, return to loot pool
        unsold_specials = execute(
            "SELECT * FROM shop_listings WHERE item_type = 'SPECIAL'"
        )
        for s in unsold_specials:
            execute_write(
                """UPDATE special_item_registry
                   SET status = 'IN_POOL', current_owner_player_id = NULL,
                       inventory_item_id = NULL, shop_listing_price = NULL,
                       last_released_method = 'UNSOLD', updated_at = ?
                   WHERE special_item_id = ?""",
                (datetime.utcnow().isoformat(), s["item_id"])
            )
            execute_write("DELETE FROM shop_listings WHERE id = ?", (s["id"],))
            logger.info("step 9: returned special item id=%d to pool (unsold)", s["item_id"])

        # Populate new daily rotation — random selection weighted by drop_chance
        _populate_shop_rotation("weapons", weapons_count)
        _populate_shop_rotation("armor",   armor_count)

        # Step 10: Populate special item shop slots = floor(player_count / 2)
        player_count  = execute_one("SELECT COUNT(*) as cnt FROM players WHERE is_banned = 0")["cnt"]
        special_slots = max(0, player_count // 2)
        if special_slots > 0:
            _populate_special_slots(special_slots)

    logger.info("steps 8-10: shop rotated (%d weapons, %d armor, %d special slots)",
                weapons_count, armor_count, special_slots if player_count else 0)


def _populate_shop_rotation(table: str, count: int):
    """Select 'count' unique items from the content table and list them in the shop."""
    items = execute(
        f"SELECT * FROM {table} WHERE is_active = 1 ORDER BY RANDOM() LIMIT ?", (count * 3,)
    )
    # Weight by drop_chance
    weighted = []
    for item in items:
        w = max(1, int(item.get("drop_chance", 0.1) * 100))
        weighted.extend([item] * w)
    random.shuffle(weighted)
    seen   = set()
    chosen = []
    for item in weighted:
        if item["id"] not in seen:
            seen.add(item["id"])
            chosen.append(item)
        if len(chosen) >= count:
            break

    item_type = "WEAPON" if table == "weapons" else "ARMOR"
    for item in chosen:
        execute_write(
            """INSERT INTO shop_listings
               (item_type, item_id, listing_source, price)
               VALUES (?, ?, 'DAILY_ROTATION', ?)""",
            (item_type, item["id"], item["credit_cost"])
        )


def _populate_special_slots(slots: int):
    """Add up to 'slots' IN_POOL special items to the shop."""
    available = execute(
        """SELECT si.id, si.credit_cost
           FROM special_items si
           JOIN special_item_registry sir ON sir.special_item_id = si.id
           WHERE sir.status = 'IN_POOL' AND si.is_active = 1
           ORDER BY RANDOM()
           LIMIT ?""",
        (slots,)
    )
    for item in available:
        # Price special items significantly higher than their base cost
        price = int(item["credit_cost"] * 2.5)
        execute_write(
            """INSERT INTO shop_listings
               (item_type, item_id, listing_source, price)
               VALUES ('SPECIAL', ?, 'DAILY_ROTATION', ?)""",
            (item["id"], price)
        )
        execute_write(
            """UPDATE special_item_registry
               SET status = 'IN_SHOP', shop_listing_price = ?, updated_at = ?
               WHERE special_item_id = ?""",
            (price, datetime.utcnow().isoformat(), item["id"])
        )


# ─────────────────────────────────────────────────────────────────────────────
# STEP 11 — Process any pending feed entries
# ─────────────────────────────────────────────────────────────────────────────

def _step11_pending_feed_entries():
    """No deferred feed entries in current design — placeholder for future use."""
    logger.info("step 11: no pending feed entries")


