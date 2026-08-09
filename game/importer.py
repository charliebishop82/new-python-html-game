"""Validate and atomically import the game-content Excel workbook."""
# Validated workbook ingestion for the complete current content schema.
# Reads the staged Excel file, validates it, diffs against current DB content,
# and applies changes atomically at midnight reset.
# Called by scheduler.py (midnight) and admin.py (manual trigger / full reset).

import os
import math
import logging
import shutil
from datetime import datetime

from openpyxl import load_workbook

from database import execute, execute_one, execute_write, exclusive_transaction
import config_defaults as cfg

logger = logging.getLogger(__name__)

REQUIRED_SHEETS = {
    "Master", "Bosses", "Minions", "Weapons",
    "Armor", "SpecialItems", "Classes", "RandomEvents", "Settings"
    , "WorldBosses", "Perks", "Contracts"
}

DAMAGE_TYPES = ("Blade", "Blunt", "Ballistic", "Energy", "Arcane", "Explosive", "Venom")

VALID_BUFF_TYPES = {
    "AC_BONUS", "DMG_REDUCTION", "ATTACK_BONUS",
    "CRIT_BONUS", "RESISTANCE_TYPE", "HP_RESTORE"
}

VALID_EFFECT_TYPES = {
    "CREDITS", "ITEM_AT_LEVEL", "BONUS_AP", "DURABILITY_RESTORE_RANDOM",
    "SPECIAL_ITEM_FROM_POOL", "HP_LOSS", "DURABILITY_LOSS_RANDOM",
    "XP_LOSS", "AP_REDUCTION_PERCENT", "PROTAGONIST_ENCOUNTER",
    "STAT_BOOST_STR", "STAT_BOOST_END", "STAT_BOOST_AGI",
    "STAT_BOOST_LCK", "STAT_BOOST_PER", "STAT_BOOST_INITIATIVE",
    "STAT_PENALTY_STR", "STAT_PENALTY_END", "STAT_PENALTY_AGI",
    "STAT_PENALTY_LCK", "STAT_PENALTY_PER", "STAT_PENALTY_INITIATIVE",
}

# Intel-sensitive columns on the bosses table — changing these clears boss_intel
INTEL_SENSITIVE_COLS = {
    "res_blade", "res_blunt", "res_ballistic", "res_energy",
    "res_arcane", "res_explosive", "res_venom",
    "weak_blade", "weak_blunt", "weak_ballistic", "weak_energy",
    "weak_arcane", "weak_explosive", "weak_venom",
    "special_attack_damage_type", "special_buff_damage_type",
}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def run_import(filepath: str = None, full_reset: bool = False) -> dict:
    """Main entry point. Returns {'success': bool, 'errors': list, 'summary': dict}."""
    if filepath is None:
        filepath = cfg.PENDING_IMPORT_PATH

    if not os.path.exists(filepath):
        return {"success": False, "errors": ["No staged import file found."], "summary": {}}

    logger.info("Starting import from %s (full_reset=%s)", filepath, full_reset)

    try:
        raw_data = parse_workbook(filepath)
    except Exception as e:
        _reject_import(filepath, f"Failed to parse workbook: {e}")
        return {"success": False, "errors": [str(e)], "summary": {}}

    errors = validate(raw_data)
    if errors:
        _reject_import(filepath, "\n".join(errors))
        return {"success": False, "errors": errors, "summary": {}}

    try:
        changes = diff_content(raw_data, full_reset)
        with exclusive_transaction():
            summary = apply_changes(changes, full_reset)
            clear_stale_intel(changes)
            auto_populate_associated_to()
        # Move the processed file out of the way
        _archive_import(filepath)
        logger.info("Import complete: %s", summary)
        return {"success": True, "errors": [], "summary": summary}
    except Exception as e:
        logger.exception("Import failed during apply_changes")
        _reject_import(filepath, str(e))
        return {"success": False, "errors": [str(e)], "summary": {}}


# ─────────────────────────────────────────────────────────────────────────────
# PARSE
# ─────────────────────────────────────────────────────────────────────────────

def parse_workbook(filepath: str) -> dict:
    """Read all sheets into a dict of {sheet_name: [row_dict, ...]}."""
    wb  = load_workbook(filepath, read_only=True, data_only=True)
    out = {}
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws      = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        rows    = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            row_dict = {h: v for h, v in zip(headers, row) if h is not None}
            # Skip rows that start with a note marker
            name_val = row_dict.get("Name") or row_dict.get("MovieName") or row_dict.get("Constant")
            if name_val and str(name_val).startswith("-"):
                continue
            if name_val and str(name_val).startswith("NOTES"):
                break
            rows.append(row_dict)
        out[sheet_name] = rows
    wb.close()
    return out


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATE
# ─────────────────────────────────────────────────────────────────────────────

def validate(raw_data: dict) -> list[str]:
    """Validate all-or-nothing. Returns list of error strings (empty = valid)."""
    errors = []

    for sheet in REQUIRED_SHEETS:
        if sheet not in raw_data:
            errors.append(f"Missing required sheet: '{sheet}'")
    if errors:
        return errors

    _validate_classes(raw_data.get("Classes", []), errors)
    _validate_bosses(raw_data.get("Bosses", []), errors)
    _validate_world_bosses(raw_data.get("WorldBosses", []), errors)
    _validate_minions(raw_data.get("Minions", []), errors)
    _validate_weapons(raw_data.get("Weapons", []), errors)
    _validate_armor(raw_data.get("Armor", []), errors)
    _validate_special_items(raw_data.get("SpecialItems", []), errors)
    _validate_perks(raw_data.get("Perks", []), errors)
    _validate_random_events(raw_data.get("RandomEvents", []), errors)
    _validate_master(raw_data, errors)
    return errors


def _require(row: dict, fields: list, sheet: str, errors: list, row_name: str = ""):
    """Provide the internal require operation used by this module."""
    for f in fields:
        if row.get(f) is None or str(row.get(f, "")).strip() == "":
            errors.append(f"[{sheet}] Row '{row_name}': missing required field '{f}'")


def _validate_classes(rows: list, errors: list):
    """Validate classes worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Classes] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name"], "Classes", errors, name)


def _validate_bosses(rows: list, errors: list):
    """Validate bosses worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Bosses] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP",
                      "SpecialAttack_Name", "SpecialAttack_Die", "SpecialAttack_DamageType",
                      "SpecialBuff_Name", "SpecialBuff_Type", "SpecialBuff_Value"], "Bosses", errors, name)
        level = r.get("Level")
        if level is not None and (int(level) < 1 or int(level) > 15):
            errors.append(f"[Bosses] '{name}': Level must be 1-15, got {level}")
        buff_type = r.get("SpecialBuff_Type", "")
        if buff_type and buff_type not in VALID_BUFF_TYPES:
            errors.append(f"[Bosses] '{name}': Invalid SpecialBuff_Type '{buff_type}'")
        atk_type = r.get("SpecialAttack_DamageType", "")
        if atk_type and atk_type not in DAMAGE_TYPES:
            errors.append(f"[Bosses] '{name}': Invalid SpecialAttack_DamageType '{atk_type}'")
        # Resistance and weakness should not overlap
        for dtype in DAMAGE_TYPES:
            d = dtype.lower()
            if r.get(f"Res_{dtype}") and r.get(f"Weak_{dtype}"):
                errors.append(f"[Bosses] '{name}': {dtype} cannot be both resistant AND weak")


def _validate_world_bosses(rows: list, errors: list):
    """Validate shared world-boss definitions without the normal level-15 cap."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[WorldBosses] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP",
                     "SpecialAttack_Name", "SpecialAttack_Die", "SpecialAttack_DamageType",
                     "SpecialBuff_Name", "SpecialBuff_Type", "SpecialBuff_Value"],
                 "WorldBosses", errors, name)
        if r.get("Level") is not None and int(r["Level"]) < 1:
            errors.append(f"[WorldBosses] '{name}': Level must be positive")
        if r.get("SpecialBuff_Type") not in VALID_BUFF_TYPES:
            errors.append(f"[WorldBosses] '{name}': invalid SpecialBuff_Type")
        if r.get("SpecialAttack_DamageType") not in DAMAGE_TYPES:
            errors.append(f"[WorldBosses] '{name}': invalid SpecialAttack_DamageType")
        for dtype in DAMAGE_TYPES:
            if r.get(f"Res_{dtype}") and r.get(f"Weak_{dtype}"):
                errors.append(f"[WorldBosses] '{name}': {dtype} cannot be resistant and weak")


def _validate_minions(rows: list, errors: list):
    """Validate minions worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Minions] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "STR", "END", "AGI", "LCK", "PER", "HP"], "Minions", errors, name)
        level = r.get("Level")
        if level is not None and (int(level) < 1 or int(level) > 15):
            errors.append(f"[Minions] '{name}': Level must be 1-15")


def _validate_weapons(rows: list, errors: list):
    """Validate weapons worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Weapons] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "Type", "DamageDie", "DamageType", "CreditCost"], "Weapons", errors, name)
        if r.get("Type") not in ("Melee", "Ranged", None):
            errors.append(f"[Weapons] '{name}': Type must be 'Melee' or 'Ranged'")
        if r.get("DamageType") and r["DamageType"] not in DAMAGE_TYPES:
            errors.append(f"[Weapons] '{name}': Invalid DamageType '{r['DamageType']}'")


def _validate_armor(rows: list, errors: list):
    """Validate armor worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Armor] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "CreditCost"], "Armor", errors, name)


def _validate_special_items(rows: list, errors: list):
    """Validate special items worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[SpecialItems] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "AssociatedTo", "AssociationType", "CreditCost"], "SpecialItems", errors, name)
        if r.get("AssociationType") not in ("Boss", "Minion", "Protagonist", "WorldBoss", None):
            errors.append(f"[SpecialItems] '{name}': invalid AssociationType")


def _validate_perks(rows: list, errors: list):
    """Validate permanent perks and their shared bonus vocabulary."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[Perks] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Level", "Description"], "Perks", errors, name)
        if r.get("Level") is not None and int(r["Level"]) < 1:
            errors.append(f"[Perks] '{name}': Level must be positive")
        dtype = r.get("BonusDamageType")
        if dtype and dtype not in DAMAGE_TYPES:
            errors.append(f"[Perks] '{name}': invalid BonusDamageType '{dtype}'")


def _validate_random_events(rows: list, errors: list):
    """Validate random events worksheet rows and report actionable errors."""
    names = set()
    for r in rows:
        name = r.get("Name", "")
        if name in names:
            errors.append(f"[RandomEvents] Duplicate name: '{name}'")
        names.add(name)
        _require(r, ["Name", "Type", "Rarity", "EffectType", "EffectAmount", "Duration"], "RandomEvents", errors, name)
        if r.get("Type") not in ("Good", "Bad", None):
            errors.append(f"[RandomEvents] '{name}': Type must be 'Good' or 'Bad'")
        if r.get("Rarity") not in ("Common", "Uncommon", "Rare", None):
            errors.append(f"[RandomEvents] '{name}': Rarity must be Common/Uncommon/Rare")
        if r.get("EffectType") and r["EffectType"] not in VALID_EFFECT_TYPES:
            errors.append(f"[RandomEvents] '{name}': Invalid EffectType '{r['EffectType']}'")


def _validate_master(raw_data: dict, errors: list):
    """Validate master worksheet rows and report actionable errors."""
    boss_names    = {r.get("Name") for r in raw_data.get("Bosses", [])}
    minion_names  = {r.get("Name") for r in raw_data.get("Minions", [])}
    weapon_names  = {r.get("Name") for r in raw_data.get("Weapons", [])}
    armor_names   = {r.get("Name") for r in raw_data.get("Armor", [])}
    special_names = {r.get("Name") for r in raw_data.get("SpecialItems", [])}

    for r in raw_data.get("Master", []):
        movie = r.get("MovieName", "?")
        _require(r, ["MovieName", "BossName", "BossWeapon", "BossArmor", "BossSpecialItem",
                      "MinionName", "MinionWeapon", "MinionArmor", "MinionSpecialItem"],
                 "Master", errors, movie)
        if r.get("BossName")          and r["BossName"]          not in boss_names:
            errors.append(f"[Master] '{movie}': BossName '{r['BossName']}' not found in Bosses sheet")
        if r.get("MinionName")        and r["MinionName"]        not in minion_names:
            errors.append(f"[Master] '{movie}': MinionName '{r['MinionName']}' not found in Minions sheet")
        if r.get("BossWeapon")        and r["BossWeapon"]        not in weapon_names:
            errors.append(f"[Master] '{movie}': BossWeapon '{r['BossWeapon']}' not found in Weapons sheet")
        if r.get("BossArmor")         and r["BossArmor"]         not in armor_names:
            errors.append(f"[Master] '{movie}': BossArmor '{r['BossArmor']}' not found in Armor sheet")
        if r.get("BossSpecialItem")   and r["BossSpecialItem"]   not in special_names:
            errors.append(f"[Master] '{movie}': BossSpecialItem '{r['BossSpecialItem']}' not found in SpecialItems sheet")
        if r.get("MinionWeapon")      and r["MinionWeapon"]      not in weapon_names:
            errors.append(f"[Master] '{movie}': MinionWeapon '{r['MinionWeapon']}' not found in Weapons sheet")
        if r.get("MinionArmor")       and r["MinionArmor"]       not in armor_names:
            errors.append(f"[Master] '{movie}': MinionArmor '{r['MinionArmor']}' not found in Armor sheet")
        if r.get("MinionSpecialItem") and r["MinionSpecialItem"] not in special_names:
            errors.append(f"[Master] '{movie}': MinionSpecialItem '{r['MinionSpecialItem']}' not found in SpecialItems sheet")
        for field, pool in [
            ("ProtagonistWeapon", weapon_names),
            ("ProtagonistArmor", armor_names),
            ("ProtagonistSpecialItem", special_names),
        ]:
            value = r.get(field)
            if value and value not in pool:
                errors.append(f"[Master] '{movie}': {field} '{value}' not found")


# ─────────────────────────────────────────────────────────────────────────────
# DIFF
# ─────────────────────────────────────────────────────────────────────────────

def diff_content(raw_data: dict, full_reset: bool = False) -> dict:
    """Compare raw_data against current DB. Returns changes dict.
    If full_reset=True, treat everything as INSERT (no existing rows)."""
    changes = {}

    changes["classes"]       = _diff_table(raw_data.get("Classes", []),       "classes",       "Name", _map_class)
    changes["bosses"]        = _diff_table(raw_data.get("Bosses", []),         "bosses",        "Name", _map_boss,  full_reset=full_reset)
    changes["world_bosses"]  = _diff_table(raw_data.get("WorldBosses", []),    "world_bosses",  "Name", _map_boss, full_reset=full_reset)
    changes["minions"]       = _diff_table(raw_data.get("Minions", []),        "minions",       "Name", _map_minion, full_reset=full_reset)
    changes["weapons"]       = _diff_table(raw_data.get("Weapons", []),        "weapons",       "Name", _map_weapon)
    changes["armor"]         = _diff_table(raw_data.get("Armor", []),          "armor",         "Name", _map_armor)
    changes["special_items"] = _diff_table(raw_data.get("SpecialItems", []),   "special_items", "Name", _map_special_item)
    changes["perks"]         = _diff_table(raw_data.get("Perks", []),          "perks",         "Name", _map_perk)
    changes["random_events"] = _diff_table(raw_data.get("RandomEvents", []),   "random_events", "Name", _map_random_event)
    changes["contracts"]     = _diff_table(raw_data.get("Contracts", []),      "contracts",     "Name", _map_contract)
    changes["settings"]      = _diff_settings(raw_data.get("Settings", []))
    changes["master_rows"]   = raw_data.get("Master", [])  # always reprocess
    changes["world_boss_rows"] = raw_data.get("WorldBosses", [])
    return changes


def _diff_table(rows: list, table: str, name_col_excel: str,
                mapper_fn, full_reset: bool = False) -> dict:
    """Compare Excel rows against DB rows matched by name.
    Returns {insert: [...], update: [...]} — never deletes."""
    existing = {}
    if not full_reset:
        db_rows = execute(f"SELECT * FROM {table}")
        existing = {r["name"]: r for r in db_rows}

    inserts = []
    updates = []
    incoming_names = set()
    for row in rows:
        name = row.get(name_col_excel)
        if not name:
            continue
        mapped = mapper_fn(row)
        mapped["name"] = str(name).strip()
        mapped["is_active"] = 1
        incoming_names.add(mapped["name"])
        if mapped["name"] in existing:
            updates.append({"db_row": existing[mapped["name"]], "new_data": mapped})
        else:
            inserts.append(mapped)
    deactivate = [r["id"] for name, r in existing.items()
                  if name not in incoming_names and r.get("is_active", 1)]
    return {"insert": inserts, "update": updates, "deactivate": deactivate}


def _diff_settings(rows: list) -> list:
    """Settings are always upsert by constant_name."""
    result = []
    for r in rows:
        name  = r.get("Constant")
        value = r.get("Value")
        desc  = r.get("Description", "")
        if name and value is not None:
            result.append({"constant_name": str(name), "value": str(value),
                           "description": str(desc) if desc else ""})
    return result


# ─────────────────────────────────────────────────────────────────────────────
# MAPPERS  (Excel row dict → DB column dict)
# ─────────────────────────────────────────────────────────────────────────────

def _b(val) -> int:
    """Convert Excel boolean/string to 0 or 1."""
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, str):
        return 1 if val.upper() in ("TRUE", "YES", "1") else 0
    return 1 if val else 0


def _i(val, default=0) -> int:
    """Provide the internal i operation used by this module."""
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _f(val, default=0.0) -> float:
    """Provide the internal f operation used by this module."""
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _s(val, default="") -> str:
    """Provide the internal s operation used by this module."""
    return str(val).strip() if val is not None else default


def _map_class(r: dict) -> dict:
    """Map a normalized class worksheet row to database fields."""
    return {
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")), "description": _s(r.get("Description")),
    }


def _map_boss(r: dict) -> dict:
    """Map a normalized boss worksheet row to database fields."""
    return {
        "level": _i(r.get("Level")),
        "str_stat": _i(r.get("STR")), "end_stat": _i(r.get("END")),
        "agi_stat": _i(r.get("AGI")), "lck_stat": _i(r.get("LCK")),
        "per_stat": _i(r.get("PER")), "max_hp": _i(r.get("HP")),
        "phase2_hp_percent": _i(r.get("Phase2_HPPercent"), 50),
        "phase3_hp_percent": _i(r.get("Phase3_HPPercent"), 25),
        "special_attack_name":        _s(r.get("SpecialAttack_Name")),
        "special_attack_die":         _s(r.get("SpecialAttack_Die")),
        "special_attack_damage_type": _s(r.get("SpecialAttack_DamageType")),
        "special_attack_flavor":      _s(r.get("SpecialAttack_Flavor")),
        "special_buff_name":          _s(r.get("SpecialBuff_Name")),
        "special_buff_type":          _s(r.get("SpecialBuff_Type")),
        "special_buff_value":         _f(r.get("SpecialBuff_Value")),
        "special_buff_damage_type":   _s(r.get("SpecialBuff_DamageType")),
        "special_buff_flavor":        _s(r.get("SpecialBuff_Flavor")),
        **{f"res_{d.lower()}":  _b(r.get(f"Res_{d}"))  for d in DAMAGE_TYPES},
        **{f"weak_{d.lower()}": _b(r.get(f"Weak_{d}")) for d in DAMAGE_TYPES},
        "drop_weapon_chance":        _f(r.get("Drop_Weapon_Chance")),
        "drop_armor_chance":         _f(r.get("Drop_Armor_Chance")),
        "drop_special_item_chance":  _f(r.get("Drop_SpecialItem_Chance")),
        "drop_credit_min":           _i(r.get("Drop_Credit_Min")),
        "drop_credit_max":           _i(r.get("Drop_Credit_Max")),
        "flavor_text":               _s(r.get("FlavorText")),
        "description":               _s(r.get("Description")),
    }


def _map_minion(r: dict) -> dict:
    """Map a normalized minion worksheet row to database fields."""
    return {
        "level": _i(r.get("Level")),
        "str_stat": _i(r.get("STR")), "end_stat": _i(r.get("END")),
        "agi_stat": _i(r.get("AGI")), "lck_stat": _i(r.get("LCK")),
        "per_stat": _i(r.get("PER")), "max_hp": _i(r.get("HP")),
        "drop_weapon_chance":       _f(r.get("Drop_Weapon_Chance")),
        "drop_armor_chance":        _f(r.get("Drop_Armor_Chance")),
        "drop_special_item_chance": _f(r.get("Drop_SpecialItem_Chance")),
        "drop_credit_min":          _i(r.get("Drop_Credit_Min")),
        "drop_credit_max":          _i(r.get("Drop_Credit_Max")),
        "flavor_text":              _s(r.get("FlavorText")),
        "description":              _s(r.get("Description")),
    }


def _map_weapon(r: dict) -> dict:
    """Map a normalized weapon worksheet row to database fields."""
    return {
        "level":       _i(r.get("Level")),
        "weapon_type": _s(r.get("Type")),
        "damage_die":  _s(r.get("DamageDie")),
        "damage_type": _s(r.get("DamageType")),
        "str_bonus":   _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus":   _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus":   _i(r.get("PER")),
        "associated_to":       _s(r.get("AssociatedTo")),
        "description":         _s(r.get("Description")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_armor(r: dict) -> dict:
    """Map a normalized armor worksheet row to database fields."""
    return {
        "level":    _i(r.get("Level")),
        "ac_bonus": _i(r.get("AC_Bonus")),
        **{f"res_{d.lower()}": _b(r.get(f"Res_{d}")) for d in DAMAGE_TYPES},
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")),
        "associated_to":       _s(r.get("AssociatedTo")),
        "description":         _s(r.get("Description")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_special_item(r: dict) -> dict:
    """Map a normalized special item worksheet row to database fields."""
    return {
        "associated_to":   _s(r.get("AssociatedTo")),
        "association_type": _s(r.get("AssociationType")),
        "description":      _s(r.get("Description")),
        "str_bonus": _i(r.get("STR")), "end_bonus": _i(r.get("END")),
        "agi_bonus": _i(r.get("AGI")), "lck_bonus": _i(r.get("LCK")),
        "per_bonus": _i(r.get("PER")),
        "initiative_bonus":    _i(r.get("InitiativeBonus")),
        "extra_attack":        _b(r.get("ExtraAttack")),
        "crit_chance_bonus":   _f(r.get("CritChanceBonus")),
        "crit_dmg_multiplier": _f(r.get("CritDmgMultiplier")),
        "ac_bonus":            _i(r.get("ACBonus")),
        **{f"res_{d.lower()}": _b(r.get(f"Res_{d}")) for d in DAMAGE_TYPES},
        "bonus_damage_type":   _s(r.get("BonusDamageType")),
        "bonus_damage_amount": _i(r.get("BonusDamageAmount")),
        "xp_multiplier":       _f(r.get("XPMultiplier")),
        "credit_multiplier":   _f(r.get("CreditMultiplier")),
        "steal_bonus":         _f(r.get("StealBonus")),
        "bonus_ap":            _i(r.get("BonusAP")),
        "hp_regen_bonus":      _i(r.get("HPRegenBonus")),
        "durability_reduction": _f(r.get("DurabilityReduction")),
        "shop_discount":       _f(r.get("ShopDiscount")),
        "sell_bonus":          _f(r.get("SellBonus")),
        "encounter_bonus":     _f(r.get("EncounterBonus")),
        "credit_cost":         _i(r.get("CreditCost")),
        "drop_chance":         _f(r.get("DropChance")),
        "starting_durability": _i(r.get("StartingDurability"), 100),
    }


def _map_perk(r: dict) -> dict:
    """Map a perk row using the same modifier fields as special items."""
    mapped = _map_special_item(r)
    for field in ("associated_to", "association_type", "credit_cost",
                  "drop_chance", "starting_durability"):
        mapped.pop(field, None)
    mapped["level"] = _i(r.get("Level"))
    return mapped


def _map_random_event(r: dict) -> dict:
    """Map a normalized random event worksheet row to database fields."""
    return {
        "event_type":    _s(r.get("Type")),
        "rarity":        _s(r.get("Rarity")),
        "flavor_text":   _s(r.get("FlavorText")),
        "effect_type":   _s(r.get("EffectType")),
        "effect_amount": _i(r.get("EffectAmount")),
        "duration":      _s(r.get("Duration")),
    }


def _map_contract(r: dict) -> dict:
    """Map an Excel-authored daily objective and its substantial completion reward."""
    return {
        "description": _s(r.get("Description")), "metric": _s(r.get("Metric")).upper(),
        "target": max(1, _i(r.get("Target"), 1)), "reward_xp": max(0, _i(r.get("RewardXP"))),
        "reward_credits": max(0, _i(r.get("RewardCredits"))),
        "reward_ap": max(0, _i(r.get("RewardAP"))), "min_level": max(1, _i(r.get("MinLevel"), 1)),
    }


# ─────────────────────────────────────────────────────────────────────────────
# APPLY CHANGES
# ─────────────────────────────────────────────────────────────────────────────

def apply_changes(changes: dict, full_reset: bool = False) -> dict:
    """Apply all inserts and updates. Must be called inside exclusive_transaction().
    Returns summary dict of counts."""
    summary = {}
    order   = ["classes", "bosses", "world_bosses", "minions", "weapons", "armor", "special_items", "perks", "random_events", "contracts"]

    for key in order:
        if key not in changes:
            continue
        tbl    = key
        data   = changes[key]
        inserts = data.get("insert", [])
        updates = data.get("update", [])
        deactivate = data.get("deactivate", [])
        for row_id in deactivate:
            execute_write(f"UPDATE {tbl} SET is_active=0 WHERE id=?", (row_id,))
            listing_type = {
                "weapons": "WEAPON",
                "armor": "ARMOR",
                "special_items": "SPECIAL",
            }.get(tbl)
            if listing_type:
                # Retired definitions remain available to existing owners and logs,
                # but must no longer be offered as active shop inventory.
                execute_write(
                    "DELETE FROM shop_listings WHERE item_type=? AND item_id=?",
                    (listing_type, row_id),
                )
        for row in inserts:
            _upsert_row(tbl, row, None)
            # Create special_item_registry row for new special items
            if tbl == "special_items":
                new_id = execute_one("SELECT id FROM special_items WHERE name = ?", (row["name"],))
                if new_id and row.get("association_type") != "WorldBoss":
                    execute_write(
                        """INSERT OR IGNORE INTO special_item_registry (special_item_id, status)
                           VALUES (?, 'IN_POOL')""",
                        (new_id["id"],)
                    )
        for item in updates:
            _upsert_row(tbl, item["new_data"], item["db_row"]["id"])
        summary[key] = {"insert": len(inserts), "update": len(updates),
                        "deactivate": len(deactivate)}

    # Settings: upsert by constant_name
    for s in changes.get("settings", []):
        execute_write(
            """INSERT OR REPLACE INTO settings (constant_name, value, description, imported_at)
               VALUES (?, ?, ?, ?)""",
            (s["constant_name"], s["value"], s["description"], datetime.utcnow().isoformat())
        )
    summary["settings"] = {"upsert": len(changes.get("settings", []))}

    # Master: process after all content tables are up to date
    _apply_master(changes.get("master_rows", []))
    summary["master"] = {"processed": len(changes.get("master_rows", []))}
    _apply_world_boss_loot()
    summary["world_boss_loot"] = {"processed": len(changes.get("world_boss_rows", []))}

    return summary


def _upsert_row(table: str, data: dict, existing_id: int | None):
    """Insert or update a content table row."""
    data["imported_at"] = datetime.utcnow().isoformat()
    if existing_id is None:
        cols   = ", ".join(data.keys())
        placeholders = ", ".join("?" * len(data))
        execute_write(
            f"INSERT OR IGNORE INTO {table} ({cols}) VALUES ({placeholders})",
            tuple(data.values())
        )
    else:
        sets = ", ".join(f"{k} = ?" for k in data)
        execute_write(
            f"UPDATE {table} SET {sets} WHERE id = ?",
            tuple(data.values()) + (existing_id,)
        )


def _apply_master(master_rows: list):
    """Process master sheet: upsert master rows, linking by name.
    Now includes protagonist FK columns."""
    execute_write("UPDATE master SET is_active=0")
    for r in master_rows:
        movie = _s(r.get('MovieName'))
        if not movie:
            continue

        def get_id(table, name):
            """Handle the get id workflow."""
            if not name:
                return None
            row = execute_one(f"SELECT id FROM {table} WHERE name = ?", (_s(name),))
            return row['id'] if row else None

        boss_id          = get_id('bosses',        r.get('BossName'))
        minion_id        = get_id('minions',       r.get('MinionName'))
        boss_weapon_id   = get_id('weapons',       r.get('BossWeapon'))
        boss_armor_id    = get_id('armor',         r.get('BossArmor'))
        boss_special_id  = get_id('special_items', r.get('BossSpecialItem'))
        min_weapon_id    = get_id('weapons',       r.get('MinionWeapon'))
        min_armor_id     = get_id('armor',         r.get('MinionArmor'))
        min_special_id   = get_id('special_items', r.get('MinionSpecialItem'))
        prot_name        = _s(r.get('ProtagonistName')) or None
        prot_description = _s(r.get('ProtagonistDescription')) or None
        prot_weapon_id   = get_id('weapons',       r.get('ProtagonistWeapon'))
        prot_armor_id    = get_id('armor',         r.get('ProtagonistArmor'))
        prot_special_id  = get_id('special_items', r.get('ProtagonistSpecialItem'))

        if not all([boss_id, minion_id, boss_weapon_id, boss_armor_id,
                    boss_special_id, min_weapon_id, min_armor_id, min_special_id]):
            logger.warning("Master row '%s': could not resolve all FK references, skipping", movie)
            continue

        now = datetime.utcnow().isoformat()
        existing = execute_one("SELECT id FROM master WHERE movie_name = ?", (movie,))
        if existing:
            execute_write(
                """UPDATE master SET
                   boss_id=?, boss_weapon_id=?, boss_armor_id=?, boss_special_item_id=?,
                   minion_id=?, minion_weapon_id=?, minion_armor_id=?, minion_special_item_id=?,
                   protagonist_name=?, protagonist_weapon_id=?, protagonist_armor_id=?,
                   protagonist_special_item_id=?, protagonist_description=?,
                   is_active=1, imported_at=?
                   WHERE movie_name=?""",
                (boss_id, boss_weapon_id, boss_armor_id, boss_special_id,
                 minion_id, min_weapon_id, min_armor_id, min_special_id,
                 prot_name, prot_weapon_id, prot_armor_id, prot_special_id,
                 prot_description,
                 now, movie)
            )
        else:
            execute_write(
                """INSERT INTO master
                   (movie_name, boss_id, boss_weapon_id, boss_armor_id, boss_special_item_id,
                    minion_id, minion_weapon_id, minion_armor_id, minion_special_item_id,
                    protagonist_name, protagonist_weapon_id, protagonist_armor_id,
                    protagonist_special_item_id, protagonist_description, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (movie, boss_id, boss_weapon_id, boss_armor_id, boss_special_id,
                 minion_id, min_weapon_id, min_armor_id, min_special_id,
                 prot_name, prot_weapon_id, prot_armor_id, prot_special_id,
                 prot_description, now)
            )


def _apply_world_boss_loot():
    """Link each imported world boss to exactly one associated reward per slot."""
    execute_write("DELETE FROM world_boss_loot")
    for boss in execute("SELECT id,name FROM world_bosses WHERE is_active=1"):
        weapon = execute_one(
            "SELECT id FROM weapons WHERE associated_to=? AND is_active=1",
            (f"{boss['name']} (WorldBoss)",)
        )
        armor = execute_one(
            "SELECT id FROM armor WHERE associated_to=? AND is_active=1",
            (f"{boss['name']} (WorldBoss)",)
        )
        special = execute_one(
            """SELECT id FROM special_items
               WHERE associated_to=? AND association_type='WorldBoss' AND is_active=1""",
            (boss["name"],)
        )
        if not all((weapon, armor, special)):
            logger.warning("World boss '%s' does not yet have all three linked rewards", boss["name"])
            continue
        execute_write(
            "INSERT INTO world_boss_loot(world_boss_id,weapon_id,armor_id,special_item_id) VALUES(?,?,?,?)",
            (boss["id"], weapon["id"], armor["id"], special["id"])
        )
def clear_stale_intel(changes: dict):
    """Clear boss_intel rows for any boss whose intel-sensitive columns changed."""
    for item in changes.get("bosses", {}).get("update", []):
        old = item["db_row"]
        new = item["new_data"]
        changed = any(
            str(old.get(col, "")) != str(new.get(col, ""))
            for col in INTEL_SENSITIVE_COLS
            if col in new
        )
        if changed:
            boss_id = old["id"]
            deleted = execute_write(
                "DELETE FROM boss_intel WHERE boss_id = ?", (boss_id,)
            )
            if deleted:
                logger.info("Cleared boss_intel for boss_id=%d (intel-sensitive columns changed)", boss_id)


# ─────────────────────────────────────────────────────────────────────────────
# AUTO-POPULATE ASSOCIATED_TO
# ─────────────────────────────────────────────────────────────────────────────

def auto_populate_associated_to():
    """Update weapons/armor associated_to field from master table.
    Format: 'MovieName (Boss)' or 'MovieName (Minion)'."""
    master_rows = execute("SELECT * FROM master")
    for m in master_rows:
        movie = execute_one("SELECT movie_name FROM master WHERE id = ?", (m["id"],))["movie_name"]
        for col, table in [
            ("boss_weapon_id",   "weapons"),
            ("boss_armor_id",    "armor"),
            ("minion_weapon_id", "weapons"),
            ("minion_armor_id",  "armor"),
        ]:
            item_id = m.get(col)
            if not item_id:
                continue
            side = "Boss" if "boss" in col else "Minion"
            execute_write(
                f"UPDATE {table} SET associated_to = ? WHERE id = ?",
                (f"{movie} ({side})", item_id)
            )


# ─────────────────────────────────────────────────────────────────────────────
# FILE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def _reject_import(filepath: str, reason: str):
    """Move the invalid file to the rejected folder and log the error."""
    os.makedirs(cfg.REJECTED_IMPORT_PATH, exist_ok=True)
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(cfg.REJECTED_IMPORT_PATH, f"rejected_{ts}.xlsx")
    try:
        shutil.move(filepath, dest)
    except Exception:
        pass
    os.makedirs(os.path.dirname(cfg.IMPORT_ERROR_LOG), exist_ok=True)
    with open(cfg.IMPORT_ERROR_LOG, "a") as f:
        f.write(f"{datetime.utcnow().isoformat()} | REJECTED | {reason}\n")
    logger.error("Import rejected: %s", reason)


def _archive_import(filepath: str):
    """Move successfully processed file to a timestamped archive."""
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(cfg.REJECTED_IMPORT_PATH, f"../applied_{ts}.xlsx")
    try:
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.move(filepath, dest)
    except Exception:
        pass


################################################################################
